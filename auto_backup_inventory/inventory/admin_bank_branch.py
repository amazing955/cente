import io
import os
import uuid
from datetime import datetime
from typing import List, Tuple

from django.contrib import admin, messages
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import path, reverse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .models import BankBranch, BranchImportLog


class BankBranchAdmin(admin.ModelAdmin):
    class Media:
        css = {
            'all': ('inventory/jazzmin_admin.css',),
        }

    list_display = ('branch_code', 'branch_name', 'region', 'district', 'status', 'created_at', 'updated_at')
    list_filter = ('status', 'region')
    search_fields = ('branch_code', 'branch_name')
    ordering = ('branch_name',)
    readonly_fields = ('created_at', 'updated_at')
    list_per_page = 25
    actions = ['export_selected_branches']
    change_list_template = 'admin/inventory/bankbranch/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.admin_site.admin_view(self.upload_excel_view), name='inventory_bankbranch_upload_excel'),
            path('preview-import/', self.admin_site.admin_view(self.preview_import_view), name='inventory_bankbranch_preview_import'),
            path('download-template/', self.admin_site.admin_view(self.download_template_view), name='inventory_bankbranch_download_template'),
            path('export-excel/', self.admin_site.admin_view(self.export_excel_view), name='inventory_bankbranch_export_excel'),
        ]
        return custom_urls + urls

    def get_list_display(self, request):
        return self.list_display

    def has_add_permission(self, request):
        return request.user.is_superuser

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser or request.user.has_perm('inventory.change_bankbranch')

    def has_view_permission(self, request, obj=None):
        return request.user.is_superuser or request.user.has_perm('inventory.view_bankbranch')

    def get_actions(self, request):
        actions = super().get_actions(request)
        if not request.user.is_superuser:
            actions.pop('delete_selected', None)
        return actions

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        return qs.filter()

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_upload_button'] = request.user.is_superuser
        extra_context['show_export_button'] = request.user.is_superuser
        return super().changelist_view(request, extra_context=extra_context)

    def export_selected_branches(self, request, queryset):
        if not request.user.is_superuser:
            messages.error(request, 'Only superusers can export selected branches.')
            return

        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Branches'
        headers = ['Branch Code', 'Branch Name', 'Region', 'District', 'Address', 'Status', 'Created At', 'Updated At']
        sheet.append(headers)
        for branch in queryset:
            sheet.append([
                branch.branch_code,
                branch.branch_name,
                branch.region,
                branch.district,
                branch.address,
                branch.status,
                branch.created_at,
                branch.updated_at,
            ])
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument/spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="selected_bank_branches.xlsx"'
        return response

    export_selected_branches.short_description = 'Export selected branches'

    def upload_excel_view(self, request):
        if not request.user.is_superuser:
            messages.error(request, 'Only superusers can upload Excel files.')
            return HttpResponseRedirect(reverse('admin:inventory_bankbranch_changelist'))

        if request.method == 'POST':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, 'Please select a file to upload.')
                return render(request, 'admin/inventory/bankbranch/upload_excel.html', {'title': 'Upload Branch Excel'})

            if not excel_file.name.lower().endswith('.xlsx'):
                messages.error(request, 'Only .xlsx files are supported.')
                return render(request, 'admin/inventory/bankbranch/upload_excel.html', {'title': 'Upload Branch Excel'})

            try:
                workbook = load_workbook(excel_file, read_only=True, data_only=True)
                sheet = workbook.active
                rows = list(sheet.iter_rows(values_only=True))
            except Exception as exc:
                messages.error(request, f'Unable to read Excel file: {exc}')
                return render(request, 'admin/inventory/bankbranch/upload_excel.html', {'title': 'Upload Branch Excel'})

            if not rows:
                messages.error(request, 'The uploaded Excel file is empty.')
                return render(request, 'admin/inventory/bankbranch/upload_excel.html', {'title': 'Upload Branch Excel'})

            preview = self._build_import_preview(rows, request.user)
            preview['filename'] = excel_file.name
            request.session['bank_branch_import_preview'] = preview
            messages.info(request, 'Preview loaded. Review the data below and click "Import" to confirm.')
            return HttpResponseRedirect(reverse('admin:inventory_bankbranch_preview_import'))

        return render(request, 'admin/inventory/bankbranch/upload_excel.html', {'title': 'Upload Branch Excel'})

    def preview_import_view(self, request):
        if not request.user.is_superuser:
            messages.error(request, 'Only superusers can import branches.')
            return HttpResponseRedirect(reverse('admin:inventory_bankbranch_changelist'))

        preview = request.session.get('bank_branch_import_preview')
        if not preview:
            messages.error(request, 'No import preview is available. Please upload an Excel file again.')
            return HttpResponseRedirect(reverse('admin:inventory_bankbranch_changelist'))

        if request.method == 'POST' and request.POST.get('confirm_import') == '1':
            result = self._apply_import(preview, request.user)
            request.session.pop('bank_branch_import_preview', None)
            messages.success(request, result['message'])
            return HttpResponseRedirect(reverse('admin:inventory_bankbranch_changelist'))

        return render(request, 'admin/inventory/bankbranch/import_preview.html', {
            'title': 'Import Preview',
            'preview': preview,
            'opts': self.model._meta,
        })

    def download_template_view(self, request):
        if not request.user.is_superuser:
            messages.error(request, 'Only superusers can download the template.')
            return HttpResponseRedirect(reverse('admin:inventory_bankbranch_changelist'))

        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Branches'
        headers = ['Branch Code', 'Branch Name', 'Region', 'District', 'Address', 'Status']
        sheet.append(headers)
        header_font = Font(bold=True)
        for cell in sheet[1]:
            cell.font = header_font
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="branch_import_template.xlsx"'
        return response

    def export_excel_view(self, request):
        if not request.user.is_superuser:
            messages.error(request, 'Only superusers can export branches.')
            return HttpResponseRedirect(reverse('admin:inventory_bankbranch_changelist'))

        queryset = self.get_queryset(request)
        output = io.BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Branches'
        headers = ['Branch Code', 'Branch Name', 'Region', 'District', 'Address', 'Status', 'Created At', 'Updated At']
        sheet.append(headers)
        for branch in queryset:
            sheet.append([
                branch.branch_code,
                branch.branch_name,
                branch.region,
                branch.district,
                branch.address,
                branch.status,
                branch.created_at,
                branch.updated_at,
            ])
        workbook.save(output)
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="bank_branches.xlsx"'
        return response

    def _build_import_preview(self, rows, user) -> dict:
        normalized_rows = self._normalize_rows(rows)
        preview = {
            'total_rows': len(normalized_rows),
            'new_branches': 0,
            'updated_branches': 0,
            'duplicate_rows': 0,
            'invalid_rows': 0,
            'rows_to_import': 0,
            'errors': [],
            'rows': [],
        }

        seen_codes = set()
        for row in normalized_rows:
            branch_code = row.get('branch_code', '')
            branch_name = row.get('branch_name', '')
            status = row.get('status', '')
            if not branch_code or not branch_name:
                preview['invalid_rows'] += 1
                preview['errors'].append(f"Row {row['row_number']}: Branch Code and Branch Name are required.")
                continue
            if branch_code in seen_codes:
                preview['duplicate_rows'] += 1
                preview['errors'].append(f"Row {row['row_number']}: Duplicate Branch Code {branch_code}.")
                continue
            seen_codes.add(branch_code)
            existing = BankBranch.objects.filter(branch_code=branch_code).first()
            if existing:
                preview['updated_branches'] += 1
            else:
                preview['new_branches'] += 1
            preview['rows'].append({
                'row_number': row['row_number'],
                'branch_code': branch_code,
                'branch_name': branch_name,
                'region': row.get('region', ''),
                'district': row.get('district', ''),
                'address': row.get('address', ''),
                'status': status or 'Active',
                'exists': existing is not None,
            })

        preview['rows_to_import'] = len(preview['rows'])
        return preview

    def _normalize_rows(self, rows) -> List[dict]:
        if not rows:
            return []

        headers = [self._normalize_header(cell) for cell in rows[0]]
        normalized_rows = []
        for index, row in enumerate(rows[1:], start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            data = {}
            for header_idx, header in enumerate(headers):
                if header_idx >= len(row):
                    break
                value = row[header_idx]
                if value is None:
                    value = ''
                else:
                    value = str(value).strip()
                data[header] = value
            data['row_number'] = index
            normalized_rows.append(data)
        return normalized_rows

    def _normalize_header(self, value) -> str:
        if value is None:
            return ''
        normalized = str(value).strip().lower()
        normalized = normalized.replace(' ', '_').replace('-', '_')
        return normalized

    def _apply_import(self, preview, user) -> dict:
        created = 0
        updated = 0
        skipped = 0
        failed = 0
        for row in preview['rows']:
            try:
                branch, created_flag = BankBranch.objects.update_or_create(
                    branch_code=row['branch_code'],
                    defaults={
                        'branch_name': row['branch_name'],
                        'region': row.get('region', ''),
                        'district': row.get('district', ''),
                        'address': row.get('address', ''),
                        'status': row.get('status', 'Active') or 'Active',
                    },
                )
                if created_flag:
                    created += 1
                else:
                    updated += 1
            except Exception:
                failed += 1
                skipped += 1

        log = BranchImportLog.objects.create(
            filename=preview.get('filename', 'imported_from_admin.xlsx'),
            uploaded_by=user,
            total_rows=preview['total_rows'],
            created_records=created,
            updated_records=updated,
            skipped_records=skipped,
            failed_records=failed,
            import_status='success' if failed == 0 else 'warning',
        )
        return {
            'message': f'Import completed successfully. {preview["total_rows"]} rows processed. {created} branches created. {updated} branches updated. {skipped} skipped. {failed} failed.',
        }


admin.site.register(BankBranch, BankBranchAdmin)
admin.site.register(BranchImportLog)

import json
import uuid
from datetime import date
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType
from django.core import mail, signing
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.http import HttpResponseForbidden
from django.test import RequestFactory, TestCase, override_settings
from django.template.loader import render_to_string
from django.urls import path, reverse
from django.utils import timezone
from openpyxl import Workbook

from .admin import CustomUserAdminForm
from .forms import BackupShipmentAssignmentForm, CustomUserCreationForm, RoleCreationForm
from .models import (
    ApplicationSetting,
    AuditLog,
    BankBranch,
    CourierProfile,
    DashboardFeatureExemption,
    DashboardFeaturePermission,
    DeliveryConfirmation,
    ExceptionCloseRequest,
    PendingApproval,
    Reconciliation,
    Shipment,
    ShipmentApprovalHistory,
    ShipmentReceipt,
    ShipmentTransportEvent,
    ShipmentException,
    Tape,
    TapeRequest,
    SchemaChangeLog,
    get_dashboard_feature_catalog,
)
from .serializer import TapeSerializer
from .views import custom_permission_denied, is_backup_administrator, is_dr_team, is_operations_manager


def forbidden_view(request):
    return HttpResponseForbidden('Forbidden')


urlpatterns = [
    path('protected-page/', forbidden_view),
]


class DrTeamAccessTests(TestCase):
    def test_is_dr_team_matches_case_insensitive_group_names(self):
        user = get_user_model().objects.create_user(
            username='dr-lowercase-user',
            email='dr-lowercase-user@example.com',
            password='StrongPass123!',
            role='user',
        )
        Group.objects.create(name='dr team')
        user.groups.add(Group.objects.get(name='dr team'))

        self.assertTrue(is_dr_team(user))


class PasswordResetFlowTests(TestCase):
    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_password_reset_requires_matching_username_and_email(self):
        user = get_user_model().objects.create_user(
            username='reset-user',
            email='reset-user@example.com',
            password='StrongPass123!',
        )

        response = self.client.post(reverse('password_reset'), {
            'username': 'reset-user',
            'email': 'different@example.com',
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'does not match the account for that username')
        self.assertEqual(len(mail.outbox), 0)

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_password_reset_sends_link_for_matching_username_and_email(self):
        get_user_model().objects.create_user(
            username='reset-user',
            email='reset-user@example.com',
            password='StrongPass123!',
        )

        response = self.client.post(reverse('password_reset'), {
            'username': 'reset-user',
            'email': 'reset-user@example.com',
        })

        self.assertRedirects(response, reverse('password_reset_done'))
        self.assertEqual(len(mail.outbox), 1)


class SupremeApproverDashboardTemplateTests(TestCase):
    def test_dashboard_template_renders_when_current_user_or_audit_user_is_missing(self):
        request = RequestFactory().get('/supreme-approver-dashboard/')
        context = {
            'dashboard_tabs': [],
            'shipments': [],
            'approval_queue': [],
            'pending_count': 0,
            'total_pending': 0,
            'approved_today': 0,
            'rejected_today': 0,
            'high_risk_transactions': [],
            'pending_tape_requests': [],
            'pending_reconciliations': [],
            'recent_audit': [
                SimpleNamespace(user=None, timestamp=timezone.now(), action='Login', module='Auth', severity='info'),
            ],
            'pending_user_changes': [],
            'pending_warehouse_release': 0,
            'pending_bulk_imports': 0,
            'pending_barcode_changes': 0,
            'pending_user_change_requests': 0,
            'pending_branch_updates': 0,
            'current_user': None,
            'current_datetime': timezone.now(),
            'last_login': timezone.now(),
            'approval_sla': '24 Hours',
            'notifications': [],
            'recent_activity': [],
            'high_risk_types': [],
            'dashboard_title': 'Enterprise Banking Approval Center',
            'system_version': 'v2.4.1',
            'database_name': 'backup_inventory',
            'server_status': 'Operational',
            'redis_status': 'Healthy',
            'websocket_status': 'Connected',
        }

        rendered = render_to_string('supreme_approver_dashboard.html', context, request=request)

        self.assertIn('Welcome back', rendered)
        self.assertIn('System', rendered)


class LoginAttemptLimitTests(TestCase):
    def test_invalid_login_attempts_lock_account_after_three_failures(self):
        user = get_user_model().objects.create_user(
            username='valid-user',
            email='valid-user@example.com',
            password='StrongPass123!',
        )

        for index in range(3):
            response = self.client.post(reverse('signin'), {'username': 'valid-user', 'password': 'wrong-pass'})
            self.assertEqual(response.status_code, 200)
            if index < 2:
                self.assertContains(response, 'Remaining attempts')
            else:
                self.assertContains(response, 'ACCOUNT TEMPORARILY LOCKED')

        user.refresh_from_db()
        self.assertGreaterEqual(user.failed_login_attempts, 3)
        self.assertIsNotNone(user.account_locked_until)

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_initial_login_logs_otp_to_terminal(self):
        user = get_user_model().objects.create_user(
            username='terminal-otp-user',
            email='terminal-otp-user@example.com',
            password='StrongPass123!',
        )

        with patch('builtins.print') as mock_print:
            response = self.client.post(reverse('signin'), {'username': 'terminal-otp-user', 'password': 'StrongPass123!'})

        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            any('2FA OTP' in str(call.args[0]) and user.username in str(call.args[0]) for call in mock_print.call_args_list)
        )

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_otp_resend_sends_new_code_and_restricts_repeated_requests(self):
        user = get_user_model().objects.create_user(
            username='otp-resend-user',
            email='otp-resend-user@example.com',
            password='StrongPass123!',
        )

        response = self.client.post(reverse('signin'), {'username': 'otp-resend-user', 'password': 'StrongPass123!'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Verification Code')

        first_otp = self.client.session['pending_2fa_otp']

        response = self.client.post(reverse('signin'), {'action': 'resend_otp'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'A new verification code has been sent')
        self.assertNotEqual(self.client.session['pending_2fa_otp'], first_otp)
        self.assertEqual(len(mail.outbox), 2)

        for _ in range(4):
            response = self.client.post(reverse('signin'), {'action': 'resend_otp'})
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, 'A new verification code has been sent')

        response = self.client.post(reverse('signin'), {'action': 'resend_otp'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'You have exceeded the maximum OTP resend attempts')


class DualApprovalWorkflowTests(TestCase):
    def test_warehouse_role_redirects_to_warehouse_dashboard_after_signin(self):
        group = Group.objects.create(name='Warehouse Ops')
        user = get_user_model().objects.create_user(
            username='warehouse-login-user',
            email='warehouse-login-user@example.com',
            password='StrongPass123!',
        )
        user.groups.add(group)

        response = self.client.post(reverse('signin'), {
            'username': 'warehouse-login-user',
            'password': 'StrongPass123!',
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Verification Code')
        otp_code = self.client.session['pending_2fa_otp']

        response = self.client.post(reverse('signin'), {'otp_code': otp_code})

        self.assertRedirects(response, reverse('warehouse-operations-dashboard'))

    def test_backup_admin_approval_requires_supreme_approver_for_final_release(self):
        backup_group = Group.objects.create(name='Backup Administrator')
        supreme_group = Group.objects.create(name='Supreme Approver')
        backup_admin = get_user_model().objects.create_user(
            username='backup-stage-user',
            email='backup-stage-user@example.com',
            password='StrongPass123!',
        )
        supreme_approver = get_user_model().objects.create_user(
            username='supreme-stage-user',
            email='supreme-stage-user@example.com',
            password='StrongPass123!',
        )
        backup_admin.groups.add(backup_group)
        supreme_approver.groups.add(supreme_group)

        tape = Tape.objects.create(
            volser='TAPE-DUAL-001',
            barcode='BAR-DUAL-001',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
        )
        courier_profile = CourierProfile.objects.create(
            courier_id='CR-DUAL-01',
            full_name='Courier One',
            vehicle_number='V-1001',
        )
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='HQ',
            destination_location='DR',
            priority_level='High',
            created_by=backup_admin,
            status='Pending',
        )

        self.client.force_login(backup_admin)
        response = self.client.post(reverse('shipment-approvals'), {
            'form_type': 'backup_admin_decision',
            'shipment_id': shipment.pk,
            'tape_id': tape.pk,
            'courier_id': courier_profile.pk,
            'decision': 'approve',
            'comments': 'First-stage approval',
        })

        shipment.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(shipment.approval_stage, 'awaiting_supreme')
        self.assertEqual(shipment.status, 'Pending')
        self.assertEqual(shipment.approved_by_backup, backup_admin)
        self.assertIsNone(shipment.approved_by_supreme)

        self.client.force_login(supreme_approver)
        response = self.client.post(reverse('shipment-approvals'), {
            'form_type': 'backup_admin_decision',
            'shipment_id': shipment.pk,
            'decision': 'approve',
            'comments': 'Final-stage approval',
        })

        shipment.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(shipment.approval_stage, 'approved')
        self.assertEqual(shipment.status, 'Approved')
        self.assertEqual(shipment.approved_by_supreme, supreme_approver)

    def test_warehouse_operations_dashboard_renders_for_warehouse_roles(self):
        group = Group.objects.create(name='Warehouse Ops')
        user = get_user_model().objects.create_user(
            username='warehouse-role-user',
            email='warehouse-role-user@example.com',
            password='StrongPass123!',
        )
        user.groups.add(group)

        self.client.force_login(user)
        response = self.client.get(reverse('warehouse-operations-dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Warehouse Operations Dashboard')


class SupremeApproverDashboardTests(TestCase):
    def test_supreme_approver_can_access_dashboard_and_view_pending_shipments(self):
        supreme_group = Group.objects.create(name='Supreme Approver')
        backup_group = Group.objects.create(name='Backup Administrator')
        supreme_user = get_user_model().objects.create_user(
            username='supreme-test-user',
            email='supreme-test-user@example.com',
            password='StrongPass123!',
        )
        supreme_user.groups.add(supreme_group)
        backup_user = get_user_model().objects.create_user(
            username='backup-test-user',
            email='backup-test-user@example.com',
            password='StrongPass123!',
        )
        backup_user.groups.add(backup_group)
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Pending',
            approval_stage='awaiting_supreme',
            approved_by_backup=backup_user,
            created_by=backup_user,
            last_updated_by=backup_user,
            approval_remarks='Pending supreme approval.',
        )

        self.client.force_login(supreme_user)
        response = self.client.get(reverse('supreme-approver-dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Pending Approvals')
        self.assertContains(response, shipment.shipment_id)

    def test_supreme_approver_review_page_is_read_only_and_approves_pending_request(self):
        supreme_group = Group.objects.create(name='Supreme Approver')
        backup_group = Group.objects.create(name='Backup Administrator')
        supreme_user = get_user_model().objects.create_user(
            username='supreme-review-user',
            email='supreme-review-user@example.com',
            password='StrongPass123!',
        )
        supreme_user.groups.add(supreme_group)
        backup_user = get_user_model().objects.create_user(
            username='backup-review-user',
            email='backup-review-user@example.com',
            password='StrongPass123!',
        )
        backup_user.groups.add(backup_group)
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Approved',
            approval_stage='awaiting_supreme',
            approved_by_backup=backup_user,
            created_by=backup_user,
            last_updated_by=backup_user,
            approval_remarks='Ready for supreme review.',
        )
        approval_request = PendingApproval.objects.create(
            transaction_type='Shipment',
            module='Shipment Workflow',
            summary='Release Shipment for secure transfer.',
            requester=backup_user,
            backup_administrator=backup_user,
            branch=backup_user.assigned_branch,
            status='Awaiting Supreme Approval',
            risk_level='High',
            related_object_id=shipment.pk,
            related_model='shipment',
            request_payload={'shipment_id': shipment.shipment_id},
        )

        self.client.force_login(supreme_user)
        response = self.client.get(reverse('approval-review', kwargs={'approval_id': approval_request.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Approval Review')
        self.assertNotContains(response, 'Create')
        self.assertNotContains(response, 'Edit')

        approval_response = self.client.post(
            reverse('approval-review', kwargs={'approval_id': approval_request.pk}),
            {'decision': 'approve', 'comment': 'Approved by supreme approver.'},
        )

        approval_request.refresh_from_db()
        shipment.refresh_from_db()
        self.assertEqual(approval_request.status, 'Approved')
        self.assertEqual(shipment.approval_stage, 'approved')
        self.assertEqual(shipment.status, 'Approved')
        self.assertEqual(approval_response.status_code, 302)

    def test_non_supreme_user_receives_forbidden(self):
        user = get_user_model().objects.create_user(
            username='regular-user',
            email='regular-user@example.com',
            password='StrongPass123!',
        )

        self.client.force_login(user)
        response = self.client.get(reverse('supreme-approver-dashboard'))

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, f"{reverse('signin')}?next=/supreme-approver-dashboard/")


class BackupDashboardSignedNavigationTests(TestCase):
    def test_operations_dashboard_renders_when_notifications_exist(self):
        operations_group = Group.objects.create(name='Operations Manager')
        operations_user = get_user_model().objects.create_user(
            username='operations-notifications-user',
            email='operations-notifications-user@example.com',
            password='StrongPass123!',
        )
        operations_user.groups.add(operations_group)
        DashboardFeaturePermission.objects.create(role=operations_group, feature_key='operations_dashboard', can_view=True)
        AuditLog.objects.create(
            name='Exception Alert Forwarded',
            action='Forwarded exception alert',
            message='exception_id=EX-100|description=Needs review',
            user=operations_user,
            severity='warning',
        )
        self.client.force_login(operations_user)

        response = self.client.get(reverse('operations-dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Operations Manager Dashboard')

    def test_operations_dashboard_supports_signed_navigation_for_history_panel(self):
        operations_group = Group.objects.create(name='Operations Manager')
        operations_user = get_user_model().objects.create_user(
            username='operations-signed-nav',
            email='operations-signed-nav@example.com',
            password='StrongPass123!',
        )
        operations_user.groups.add(operations_group)
        DashboardFeaturePermission.objects.create(role=operations_group, feature_key='operations_dashboard', can_view=True)
        self.client.force_login(operations_user)

        signed_token = signing.dumps(
            {'feature': 'operations_dashboard', 'params': {'show_admin': '1', 'show_admin_history': '1'}, 'dashboard': 'operations'},
            salt='inventory-dashboard-navigation',
            compress=True,
        )

        response = self.client.get(reverse('operations-dashboard-navigation', kwargs={'signed_token': signed_token}))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['show_admin_history'])
        self.assertEqual(response.context['active_feature_key'], 'operations_dashboard')

    def test_backup_admin_can_access_signed_navigation_and_feature_module(self):
        backup_group = Group.objects.create(name='Backup Administrator')
        backup_admin = get_user_model().objects.create_user(
            username='backup-admin-access-check',
            email='backup-admin-access-check@example.com',
            password='StrongPass123!',
        )
        backup_admin.groups.add(backup_group)
        self.client.force_login(backup_admin)

        signed_token = signing.dumps(
            {'feature': 'tape_inventory', 'params': {'show_tape_inventory': '1'}},
            salt='inventory-dashboard-navigation',
            compress=True,
        )

        nav_response = self.client.get(reverse('backup-dashboard-navigation', kwargs={'signed_token': signed_token}))
        feature_response = self.client.get(reverse('feature-module', kwargs={'feature_key': 'tape_inventory'}))

        self.assertEqual(nav_response.status_code, 200)
        self.assertEqual(feature_response.status_code, 200)
        self.assertEqual(nav_response.context['active_feature_key'], 'tape_inventory')
        self.assertEqual(feature_response.context['active_feature_key'], 'tape_inventory')

    def test_signed_navigation_redirects_to_dashboard_panel(self):
        backup_admin = get_user_model().objects.create_superuser(
            username='backup-admin-signed-nav',
            email='backup-admin-signed-nav@example.com',
            password='StrongPass123!',
        )
        self.client.force_login(backup_admin)

        signed_token = signing.dumps(
            {'feature': 'tape_inventory', 'params': {'show_tape_inventory': '1'}},
            salt='inventory-dashboard-navigation',
            compress=True,
        )

        response = self.client.get(reverse('backup-dashboard-navigation', kwargs={'signed_token': signed_token}))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['active_feature_key'], 'tape_inventory')
        self.assertTrue(response.context['show_tape_inventory_panel'])

    def test_tampered_signed_navigation_returns_bad_request(self):
        backup_admin = get_user_model().objects.create_superuser(
            username='backup-admin-tampered-nav',
            email='backup-admin-tampered-nav@example.com',
            password='StrongPass123!',
        )
        self.client.force_login(backup_admin)

        signed_token = signing.dumps(
            {'feature': 'tape_inventory', 'params': {'show_tape_inventory': '1'}},
            salt='inventory-dashboard-navigation',
            compress=True,
        )
        tampered_token = signed_token[:-1] + ('A' if signed_token[-1] != 'A' else 'B')

        response = self.client.get(reverse('backup-dashboard-navigation', kwargs={'signed_token': tampered_token}))

        self.assertEqual(response.status_code, 400)


class ReconciliationInitiationWorkflowTests(TestCase):
    def test_shift_reconciliation_allows_multiple_operator_selection(self):
        backup_admin = get_user_model().objects.create_superuser(
            username='backup-admin-shift-multi',
            email='backup-admin-shift-multi@example.com',
            password='StrongPass123!',
        )
        first_operator = get_user_model().objects.create_user(
            username='operator-one',
            email='operator-one@example.com',
            password='StrongPass123!',
            role='operations_manager',
        )
        second_operator = get_user_model().objects.create_user(
            username='operator-two',
            email='operator-two@example.com',
            password='StrongPass123!',
            role='operations_manager',
        )
        reconciliation = Reconciliation.objects.create(
            location='Head Office',
            performed_by=backup_admin,
            status='Open',
        )

        self.client.force_login(backup_admin)
        response = self.client.post(reverse('backup-dashboard'), {
            'form_type': 'shift_reconciliation',
            'reconciliation_pk': str(reconciliation.id),
            'operator_pks': [str(first_operator.id), str(second_operator.id)],
        })

        self.assertEqual(response.status_code, 302)
        reconciliation.refresh_from_db()
        self.assertEqual(reconciliation.assigned_operator, first_operator)
        self.assertTrue(
            AuditLog.objects.filter(
                name='Reconciliation Shifted',
                action__icontains=first_operator.username,
            ).exists()
        )

    def test_dr_team_can_request_reconciliation_from_investigation_dashboard(self):
        branch = BankBranch.objects.create(branch_code='KLA-010', branch_name='Kampala North Branch', status='Active')
        dr_group = Group.objects.create(name='DR Team')
        user = get_user_model().objects.create_user(
            username='dr-requester',
            email='dr-requester@example.com',
            password='StrongPass123!',
            role='auditor',
            assigned_branch=branch,
        )
        user.groups.add(dr_group)

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Approved',
            source_location='Kampala North Branch',
            destination_location='Mbarara Branch',
            requesting_branch=branch,
            created_by=user,
        )
        exception = ShipmentException.objects.create(
            shipment=shipment,
            exception_type='Missing Tape',
            description='Tape missing during handover.',
            reported_by=user,
            severity='High',
            status='Open',
        )

        self.client.force_login(user)
        response = self.client.post(reverse('initiate-reconciliation-request'), {
            'exception_id': exception.exception_id,
            'requester_name': user.get_full_name() or user.username,
            'comment': 'Please open a reconciliation for review.',
        })

        self.assertEqual(response.status_code, 302)
        self.assertTrue(AuditLog.objects.filter(name='Reconciliation Requested').exists())
        alert = AuditLog.objects.filter(name='Reconciliation Requested').latest('timestamp')
        self.assertEqual(alert.user, user)
        self.assertIn(exception.exception_id, alert.message or alert.action)

    def test_backup_admin_alert_link_can_open_reconciliation_request_context(self):
        backup_admin = get_user_model().objects.create_superuser(
            username='backup-admin-alert',
            email='backup-admin-alert@example.com',
            password='StrongPass123!',
        )
        requester = get_user_model().objects.create_user(
            username='dr-alert-requester',
            email='dr-alert-requester@example.com',
            password='StrongPass123!',
        )
        alert = AuditLog.objects.create(
            name='Reconciliation Requested',
            action='Reconciliation requested for exception EXC-TEST-001',
            message='exception_id=EXC-TEST-001|requester_name=DR Team|comment=Please open a reconciliation for review.',
            user=requester,
            severity='warning',
        )

        self.client.force_login(backup_admin)
        response = self.client.get(reverse('backup-dashboard'), {
            'show_alerts': '1',
            'show_reconciliation': '1',
            'show_add_reconciliation': '1',
            'reconciliation_request_id': alert.id,
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Reconciliation request from DR Team')
        self.assertContains(response, 'Please open a reconciliation for review.')

    def test_backup_admin_can_forward_exception_alert_to_dr_team(self):
        dr_group = Group.objects.create(name='DR Team')
        backup_admin = get_user_model().objects.create_superuser(
            username='backup-admin-forward',
            email='backup-admin-forward@example.com',
            password='StrongPass123!',
        )
        dr_user = get_user_model().objects.create_user(
            username='dr-forward-user',
            email='dr-forward-user@example.com',
            password='StrongPass123!',
            role='auditor',
        )
        dr_user.groups.add(dr_group)

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Approved',
            source_location='Kampala North Branch',
            destination_location='Mbarara Branch',
            requesting_branch=BankBranch.objects.create(branch_code='KLA-001', branch_name='Kampala North Branch', status='Active'),
            created_by=backup_admin,
        )
        ShipmentException.objects.create(
            shipment=shipment,
            exception_id='EXC-TEST-002',
            exception_type='Missing Tape',
            description='Tape missing during handover.',
            reported_by=backup_admin,
            severity='High',
            status='Open',
        )

        alert = AuditLog.objects.create(
            name='Shipment Exception Reported',
            action='Reported exception EXC-TEST-002 for shipment SHP-0001',
            message='exception_id=EXC-TEST-002|description=Tape missing during handover.',
            user=backup_admin,
            severity='warning',
        )

        self.client.force_login(backup_admin)
        response = self.client.post(reverse('backup-dashboard'), {
            'form_type': 'forward_exception_to_dr_team',
            'alert_id': str(alert.id),
        })

        self.assertRedirects(response, reverse('backup-dashboard') + '?show_alerts=1')
        exception = ShipmentException.objects.get(exception_id='EXC-TEST-002')
        self.assertEqual(exception.status, 'Investigating')
        self.assertTrue(
            AuditLog.objects.filter(
                name='Exception Alert Forwarded',
                user=dr_user,
                message__contains='exception_id=EXC-TEST-002'
            ).exists()
        )

    def test_dr_dashboard_shows_forwarded_exception_notification_without_exception_id(self):
        dr_group = Group.objects.create(name='DR Team')
        dr_user = get_user_model().objects.create_user(
            username='dr-forward-display',
            email='dr-forward-display@example.com',
            password='StrongPass123!',
            role='auditor',
        )
        dr_user.groups.add(dr_group)

        AuditLog.objects.create(
            name='Exception Alert Forwarded',
            action='Forwarded exception alert for review',
            message='description=Pending exception details|target_url=/investigation-dashboard/',
            user=dr_user,
            severity='warning',
        )

        self.client.force_login(dr_user)
        response = self.client.get(reverse('investigation-dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Forwarded Exception Notifications')
        self.assertContains(response, 'Exception ID not yet available')

    def test_closed_exceptions_do_not_appear_in_dr_notifications(self):
        dr_group = Group.objects.create(name='DR Team')
        dr_user = get_user_model().objects.create_user(
            username='dr-closed-notify',
            email='dr-closed-notify@example.com',
            password='StrongPass123!',
            role='auditor',
        )
        dr_user.groups.add(dr_group)

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Approved',
            source_location='Kampala North Branch',
            destination_location='Mbarara Branch',
            requesting_branch=BankBranch.objects.create(branch_code='KLA-002', branch_name='Kampala South Branch', status='Active'),
            created_by=dr_user,
        )
        exception = ShipmentException.objects.create(
            shipment=shipment,
            exception_id='EXC-CLOSED-001',
            exception_type='Missing Tape',
            description='Closed after investigation.',
            reported_by=dr_user,
            severity='High',
            status='Closed',
        )
        AuditLog.objects.create(
            name='Exception Alert Forwarded',
            action='Forwarded closed exception alert for review',
            message=f'exception_id={exception.exception_id}|description=Closed exception',
            user=dr_user,
            severity='warning',
        )

        self.client.force_login(dr_user)
        response = self.client.get(reverse('investigation-dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'Forwarded closed exception alert for review')

    def test_dr_team_can_close_exception_with_investigation_results(self):
        # Setup DR Team group and user
        dr_group = Group.objects.create(name='DR Team')
        branch = BankBranch.objects.create(branch_code='KMP-001', branch_name='Kampala', status='Active')
        dr_user = get_user_model().objects.create_user(
            username='dr-investigator',
            email='dr-investigator@example.com',
            password='pass1234',
            role='auditor',
            assigned_branch=branch,
        )
        dr_user.groups.add(dr_group)

        # Create test data
        tape = Tape.objects.create(
            volser='TAPE-CLOSE-001',
            barcode='BC-CLOSE-001',
            tape_type='LTO-8',
            retention_end_date=date(2028, 1, 1),
            current_location='Vault',
        )
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Approved',
            source_location='Kampala',
            destination_location='Mbarara',
            requesting_branch=branch,
            releasing_custodian='Jane',
            receiving_custodian='John',
            created_by=dr_user,
        )
        shipment.tapes.add(tape)
        exception = ShipmentException.objects.create(
            exception_id='EXC-CLOSE-001',
            shipment=shipment,
            tape=tape,
            exception_type='Discrepancy',
            description='Test discrepancy',
            reported_by=dr_user,
            status='Investigating',
        )

        # Request to close the exception
        self.client.force_login(dr_user)
        response = self.client.post(reverse('close-exception'), {
            'exception_id': 'EXC-CLOSE-001',
            'investigator_name': 'DR Investigator',
            'investigation_results': 'Investigation completed. Tape found in vault.',
        })

        # Verify redirect and success message
        self.assertEqual(response.status_code, 302)
        self.assertIn('investigation-dashboard', response.url)

        # Verify exception status is still 'Investigating' (not yet closed)
        exception.refresh_from_db()
        self.assertEqual(exception.status, 'Investigating')

        # Verify ExceptionCloseRequest was created with Pending status
        close_request = ExceptionCloseRequest.objects.filter(exception=exception).first()
        self.assertIsNotNone(close_request)
        self.assertEqual(close_request.status, 'Pending')
        self.assertEqual(close_request.requested_by, dr_user)
        self.assertIn('Tape found in vault', close_request.investigation_results)

        # Verify AuditLog entry was created for the request
        audit_log = AuditLog.objects.filter(
            name='Exception Close Requested',
            action__icontains='EXC-CLOSE-001'
        ).first()
        self.assertIsNotNone(audit_log)
        self.assertEqual(audit_log.severity, 'warning')

    def test_admin_can_approve_close_exception_request(self):
        # Setup users and groups
        dr_group = Group.objects.create(name='DR Team')
        admin_group = Group.objects.create(name='Backup Administrator')
        
        branch = BankBranch.objects.create(branch_code='KMP-002', branch_name='Kampala', status='Active')
        
        dr_user = get_user_model().objects.create_user(
            username='dr-user-2',
            email='dr-user-2@example.com',
            password='pass1234',
            role='auditor',
            assigned_branch=branch,
        )
        dr_user.groups.add(dr_group)
        
        admin_user = get_user_model().objects.create_user(
            username='admin-approver',
            email='admin-approver@example.com',
            password='pass1234',
            role='admin',
            assigned_branch=branch,
        )
        admin_user.groups.add(admin_group)

        # Create exception and close request
        tape = Tape.objects.create(
            volser='TAPE-APPROVE-001',
            barcode='BC-APPROVE-001',
            tape_type='LTO-8',
            retention_end_date=date(2028, 1, 1),
            current_location='Vault',
        )
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Approved',
            source_location='Kampala',
            destination_location='Mbarara',
            requesting_branch=branch,
            releasing_custodian='Jane',
            receiving_custodian='John',
            created_by=dr_user,
        )
        shipment.tapes.add(tape)
        exception = ShipmentException.objects.create(
            exception_id='EXC-APPROVE-001',
            shipment=shipment,
            tape=tape,
            exception_type='Discrepancy',
            description='Test discrepancy',
            reported_by=dr_user,
            status='Investigating',
        )
        close_request = ExceptionCloseRequest.objects.create(
            exception=exception,
            requested_by=dr_user,
            investigation_results='Tape found in vault.',
            status='Pending'
        )

        # Admin approves the close request
        self.client.force_login(admin_user)
        response = self.client.post(reverse('approve-close-exception', kwargs={'close_request_id': close_request.id}), {
            'action': 'approve',
            'approval_comment': 'Investigation findings verified.',
        })

        # Verify redirect
        self.assertEqual(response.status_code, 302)
        self.assertIn('backup-dashboard', response.url)

        # Verify close request status is now 'Approved'
        close_request.refresh_from_db()
        self.assertEqual(close_request.status, 'Approved')
        self.assertEqual(close_request.approved_by, admin_user)
        self.assertIn('verified', close_request.approval_comment)

        # Verify exception status is now 'Closed'
        exception.refresh_from_db()
        self.assertEqual(exception.status, 'Closed')

        # Verify success AuditLog entry was created
        audit_log = AuditLog.objects.filter(
            name='Exception Closed',
            action__icontains='EXC-APPROVE-001'
        ).first()
        self.assertIsNotNone(audit_log)
        self.assertEqual(audit_log.severity, 'success')




class DjangoAdminCourierUserCreationTests(TestCase):
    def test_admin_add_user_page_renders_without_server_error(self):
        superuser = get_user_model().objects.create_superuser(
            username='admin-test-user',
            email='admin-test-user@example.com',
            password='StrongPass123!',
        )
        self.client.force_login(superuser)

        response = self.client.get(reverse('admin:inventory_customuser_add'))

        self.assertEqual(response.status_code, 200)

    def test_admin_form_exposes_vehicle_number_for_courier_users(self):
        courier_group = Group.objects.create(name='Courier')
        form = CustomUserAdminForm(data={
            'username': 'courier-admin-user',
            'email': 'courier-admin-user@example.com',
            'first_name': 'Courier',
            'last_name': 'User',
            'role': 'user',
            'groups': [courier_group.pk],
        })

        self.assertIn('vehicle_number', form.fields)
        self.assertFalse(form.is_valid())
        self.assertIn('__all__', form.errors)
        self.assertIn('Vehicle Number is required for courier accounts.', str(form.errors['__all__']))


class ExceptionInvestigationApiTests(TestCase):
    def test_open_exception_investigation_endpoint_returns_comprehensive_payload(self):
        branch = BankBranch.objects.create(branch_code='KLA-001', branch_name='Kampala Branch', status='Active')
        dr_group = Group.objects.create(name='DR Team')
        user = get_user_model().objects.create_user(
            username='investigator',
            email='investigator@example.com',
            password='pass1234',
            role='auditor',
            assigned_branch=branch,
        )
        user.groups.add(dr_group)
        tape = Tape.objects.create(
            volser='TAPE-INV-001',
            barcode='BC-INV-001',
            tape_type='LTO-8',
            retention_end_date=date(2028, 1, 1),
            current_location='Kampala Vault',
        )
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Approved',
            source_location='Kampala Branch',
            destination_location='Mbarara Branch',
            requesting_branch=branch,
            releasing_custodian='Jane Doe',
            receiving_custodian='John Doe',
            created_by=user,
        )
        shipment.tapes.add(tape)

        courier_user = get_user_model().objects.create_user(
            username='courier-investigator',
            email='courier-investigator@example.com',
            password='pass1234',
            role='user',
        )
        courier_profile = CourierProfile.objects.create(
            courier_id='CR-INV-001',
            full_name='Courier One',
            vehicle_number='V-100',
            user=courier_user,
            email='courier-investigator@example.com',
        )
        ShipmentTransportEvent.objects.create(
            shipment=shipment,
            courier=courier_profile,
            event_type='Picked Up',
            comments='Collected by courier',
        )
        ShipmentReceipt.objects.create(
            shipment=shipment,
            courier=courier_profile,
            pickup_location='Kampala Branch',
            custody_confirmed=True,
            manifest_verified=True,
            tape_count_matched=True,
            notes='Pickup received',
        )
        ShipmentApprovalHistory.objects.create(
            shipment=shipment,
            action='Approved',
            comments='Ready for dispatch',
            user=user,
        )
        exception = ShipmentException.objects.create(
            shipment=shipment,
            tape=tape,
            exception_type='Missing Tape',
            description='Tape missing during handover.',
            reported_by=user,
            severity='High',
            status='Open',
        )

        auth_response = self.client.post(
            reverse('token_obtain_pair'),
            {'username': user.username, 'password': 'pass1234'},
            content_type='application/json',
        )
        self.assertEqual(auth_response.status_code, 200)
        access_token = auth_response.json()['access']

        response = self.client.get(
            reverse('exception-investigation', kwargs={'exception_id': exception.exception_id}),
            HTTP_AUTHORIZATION=f'Bearer {access_token}',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['exception']['id'], exception.exception_id)
        self.assertEqual(payload['exception']['status'], 'Open')
        self.assertEqual(payload['tape']['volser'], tape.volser)
        self.assertEqual(payload['shipment']['shipment_id'], shipment.shipment_id)
        self.assertGreaterEqual(len(payload['timeline']), 1)
        self.assertGreaterEqual(len(payload['chain_of_custody']), 1)
        self.assertIn('audit_logs', payload)
        self.assertIn('notifications', payload)
        self.assertTrue(AuditLog.objects.filter(action__icontains=exception.exception_id).exists())

    def test_double_slash_investigation_url_is_supported(self):
        branch = BankBranch.objects.create(branch_code='KLA-002', branch_name='Kampala East Branch', status='Active')
        dr_group = Group.objects.create(name='DR Team')
        user = get_user_model().objects.create_user(
            username='investigator-double',
            email='investigator-double@example.com',
            password='pass1234',
            role='auditor',
            assigned_branch=branch,
        )
        user.groups.add(dr_group)
        tape = Tape.objects.create(
            volser='TAPE-INV-002',
            barcode='BC-INV-002',
            tape_type='LTO-8',
            retention_end_date=date(2028, 1, 1),
            current_location='Kampala Vault',
        )
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Approved',
            source_location='Kampala Branch',
            destination_location='Mbarara Branch',
            requesting_branch=branch,
            created_by=user,
        )
        shipment.tapes.add(tape)
        exception = ShipmentException.objects.create(
            shipment=shipment,
            tape=tape,
            exception_type='Missing Tape',
            description='Tape missing during handover.',
            reported_by=user,
            severity='High',
            status='Open',
        )

        auth_response = self.client.post(
            reverse('token_obtain_pair'),
            {'username': user.username, 'password': 'pass1234'},
            content_type='application/json',
        )
        self.assertEqual(auth_response.status_code, 200)
        access_token = auth_response.json()['access']

        response = self.client.get(
            f'//api/investigation/{exception.exception_id}/',
            HTTP_AUTHORIZATION=f'Bearer {access_token}',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['exception']['id'], exception.exception_id)

    def test_investigation_api_allows_forwarded_exceptions_in_investigating_state(self):
        branch = BankBranch.objects.create(branch_code='KLA-004', branch_name='Kampala Central Branch', status='Active')
        dr_group = Group.objects.create(name='DR Team')
        user = get_user_model().objects.create_user(
            username='investigator-forwarded',
            email='investigator-forwarded@example.com',
            password='pass1234',
            role='auditor',
            assigned_branch=branch,
        )
        user.groups.add(dr_group)
        tape = Tape.objects.create(
            volser='TAPE-INV-003',
            barcode='BC-INV-003',
            tape_type='LTO-8',
            retention_end_date=date(2028, 1, 1),
            current_location='Kampala Vault',
        )
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Approved',
            source_location='Kampala Branch',
            destination_location='Mbarara Branch',
            requesting_branch=branch,
            created_by=user,
        )
        shipment.tapes.add(tape)
        exception = ShipmentException.objects.create(
            shipment=shipment,
            tape=tape,
            exception_type='Missing Tape',
            description='Tape missing during handover.',
            reported_by=user,
            severity='High',
            status='Investigating',
        )

        auth_response = self.client.post(
            reverse('token_obtain_pair'),
            {'username': user.username, 'password': 'pass1234'},
            content_type='application/json',
        )
        self.assertEqual(auth_response.status_code, 200)
        access_token = auth_response.json()['access']

        response = self.client.get(
            reverse('exception-investigation', kwargs={'exception_id': exception.exception_id}),
            HTTP_AUTHORIZATION=f'Bearer {access_token}',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['exception']['id'], exception.exception_id)
        self.assertEqual(response.json()['exception']['status'], 'Investigating')


class JwtAuthenticationAndAuthorizationTests(TestCase):
    def test_jwt_obtain_refresh_and_blacklist_flow_with_role_based_access(self):
        branch = BankBranch.objects.create(branch_code='KLA-003', branch_name='Kampala West Branch', status='Active')
        dr_group = Group.objects.create(name='DR Team')
        auditor = get_user_model().objects.create_user(
            username='jwt-auditor',
            email='jwt-auditor@example.com',
            password='StrongPass123!',
            role='auditor',
            assigned_branch=branch,
        )
        auditor.groups.add(dr_group)
        regular_user = get_user_model().objects.create_user(
            username='jwt-regular',
            email='jwt-regular@example.com',
            password='StrongPass123!',
            role='user',
            assigned_branch=branch,
        )
        tape = Tape.objects.create(
            volser='TAPE-JWT-001',
            barcode='BC-JWT-001',
            tape_type='LTO-8',
            retention_end_date=date(2028, 1, 1),
            current_location='Kampala Vault',
        )
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Approved',
            source_location='Kampala Branch',
            destination_location='Mbarara Branch',
            requesting_branch=branch,
            created_by=auditor,
        )
        shipment.tapes.add(tape)
        exception = ShipmentException.objects.create(
            shipment=shipment,
            tape=tape,
            exception_type='Missing Tape',
            description='Tape missing during handover.',
            reported_by=auditor,
            severity='High',
            status='Open',
        )

        obtain_response = self.client.post(
            reverse('token_obtain_pair'),
            {'username': auditor.username, 'password': 'StrongPass123!'},
            content_type='application/json',
        )
        self.assertEqual(obtain_response.status_code, 200)
        tokens = obtain_response.json()
        self.assertIn('access', tokens)
        self.assertIn('refresh', tokens)

        anonymous_response = self.client.get(reverse('exception-investigation', kwargs={'exception_id': exception.exception_id}))
        self.assertEqual(anonymous_response.status_code, 401)

        auditor_response = self.client.get(
            reverse('exception-investigation', kwargs={'exception_id': exception.exception_id}),
            HTTP_AUTHORIZATION=f"Bearer {tokens['access']}",
        )
        self.assertEqual(auditor_response.status_code, 200)
        self.assertEqual(auditor_response.json()['exception']['id'], exception.exception_id)

        regular_obtain_response = self.client.post(
            reverse('token_obtain_pair'),
            {'username': regular_user.username, 'password': 'StrongPass123!'},
            content_type='application/json',
        )
        regular_tokens = regular_obtain_response.json()
        regular_response = self.client.get(
            reverse('exception-investigation', kwargs={'exception_id': exception.exception_id}),
            HTTP_AUTHORIZATION=f"Bearer {regular_tokens['access']}",
        )
        self.assertEqual(regular_response.status_code, 403)

        refresh_response = self.client.post(
            reverse('token_refresh'),
            {'refresh': tokens['refresh']},
            content_type='application/json',
        )
        self.assertEqual(refresh_response.status_code, 200)

        blacklist_response = self.client.post(
            reverse('token_blacklist'),
            {'refresh': refresh_response.json()['refresh']},
            content_type='application/json',
            HTTP_AUTHORIZATION=f"Bearer {tokens['access']}",
        )
        self.assertEqual(blacklist_response.status_code, 200)

        reusing_refresh_response = self.client.post(
            reverse('token_refresh'),
            {'refresh': refresh_response.json()['refresh']},
            content_type='application/json',
        )
        self.assertEqual(reusing_refresh_response.status_code, 401)


class CourierActivityLogTests(TestCase):
    def test_courier_profile_requires_vehicle_number(self):
        with self.assertRaises(ValidationError):
            CourierProfile.objects.create(
                courier_id='CR-TEST-001',
                full_name='Test Courier',
                email='test-courier@example.com',
                active_status=True,
            )

    def test_activity_log_shows_exception_activity_for_courier(self):
        courier_group = Group.objects.create(name='Courier')
        courier_user = get_user_model().objects.create_user(
            username='activity-log-courier',
            email='activity-log-courier@example.com',
            password='pass1234',
            first_name='Activity',
            last_name='Courier',
        )
        courier_user.groups.add(courier_group)

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='Nairobi Branch',
            destination_location='Kampala Branch',
            status='Approved',
            releasing_custodian='Ops User',
        )
        ShipmentException.objects.create(
            shipment=shipment,
            exception_type='Missing Tape',
            description='One tape was not present during handover.',
            reported_by=courier_user,
            severity='High',
        )

        self.client.force_login(courier_user)
        response = self.client.get(reverse('activity-log'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Missing Tape')
        self.assertContains(response, 'One tape was not present during handover.')


class OperationsManagerBranchAssignmentTests(TestCase):
    def test_operations_manager_creation_requires_assigned_branch(self):
        branch = BankBranch.objects.create(branch_code='KMA-001', branch_name='Kampala Main Branch', status='Active')
        form = CustomUserCreationForm(data={
            'username': 'opsmanager',
            'email': 'opsmanager@example.com',
            'password1': 'StrongPass123!',
            'password2': 'StrongPass123!',
            'first_name': 'Ops',
            'last_name': 'Manager',
            'role': 'operations_manager',
        })

        self.assertFalse(form.is_valid())
        self.assertIn('assigned_branch', form.errors)
        self.assertIn('Assigned Branch is required', str(form.errors['assigned_branch'][0]))

        form = CustomUserCreationForm(data={
            'username': 'opsmanager2',
            'email': 'opsmanager2@example.com',
            'password1': 'StrongPass123!',
            'password2': 'StrongPass123!',
            'first_name': 'Ops',
            'last_name': 'Manager',
            'role': 'operations_manager',
            'assigned_branch': branch.pk,
        })

        self.assertTrue(form.is_valid(), form.errors)

    def test_operations_manager_shipment_uses_assigned_branch_from_user_profile(self):
        branch = BankBranch.objects.create(branch_code='MBA-002', branch_name='Mbarara Branch', status='Active')
        operations_group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='opsuser',
            email='opsuser@example.com',
            password='pass1234',
            role='operations_manager',
        )
        user.groups.add(operations_group)
        user.assigned_branch = branch
        user.save(update_fields=['assigned_branch'])

        self.client.force_login(user)
        response = self.client.post(reverse('start-shipment-request'), {
            'form_type': 'submit_shipment_request',
            'branch_name': 'Some Other Branch',
            'requester_name': 'Ops User',
            'request_details': 'Move tapes to the branch',
        })

        self.assertEqual(response.status_code, 302)
        shipment = Shipment.objects.filter(created_by=user).latest('created_at')
        self.assertEqual(shipment.requesting_branch, branch)
        self.assertEqual(shipment.source_location, branch.branch_name)
        self.assertEqual(shipment.destination_location, branch.branch_name)


class AuditorDashboardTests(TestCase):
    def setUp(self):
        self.group = Group.objects.create(name='IT Compliance Auditor')
        self.user = get_user_model().objects.create_user(
            username='auditor',
            email='auditor@example.com',
            password='pass1234',
        )
        self.user.groups.add(self.group)

    def test_auditor_dashboard_renders_read_only_compliance_sections(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('auditor-dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'IT Compliance Auditor')
        self.assertContains(response, 'Compliance Health')
        self.assertContains(response, 'Audit Trail Review')
        self.assertContains(response, 'Read-only view')

    def test_user_with_auditor_role_can_access_dashboard(self):
        role_user = get_user_model().objects.create_user(
            username='auditor-role',
            email='auditor-role@example.com',
            password='pass1234',
            role='auditor',
        )
        self.client.force_login(role_user)
        response = self.client.get(reverse('auditor-dashboard'))

        self.assertEqual(response.status_code, 200)

    @override_settings(ROOT_URLCONF='inventory.tests')
    def test_custom_permission_denied_view_uses_403_template(self):
        response = self.client.get('/protected-page/')

        self.assertEqual(response.status_code, 403)
        self.assertContains(response, 'Access denied', status_code=403)
        self.assertContains(response, 'Return home', status_code=403)

    def test_auditor_reports_page_renders_report_module(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse('auditor-reports'),
            {
                'show_reports': '1',
                'report_category': 'inventory',
                'report_period': '2026-06',
                'report_type': 'monthly',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Report types')
        self.assertContains(response, 'Export CSV')

    def test_reports_load_inside_dashboard_when_requested(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse('auditor-dashboard'),
            {
                'view': 'reports',
                'show_reports': '1',
                'report_category': 'inventory',
                'report_period': '2026-06',
                'report_type': 'monthly',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Compliance Reports')
        self.assertContains(response, 'Report types')
        self.assertContains(response, 'Export CSV')
        self.assertContains(response, 'Compliance Health')
        self.assertContains(response, 'Audit Trail Review')

    def test_auditor_sections_render_inside_dashboard_shell(self):
        self.client.force_login(self.user)
        cases = [
            ('exceptions', 'Exception Review'),
            ('shipments', 'Shipment Compliance Review'),
            ('retention', 'Retention Compliance Review'),
            ('reconciliation', 'Reconciliation Review'),
        ]

        for view_name, expected_text in cases:
            with self.subTest(view=view_name):
                response = self.client.get(reverse('auditor-dashboard'), {'view': view_name})
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, expected_text)
                self.assertContains(response, 'IT Compliance Auditor Dashboard')

    def test_audit_logs_render_inside_auditor_dashboard(self):
        self.client.force_login(self.user)
        AuditLog.objects.create(
            name='System',
            action='Tape inventory reconciliation completed',
            user=self.user,
            severity='info',
        )

        response = self.client.get(reverse('auditor-dashboard'), {'view': 'audit-logs'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Tape inventory reconciliation completed')
        self.assertContains(response, 'Audit Trail Review')

    def test_audit_logs_filter_form_submits_audit_logs_view_param(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse('auditor-dashboard'), {'view': 'audit-logs'})

        self.assertContains(response, 'name="view"')
        self.assertContains(response, 'value="audit-logs"')

    def test_audit_logs_view_renders_enhanced_filter_controls(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse('auditor-dashboard'), {'view': 'audit-logs', 'search': 'reconciliation', 'severity': 'info'})

        self.assertContains(response, 'name="search"')
        self.assertContains(response, 'name="log_type"')
        self.assertContains(response, 'name="module"')
        self.assertContains(response, 'name="severity"')
        self.assertContains(response, 'name="user"')
        self.assertContains(response, 'name="status"')
        self.assertContains(response, 'name="date_range"')
        self.assertContains(response, 'name="per_page"')

    def test_audit_logs_export_generates_csv(self):
        self.client.force_login(self.user)
        AuditLog.objects.create(
            name='Reports',
            action='Compliance report exported',
            user=self.user,
            severity='success',
        )

        response = self.client.get(reverse('auditor-dashboard'), {'view': 'audit-logs', 'export': 'csv'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        self.assertContains(response, 'Compliance report exported')

    def test_unknown_page_uses_custom_404_template(self):
        response = self.client.get('/definitely-not-a-real-page/')

        self.assertEqual(response.status_code, 404)
        self.assertContains(response, 'Page not found', status_code=404)
        self.assertContains(response, 'Return home', status_code=404)

    def test_user_with_auditor_group_name_can_access_dashboard(self):
        group = Group.objects.create(name='Auditor')
        group_user = get_user_model().objects.create_user(
            username='auditor-group',
            email='auditor-group@example.com',
            password='pass1234',
        )
        group_user.groups.add(group)
        self.client.force_login(group_user)
        response = self.client.get(reverse('auditor-dashboard'))

        self.assertEqual(response.status_code, 200)

    def test_auditor_can_submit_pending_shipment_request(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse('auditor-dashboard'),
            {
                'form_type': 'submit_shipment_request',
                'branch_name': 'Nairobi Branch',
                'request_details': 'Need secure transfer for tape inventory.',
            },
        )

        self.assertEqual(response.status_code, 302)
        shipment = Shipment.objects.get(created_by=self.user)
        self.assertEqual(shipment.status, 'Pending')
        self.assertEqual(shipment.source_location, 'Nairobi Branch')
        self.assertEqual(shipment.approval_remarks, 'Need secure transfer for tape inventory.')
        self.assertEqual(shipment.created_by, self.user)


class BackupDashboardSettingsTests(TestCase):
    def test_backup_dashboard_audit_logs_renders_enhanced_filters(self):
        user = get_user_model().objects.create_superuser(
            username='backup-audit-admin',
            email='backup-audit-admin@example.com',
            password='pass1234',
        )
        AuditLog.objects.create(
            name='Reports',
            action='Backup audit log review requested',
            user=user,
            severity='warning',
        )

        self.client.force_login(user)
        response = self.client.get(reverse('backup-dashboard'), {'show_audit': '1', 'search': 'review', 'severity': 'warning'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="search"')
        self.assertContains(response, 'name="log_type"')
        self.assertContains(response, 'name="module"')
        self.assertContains(response, 'name="severity"')
        self.assertContains(response, 'name="user"')
        self.assertContains(response, 'name="status"')
        self.assertContains(response, 'name="date_range"')
        self.assertContains(response, 'name="per_page"')

    def test_backup_dashboard_settings_save_reconciliation_schedule(self):
        user = get_user_model().objects.create_superuser(
            username='backup-settings-admin',
            email='backup-settings-admin@example.com',
            password='pass1234',
        )
        application_settings = ApplicationSetting.objects.create()

        self.client.force_login(user)
        response = self.client.post(
            reverse('backup-dashboard'),
            {
                'form_type': 'system_settings',
                'backup_retention_days': '180',
                'shipment_notification_enabled': 'on',
                'email_alerts_enabled': 'on',
                'allow_offsite_transfers': 'on',
                'max_tapes_per_shipment': '60',
                'audit_logging_level': 'warning',
                'audit_retention_years': '8',
                'default_dashboard_section': 'reports',
                'maintenance_window_start': '02:00',
                'maintenance_window_end': '04:00',
                'next_reconciliation_date': '2026-08-15',
                'reconciliation_alert_start_days_before': '5',
                'reconciliation_alert_duration_days': '10',
            },
        )

        self.assertEqual(response.status_code, 302)
        application_settings.refresh_from_db()
        self.assertEqual(application_settings.next_reconciliation_date, date(2026, 8, 15))
        self.assertEqual(application_settings.reconciliation_alert_start_days_before, 5)
        self.assertEqual(application_settings.reconciliation_alert_duration_days, 10)


class ShipmentFormTests(TestCase):
    def test_add_shipment_form_prefills_releasing_custodian_with_current_user(self):
        backup_group = Group.objects.create(name='Backup Administrator')
        user = get_user_model().objects.create_user(
            username='backup-admin',
            email='backup-admin@example.com',
            password='pass1234',
            first_name='Jane',
            last_name='Doe',
        )
        user.groups.add(backup_group)

        self.client.force_login(user)
        response = self.client.get(reverse('backup-dashboard'), {'show_add_shipment': '1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Jane Doe')


class ExcelSchemaSynchronizationTests(TestCase):
    def test_uploading_excel_with_new_columns_shows_schema_preview(self):
        user = get_user_model().objects.create_superuser(
            username='excel-preview-admin',
            email='excel-preview-admin@example.com',
            password='pass1234',
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['volser', 'barcode', 'tape_type', 'status', 'current_location', 'retention_end_date', 'manufacturer', 'Security Tag', 'Vault Number'])
        sheet.append(['TAPE-001', 'BC-001', 'LTO-8', 'Active', 'Vault A', '2026-12-31', 'IBM', 'SEC-001', 'V1'])

        excel_file = SimpleUploadedFile(
            'inventory_template.xlsx',
            b'',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        excel_file.file = BytesIO()
        workbook.save(excel_file.file)
        excel_file.file.seek(0)
        excel_file.content = excel_file.file.getvalue()

        self.client.force_login(user)
        response = self.client.post(
            reverse('backup-dashboard'),
            {'form_type': 'upload_inventory_excel', 'inventory_file': excel_file},
            format='multipart',
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('schema_preview', response.context)
        self.assertEqual(response.context['schema_preview']['table_name'], 'inventory_tape')
        self.assertGreaterEqual(len(response.context['schema_preview']['new_columns']), 1)
        self.assertContains(response, 'Schema Synchronization Preview')
        self.assertContains(response, 'Approve & Synchronize')

    def test_approving_schema_sync_adds_columns_and_imports_rows(self):
        user = get_user_model().objects.create_superuser(
            username='excel-sync-admin',
            email='excel-sync-admin@example.com',
            password='pass1234',
        )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['volser', 'barcode', 'tape_type', 'status', 'current_location', 'retention_end_date', 'manufacturer', 'Security Tag', 'Vault Number'])
        sheet.append(['TAPE-100', 'BC-100', 'LTO-8', 'Active', 'Vault A', '2026-12-31', 'IBM', 'SEC-100', 'V2'])

        excel_file = SimpleUploadedFile(
            'inventory_template.xlsx',
            b'',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        excel_file.file = BytesIO()
        workbook.save(excel_file.file)
        excel_file.file.seek(0)
        excel_file.content = excel_file.file.getvalue()

        self.client.force_login(user)
        self.client.post(
            reverse('backup-dashboard'),
            {'form_type': 'upload_inventory_excel', 'inventory_file': excel_file},
            format='multipart',
        )

        response = self.client.post(
            reverse('backup-dashboard'),
            {'form_type': 'approve_excel_schema_sync'},
        )

        self.assertEqual(response.status_code, 302)
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT column_name FROM information_schema.columns WHERE table_schema = current_schema() AND table_name = %s ORDER BY ordinal_position",
                ['inventory_tape'],
            )
            columns = [row[0] for row in cursor.fetchall()]
        self.assertIn('security_tag', columns)
        self.assertIn('vault_number', columns)
        self.assertTrue(SchemaChangeLog.objects.filter(synchronization_status='applied').exists())
        self.assertTrue(Tape.objects.filter(volser='TAPE-100').exists())


class DashboardFeatureNavigationTests(TestCase):
    def test_role_form_includes_django_admin_permissions_as_assignable_features(self):
        content_type = ContentType.objects.get(app_label='inventory', model='tape')
        permission = Permission.objects.get(content_type=content_type, codename='view_tape')

        form = RoleCreationForm()
        feature_choices = dict(form.fields['features'].choices)

        self.assertIn(f'{permission.content_type.app_label}.{permission.codename}', feature_choices)

    def test_dashboard_context_processor_exposes_permitted_features(self):
        group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='nav-user',
            email='nav-user@example.com',
            password='pass1234',
        )
        user.groups.add(group)

        DashboardFeaturePermission.objects.create(role=group, feature_key='operations_dashboard', can_view=True)

        self.client.force_login(user)
        response = self.client.get(reverse('operations-dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(any(entry['key'] == 'operations_dashboard' for entry in response.context['dashboard_features']))

    def test_dashboard_context_processor_exposes_api_navigation_urls_for_permitted_features(self):
        group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='nav-api-user',
            email='nav-api-user@example.com',
            password='pass1234',
        )
        user.groups.add(group)

        DashboardFeaturePermission.objects.create(role=group, feature_key='operations_dashboard', can_view=True)

        self.client.force_login(user)
        response = self.client.get(reverse('operations-dashboard'))

        self.assertEqual(response.status_code, 200)
        entry = next((item for item in response.context['dashboard_features'] if item['key'] == 'operations_dashboard'), None)
        self.assertIsNotNone(entry)
        self.assertIn('api_url', entry)
        self.assertEqual(entry['api_url'], reverse('api-feature-navigation', kwargs={'feature_key': 'operations_dashboard'}))

    def test_dashboard_feature_exemptions_hide_feature_for_individual_user(self):
        group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='nav-exempt',
            email='nav-exempt@example.com',
            password='pass1234',
        )
        user.groups.add(group)

        DashboardFeaturePermission.objects.create(role=group, feature_key='operations_dashboard', can_view=True)
        DashboardFeatureExemption.objects.create(user=user, feature_key='operations_dashboard', is_active=True, reason='Temporary exemption')

        self.client.force_login(user)
        response = self.client.get(reverse('operations-dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(any(entry['key'] == 'operations_dashboard' for entry in response.context['dashboard_features']))

    def test_feature_catalog_uses_feature_module_routes_instead_of_dashboard_urls(self):
        features = get_dashboard_feature_catalog()

        self.assertTrue(features)
        for feature in features:
            self.assertEqual(feature['url_name'], 'feature-module')
            self.assertEqual(feature['url_kwargs']['feature_key'], feature['key'])

    def test_feature_navigation_api_returns_standalone_page_payload_for_features(self):
        group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='nav-fragment-user',
            email='nav-fragment-user@example.com',
            password='pass1234',
        )
        user.groups.add(group)

        DashboardFeaturePermission.objects.create(role=group, feature_key='shipment_approvals', can_view=True)

        self.client.force_login(user)
        response = self.client.get(reverse('api-feature-navigation', kwargs={'feature_key': 'shipment_approvals'}))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['load_mode'], 'page')
        self.assertIn('/operations-dashboard/nav/', payload['target_url'])
        self.assertEqual(payload['target_url'], payload['fragment_url'])

    def test_feature_permission_does_not_confer_backup_admin_role(self):
        group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='feature-role-ops',
            email='feature-role-ops@example.com',
            password='pass1234',
        )
        user.groups.add(group)
        DashboardFeaturePermission.objects.create(role=group, feature_key='tape_inventory', can_view=True)

        self.assertTrue(is_operations_manager(user))
        self.assertFalse(is_backup_administrator(user))

    def test_feature_page_hides_sidebar_when_feature_key_is_present(self):
        group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='feature-page-sidebar-user',
            email='feature-page-sidebar-user@example.com',
            password='pass1234',
        )
        user.groups.add(group)

        self.client.force_login(user)
        response = self.client.get(reverse('shipment-approvals'), {'feature_key': 'shipment_approvals'})

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'class="app-sidebar')
        self.assertNotContains(response, 'id="sidebarPanel"')

    def test_feature_permission_does_not_grant_dashboard_access_for_custom_group_name(self):
        group = Group.objects.create(name='Warehouse Ops')
        user = get_user_model().objects.create_user(
            username='custom-group-user',
            email='custom-group-user@example.com',
            password='pass1234',
        )
        user.groups.add(group)

        DashboardFeaturePermission.objects.create(role=group, feature_key='operations_dashboard', can_view=True)

        self.client.force_login(user)
        response = self.client.get(reverse('operations-dashboard'))

        self.assertEqual(response.status_code, 302)

    def test_feature_module_allows_feature_access_without_dashboard_role(self):
        group = Group.objects.create(name='Warehouse Ops')
        user = get_user_model().objects.create_user(
            username='feature-module-user',
            email='feature-module-user@example.com',
            password='pass1234',
        )
        user.groups.add(group)

        DashboardFeaturePermission.objects.create(role=group, feature_key='shipment_approvals', can_view=True)

        self.client.force_login(user)
        response = self.client.get(reverse('feature-module', kwargs={'feature_key': 'shipment_approvals'}))

        self.assertEqual(response.status_code, 200)

    def test_feature_module_renders_backup_feature_without_dashboard_sidebar(self):
        group = Group.objects.create(name='Backup Administrator')
        user = get_user_model().objects.create_user(
            username='backup-feature-shell-user',
            email='backup-feature-shell-user@example.com',
            password='pass1234',
        )
        user.groups.add(group)

        DashboardFeaturePermission.objects.create(role=group, feature_key='tape_inventory', can_view=True)

        self.client.force_login(user)
        response = self.client.get(reverse('feature-module', kwargs={'feature_key': 'tape_inventory'}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Back to dashboard')
        self.assertNotContains(response, 'class="sidebar"')
        self.assertEqual(response.context['active_feature_key'], 'tape_inventory')

    def test_feature_link_opens_selected_section_inside_backup_dashboard(self):
        group = Group.objects.create(name='Backup Administrator')
        user = get_user_model().objects.create_user(
            username='backup-feature-user',
            email='backup-feature-user@example.com',
            password='pass1234',
        )
        user.groups.add(group)

        DashboardFeaturePermission.objects.create(role=group, feature_key='tape_inventory', can_view=True)

        self.client.force_login(user)
        response = self.client.get(reverse('backup-dashboard'), {'feature_key': 'tape_inventory'})

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['show_tape_inventory_panel'])
        self.assertEqual(response.context['active_feature_key'], 'tape_inventory')

    def test_feature_link_opens_selected_section_inside_operations_dashboard(self):
        group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='operations-feature-user',
            email='operations-feature-user@example.com',
            password='pass1234',
        )
        user.groups.add(group)

        DashboardFeaturePermission.objects.create(role=group, feature_key='exception_management', can_view=True)

        self.client.force_login(user)
        response = self.client.get(reverse('operations-dashboard'), {'feature_key': 'exception_management'})

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['show_exception_panel'])
        self.assertEqual(response.context['active_feature_key'], 'exception_management')


class ApiEndpointTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='api-user',
            email='api-user@example.com',
            password='pass1234',
        )
        self.client.force_login(self.user)

    def test_dashboard_summary_api_returns_counts(self):
        Tape.objects.create(
            volser='TAPE-001',
            barcode='BAR-001',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
        )
        Shipment.objects.create(
            shipment_id='SHP-1001',
            source_location='HQ',
            destination_location='DR',
            shipment_type='Off-Site Transfer',
            priority_level='Normal',
            status='Pending',
            created_by=self.user,
        )
        AuditLog.objects.create(
            name='System',
            action='API endpoint tested',
            user=self.user,
            severity='info',
        )

        response = self.client.get('/api/dashboard-summary/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['tape_count'], 1)
        self.assertEqual(response.json()['shipment_count'], 1)
        self.assertEqual(response.json()['audit_log_count'], 1)

    def test_tape_list_api_returns_tape_payload(self):
        Tape.objects.create(
            volser='TAPE-002',
            barcode='BAR-002',
            tape_type='LTO-9',
            retention_end_date=date(2031, 1, 1),
        )

        response = self.client.get('/api/tapes/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['count'], 1)
        self.assertEqual(response.json()['results'][0]['volser'], 'TAPE-002')

    def test_tape_serializer_returns_expected_fields(self):
        tape = Tape.objects.create(
            volser='TAPE-003',
            barcode='BAR-003',
            tape_type='LTO-8',
            retention_end_date=date(2032, 1, 1),
            status='Active',
            current_location='Vault A',
        )

        payload = TapeSerializer(tape).data

        self.assertEqual(payload['volser'], 'TAPE-003')
        self.assertEqual(payload['status'], 'Active')
        self.assertEqual(payload['location'], 'Vault A')

    def test_tape_creation_via_api_returns_created_payload(self):
        response = self.client.post(
            '/api/tapes/',
            json.dumps({
                'volser': 'TAPE-004',
                'tape_type': 'LTO-8',
                'retention_end_date': '2035-01-01',
                'current_location': 'Vault B',
                'status': 'Active',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(Tape.objects.count(), 1)
        self.assertEqual(response.json()['result']['volser'], 'TAPE-004')

    def test_shipments_api_returns_collection(self):
        Shipment.objects.create(
            shipment_id='SHP-2001',
            source_location='HQ',
            destination_location='DR',
            shipment_type='Off-Site Transfer',
            priority_level='Normal',
            status='Pending',
            created_by=self.user,
        )

        response = self.client.get('/api/shipments/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['count'], 1)
        self.assertEqual(response.json()['results'][0]['shipment_id'], 'SHP-2001')

    def test_audit_logs_api_returns_collection(self):
        AuditLog.objects.create(
            name='System',
            action='Audit log fetched through API',
            user=self.user,
            severity='info',
        )

        response = self.client.get('/api/audit-logs/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['count'], 1)
        self.assertEqual(response.json()['results'][0]['action'], 'Audit log fetched through API')


class TapeStatusProtectionTests(TestCase):
    def test_scratch_status_change_is_rejected_for_tape_on_legal_hold(self):
        backup_group = Group.objects.create(name='Backup Administrator')
        backup_admin = get_user_model().objects.create_user(
            username='backup-admin-protect',
            email='backup-admin-protect@example.com',
            password='pass1234',
        )
        backup_admin.groups.add(backup_group)
        tape = Tape.objects.create(
            volser='TAPE-200',
            barcode='BAR-200',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Vault A',
            legal_hold=True,
        )

        self.client.force_login(backup_admin)
        response = self.client.post(
            reverse('backup-dashboard'),
            {
                'form_type': 'tape_action',
                'selected_tape': tape.pk,
                'action': 'edit_details',
                'volser': tape.volser,
                'barcode': tape.barcode,
                'current_location': tape.current_location,
                'retention_end_date': '2030-01-01',
                'status': 'Scratch Eligible',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cannot mark tape as Scratch')
        tape.refresh_from_db()
        self.assertEqual(tape.status, 'Active')


class TapeRequestWorkflowTests(TestCase):
    def test_operator_request_can_be_approved_into_a_shipment(self):
        operations_group = Group.objects.create(name='Operations Manager')
        backup_group = Group.objects.create(name='Backup Administrator')
        operator = get_user_model().objects.create_user(
            username='operator-one',
            email='operator-one@example.com',
            password='pass1234',
            first_name='Op',
            last_name='User',
        )
        operator.groups.add(operations_group)
        backup_admin = get_user_model().objects.create_user(
            username='backup-admin',
            email='backup-admin@example.com',
            password='pass1234',
            first_name='Backup',
            last_name='Admin',
        )
        backup_admin.groups.add(backup_group)
        tape = Tape.objects.create(
            volser='TAPE-100',
            barcode='BAR-100',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Vault A',
        )

        self.client.force_login(operator)
        response = self.client.post(
            reverse('operations-dashboard'),
            {
                'form_type': 'submit_tape_request',
                'tape': tape.pk,
                'quantity': 1,
                'destination_location': 'DR Site',
                'receiving_organization': 'Ops Team',
                'reason': 'Need this tape for a scheduled restore.',
            },
        )

        self.assertEqual(response.status_code, 302)
        request = TapeRequest.objects.get(requested_by=operator)
        self.assertEqual(request.status, 'Pending')

        self.client.force_login(backup_admin)
        response = self.client.post(
            reverse('backup-dashboard'),
            {
                'form_type': 'approve_tape_request',
                'request_id': request.pk,
                'approval_notes': 'Approved for immediate handoff.',
            },
        )

        self.assertEqual(response.status_code, 302)
        request.refresh_from_db()
        self.assertEqual(request.status, 'Approved')
        self.assertTrue(request.shipment_id is not None)
        shipment = Shipment.objects.get(pk=request.shipment_id)
        self.assertEqual(shipment.status, 'Approved')
        self.assertEqual(shipment.destination_location, 'DR Site')
        self.assertEqual(shipment.number_of_tapes, 1)


class ShipmentWorkflowTests(TestCase):
    def test_start_shipment_request_renders_embedded_fragment_when_requested_partially(self):
        operations_group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='operator-fragment',
            email='operator-fragment@example.com',
            password='pass1234',
        )
        user.groups.add(operations_group)

        self.client.force_login(user)
        response = self.client.get(reverse('start-shipment-request'), {'partial': '1'})

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'start_shipment_request_fragment.html')
        self.assertContains(response, 'Start a shipment request')
        self.assertNotContains(response, '<!DOCTYPE html>')

    def test_start_shipment_request_renders_dual_panel_tape_selector(self):
        operations_group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='operator-selector',
            email='operator-selector@example.com',
            password='pass1234',
        )
        user.groups.add(operations_group)
        Tape.objects.create(
            volser='VOL-100',
            barcode='BC-100',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Vault A',
        )

        self.client.force_login(user)
        response = self.client.get(reverse('start-shipment-request'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Available Tapes')
        self.assertContains(response, 'Selected Tapes')
        self.assertContains(response, 'Search Tape')
        self.assertContains(response, 'Select All')
        self.assertContains(response, 'Clear All')

    def test_operator_shipment_request_stores_branch_name_as_destination(self):
        operations_group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='operator-destination',
            email='operator-destination@example.com',
            password='pass1234',
        )
        user.groups.add(operations_group)

        self.client.force_login(user)
        response = self.client.post(
            reverse('operations-dashboard'),
            {
                'form_type': 'submit_shipment_request',
                'branch_name': 'Nairobi Branch',
                'requester_name': 'Operator One',
                'request_details': 'Move tapes to Nairobi Branch.',
            },
        )

        self.assertEqual(response.status_code, 302)
        shipment = Shipment.objects.get(created_by=user)
        self.assertEqual(shipment.destination_location, 'Nairobi Branch')

    def test_operator_shipment_request_flow_reaches_courier_acceptance(self):
        operations_group = Group.objects.create(name='Operations Manager')
        backup_group = Group.objects.create(name='Backup Administrator')
        courier_group = Group.objects.create(name='Courier')

        operator = get_user_model().objects.create_user(
            username='operator-flow',
            email='operator-flow@example.com',
            password='pass1234',
            first_name='Op',
            last_name='User',
        )
        operator.groups.add(operations_group)

        backup_admin = get_user_model().objects.create_user(
            username='backup-flow',
            email='backup-flow@example.com',
            password='pass1234',
            first_name='Backup',
            last_name='Admin',
        )
        backup_admin.groups.add(backup_group)

        courier_user = get_user_model().objects.create_user(
            username='courier-flow',
            email='courier-flow@example.com',
            password='pass1234',
            first_name='Courier',
            last_name='Guy',
        )
        courier_user.groups.add(courier_group)
        courier_profile = CourierProfile.objects.create(
            user=courier_user,
            courier_id='CR-100',
            full_name='Courier Guy',
            phone_number='555-1000',
            email='courier-flow@example.com',
            vehicle_number='VEH-100',
        )

        tape = Tape.objects.create(
            volser='TAPE-900',
            barcode='BAR-900',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Vault A',
        )

        self.client.force_login(operator)
        response = self.client.post(
            reverse('operations-dashboard'),
            {
                'form_type': 'submit_shipment_request',
                'branch_name': 'Nairobi Branch',
                'request_details': 'Need transfer of tapes to DR site.',
            },
        )

        self.assertEqual(response.status_code, 302)
        shipment = Shipment.objects.get(created_by=operator)
        self.assertEqual(shipment.status, 'Pending')
        self.assertEqual(shipment.source_location, 'Nairobi Branch')

        self.client.force_login(backup_admin)
        response = self.client.post(
            reverse('shipment-approvals'),
            {
                'form_type': 'backup_admin_decision',
                'shipment_id': shipment.pk,
                'tape_id': tape.pk,
                'courier_id': courier_profile.pk,
                'decision': 'approve',
                'comments': 'Approved for dispatch.',
            },
        )

        shipment.refresh_from_db()
        self.assertEqual(shipment.status, 'Approved')
        self.assertEqual(shipment.tapes.count(), 1)
        self.assertEqual(shipment.courier_name, 'Courier Guy')

        self.client.force_login(courier_user)
        response = self.client.post(
            reverse('pickup-confirmation', args=[shipment.pk]),
            {
                'manifest_reference': 'MANIFEST-1',
                'pickup_date': '2026-06-26',
                'pickup_time': '09:00',
                'pickup_location': 'Vault A',
                'notes': 'Pickup confirmed',
                'all_tapes_scanned': 'on',
                'manifest_verified': 'on',
                'tape_count_matched': 'on',
                'no_damaged_tapes': 'on',
                'custody_accepted': 'on',
            },
        )

        shipment.refresh_from_db()
        self.assertEqual(shipment.status, 'Picked Up')




class OperationsDashboardCustodyGovernanceTests(TestCase):
    def test_custody_governance_cards_use_receipt_and_delivery_records(self):
        operations_group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='ops-custody-metrics',
            email='ops-custody-metrics@example.com',
            password='pass1234',
        )
        user.groups.add(operations_group)

        courier = CourierProfile.objects.create(
            courier_id='CR-200',
            full_name='Test Courier',
            email='courier@example.com',
        )

        open_transfer = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Picked Up',
            source_location='Vault A',
            destination_location='Nairobi Branch',
            created_by=user,
        )
        ShipmentReceipt.objects.create(
            shipment=open_transfer,
            courier=courier,
            pickup_location='Vault A',
            custody_confirmed=True,
            custody_accepted=True,
        )

        missing_handoff = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Dispatched',
            source_location='Vault A',
            destination_location='Nairobi Branch',
            created_by=user,
        )

        delivered_without_confirmation = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Delivered',
            source_location='Vault A',
            destination_location='Nairobi Branch',
            created_by=user,
        )

        completed_transfer = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Delivered',
            source_location='Vault A',
            destination_location='Nairobi Branch',
            created_by=user,
        )
        DeliveryConfirmation.objects.create(
            shipment=completed_transfer,
            courier=courier,
            destination_location='Nairobi Branch',
            receiving_custodian='Ops Lead',
            delivery_status='Delivered',
            manifest_matched=True,
            all_tapes_delivered=True,
            discrepancies_resolved=True,
        )

        self.client.force_login(user)
        response = self.client.get(reverse('operations-dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['custody_transfers_open'], 1)
        self.assertEqual(response.context['custody_transfers_completed'], 1)
        self.assertEqual(response.context['missing_handoffs'], 1)
        self.assertEqual(response.context['unverified_deliveries'], 1)


class OperationsDashboardNotificationsTests(TestCase):
    def test_notification_view_links_render_with_target_url_hooks(self):
        operations_group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='ops-notifications',
            email='ops-notifications@example.com',
            password='pass1234',
        )
        user.groups.add(operations_group)
        AuditLog.objects.create(
            name='Shipment Request Submitted',
            action='Shipment request was submitted for review.',
            message=f'target_url={reverse("pickup-confirmation", args=[uuid.uuid4()])}',
            user=user,
            severity='warning',
            is_read=False,
        )

        self.client.force_login(user)
        response = self.client.get(reverse('operations-dashboard'), {'show_notifications': '1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'notification-view-link')
        self.assertContains(response, 'data-target-url=')
        self.assertContains(response, 'pickup-confirmation')


class OperationsDashboardReportsTests(TestCase):
    def setUp(self):
        self.group = Group.objects.create(name='Operations Manager')
        self.user = get_user_model().objects.create_user(
            username='ops-reports',
            email='ops-reports@example.com',
            password='pass1234',
        )
        self.user.groups.add(self.group)

    def test_operations_dashboard_reports_panel_uses_shared_report_module(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse('operations-dashboard'),
            {
                'show_reports': '1',
                'report_category': 'inventory',
                'report_period': '2026-06',
                'report_type': 'monthly',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Report types')
        self.assertContains(response, 'Export CSV')

    def test_operations_dashboard_inventory_report_search_filters_table_rows(self):
        Tape.objects.create(
            volser='TAPE-100',
            barcode='BAR-100',
            tape_type='LTO-8',
            status='Active',
            current_location='Vault A',
            retention_end_date=date(2030, 1, 1),
        )
        Tape.objects.create(
            volser='TAPE-200',
            barcode='BAR-200',
            tape_type='LTO-8',
            status='Active',
            current_location='Vault B',
            retention_end_date=date(2030, 1, 1),
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse('operations-dashboard'),
            {
                'show_reports': 'reports',
                'report_category': 'inventory',
                'report_period': '2026-06',
                'report_type': 'monthly',
                'report_search_inventory': 'TAPE-100',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'TAPE-100')
        self.assertNotContains(response, 'TAPE-200')

    def test_operations_dashboard_report_only_view_renders_report_table_without_full_dashboard(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse('operations-dashboard'),
            {
                'show_reports': 'reports',
                'report_category': 'inventory',
                'report_period': '2026-06',
                'report_type': 'monthly',
                'report_only': '1',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Report types')
        self.assertContains(response, 'Inventory Report')
        self.assertContains(response, 'showReportsPanel();')


class TwoFactorLoginTests(TestCase):
    def test_signin_sends_terminal_otp_and_completes_login(self):
        operations_group = Group.objects.create(name='Operations Manager')
        user = get_user_model().objects.create_user(
            username='otp-user',
            email='otp-user@example.com',
            password='pass1234',
        )
        user.groups.add(operations_group)

        with patch('builtins.print') as mocked_print:
            response = self.client.post(reverse('signin'), {'username': 'otp-user', 'password': 'pass1234'})

            self.assertEqual(response.status_code, 200)
            self.assertContains(response, 'Verification Code')
            self.assertTrue(self.client.session.get('pending_2fa_user_id'))
            otp_code = self.client.session.get('pending_2fa_otp')
            self.assertTrue(otp_code)
            mocked_print.assert_called()

            final_response = self.client.post(reverse('signin'), {'otp_code': otp_code})

            self.assertEqual(final_response.status_code, 302)
            self.assertRedirects(final_response, reverse('operations-dashboard'))
            self.assertIn('_auth_user_id', self.client.session)

    def test_signin_redirects_dr_team_to_investigation_dashboard(self):
        dr_team_group = Group.objects.create(name='DR Team')
        user = get_user_model().objects.create_user(
            username='dr-otp-user',
            email='dr-otp-user@example.com',
            password='pass1234',
        )
        user.groups.add(dr_team_group)

        with patch('builtins.print'):
            response = self.client.post(reverse('signin'), {'username': 'dr-otp-user', 'password': 'pass1234'})
            self.assertEqual(response.status_code, 200)
            otp_code = self.client.session.get('pending_2fa_otp')
            self.assertTrue(otp_code)

            final_response = self.client.post(reverse('signin'), {'otp_code': otp_code})

            self.assertEqual(final_response.status_code, 302)
            self.assertRedirects(final_response, reverse('investigation-dashboard'))
            self.assertIn('_auth_user_id', self.client.session)


class BackupDashboardShipmentApprovalTests(TestCase):
    def test_backup_dashboard_alert_bell_counts_only_unread_alerts(self):
        backup_group = Group.objects.create(name='Backup Administrator')
        backup_admin = get_user_model().objects.create_user(
            username='backup-bell-count',
            email='backup-bell-count@example.com',
            password='pass1234',
            first_name='Backup',
            last_name='Admin',
        )
        backup_admin.groups.add(backup_group)

        AuditLog.objects.create(
            name='Shipment Rejected',
            action='An alert arrived for review.',
            user=backup_admin,
            severity='warning',
            is_read=False,
        )
        AuditLog.objects.create(
            name='Shipment Approved',
            action='A previous alert was already reviewed.',
            user=backup_admin,
            severity='error',
            is_read=True,
        )

        self.client.force_login(backup_admin)
        response = self.client.get(reverse('backup-dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['alert_count'], 1)

    def test_backup_dashboard_marks_alerts_read_when_panel_is_opened(self):
        backup_group = Group.objects.create(name='Backup Administrator')
        backup_admin = get_user_model().objects.create_user(
            username='backup-new-badge',
            email='backup-new-badge@example.com',
            password='pass1234',
            first_name='Backup',
            last_name='Admin',
        )
        backup_admin.groups.add(backup_group)

        alert = AuditLog.objects.create(
            name='Shipment Rejected',
            action='Shipment REQ-100 was rejected by the backup administrator.',
            user=backup_admin,
            severity='warning',
            is_read=False,
        )

        self.client.force_login(backup_admin)
        initial_response = self.client.get(reverse('backup-dashboard'))
        self.assertEqual(initial_response.status_code, 200)
        self.assertContains(initial_response, '<span class="badge bg-primary">New</span>')
        self.assertContains(initial_response, alert.action)

        panel_response = self.client.get(reverse('backup-dashboard'), {'show_alerts': '1'})
        self.assertEqual(panel_response.status_code, 200)
        self.assertNotContains(panel_response, '<span class="badge bg-primary">New</span>')
        alert.refresh_from_db()
        self.assertTrue(alert.is_read)
        self.assertIsNotNone(alert.read_at)

    def test_backup_dashboard_alert_renders_clickable_target_url(self):
        backup_group = Group.objects.create(name='Backup Administrator')
        backup_admin = get_user_model().objects.create_user(
            username='backup-clickable-alert',
            email='backup-clickable-alert@example.com',
            password='pass1234',
        )
        backup_admin.groups.add(backup_group)

        target_url = reverse('shipment-detail', args=[uuid.uuid4()])
        AuditLog.objects.create(
            name='Pickup Confirmed',
            action='Pickup confirmed for shipment SH-12345.',
            message=f'target_url={target_url}',
            user=backup_admin,
            severity='warning',
            is_read=False,
        )

        self.client.force_login(backup_admin)
        response = self.client.get(reverse('backup-dashboard'), {'show_alerts': '1'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, target_url)
        self.assertContains(response, 'text-decoration-none text-dark w-100')

    def test_backup_dashboard_alert_lists_requested_tapes_and_locations(self):
        backup_group = Group.objects.create(name='Backup Administrator')
        backup_admin = get_user_model().objects.create_user(
            username='backup-tape-list',
            email='backup-tape-list@example.com',
            password='pass1234',
        )
        backup_admin.groups.add(backup_group)

        tape = Tape.objects.create(
            volser='TAPE-ALERT-001',
            barcode='BAR-ALERT-001',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Vault A',
        )
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='Nairobi Branch',
            destination_location='Mombasa Branch',
            status='Pending',
            releasing_custodian='Ops User',
            created_by=backup_admin,
        )
        shipment.tapes.add(tape)
        shipment.number_of_tapes = shipment.tapes.count()
        shipment.save(update_fields=['number_of_tapes'])

        self.client.force_login(backup_admin)
        response = self.client.get(reverse('backup-dashboard'), {'show_alerts': '1', 'approve_shipment': shipment.pk})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Requested Tapes')
        self.assertContains(response, tape.volser)
        self.assertContains(response, tape.current_location)

    def test_backup_admin_can_approve_pending_shipment_from_dashboard_with_barcode_and_courier(self):
        operations_group = Group.objects.create(name='Operations Manager')
        backup_group = Group.objects.create(name='Backup Administrator')
        courier_group = Group.objects.create(name='Courier')

        operator = get_user_model().objects.create_user(
            username='operator-approve',
            email='operator-approve@example.com',
            password='pass1234',
            first_name='Op',
            last_name='User',
        )
        operator.groups.add(operations_group)

        backup_admin = get_user_model().objects.create_user(
            username='backup-approve',
            email='backup-approve@example.com',
            password='pass1234',
            first_name='Backup',
            last_name='Admin',
        )
        backup_admin.groups.add(backup_group)

        courier_user = get_user_model().objects.create_user(
            username='courier-approve',
            email='courier-approve@example.com',
            password='pass1234',
            first_name='Courier',
            last_name='Guy',
        )
        courier_user.groups.add(courier_group)
        courier_profile = CourierProfile.objects.create(
            user=courier_user,
            courier_id='CR-200',
            full_name='Courier Guy',
            phone_number='555-2000',
            email='courier-approve@example.com',
        )

        tape = Tape.objects.create(
            volser='TAPE-777',
            barcode='BAR-777',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Vault A',
        )

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='Nairobi Branch',
            status='Pending',
            releasing_custodian='Op User',
            created_by=operator,
        )

        self.client.force_login(backup_admin)
        response = self.client.post(
            reverse('backup-dashboard'),
            {
                'form_type': 'backup_admin_assignment',
                'shipment_id': shipment.pk,
                'barcode': tape.barcode,
                'courier': courier_profile.pk,
                'decision': 'approve',
                'comments': 'Approved for dispatch.',
            },
        )

        self.assertEqual(response.status_code, 302)
        shipment.refresh_from_db()
        self.assertEqual(shipment.status, 'Approved')
        self.assertTrue(shipment.tapes.filter(pk=tape.pk).exists())
        self.assertEqual(shipment.courier_name, 'Courier Guy')
        self.assertTrue(
            shipment.created_by and
            ShipmentApprovalHistory.objects.filter(shipment=shipment, action='Approved').exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(user=operator, action__icontains='approved').exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(user=courier_user, action__icontains='approved').exists()
        )

    def test_backup_admin_assignment_sends_email_to_profile_only_courier(self):
        backup_group = Group.objects.create(name='Backup Administrator')
        courier_group = Group.objects.create(name='Courier')

        operator = get_user_model().objects.create_user(
            username='operator-profile-email',
            email='operator-profile-email@example.com',
            password='pass1234',
        )
        operator.groups.add(backup_group)

        courier_profile = CourierProfile.objects.create(
            user=None,
            courier_id='CR-EMAIL-001',
            full_name='Profile Courier',
            email='profile-courier@example.com',
            vehicle_number='KDA 321B',
            active_status=True,
        )

        tape = Tape.objects.create(
            volser='TAPE-888',
            barcode='BAR-888',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Vault B',
        )

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='Nairobi Branch',
            status='Pending',
            releasing_custodian='Op User',
            created_by=operator,
        )

        self.client.force_login(operator)
        with patch('inventory.views.send_mail') as mock_send_mail:
            response = self.client.post(
                reverse('backup-dashboard'),
                {
                    'form_type': 'backup_admin_assignment',
                    'shipment_id': shipment.pk,
                    'barcode': tape.barcode,
                    'courier': courier_profile.pk,
                    'decision': 'approve',
                    'comments': 'Approved for dispatch.',
                },
            )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(mock_send_mail.called)
        args, kwargs = mock_send_mail.call_args
        self.assertIn('Shipment', args[0])
        self.assertEqual(args[3], ['profile-courier@example.com'])
        self.assertEqual(kwargs.get('fail_silently'), True)

        # Confirm notification payload includes actionable pickup URL metadata.
        self.assertTrue(
            AuditLog.objects.filter(
                name='Shipment Assigned',
                message__contains='target_url=',
            ).exists()
        )

    def test_assignment_form_adds_notification_payload_for_courier_profile(self):
        backup_group = Group.objects.create(name='Backup Administrator')
        courier_group = Group.objects.create(name='Courier')

        operator = get_user_model().objects.create_user(
            username='operator-profile-notify',
            email='operator-profile-notify@example.com',
            password='pass1234',
        )
        operator.groups.add(backup_group)

        courier_profile = CourierProfile.objects.create(
            user=None,
            courier_id='CR-NOTIFY-001',
            full_name='Notification Courier',
            email='notification-courier@example.com',
            vehicle_number='KDA 654B',
            active_status=True,
        )

        tape = Tape.objects.create(
            volser='TAPE-999',
            barcode='BAR-999',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Vault B',
        )

        shipment = Shipment.objects.create(
            shipment_type='Return',
            source_location='Nairobi Branch',
            status='Pending',
            releasing_custodian='Op User',
            created_by=operator,
        )

        self.client.force_login(operator)
        response = self.client.post(
            reverse('backup-dashboard'),
            {
                'form_type': 'backup_admin_assignment',
                'shipment_id': shipment.pk,
                'barcode': tape.barcode,
                'courier': courier_profile.pk,
                'decision': 'approve',
                'comments': 'Ready for pickup and return review.',
            },
        )

        self.assertEqual(response.status_code, 302)
        notification_log = AuditLog.objects.filter(name='Shipment Assigned', message__contains='target_url=').latest('timestamp')
        self.assertIsNotNone(notification_log)
        self.assertIn(reverse('pickup-confirmation', args=[shipment.pk]), notification_log.message)

    def test_supreme_approval_notifies_backup_admin_with_print_ready_link(self):
        backup_group = Group.objects.create(name='Backup Administrator')
        supreme_group = Group.objects.create(name='Supreme Approver')
        backup_admin = get_user_model().objects.create_user(
            username='backup-print-notify',
            email='backup-print-notify@example.com',
            password='pass1234',
            first_name='Backup',
            last_name='Admin',
        )
        backup_admin.groups.add(backup_group)
        supreme_approver = get_user_model().objects.create_user(
            username='supreme-print-notify',
            email='supreme-print-notify@example.com',
            password='pass1234',
            first_name='Supreme',
            last_name='Approver',
        )
        supreme_approver.groups.add(supreme_group)

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='Nairobi Branch',
            destination_location='Mombasa Branch',
            status='Pending',
            approval_stage='awaiting_supreme',
            approved_by_backup=backup_admin,
            created_by=backup_admin,
            last_updated_by=backup_admin,
            approval_remarks='Ready for final approval.',
        )

        self.client.force_login(supreme_approver)
        response = self.client.post(
            reverse('shipment-approvals'),
            {
                'form_type': 'backup_admin_decision',
                'shipment_id': shipment.pk,
                'decision': 'approve',
                'comments': 'Final approval granted.',
            },
        )

        self.assertEqual(response.status_code, 302)
        shipment.refresh_from_db()
        self.assertEqual(shipment.approval_stage, 'approved')
        self.assertEqual(shipment.status, 'Approved')
        notification = AuditLog.objects.filter(user=backup_admin, name='Shipment Approval Ready for Printing').latest('timestamp')
        self.assertIn('target_url=', notification.message)
        self.assertIn(reverse('approval-form-preview', args=[shipment.pk]), notification.message)

    def test_assignment_form_includes_courier_group_user_without_existing_profile(self):
        courier_group = Group.objects.create(name='Courier')
        courier_user = get_user_model().objects.create_user(
            username='carrier-no-profile',
            email='carrier-no-profile@example.com',
            password='pass1234',
            first_name='Carrier',
            last_name='User',
        )
        courier_user.groups.add(courier_group)

        form = BackupShipmentAssignmentForm()
        choices = dict(form.fields['courier'].choices)

        self.assertIn(f'user:{courier_user.pk}', choices)
        self.assertEqual(choices[f'user:{courier_user.pk}'], 'Carrier User')

    def test_assignment_form_includes_user_with_courier_role_even_without_group(self):
        courier_user = get_user_model().objects.create_user(
            username='role-courier',
            email='role-courier@example.com',
            password='pass1234',
            first_name='Courier',
            last_name='Role',
        )
        courier_user.role = 'courier'
        courier_user.save(update_fields=['role'])

        form = BackupShipmentAssignmentForm()
        choices = dict(form.fields['courier'].choices)

        self.assertIn(f'user:{courier_user.pk}', choices)
        self.assertEqual(choices[f'user:{courier_user.pk}'], 'Courier Role')

    def test_assignment_form_renders_courier_options_in_select(self):
        courier_group = Group.objects.create(name='Courier')
        courier_user = get_user_model().objects.create_user(
            username='rendered-courier',
            email='rendered-courier@example.com',
            password='pass1234',
            first_name='Rendered',
            last_name='Courier',
        )
        courier_user.groups.add(courier_group)

        form = BackupShipmentAssignmentForm()
        rendered = form['courier'].as_widget()

        self.assertIn('Rendered Courier', rendered)

    def test_assigned_shipments_page_shows_shipments_for_courier_user(self):
        courier_group = Group.objects.create(name='Courier')
        courier_user = get_user_model().objects.create_user(
            username='assigned-courier',
            email='assigned-courier@example.com',
            password='pass1234',
            first_name='Assigned',
            last_name='Courier',
        )
        courier_user.groups.add(courier_group)

        operator_user = get_user_model().objects.create_user(
            username='operator-assignee',
            email='operator-assignee@example.com',
            password='pass1234',
            first_name='Operator',
            last_name='Assignee',
        )

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='Nairobi Branch',
            destination_location='Mombasa Branch',
            status='Approved',
            releasing_custodian='Ops User',
            created_by=operator_user,
            courier_name='Assigned Courier',
            courier_contact='assigned-courier@example.com',
        )

        self.client.force_login(courier_user)
        response = self.client.get(reverse('assigned-shipments'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, shipment.shipment_id)
        self.assertContains(response, 'Mombasa Branch')

    def test_manifest_detail_renders_when_approval_user_is_missing(self):
        courier_group = Group.objects.create(name='Courier')
        courier_user = get_user_model().objects.create_user(
            username='manifest-no-approval',
            email='manifest-no-approval@example.com',
            password='pass1234',
            first_name='Manifest',
            last_name='Courier',
        )
        courier_user.groups.add(courier_group)

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='Nairobi Branch',
            destination_location='Mombasa Branch',
            status='Approved',
            releasing_custodian='Ops User',
        )

        self.client.force_login(courier_user)
        response = self.client.get(reverse('manifest-detail', args=[shipment.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Manifest Details')
        self.assertContains(response, 'N/A')

    def test_pickup_confirmation_works_for_courier_group_user_without_profile(self):
        courier_group = Group.objects.create(name='Courier')
        courier_user = get_user_model().objects.create_user(
            username='pickup-no-profile',
            email='pickup-no-profile@example.com',
            password='pass1234',
            first_name='Pickup',
            last_name='Courier',
        )
        courier_user.groups.add(courier_group)

        operator = get_user_model().objects.create_user(
            username='pickup-operator',
            email='pickup-operator@example.com',
            password='pass1234',
            first_name='Pickup',
            last_name='Operator',
        )

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='Nairobi Branch',
            status='Approved',
            releasing_custodian='Ops User',
            created_by=operator,
        )

        self.client.force_login(courier_user)
        response = self.client.post(
            reverse('pickup-confirmation', args=[shipment.pk]),
            {
                'manifest_reference': 'MANIFEST-2',
                'pickup_date': '2026-06-26',
                'pickup_time': '09:00',
                'pickup_location': 'Vault A',
                'notes': 'Pickup confirmed',
                'all_tapes_scanned': 'on',
                'manifest_verified': 'on',
                'tape_count_matched': 'on',
                'no_damaged_tapes': 'on',
                'custody_accepted': 'on',
            },
        )

        shipment.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(shipment.status, 'Picked Up')
        self.assertTrue(CourierProfile.objects.filter(user=courier_user).exists())

    def test_courier_dashboard_activity_log_shows_pickup_for_specific_user(self):
        courier_group = Group.objects.create(name='Courier')
        courier_user = get_user_model().objects.create_user(
            username='activity-courier',
            email='activity-courier@example.com',
            password='pass1234',
            first_name='Activity',
            last_name='Courier',
        )
        courier_user.groups.add(courier_group)

        operator = get_user_model().objects.create_user(
            username='activity-operator',
            email='activity-operator@example.com',
            password='pass1234',
            first_name='Activity',
            last_name='Operator',
        )

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='Nairobi Branch',
            status='Approved',
            releasing_custodian='Ops User',
            created_by=operator,
        )

        self.client.force_login(courier_user)
        self.client.post(
            reverse('pickup-confirmation', args=[shipment.pk]),
            {
                'manifest_reference': 'MANIFEST-ACT',
                'pickup_date': '2026-06-26',
                'pickup_time': '09:00',
                'pickup_location': 'Vault A',
                'notes': 'Pickup confirmed',
                'all_tapes_scanned': 'on',
                'manifest_verified': 'on',
                'tape_count_matched': 'on',
                'no_damaged_tapes': 'on',
                'custody_accepted': 'on',
            },
        )

        dashboard_response = self.client.get(reverse('courier-dashboard'))
        self.assertEqual(dashboard_response.status_code, 200)
        self.assertContains(dashboard_response, 'Picked Up')
        self.assertTrue(
            ShipmentTransportEvent.objects.filter(
                shipment=shipment,
                courier__user=courier_user,
                event_type='Picked Up',
            ).exists()
        )

    def test_courier_dashboard_shows_notification_alerts_to_courier(self):
        courier_group = Group.objects.create(name='Courier')
        courier_user = get_user_model().objects.create_user(
            username='courier-notify',
            email='courier-notify@example.com',
            password='pass1234',
            first_name='Courier',
            last_name='Notify',
        )
        courier_user.groups.add(courier_group)

        target_url = reverse('pickup-confirmation', args=[uuid.uuid4()])
        AuditLog.objects.create(
            name='Shipment Assigned',
            action='Shipment SH-12345 assigned to you.',
            message=f'target_url={target_url}',
            user=courier_user,
            severity='warning',
            is_read=False,
        )

        self.client.force_login(courier_user)
        response = self.client.get(reverse('courier-dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Notifications')
        self.assertContains(response, 'Shipment SH-12345 assigned to you.')
        self.assertContains(response, target_url)

    def test_pickup_confirmation_autofills_manifest_and_notifies_backup_admin(self):
        backup_group = Group.objects.create(name='Backup Administrator')
        backup_admin = get_user_model().objects.create_user(
            username='backup-notify',
            email='backup-notify@example.com',
            password='pass1234',
            first_name='Backup',
            last_name='Admin',
        )
        backup_admin.groups.add(backup_group)

        courier_group = Group.objects.create(name='Courier')
        courier_user = get_user_model().objects.create_user(
            username='pickup-courier',
            email='pickup-courier@example.com',
            password='pass1234',
            first_name='Pickup',
            last_name='Courier',
        )
        courier_user.groups.add(courier_group)

        courier_profile = CourierProfile.objects.create(
            user=courier_user,
            courier_id='CR-1001',
            full_name='Pickup Courier',
            email='pickup-courier@example.com',
            vehicle_number='KDA 123A',
            active_status=True,
        )

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='Nairobi Vault',
            destination_location='Kampala Branch',
            status='Approved',
            releasing_custodian='Ops User',
            created_by=backup_admin,
        )

        self.client.force_login(courier_user)
        get_response = self.client.get(reverse('pickup-confirmation', args=[shipment.pk]))
        self.assertEqual(get_response.status_code, 200)
        self.assertEqual(
            get_response.context['form'].initial['manifest_reference'],
            f'MANIFEST-{shipment.shipment_id[:8].upper()}',
        )
        self.assertEqual(get_response.context['form'].initial['pickup_location'], 'Nairobi Vault')

        post_response = self.client.post(
            reverse('pickup-confirmation', args=[shipment.pk]),
            {
                'manifest_reference': 'MANIFEST-001',
                'pickup_date': '2026-06-26',
                'pickup_time': '09:00',
                'pickup_location': 'Nairobi Vault',
                'notes': 'Pickup confirmed',
                'all_tapes_scanned': 'on',
                'manifest_verified': 'on',
                'tape_count_matched': 'on',
                'no_damaged_tapes': 'on',
                'custody_accepted': 'on',
            },
        )

        self.assertEqual(post_response.status_code, 302)
        shipment.refresh_from_db()
        self.assertEqual(shipment.vehicle_number, 'KDA 123A')
        self.assertTrue(
            AuditLog.objects.filter(
                user=backup_admin,
                action__icontains='MANIFEST-001',
            ).filter(
                action__icontains='KDA 123A',
            ).exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(
                user=backup_admin,
                message__icontains='target_url=',
            ).exists()
        )

    def test_backup_admin_can_reject_pending_shipment_with_comment_and_notify_operator(self):
        operations_group = Group.objects.create(name='Operations Manager')
        backup_group = Group.objects.create(name='Backup Administrator')

        operator = get_user_model().objects.create_user(
            username='operator-reject',
            email='operator-reject@example.com',
            password='pass1234',
            first_name='Op',
            last_name='User',
        )
        operator.groups.add(operations_group)

        backup_admin = get_user_model().objects.create_user(
            username='backup-reject',
            email='backup-reject@example.com',
            password='pass1234',
            first_name='Backup',
            last_name='Admin',
        )
        backup_admin.groups.add(backup_group)

        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='Nairobi Branch',
            status='Pending',
            releasing_custodian='Op User',
            created_by=operator,
        )

        self.client.force_login(backup_admin)
        response = self.client.post(
            reverse('backup-dashboard'),
            {
                'form_type': 'backup_admin_assignment',
                'shipment_id': shipment.pk,
                'submit_action': 'reject',
                'comments': 'Please resubmit with the full shipment details.',
            },
        )

        self.assertEqual(response.status_code, 302)
        shipment.refresh_from_db()
        self.assertEqual(shipment.status, 'Rejected')
        self.assertEqual(shipment.approval_remarks, 'Please resubmit with the full shipment details.')
        self.assertTrue(
            ShipmentApprovalHistory.objects.filter(shipment=shipment, action='Rejected').exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(user=operator, action__icontains='rejected').exists()
        )
        self.assertTrue(
            AuditLog.objects.filter(user=operator, action__icontains='Please resubmit').exists()
        )


class ShipmentOperationsWorkflowTests(TestCase):
    def test_operations_manager_can_receive_tapes_and_mark_shipment_completed(self):
        operations_group = Group.objects.create(name='Operations Manager')
        operator = get_user_model().objects.create_user(
            username='ops-complete',
            email='ops-complete@example.com',
            password='pass1234',
            first_name='Ops',
            last_name='User',
        )
        operator.groups.add(operations_group)
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            source_location='Vault A',
            destination_location='DR Site',
            receiving_organization='Ops Team',
            status='Approved',
            priority_level='High',
            releasing_custodian='Jane Doe',
        )

        self.client.force_login(operator)
        response = self.client.post(
            reverse('shipment-detail', args=[shipment.pk]),
            {
                'form_type': 'operator_receipt_completion',
                'receiving_custodian': 'Ops User',
                'receipt_notes': 'Tapes received and verified at the receiving site.',
            },
        )

        self.assertEqual(response.status_code, 200)
        shipment.refresh_from_db()
        self.assertEqual(shipment.status, 'Completed')
        self.assertEqual(shipment.received_by, 'Ops User')
        self.assertEqual(shipment.delivery_status, 'Delivered')
        self.assertEqual(shipment.delivery_notes, 'Tapes received and verified at the receiving site.')


class InventoryImportTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username='admin-import',
            email='admin-import@example.com',
            password='pass1234',
        )

    def test_excel_upload_creates_pending_import_approval(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Inventory'
        sheet.append(['volser', 'barcode', 'tape_type', 'status', 'current_location', 'retention_end_date', 'manufacturer'])
        sheet.append(['ABC123', 'BAR001', 'LTO-8', 'Active', 'Room A', '2030-01-01', 'IBM'])
        sheet.append(['XYZ999', 'BAR002', 'LTO-9', 'Damaged', 'Room B', '2029-12-31', 'Quantum'])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        self.client.force_login(self.user)
        response = self.client.post(
            reverse('backup-dashboard'),
            {
                'form_type': 'upload_inventory_excel',
                'inventory_file': SimpleUploadedFile('inventory.xlsx', buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(PendingApproval.objects.filter(transaction_type='Import Inventory Excel', requester=self.user, status='Pending').exists())
        self.assertEqual(Tape.objects.count(), 0)


class InventoryReportExportTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username='admin',
            email='admin@example.com',
            password='pass1234',
        )

    def test_custody_reports_page_renders_without_template_error(self):
        Shipment.objects.create(
            shipment_date=date(2026, 6, 15),
            shipment_type='Off-Site Transfer',
            source_location='Vault A',
            destination_location='Vault B',
            releasing_custodian='Alice',
            receiving_custodian='Bob',
            approval_remarks='Transfer approved',
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse('backup-dashboard'),
            {
                'show_reports': '1',
                'report_category': 'custody',
                'report_period': '2026-06',
                'report_type': 'monthly',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Previous Custodian')
        self.assertContains(response, 'Transfer Date')

    def test_reconciliation_reports_page_renders_with_table_controls(self):
        Reconciliation.objects.create(
            reconciliation_date=date(2026, 6, 15),
            location='Vault A',
            status='Completed',
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse('backup-dashboard'),
            {
                'show_reports': '1',
                'report_category': 'reconciliation',
                'report_period': '2026-06',
                'report_type': 'monthly',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Reconciliation ID')
        self.assertContains(response, 'Search')

    def test_inventory_reports_page_renders_without_template_error(self):
        self.client.force_login(self.user)
        response = self.client.get(
            reverse('backup-dashboard'),
            {
                'show_reports': '1',
                'report_category': 'inventory',
                'report_period': '2026-06',
                'report_type': 'monthly',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Month')
        self.assertContains(response, 'Export CSV')

    def test_inventory_report_export_includes_selected_month_data(self):
        Tape.objects.create(
            volser='ABC123',
            barcode='BAR001',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Room A',
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse('backup-dashboard'),
            {
                'show_reports': '1',
                'report_category': 'inventory',
                'report_period': '2026-06',
                'report_type': 'monthly',
                'export_csv': '1',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        self.assertContains(response, 'ABC123')
        self.assertContains(response, 'BAR001')
        self.assertContains(response, 'VolSER')

    def test_inventory_report_export_accepts_non_numeric_export_flag(self):
        Tape.objects.create(
            volser='ABC123',
            barcode='BAR001',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Room A',
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse('backup-dashboard'),
            {
                'show_reports': 'reports',
                'report_category': 'inventory',
                'report_period_inventory': '2026-06',
                'report_type': 'monthly',
                'export_csv_inventory': 'csv',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        self.assertContains(response, 'ABC123')

    def test_inventory_search_controls_are_scoped_to_inventory_table(self):
        Tape.objects.create(
            volser='ABC123',
            barcode='BAR001',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Room A',
        )
        Tape.objects.create(
            volser='XYZ999',
            barcode='BAR002',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Room B',
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse('backup-dashboard'),
            {
                'show_reports': '1',
                'report_category': 'inventory',
                'report_period': '2026-06',
                'report_type': 'monthly',
                'report_search_inventory': 'ABC123',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ABC123')
        self.assertNotContains(response, 'XYZ999')

    def test_inventory_pdf_export_uses_scoped_export_flag(self):
        Tape.objects.create(
            volser='ABC123',
            barcode='BAR001',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Room A',
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse('backup-dashboard'),
            {
                'show_reports': '1',
                'report_category': 'inventory',
                'report_period_inventory': '2026-06',
                'report_type': 'monthly',
                'export_pdf_inventory': '1',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_inventory_excel_export_uses_scoped_export_flag(self):
        Tape.objects.create(
            volser='ABC123',
            barcode='BAR001',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Room A',
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse('backup-dashboard'),
            {
                'show_reports': '1',
                'report_category': 'inventory',
                'report_period_inventory': '2026-06',
                'report_type': 'monthly',
                'export_excel_inventory': '1',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_inventory_report_share_sends_email_for_selected_report(self):
        Tape.objects.create(
            volser='ABC123',
            barcode='BAR001',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Room A',
        )

        self.client.force_login(self.user)
        with patch('inventory.views.EmailMessage') as mock_email_cls:
            mock_instance = mock_email_cls.return_value
            mock_instance.send.return_value = 1

            response = self.client.get(
                reverse('backup-dashboard'),
                {
                    'show_reports': '1',
                    'report_category': 'inventory',
                    'report_period_inventory': '2026-06',
                    'report_type': 'monthly',
                    'share_report': '1',
                    'share_email': 'recipient@example.com',
                },
            )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(mock_email_cls.called)
        self.assertEqual(mock_email_cls.call_args.args[2], self.user.email)
        self.assertEqual(mock_email_cls.call_args.args[3], ['recipient@example.com'])
        self.assertTrue(mock_instance.attach.called)
        mock_instance.send.assert_called_once_with(fail_silently=False)

    def test_inventory_report_share_handles_smtp_errors_without_500(self):
        Tape.objects.create(
            volser='ABC123',
            barcode='BAR001',
            tape_type='LTO-8',
            retention_end_date=date(2030, 1, 1),
            status='Active',
            current_location='Room A',
        )

        self.client.force_login(self.user)
        with patch('inventory.views.EmailMessage') as mock_email_cls:
            mock_instance = mock_email_cls.return_value
            mock_instance.send.side_effect = Exception('Authentication Required')

            response = self.client.get(
                reverse('backup-dashboard'),
                {
                    'show_reports': '1',
                    'report_category': 'inventory',
                    'report_period_inventory': '2026-06',
                    'report_type': 'monthly',
                    'share_report': '1',
                    'share_email': 'recipient@example.com',
                },
            )

        self.assertEqual(response.status_code, 302)
        follow_response = self.client.get(response.url)
        self.assertContains(follow_response, 'Report sharing failed')


class ShipmentDetailConfirmationTests(TestCase):
    def test_confirm_shipment_review_marks_shipment_as_completed(self):
        user = get_user_model().objects.create_superuser(
            username='shipment-reviewer',
            email='shipment-reviewer@example.com',
            password='StrongPass123!',
        )
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Approved',
            source_location='Acacia',
            destination_location='Acacia',
            releasing_custodian='Mukendi Olivier',
            receiving_organization='Pending review',
            courier_name='mars brunos',
            tracking_number='TRK-SHP-CD5A',
            created_by=user,
        )

        self.client.force_login(user)
        response = self.client.post(
            reverse('shipment-detail', args=[shipment.pk]),
            {'form_type': 'confirm_shipment_review'},
        )

        self.assertEqual(response.status_code, 200)
        shipment.refresh_from_db()
        self.assertEqual(shipment.status, 'Completed')
        self.assertEqual(shipment.delivery_status, 'Delivered')
        self.assertTrue(AuditLog.objects.filter(name='Shipment Review Confirmed').exists())

    def test_confirm_shipment_review_creates_notification_with_target_url(self):
        user = get_user_model().objects.create_superuser(
            username='shipment-review-notify',
            email='shipment-review-notify@example.com',
            password='StrongPass123!',
        )
        shipment = Shipment.objects.create(
            shipment_type='Off-Site Transfer',
            status='Approved',
            source_location='Acacia',
            destination_location='Acacia',
            releasing_custodian='Mukendi Olivier',
            receiving_organization='Pending review',
            courier_name='mars brunos',
            tracking_number='TRK-SHP-CD6B',
            created_by=user,
        )

        self.client.force_login(user)
        response = self.client.post(
            reverse('shipment-detail', args=[shipment.pk]),
            {'form_type': 'confirm_shipment_review'},
        )

        self.assertEqual(response.status_code, 200)
        audit_log = AuditLog.objects.filter(name='Shipment Review Confirmed').latest('timestamp')
        self.assertEqual(audit_log.user, user)
        self.assertIn('target_url=', audit_log.message)
        self.assertEqual(audit_log.severity, 'warning')

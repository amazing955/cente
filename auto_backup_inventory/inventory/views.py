from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group
from django.conf import settings
import base64
import inspect
import logging
import random
import re
import string
from urllib.parse import urlencode
from django.core import signing
from django.core.cache import cache
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.mail import EmailMessage, send_mail
from django.core.paginator import Paginator
from django.template.loader import render_to_string
import csv
import json  
import uuid
from io import BytesIO, StringIO
from datetime import date, datetime, timedelta
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from django.contrib import messages
from django.db import connection, transaction
from django.db.models import Q, Count
from django.http import Http404, HttpResponse, HttpResponseBadRequest, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone

from .authentication import AuditedJWTAuthentication
from .permissions import InvestigationPermission
from .serializer import AuditLogSerializer, ShipmentSerializer, TapeSerializer
from .utils import create_pending_approval

logger = logging.getLogger(__name__)
DASHBOARD_NAVIGATION_SALT = 'inventory-dashboard-navigation'
DASHBOARD_NAVIGATION_MAX_AGE = 1800


def _get_backup_administrator():
    UserModel = get_user_model()
    return UserModel.objects.filter(is_active=True, groups__name='Backup Administrator').first()


def _build_pending_tape_approval(request, transaction_type, summary, payload_fields, object_id='', priority='Medium', risk_level='Medium'):
    backup_admin = _get_backup_administrator()
    request_payload = {k: (str(v) if not isinstance(v, (dict, list)) else v) for k, v in payload_fields.items()}
    action_type = 'update' if object_id else 'create'
    requested_change = {
        'action': action_type,
        'model': 'Tape',
        'fields': request_payload,
    }
    if object_id:
        requested_change['object_id'] = str(object_id)
    return create_pending_approval(
        transaction_type=transaction_type,
        module='Inventory',
        summary=summary,
        requester=request.user,
        backup_administrator=backup_admin,
        priority=priority,
        risk_level=risk_level,
        status='Pending',
        request_payload=request_payload,
        requested_changes=[requested_change],
        related_model='tape',
        related_object_id=str(object_id) if object_id else '',
        audit_history=[{
            'action': 'Requested',
            'user': request.user.get_full_name() or request.user.username,
            'timestamp': timezone.localtime().isoformat(),
        }],
    )


def _build_pending_inventory_import_approval(request, workbook, file_name='inventory.xlsx'):
    requested_changes = _parse_inventory_excel_workbook(workbook)
    if not requested_changes:
        return None

    backup_admin = _get_backup_administrator()
    request_payload = {
        'file_name': file_name,
        'row_count': len(requested_changes),
    }

    return create_pending_approval(
        transaction_type='Import Inventory Excel',
        module='Inventory',
        summary=f'Import {len(requested_changes)} tape rows from {file_name}',
        requester=request.user,
        backup_administrator=backup_admin,
        priority='Medium',
        risk_level='High',
        status='Pending',
        request_payload=request_payload,
        requested_changes=requested_changes,
        related_model='tape',
        related_object_id='',
        audit_history=[{
            'action': 'Requested',
            'user': request.user.get_full_name() or request.user.username,
            'timestamp': timezone.localtime().isoformat(),
        }],
    )


def _parse_inventory_excel_workbook(workbook):
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
    header_map = {normalize_excel_column_name(name): idx for idx, name in enumerate(headers) if name}
    changes = []
    for row in rows[1:]:
        if not any(cell not in (None, '') for cell in row):
            continue
        values = {
            normalize_excel_column_name(header): (row[idx] if idx < len(row) else '')
            for header, idx in ((header, idx) for idx, header in enumerate(headers) if header)
        }
        volser = str(values.get('volser', '')).strip()
        barcode = str(values.get('barcode', '')).strip()
        tape_type = str(values.get('tape_type', '')).strip()
        status = str(values.get('status', '')).strip() or 'Active'
        current_location = str(values.get('current_location', '')).strip()
        retention_end_date_value = values.get('retention_end_date', '')
        manufacturer = str(values.get('manufacturer', '')).strip()

        if not volser or not barcode or not tape_type:
            continue

        try:
            retention_end_date = parse_date(str(retention_end_date_value)) if retention_end_date_value not in (None, '') else date.today() + timedelta(days=365)
        except Exception:
            retention_end_date = date.today() + timedelta(days=365)

        normalized_fields = {}
        for field_name, field_value in values.items():
            if field_value in (None, ''):
                normalized_fields[field_name] = ''
                continue
            if field_name == 'retention_end_date':
                try:
                    parsed_date = parse_date(str(field_value).strip())
                    normalized_fields[field_name] = parsed_date.isoformat() if parsed_date else date.today().isoformat()
                except Exception:
                    normalized_fields[field_name] = date.today().isoformat()
                continue
            if isinstance(field_value, datetime):
                normalized_fields[field_name] = field_value.isoformat()
                continue
            if isinstance(field_value, date) and not isinstance(field_value, datetime):
                normalized_fields[field_name] = field_value.isoformat()
                continue
            normalized_fields[field_name] = str(field_value).strip()

        normalized_fields['status'] = normalized_fields.get('status') or 'Active'
        if normalized_fields['status'] not in dict(Tape.STATUS_CHOICES):
            normalized_fields['status'] = 'Active'

        existing_tape = Tape.objects.filter(volser__iexact=volser).first()
        change_payload = {
            'action': 'update' if existing_tape else 'create',
            'model': 'Tape',
            'fields': normalized_fields,
        }
        if existing_tape:
            change_payload['object_id'] = str(existing_tape.pk)
        changes.append(change_payload)
    return changes


def _create_bulk_tape_import_pending_approval(request, workbook, file_name='inventory_import.xlsx'):
    requested_changes = _parse_inventory_excel_workbook(workbook)
    if not requested_changes:
        return None
    backup_admin = _get_backup_administrator()
    request_payload = {
        'file_name': file_name,
        'row_count': len(requested_changes),
    }
    return create_pending_approval(
        transaction_type='Bulk Tape Import',
        module='Inventory',
        summary=f'Bulk tape import from {file_name} ({len(requested_changes)} rows)',
        requester=request.user,
        backup_administrator=backup_admin,
        priority='High',
        risk_level='High',
        status='Pending',
        request_payload=request_payload,
        requested_changes=requested_changes,
        related_model='tape',
        related_object_id='',
        audit_history=[{
            'action': 'Requested Bulk Tape Import',
            'user': request.user.get_full_name() or request.user.username,
            'timestamp': timezone.localtime().isoformat(),
        }],
    )


def _create_schema_sync_pending_approval(request, pending_preview):
    file_bytes = base64.b64decode(pending_preview.get('file_bytes', ''))
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    new_columns = pending_preview.get('preview', {}).get('new_columns', [])
    if not new_columns:
        return None
    sql_statements = []
    for column in new_columns:
        column_name = column['name']
        column_type = column['detected_type']
        if not is_safe_column_name(column_name):
            return None
        sql_statements.append(f'ALTER TABLE inventory_tape ADD COLUMN IF NOT EXISTS {column_name} {column_type};')
    requested_changes = [
        {
            'action': 'schema_sync',
            'model': 'Tape',
            'sql_statements': sql_statements,
            'import_changes': _parse_inventory_excel_workbook(workbook),
        }
    ]
    backup_admin = _get_backup_administrator()
    request_payload = {
        'file_name': pending_preview.get('file_name', 'unknown.xlsx'),
        'new_columns': new_columns,
        'sql_statements': sql_statements,
    }
    return create_pending_approval(
        transaction_type='Tape Schema Synchronization',
        module='Inventory',
        summary=f'Schema synchronization for {pending_preview.get("file_name", "inventory upload")}',
        requester=request.user,
        backup_administrator=backup_admin,
        priority='Critical',
        risk_level='Critical',
        status='Pending',
        request_payload=request_payload,
        requested_changes=requested_changes,
        related_model='tape',
        related_object_id='',
        audit_history=[{
            'action': 'Requested Tape Schema Synchronization',
            'user': request.user.get_full_name() or request.user.username,
            'timestamp': timezone.localtime().isoformat(),
        }],
    )


def _build_investigation_log_entry(exception, request):
    return {
        'timestamp': timezone.now().isoformat(),
        'message': f"Exception investigation requested for {exception.exception_id}",
        'severity': 'info',
        'source': 'API',
        'actor': getattr(request.user, 'username', None) or 'anonymous',
    }


def _build_reconciliation_request_message(exception_id, requester_name, comment):
    encoded_parts = [
        f"exception_id={exception_id or ''}",
        f"requester_name={requester_name or ''}",
        f"comment={comment or ''}",
    ]
    return '|'.join(encoded_parts)


def _parse_reconciliation_request_message(message):
    payload = {}
    for part in (message or '').split('|'):
        if '=' not in part:
            continue
        key, value = part.split('=', 1)
        payload[key] = value
    return payload


def _parse_auditlog_metadata(message):
    metadata = {}
    for part in (message or '').split('|'):
        if '=' not in part:
            continue
        key, value = part.split('=', 1)
        metadata[key] = value
    return metadata


def build_signed_dashboard_navigation_token(feature_key, params=None, dashboard='backup'):
    payload = {
        'feature': feature_key,
        'dashboard': dashboard,
        'params': params or {},
    }
    return signing.dumps(payload, salt=DASHBOARD_NAVIGATION_SALT, compress=True)


def build_dashboard_navigation_url(feature_key, params=None, dashboard='backup'):
    signed_token = build_signed_dashboard_navigation_token(feature_key, params=params, dashboard=dashboard)
    if dashboard == 'operations':
        return reverse('operations-dashboard-navigation', kwargs={'signed_token': signed_token})
    return reverse('backup-dashboard-navigation', kwargs={'signed_token': signed_token})


def _apply_signed_dashboard_navigation(request, payload):
    feature_key = payload.get('feature')
    params = payload.get('params') or {}
    if not isinstance(params, dict):
        logger.warning('Dashboard navigation token had invalid params payload', extra={'path': request.path})
        return HttpResponseBadRequest('Dashboard navigation link is invalid.')

    if not feature_key:
        logger.warning('Dashboard navigation token was missing a feature key', extra={'path': request.path})
        return HttpResponseBadRequest('Dashboard navigation link is invalid.')

    if not has_dashboard_feature_access(request.user, feature_key):
        raise PermissionDenied

    request.GET = request.GET.copy()
    request.GET['feature_key'] = feature_key
    request.POST = request.POST.copy() if request.method == 'POST' else request.POST
    if request.method == 'POST':
        request.POST['feature_key'] = feature_key
    request.signed_dashboard_navigation = True

    for key, value in params.items():
        if value is None:
            continue
        request.GET[key] = str(value)
        if request.method == 'POST':
            request.POST[key] = str(value)

    return None


@login_required(login_url='signin')
def backup_dashboard_navigation(request, signed_token):
    try:
        payload = signing.loads(signed_token, salt=DASHBOARD_NAVIGATION_SALT, max_age=DASHBOARD_NAVIGATION_MAX_AGE)
    except signing.SignatureExpired:
        logger.warning('Dashboard navigation token expired', extra={'path': request.path})
        return HttpResponseBadRequest('Dashboard navigation link has expired.')
    except signing.BadSignature:
        logger.warning('Dashboard navigation token failed verification', extra={'path': request.path})
        return HttpResponseBadRequest('Dashboard navigation link is invalid.')
    except Exception as exc:
        logger.warning('Dashboard navigation token could not be parsed', extra={'path': request.path, 'error': str(exc)})
        return HttpResponseBadRequest('Dashboard navigation link is invalid.')

    if not isinstance(payload, dict):
        logger.warning('Dashboard navigation token payload was malformed', extra={'path': request.path})
        return HttpResponseBadRequest('Dashboard navigation link is invalid.')

    invalid_payload = _apply_signed_dashboard_navigation(request, payload)
    if invalid_payload is not None:
        return invalid_payload

    return backup_dashboard(request)


@login_required(login_url='signin')
def operations_dashboard_navigation(request, signed_token):
    try:
        payload = signing.loads(signed_token, salt=DASHBOARD_NAVIGATION_SALT, max_age=DASHBOARD_NAVIGATION_MAX_AGE)
    except signing.SignatureExpired:
        logger.warning('Operations dashboard navigation token expired', extra={'path': request.path})
        return HttpResponseBadRequest('Dashboard navigation link has expired.')
    except signing.BadSignature:
        logger.warning('Operations dashboard navigation token failed verification', extra={'path': request.path})
        return HttpResponseBadRequest('Dashboard navigation link is invalid.')
    except Exception as exc:
        logger.warning('Operations dashboard navigation token could not be parsed', extra={'path': request.path, 'error': str(exc)})
        return HttpResponseBadRequest('Dashboard navigation link is invalid.')

    if not isinstance(payload, dict):
        logger.warning('Operations dashboard navigation token payload was malformed', extra={'path': request.path})
        return HttpResponseBadRequest('Dashboard navigation link is invalid.')

    invalid_payload = _apply_signed_dashboard_navigation(request, payload)
    if invalid_payload is not None:
        return invalid_payload

    return operations_dashboard(request)


def _parse_exception_metadata(alert):
    exception_id = None
    description = None

    if not alert:
        return None, None

    if alert.message:
        metadata = _parse_auditlog_metadata(alert.message)
        exception_id = metadata.get('exception_id')
        description = metadata.get('description') or metadata.get('exception_description')

    if not exception_id and alert.action:
        match = re.search(r'exception\s+([A-Z0-9\-]+)', alert.action, re.IGNORECASE)
        if match:
            exception_id = match.group(1)

    if exception_id:
        exception = ShipmentException.objects.filter(exception_id=exception_id).first()
        if exception:
            description = description or getattr(exception, 'description', None)

    return exception_id, description


def _build_forwarded_notification_context(alert):
    metadata = _parse_auditlog_metadata(alert.message)
    exception_id = metadata.get('exception_id')
    target_url = metadata.get('target_url')
    if not target_url:
        if exception_id:
            target_url = reverse('investigation-dashboard', kwargs={'exception_id': exception_id})
        else:
            target_url = reverse('investigation-dashboard')

    description = metadata.get('description') or metadata.get('exception_description') or alert.action or alert.message or alert.name
    return {
        'id': str(alert.id),
        'action': alert.action or alert.name or 'Forwarded exception alert',
        'exception_id': exception_id,
        'description': description,
        'target_url': target_url,
        'timestamp': alert.timestamp,
        'severity': alert.severity,
        'message': alert.message,
    }


def investigation_dashboard_page(request, exception_id=None):
    requester_name = ''
    forwarded_notifications = []
    if getattr(request.user, 'is_authenticated', False):
        requester_name = request.user.get_full_name() or request.user.username
        forwarded_alerts = AuditLog.objects.filter(user=request.user, name='Exception Alert Forwarded').order_by('-timestamp')[:12]
        filtered_alerts = []
        for alert in forwarded_alerts:
            metadata = _parse_auditlog_metadata(alert.message)
            exception_id = metadata.get('exception_id')
            if not exception_id:
                filtered_alerts.append(alert)
                continue

            exception_obj = ShipmentException.objects.filter(exception_id=exception_id).first()
            if exception_obj is None or exception_obj.status != 'Closed':
                filtered_alerts.append(alert)

        forwarded_notifications = [_build_forwarded_notification_context(alert) for alert in filtered_alerts]
    return render(request, 'DRdashboard.html', {
        'exception_id': exception_id or '',
        'requester_name': requester_name,
        'reconciliation_requested': request.GET.get('reconciliation_requested') == '1',
        'forwarded_notifications': forwarded_notifications,
    })


@login_required(login_url='signin')
def initiate_reconciliation_request(request):
    if request.method != 'POST':
        return redirect('investigation-dashboard')

    if not (request.user.is_superuser or is_dr_team(request.user)):
        messages.error(request, 'Only DR Team members can initiate reconciliation requests.')
        return redirect('investigation-dashboard')

    exception_id = (request.POST.get('exception_id') or '').strip()
    requester_name = (request.POST.get('requester_name') or request.user.get_full_name() or request.user.username).strip()
    comment = (request.POST.get('comment') or '').strip()

    message = _build_reconciliation_request_message(exception_id, requester_name, comment)
    AuditLog.objects.create(
        name='Reconciliation Requested',
        action=f'Reconciliation requested for exception {exception_id or "unknown"}',
        message=message,
        user=request.user,
        severity='warning',
    )

    messages.success(request, 'Reconciliation request sent to the backup administrator.')
    redirect_url = reverse('investigation-dashboard', kwargs={'exception_id': exception_id}) if exception_id else reverse('investigation-dashboard')
    return redirect(f'{redirect_url}?reconciliation_requested=1')


@login_required(login_url='signin')
def close_exception(request):
    if request.method != 'POST':
        return redirect('investigation-dashboard')

    if not (request.user.is_superuser or is_dr_team(request.user)):
        messages.error(request, 'Only DR Team members can request exception closure.')
        return redirect('investigation-dashboard')

    exception_id = (request.POST.get('exception_id') or '').strip()
    investigation_results = (request.POST.get('investigation_results') or '').strip()

    if not exception_id:
        messages.error(request, 'Exception ID is required.')
        return redirect('investigation-dashboard')

    try:
        exception = ShipmentException.objects.get(exception_id=exception_id)
    except ShipmentException.DoesNotExist:
        messages.error(request, f'Exception {exception_id} not found.')
        return redirect('investigation-dashboard')

    # Create a close request instead of directly closing
    close_request = ExceptionCloseRequest.objects.create(
        exception=exception,
        requested_by=request.user,
        investigation_results=investigation_results,
        status='Pending'
    )

    # Create audit log entry
    AuditLog.objects.create(
        name='Exception Close Requested',
        action=f'Close request initiated for exception {exception_id}',
        message=f'investigator={request.user.get_full_name() or request.user.username}|results_preview={investigation_results[:100]}...',
        user=request.user,
        severity='warning',
    )

    # Notify backup administrators
    backup_admin_group = Group.objects.filter(name='Backup Administrator').first()
    if backup_admin_group:
        admins = backup_admin_group.user_set.filter(is_active=True)
        # In a real scenario, you'd send emails here using Django's email backend
        # For now, we'll just create audit logs for the admins to see
        for admin in admins:
            AuditLog.objects.create(
                name='Exception Close Request Notification',
                action=f'Close request {close_request.id} awaiting approval for exception {exception_id}',
                message=f'requested_by={request.user.get_full_name() or request.user.username}|close_request_id={close_request.id}',
                user=admin,
                severity='warning',
            )

    messages.success(request, f'Exception closure request submitted for admin review. Reference ID: {close_request.id}')
    redirect_url = reverse('investigation-dashboard', kwargs={'exception_id': exception_id})
    return redirect(redirect_url)


@login_required(login_url='signin')
def approve_close_exception(request, close_request_id):
    """Admin endpoint to approve a close exception request.
    GET: render approval form where admin can add a remark and choose approve/reject.
    POST: process the approval or rejection.
    """
    # Check if user is backup admin
    if not (request.user.is_superuser or is_backup_administrator(request.user)):
        messages.error(request, 'Only backup administrators can approve exception closures.')
        return redirect('backup-dashboard')

    try:
        close_request = ExceptionCloseRequest.objects.get(id=close_request_id)
    except ExceptionCloseRequest.DoesNotExist:
        messages.error(request, 'Close request not found.')
        return redirect('backup-dashboard')

    if request.method == 'GET':
        # Render a simple approval form for the admin
        return render(request, 'approve_close_request.html', {
            'close_request': close_request,
        })

    if close_request.status != 'Pending':
        messages.error(request, f'This request has already been {close_request.status.lower()}.')
        return redirect('backup-dashboard')

    approval_comment = (request.POST.get('approval_comment') or '').strip()
    action = (request.POST.get('action') or 'approve').strip().lower()

    if action == 'approve':
        # Update close request status
        close_request.status = 'Approved'
        close_request.approval_comment = approval_comment
        close_request.approved_by = request.user
        close_request.reviewed_at = timezone.now()
        close_request.save()

        # Update exception status to Closed
        close_request.exception.status = 'Closed'
        close_request.exception.save(update_fields=['status'])

        # Create audit log entry
        AuditLog.objects.create(
            name='Exception Closed',
            action=f'Exception {close_request.exception.exception_id} closed by admin approval',
            message=f'approved_by={request.user.get_full_name() or request.user.username}|approval_comment={approval_comment}',
            user=request.user,
            severity='success',
        )

        messages.success(request, f'Exception {close_request.exception.exception_id} has been closed and marked as completed.')
    
    elif action == 'reject':
        # Reject the close request
        close_request.status = 'Rejected'
        close_request.approval_comment = approval_comment
        close_request.approved_by = request.user
        close_request.reviewed_at = timezone.now()
        close_request.save()

        # Create audit log entry
        AuditLog.objects.create(
            name='Exception Close Request Rejected',
            action=f'Close request rejected for exception {close_request.exception.exception_id}',
            message=f'rejected_by={request.user.get_full_name() or request.user.username}|reason={approval_comment}',
            user=request.user,
            severity='warning',
        )

        messages.warning(request, f'Exception closure request for {close_request.exception.exception_id} has been rejected.')

    return redirect('backup-dashboard')



@login_required(login_url='signin')
def approval_form_preview(request, shipment_pk):
    """Render a printable approval form preview for a shipment.
    Accessible to backup administrators, the backup approver, the shipment creator, or superusers.
    """
    shipment = get_object_or_404(Shipment.objects.prefetch_related('tapes'), pk=shipment_pk)

    # Permission check: allow backup admins, the backup approver, the creator, or superusers
    allowed = (
        request.user.is_superuser or
        is_backup_administrator(request.user) or
        (getattr(shipment, 'approved_by_backup', None) == request.user) or
        (getattr(shipment, 'created_by', None) == request.user)
    )
    if not allowed:
        raise PermissionDenied

    # support alternative printable voucher format via query `?format=voucher`
    # generate asset description from package/tapes when not explicitly provided
    try:
        tapes = list(shipment.tapes.all())
    except Exception:
        tapes = []

    # prefer an explicit description if present (some Shipment models may not have a `description` field)
    desc = getattr(shipment, 'description', None) or getattr(shipment, 'approval_remarks', None)
    if desc:
        asset_description = desc
    elif tapes:
        count = len(tapes)
        src = (getattr(shipment, 'source_location', '') or 'Unknown').title()
        dst = (getattr(shipment, 'destination_location', '') or 'Unknown').title()
        # concise phrasing: "<N> tapes from BranchName to Destination Name"
        asset_description = f"{count} tape{'s' if count != 1 else ''} from {src} to {dst}"
    else:
        asset_description = 'No tapes assigned.'

    context = {
        'shipment': shipment,
        'asset_description': asset_description,
    }

    if request.GET.get('format') == 'voucher':
        return render(request, 'approval_form_voucher.html', context)

    return render(request, 'approval_form_preview.html', context)


def _authenticate_api_user(request):
    if getattr(request.user, 'is_authenticated', False):
        return request.user

    auth_backend = AuditedJWTAuthentication()
    try:
        auth_result = auth_backend.authenticate(request)
    except Exception:
        return None

    if not auth_result:
        return None

    user, _ = auth_result
    request.user = user
    return user


def _require_authenticated_api_user(request, require_investigation_role=False):
    user = _authenticate_api_user(request)
    if not user:
        return None, JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)

    if require_investigation_role and not InvestigationPermission().has_permission(request, None):
        return user, JsonResponse({'detail': 'You do not have permission to perform this action.'}, status=403)

    return user, None


def exception_investigation_view(request, exception_id):
    if request.method not in {'GET'}:
        return JsonResponse({'detail': 'Only GET requests are allowed.'}, status=405)

    user, auth_error = _require_authenticated_api_user(request, require_investigation_role=True)
    if auth_error is not None:
        return auth_error

    exception_obj = ShipmentException.objects.select_related(
        'shipment', 'shipment__requesting_branch', 'shipment__created_by', 'shipment__approved_by', 'shipment__last_updated_by',
        'tape', 'reported_by'
    ).filter(exception_id=exception_id).first()

    if not exception_obj:
        return JsonResponse({'detail': 'Exception not found.'}, status=404)

    if exception_obj.status not in {'Open', 'Investigating'}:
        return JsonResponse({'detail': 'Exception is not open or under investigation.'}, status=400)

    shipment = exception_obj.shipment
    tape = exception_obj.tape

    transport_events = list(
        ShipmentTransportEvent.objects.select_related('courier', 'courier__user').filter(shipment=shipment).order_by('-event_date', '-event_time', '-created_at')
    )
    receipts = list(
        ShipmentReceipt.objects.select_related('courier', 'courier__user').filter(shipment=shipment).order_by('-created_at')
    )
    approvals = list(
        ShipmentApprovalHistory.objects.select_related('user').filter(shipment=shipment).order_by('-created_at')
    )
    related_exceptions = list(
        ShipmentException.objects.select_related('tape', 'reported_by').filter(shipment=shipment).exclude(pk=exception_obj.pk).order_by('-reported_date')
    )
    audit_logs = list(
        AuditLog.objects.select_related('user').filter(action__icontains=exception_obj.exception_id).order_by('-timestamp')[:20]
    )

    timeline = []
    timeline.append({
        'type': 'exception_reported',
        'timestamp': exception_obj.reported_date.isoformat() if exception_obj.reported_date else None,
        'message': f"Exception {exception_obj.exception_id} reported for shipment {shipment.shipment_id}",
        'actor': getattr(exception_obj.reported_by, 'username', None) or 'system',
    })
    for event in transport_events:
        timeline.append({
            'type': 'transport_event',
            'timestamp': datetime.combine(event.event_date, event.event_time).isoformat() if event.event_date and event.event_time else None,
            'message': f"{event.event_type} - {event.comments or 'No notes'}",
            'actor': getattr(event.courier, 'full_name', None) or 'courier',
        })
    for receipt in receipts:
        timeline.append({
            'type': 'receipt',
            'timestamp': receipt.created_at.isoformat() if receipt.created_at else None,
            'message': f"Receipt captured: {receipt.notes or 'No notes'}",
            'actor': getattr(receipt.courier, 'full_name', None) or 'courier',
        })
    for approval in approvals:
        timeline.append({
            'type': 'approval',
            'timestamp': approval.created_at.isoformat() if approval.created_at else None,
            'message': approval.comments or 'Approval history updated',
            'actor': getattr(approval.user, 'username', None) or 'system',
        })

    # Add explicit pickup and delivery timeline entries when available
    try:
        # detect a pickup transport event (common event_type labels)
        pickup_event = next((e for e in transport_events if e.event_type and e.event_type.lower() in ('picked up', 'pickup', 'pickup_confirmed', 'picked_up')), None)
        if pickup_event:
            pickup_ts = None
            if getattr(pickup_event, 'event_date', None) and getattr(pickup_event, 'event_time', None):
                pickup_ts = datetime.combine(pickup_event.event_date, pickup_event.event_time).isoformat()
            timeline.append({
                'type': 'picked_up',
                'timestamp': pickup_ts,
                'message': f"Shipment picked up - {pickup_event.comments or 'No notes'}",
                'actor': getattr(pickup_event.courier, 'full_name', None) or 'courier',
            })

        # use shipment delivery_date/delivery_time if present
        if getattr(shipment, 'delivery_date', None) or getattr(shipment, 'delivery_time', None):
            delivered_ts = None
            if getattr(shipment, 'delivery_date', None) and getattr(shipment, 'delivery_time', None):
                delivered_ts = datetime.combine(shipment.delivery_date, shipment.delivery_time).isoformat()
            elif getattr(shipment, 'delivery_date', None):
                delivered_ts = datetime.combine(shipment.delivery_date, datetime.min.time()).isoformat()
            timeline.append({
                'type': 'delivered',
                'timestamp': delivered_ts,
                'message': f"Shipment delivered - {getattr(shipment, 'delivery_status', '') or 'Delivered'}",
                'actor': getattr(shipment, 'received_by', None) or 'operator',
            })
    except Exception:
        # defensive: don't break the investigation view if ancillary fields are missing
        pass

    chain_of_custody = []
    if shipment.releasing_custodian:
        chain_of_custody.append({'role': 'releasing_custodian', 'value': shipment.releasing_custodian})
    if shipment.receiving_custodian:
        chain_of_custody.append({'role': 'receiving_custodian', 'value': shipment.receiving_custodian})
    for receipt in receipts:
        chain_of_custody.append({
            'role': 'receipt',
            'value': receipt.pickup_location,
            'confirmed': receipt.custody_confirmed,
            'courier': getattr(receipt.courier, 'full_name', None),
        })
    for event in transport_events:
        chain_of_custody.append({
            'role': 'transport_event',
            'value': event.event_type,
            'courier': getattr(event.courier, 'full_name', None),
        })

    payload = {
        'exception': {
            'id': exception_obj.exception_id,
            'type': exception_obj.exception_type,
            'status': exception_obj.status,
            'severity': exception_obj.severity,
            'description': exception_obj.description,
            'reported_by': getattr(exception_obj.reported_by, 'username', None),
            'reported_date': exception_obj.reported_date.isoformat() if exception_obj.reported_date else None,
        },
        'tape': {
            'id': str(tape.id) if tape else None,
            'volser': tape.volser if tape else None,
            'barcode': tape.barcode if tape else None,
            'status': tape.status if tape else None,
            'current_location': tape.current_location if tape else None,
            'retention_end_date': tape.retention_end_date.isoformat() if tape and tape.retention_end_date else None,
        },
        'shipment': {
            'id': str(shipment.id),
            'shipment_id': shipment.shipment_id,
            'status': shipment.status,
            'shipment_type': shipment.shipment_type,
            'priority_level': shipment.priority_level,
            'source_location': shipment.source_location,
            'destination_location': shipment.destination_location,
            'requesting_branch': getattr(shipment.requesting_branch, 'branch_name', None),
            'releasing_custodian': shipment.releasing_custodian,
            'receiving_custodian': shipment.receiving_custodian,
            'tracking_number': shipment.tracking_number,
            'vehicle_number': shipment.vehicle_number,
            'created_by': getattr(shipment.created_by, 'username', None),
            'approved_by': getattr(shipment.approved_by, 'username', None),
        },
        'timeline': sorted(timeline, key=lambda item: item['timestamp'] or '', reverse=True),
        'chain_of_custody': chain_of_custody,
        'logs': [
            {
                'timestamp': entry.timestamp.isoformat() if entry.timestamp else None,
                'message': entry.message,
                'severity': entry.severity,
                'user': getattr(entry.user, 'username', None),
            }
            for entry in audit_logs
        ],
        'audit_logs': [
            {
                'id': str(entry.id),
                'timestamp': entry.timestamp.isoformat() if entry.timestamp else None,
                'name': entry.name,
                'action': entry.action,
                'message': entry.message,
                'severity': entry.severity,
                'user': getattr(entry.user, 'username', None),
            }
            for entry in audit_logs
        ],
        'user_interactions': [
            {
                'username': getattr(exception_obj.reported_by, 'username', None),
                'role': getattr(exception_obj.reported_by, 'role', None),
                'assigned_branch': getattr(getattr(exception_obj.reported_by, 'assigned_branch', None), 'branch_name', None),
            }
        ],
        'notifications': [
            {
                'type': 'audit_log',
                'message': 'Investigation access recorded',
            }
        ],
        'related_exceptions': [
            {
                'id': exc.exception_id,
                'type': exc.exception_type,
                'status': exc.status,
                'severity': exc.severity,
            }
            for exc in related_exceptions
        ],
        'users': [
            {
                'username': getattr(exception_obj.reported_by, 'username', None),
                'role': getattr(exception_obj.reported_by, 'role', None),
                'branch': getattr(getattr(exception_obj.reported_by, 'assigned_branch', None), 'branch_name', None),
            },
            {
                'username': getattr(shipment.created_by, 'username', None),
                'role': getattr(shipment.created_by, 'role', None),
                'branch': getattr(getattr(shipment.created_by, 'assigned_branch', None), 'branch_name', None),
            },
        ],
        'branch_details': {
            'branch_name': getattr(shipment.requesting_branch, 'branch_name', None),
            'branch_code': getattr(shipment.requesting_branch, 'branch_code', None),
            'region': getattr(shipment.requesting_branch, 'region', None),
        },
        'courier_details': {
            'name': getattr(transport_events[0].courier, 'full_name', None) if transport_events else None,
            'vehicle_number': getattr(transport_events[0].courier, 'vehicle_number', None) if transport_events else None,
        },
    }

    AuditLog.objects.create(
        name='Exception Investigation',
        action=f'Investigated exception {exception_obj.exception_id}',
        user=user,
        message='Exception investigation API accessed',
        severity='info',
    )

    return JsonResponse(payload)


def custom_exception_handler(exc, context):
    from rest_framework.views import exception_handler

    response = exception_handler(exc, context)
    if response is None:
        return None
    return response


def custom_page_not_found(request, exception=None):
    return render(request, '404.html', status=404)


def custom_permission_denied(request, exception=None):
    return render(request, '403.html', status=403)


def custom_bad_request(request, exception=None):
    return render(request, '400.html', status=400)


def custom_server_error(request):
    return render(request, '500.html', status=500)

from django.utils.dateparse import parse_date


# REST API helpers used by the dashboard and external integrations.
def api_dashboard_summary(request):
    user, auth_error = _require_authenticated_api_user(request)
    if auth_error is not None:
        return auth_error

    summary = {
        'tape_count': Tape.objects.count(),
        'shipment_count': Shipment.objects.count(),
        'audit_log_count': AuditLog.objects.count(),
        'pending_shipments': Shipment.objects.filter(status__iexact='Pending').count(),
    }
    return JsonResponse(summary)


def api_tape_list(request):
    user, auth_error = _require_authenticated_api_user(request)
    if auth_error is not None:
        return auth_error

    if request.method == 'POST':
        try:
            payload = json.loads(request.body.decode('utf-8')) if request.body else {}
        except json.JSONDecodeError:
            return JsonResponse({'detail': 'Invalid JSON payload.'}, status=400)

        form = AddTapeForm(payload)
        if form.is_valid():
            # Do not commit immediately — create a PendingApproval for backup review
            data = form.cleaned_data
            payload_fields = {k: (str(v) if not isinstance(v, (dict, list)) else v) for k, v in data.items()}
            pa = _build_pending_tape_approval(
                request,
                transaction_type='Add Tape',
                summary=f"Register tape {payload_fields.get('volser') or payload_fields.get('barcode')}",
                payload_fields=payload_fields,
                priority='Medium',
                risk_level='High',
            )
            if pa:
                return JsonResponse({'result': 'pending', 'pending_id': str(pa.pk)}, status=202)
            return JsonResponse({'detail': 'Failed to create pending approval.'}, status=500)

        return JsonResponse({'detail': 'Validation failed.', 'errors': form.errors}, status=400)

    tapes = Tape.objects.order_by('-date_registered')[:20]
    payload = {
        'count': tapes.count(),
        'results': [TapeSerializer(tape).data for tape in tapes],
    }
    return JsonResponse(payload)


def api_shipment_list(request):
    user, auth_error = _require_authenticated_api_user(request)
    if auth_error is not None:
        return auth_error

    shipments = Shipment.objects.order_by('-shipment_date', '-created_at')[:20]
    payload = {
        'count': shipments.count(),
        'results': [ShipmentSerializer(shipment).data for shipment in shipments],
    }
    return JsonResponse(payload)


def api_audit_log_list(request):
    user, auth_error = _require_authenticated_api_user(request)
    if auth_error is not None:
        return auth_error

    logs = AuditLog.objects.order_by('-timestamp')[:20]
    payload = {
        'count': logs.count(),
        'results': [AuditLogSerializer(log).data for log in logs],
    }
    return JsonResponse(payload)


def build_audit_log_queryset(request):
    queryset = AuditLog.objects.select_related('user').all().order_by('-timestamp')

    search = (request.GET.get('search') or '').strip()
    log_type = (request.GET.get('log_type') or '').strip()
    module = (request.GET.get('module') or '').strip()
    severity = (request.GET.get('severity') or '').strip()
    user_filter = (request.GET.get('user') or '').strip()
    status = (request.GET.get('status') or '').strip()
    date_range = (request.GET.get('date_range') or '').strip()
    date_from = (request.GET.get('date_from') or '').strip()
    date_to = (request.GET.get('date_to') or '').strip()
    per_page = int(request.GET.get('per_page') or 25)

    if search:
        queryset = queryset.filter(
            Q(name__icontains=search)
            | Q(action__icontains=search)
            | Q(message__icontains=search)
            | Q(user__username__icontains=search)
        )
    if log_type:
        queryset = queryset.filter(action__icontains=log_type)
    if module:
        queryset = queryset.filter(name__icontains=module)
    if severity:
        queryset = queryset.filter(severity__iexact=severity)
    if user_filter:
        queryset = queryset.filter(user__username__icontains=user_filter)
    if status:
        queryset = queryset.filter(action__icontains=status)

    if date_range:
        today = timezone.localdate()
        if date_range == 'today':
            queryset = queryset.filter(timestamp__date=today)
        elif date_range == 'yesterday':
            queryset = queryset.filter(timestamp__date=today - timedelta(days=1))
        elif date_range == 'last_7_days':
            start = today - timedelta(days=7)
            queryset = queryset.filter(timestamp__date__gte=start, timestamp__date__lte=today)
        elif date_range == 'last_30_days':
            start = today - timedelta(days=30)
            queryset = queryset.filter(timestamp__date__gte=start, timestamp__date__lte=today)
        elif date_range == 'this_month':
            queryset = queryset.filter(timestamp__year=today.year, timestamp__month=today.month)
        elif date_range == 'this_year':
            queryset = queryset.filter(timestamp__year=today.year)

    if date_from:
        queryset = queryset.filter(timestamp__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(timestamp__date__lte=date_to)

    return queryset, {
        'search': search,
        'log_type': log_type,
        'module': module,
        'severity': severity,
        'user': user_filter,
        'status': status,
        'date_range': date_range,
        'date_from': date_from,
        'date_to': date_to,
        'per_page': per_page,
    }


def export_audit_logs(queryset, export_format):
    headers = ['Timestamp', 'User', 'Log Type', 'Module', 'Status', 'Severity', 'Action']
    rows = []
    for entry in queryset:
        rows.append([
            entry.timestamp.strftime('%Y-%m-%d %H:%M:%S') if entry.timestamp else '',
            entry.user.username if entry.user else 'System',
            entry.log_type,
            entry.module,
            entry.status or '-',
            entry.severity or 'info',
            entry.action or '',
        ])

    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="audit_logs.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    if export_format == 'excel':
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Audit Logs'
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="audit_logs.xlsx"'
        return response

    if export_format == 'pdf':
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesizes=landscape(letter), title='Audit Logs')
        styles = getSampleStyleSheet()
        table_rows = [headers] + rows
        table = Table(table_rows, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        doc.build([Paragraph('Audit Logs', styles['Title']), Spacer(1, 12), table])
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="audit_logs.pdf"'
        return response

    return None


def api_feature_navigation(request, feature_key):
    if not request.user.is_authenticated:
        return JsonResponse({'detail': 'Authentication required.'}, status=401)

    if not has_dashboard_feature_access(request.user, feature_key):
        return JsonResponse({'detail': 'Feature access denied.'}, status=403)

    feature = next((item for item in get_dashboard_feature_catalog() if item['key'] == feature_key), None)
    if not feature:
        return JsonResponse({'detail': 'Feature not found.'}, status=404)

    dashboard = 'operations' if feature.get('scope') == 'operations' else 'backup'
    signed_token = build_signed_dashboard_navigation_token(feature_key, {f'show_{feature_key}': '1'}, dashboard=dashboard)
    target_url = build_dashboard_navigation_url(feature_key, {f'show_{feature_key}': '1'}, dashboard=dashboard)

    return JsonResponse({
        'feature_key': feature_key,
        'target_url': target_url,
        'fragment_url': target_url,
        'load_mode': 'page',
        'scope': feature.get('scope'),
    })

from .forms import *
from .models import *


def is_valid_uuid(value):
    if not value:
        return False
    try:
        uuid.UUID(str(value))
        return True
    except (ValueError, TypeError, AttributeError):
        return False


def get_object_by_uuid_pk(model, pk_value):
    if not is_valid_uuid(pk_value):
        return None
    return model.objects.filter(pk=pk_value).first()


def _resolve_courier_selection(courier_selection):
    courier_profile = None
    courier_user = None
    courier_name = ''
    courier_contact = ''

    if not courier_selection:
        return courier_profile, courier_user, courier_name, courier_contact

    if isinstance(courier_selection, str):
        if courier_selection.startswith('profile:'):
            profile_id = courier_selection.split(':', 1)[1]
            courier_profile = CourierProfile.objects.filter(pk=profile_id).first()
        elif courier_selection.startswith('user:'):
            user_id = courier_selection.split(':', 1)[1]
            courier_user = get_user_model().objects.filter(pk=user_id).first()
            courier_profile = getattr(courier_user, 'courier_profile', None) if courier_user else None
        else:
            courier_profile = CourierProfile.objects.filter(pk=courier_selection).first()
            if not courier_profile:
                courier_user = get_user_model().objects.filter(pk=courier_selection).first()
                courier_profile = getattr(courier_user, 'courier_profile', None) if courier_user else None

    if courier_profile:
        courier_name = courier_profile.full_name
        courier_contact = courier_profile.phone_number or courier_profile.email or ''
    elif courier_user:
        courier_name = courier_user.get_full_name() or courier_user.username
        courier_contact = courier_user.email or ''

    return courier_profile, courier_user, courier_name, courier_contact


def feature_module(request, feature_key):
    if not request.user.is_authenticated:
        return redirect('signin')

    if not has_dashboard_feature_access(request.user, feature_key):
        raise PermissionDenied

    feature = next((item for item in get_dashboard_feature_catalog() if item['key'] == feature_key), None)
    if not feature:
        raise Http404

    configured_url = feature.get('url')
    if configured_url:
        return redirect(configured_url)

    configured_url_name = feature.get('url_name')
    if configured_url_name:
        try:
            return redirect(reverse(configured_url_name, kwargs=feature.get('url_kwargs', {})))
        except Exception:
            pass

    def render_feature_view(view_func):
        request.GET = request.GET.copy()
        request.GET['feature_key'] = feature_key
        request.POST = request.POST.copy() if request.method == 'POST' else request.POST
        if request.method == 'POST':
            request.POST['feature_key'] = feature_key
        return inspect.unwrap(view_func)(request)

    if feature_key == 'shipment_approvals':
        return render_feature_view(shipment_approvals)

    if feature_key == 'warehouse_operations_dashboard':
        return render_feature_view(warehouse_operations_dashboard)

    if feature_key == 'add_tape':
        return render_feature_view(add_tape)

    if is_backup_administrator(request.user) and feature_key in {
        'backup_dashboard', 'add_tape', 'tape_inventory', 'shipments', 'reconciliation', 'reports', 'audit_logs'
    }:
        return render_feature_view(backup_dashboard)

    if is_operations_manager(request.user) and feature_key in {
        'operations_dashboard', 'shipment_approvals', 'exception_management'
    }:
        return render_feature_view(operations_dashboard)

    if is_warehouse_staff(request.user) and feature_key == 'warehouse_operations_dashboard':
        return render_feature_view(warehouse_operations_dashboard)

    if is_it_compliance_auditor(request.user) and feature_key in {
        'audit_logs', 'reports', 'exception_management', 'operations_dashboard'
    }:
        return render_feature_view(audit_logs_view)

    if is_courier(request.user):
        return render_feature_view(assigned_shipments)

    return render(request, 'dashboard.html', {'feature_key': feature_key, 'dashboard_features': []})


def get_feature_panel_state(feature_key):
    if not feature_key:
        return {}

    mapping = {
        'add_tape': {'show_add_tape_panel': True},
        'tape_inventory': {'show_tape_inventory_panel': True},
        'shipments': {'show_shipments_panel': True},
        'reconciliation': {'show_reconciliation_panel': True},
        'reports': {'show_reports_panel': True},
        'audit_logs': {'show_audit_panel': True},
        'analytics': {'show_analytics_panel': True},
        'shipment_monitoring': {'show_monitoring_panel': True},
        'exception_management': {'show_exception_panel': True},
        'custody_governance': {'show_custody_panel': True},
        'reconciliation_review': {'show_reconciliation_panel': True},
        'compliance_monitoring': {'show_compliance_panel': True},
        'backup_dashboard': {},
        'operations_dashboard': {},
    }
    return mapping.get(feature_key, {})


def is_courier(user):
    return user.is_authenticated and (
        user.is_superuser or
        user.groups.filter(name='Courier').exists() or
        getattr(user, 'courier_profile', None) is not None
    )


def get_courier_profile(user):
    return getattr(user, 'courier_profile', None)


def ensure_courier_profile(user):
    profile = get_courier_profile(user)
    if profile:
        return profile

    if not user or not user.is_authenticated:
        return None

    profile = CourierProfile.objects.filter(user=user).first()
    if profile:
        return profile

    full_name = user.get_full_name() or user.username or 'Courier User'
    courier_id = f'CR-{uuid.uuid4().hex[:8].upper()}'
    profile = CourierProfile.objects.create(
        user=user,
        courier_id=courier_id,
        full_name=full_name,
        phone_number='',
        email=user.email or '',
        vehicle_number='Not provided',
        active_status=True,
    )
    return profile


def get_courier_shipments(user):
    courier = get_courier_profile(user)
    if not user.is_authenticated:
        return Shipment.objects.none()

    queryset = Shipment.objects.filter(
        Q(created_by=user) |
        Q(approved_by=user)
    ).distinct()

    if courier:
        queryset = queryset | Shipment.objects.filter(
            Q(courier_name__iexact=courier.full_name) |
            Q(courier_contact__iexact=courier.phone_number) |
            Q(receipts__courier=courier) |
            Q(deliveries__courier=courier)
        ).distinct()
    else:
        queryset = queryset | Shipment.objects.filter(
            Q(courier_name__icontains=user.get_full_name() or user.username) |
            Q(courier_contact__icontains=user.email)
        ).distinct()

    return queryset.order_by('-shipment_date')

ADMIN_FEATURE_TABS = [
    {
        'tab_id': 'overview',
        'label': 'Overview',
        'feature_keys': [
            'Inventory Overview',
            'View Tape Records',
            'Add Tape',
            'Edit Tape Details',
            'Scan Barcode/RFID',
            'Update Tape Location',
            'Mark Tape as Damaged',
            'Initiate Shipment Requests',
            'Perform Reconciliation',
            'View Inventory Reports',
            'View Audit History',
        ],
    },
    {
        'tab_id': 'users',
        'label': 'Users',
        'feature_keys': ['User Management'],
    },
    {
        'tab_id': 'roles',
        'label': 'Roles',
        'feature_keys': ['Security Controls', 'User Management'],
    },
    {
        'tab_id': 'inventory',
        'label': 'Tape Inventory',
        'feature_keys': [
            'Tape Management',
            'View Tape Records',
            'Add Tape',
            'Edit Tape Details',
            'Scan Barcode/RFID',
            'Update Tape Location',
            'Mark Tape as Damaged',
        ],
    },
    {
        'tab_id': 'shipments',
        'label': 'Shipments',
        'feature_keys': [
            'Shipment Tracking',
            'Initiate Shipment Requests',
            'Perform Reconciliation',
        ],
    },
    {
        'tab_id': 'reports',
        'label': 'Reports',
        'feature_keys': ['Reporting', 'View Inventory Reports'],
    },
    {
        'tab_id': 'audit',
        'label': 'Audit Logs',
        'feature_keys': ['Audit Logging', 'View Audit History'],
    },
    {
        'tab_id': 'approvals',
        'label': 'Approvals',
        'feature_keys': ['Approvals', 'User Management', 'Security Controls'],
    },
]

BACKUP_FEATURE_TABS = [
    {
        'tab_id': 'overview',
        'label': 'Overview',
        'feature_keys': [
            'Inventory Overview',
            'View Tape Records',
            'Add Tape',
            'Edit Tape Details',
            'Scan Barcode/RFID',
            'Update Tape Location',
            'Mark Tape as Damaged',
            'Initiate Shipment Requests',
            'Perform Reconciliation',
            'View Inventory Reports',
            'View Audit History',
        ],
    },
    {
        'tab_id': 'inventory',
        'label': 'Inventory',
        'feature_keys': [
            'Tape Management',
            'View Tape Records',
            'Add Tape',
            'Edit Tape Details',
            'Scan Barcode/RFID',
            'Update Tape Location',
            'Mark Tape as Damaged',
        ],
    },
    {
        'tab_id': 'shipments',
        'label': 'Shipments',
        'feature_keys': [
            'Shipment Tracking',
            'Initiate Shipment Requests',
            'Perform Reconciliation',
        ],
    },
    {
        'tab_id': 'audit',
        'label': 'Audit Logs',
        'feature_keys': ['Audit Logging', 'View Audit History'],
    },
    {
        'tab_id': 'approvals',
        'label': 'Approvals',
        'feature_keys': ['Approvals', 'User Management', 'Security Controls'],
    },
]

OPERATIONS_FEATURE_TABS = [
    {
        'tab_id': 'dashboard',
        'label': 'Dashboard',
    },
    {
        'tab_id': 'shipment_approvals',
        'label': 'Shipment Approvals',
    },
    {
        'tab_id': 'shipment_monitoring',
        'label': 'Shipment Monitoring',
    },
    {
        'tab_id': 'exception_management',
        'label': 'Exception Management',
    },
    {
        'tab_id': 'custody_governance',
        'label': 'Custody Governance',
    },
    {
        'tab_id': 'reconciliation_review',
        'label': 'Reconciliation Review',
    },
    {
        'tab_id': 'compliance_monitoring',
        'label': 'Compliance Monitoring',
    },
    {
        'tab_id': 'reports',
        'label': 'Reports',
    },
    {
        'tab_id': 'analytics',
        'label': 'Analytics',
    },
    {
        'tab_id': 'notifications',
        'label': 'Notifications',
    },
    {
        'tab_id': 'settings',
        'label': 'Settings',
    },
]


def unique_features(features):
    seen = set()
    ordered = []
    for feature in features:
        if feature not in seen:
            seen.add(feature)
            ordered.append(feature)
    return ordered


def get_dashboard_tabs(user, feature_tabs, preserve_empty_tabs=False):
    normalized_tabs = []
    user_features = getattr(user, 'feature_names', [])
    for tab in feature_tabs:
        normalized_tab = tab.copy()
        feature_keys = unique_features(tab.get('feature_keys', []))
        if not user.is_superuser:
            feature_keys = [feature for feature in feature_keys if feature in user_features]
        normalized_tab['feature_keys'] = feature_keys
        if user.is_superuser or preserve_empty_tabs or feature_keys:
            normalized_tabs.append(normalized_tab)
    return normalized_tabs


def get_last_six_month_counts(tapes_queryset):
    today = timezone.localdate()

    def month_start(base_date, offset_months):
        year = base_date.year + (base_date.month - 1 + offset_months) // 12
        month = (base_date.month - 1 + offset_months) % 12 + 1
        return date(year, month, 1)

    month_starts = [month_start(today, offset) for offset in range(-5, 1)]
    labels = [month.strftime('%b') for month in month_starts]
    counts = []
    for month in month_starts:
        next_month = month_start(month, 1)
        counts.append(
            tapes_queryset.filter(
                date_registered__gte=month,
                date_registered__lt=next_month
            ).count()
        )
    return labels, counts


def get_first_day_of_month(month_string):
    if not month_string:
        return None
    try:
        year, month = [int(part) for part in month_string.split('-')]
        return date(year, month, 1)
    except (ValueError, TypeError):
        return None


def get_report_categories():
    return [
        {'slug': 'inventory', 'name': 'Inventory Report', 'description': 'Tape inventory status, counts, and current holdings.'},
        {'slug': 'shipment', 'name': 'Shipment Report', 'description': 'Shipment volume, status, and delivery performance.'},
        {'slug': 'custody', 'name': 'Custody Report', 'description': 'Custody transfer, acceptance, and compliance metrics.'},
        {'slug': 'reconciliation', 'name': 'Reconciliation Report', 'description': 'Reconciliation sessions, discrepancies, and resolution progress.'},
        {'slug': 'retention', 'name': 'Retention Report', 'description': 'Retention expiry and retention action counts for the month.'},
        {'slug': 'compliance', 'name': 'Compliance Report', 'description': 'Compliance alerts, audit readiness, and control checks.'},
        {'slug': 'exception', 'name': 'Exception Report', 'description': 'Exception counts, issue categories, and unresolved incidents.'},
        {'slug': 'audit_trail', 'name': 'Audit Trail Report', 'description': 'Audit events, changed records, and system activity.'},
        {'slug': 'management_summary', 'name': 'Management Summary Report', 'description': 'Executive summary of key program metrics and trends.'},
    ]


def get_latest_custodian_for_tape(tape):
    latest_shipment = tape.shipments.order_by('-shipment_date', '-created_at').first()
    if not latest_shipment:
        return None
    if latest_shipment.receiving_custodian:
        return latest_shipment.receiving_custodian
    if latest_shipment.releasing_custodian:
        return latest_shipment.releasing_custodian
    return None


def get_next_month(first_day):
    if first_day.month == 12:
        return date(first_day.year + 1, 1, 1)
    return date(first_day.year, first_day.month + 1, 1)


def sort_report_rows(rows, sort_key=None, sort_order='asc'):
    if not sort_key:
        return rows

    def sort_value(row):
        value = row.get(sort_key)
        if value is None:
            return ''
        if hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d')
        return str(value).lower()

    return sorted(rows, key=sort_value, reverse=sort_order == 'desc')


def paginate_report_rows(request, rows, page_size=10, page_param='report_page'):
    paginator = Paginator(rows, page_size)
    page_number = request.GET.get(page_param, '1')
    page_obj = paginator.get_page(page_number)
    return paginator, page_obj


def get_scoped_report_param(request, report_category, param_name, default=''):
    if report_category:
        scoped_value = request.GET.get(f'{param_name}_{report_category}')
        if scoped_value is not None:
            return (scoped_value or '').strip()
    return (request.GET.get(param_name, default) or '').strip()


def get_scoped_report_flag(request, report_category, param_name):
    def is_enabled(value):
        if value is None:
            return False
        return str(value).strip().lower() not in {'', '0', 'false', 'no', 'off'}

    if report_category:
        scoped_value = request.GET.get(f'{param_name}_{report_category}')
        if scoped_value is not None:
            return is_enabled(scoped_value)
    return is_enabled(request.GET.get(param_name))


def generate_daily_report_data(report_date):
    return {
        'period': report_date.strftime('%Y-%m-%d'),
        'tapes_registered': Tape.objects.filter(date_registered=report_date).count(),
        'active_tapes': Tape.objects.filter(status='Active').count(),
        'off_site_tapes': Tape.objects.filter(status='Off-Site').count(),
        'missing_tapes': Tape.objects.filter(status='Missing').count(),
        'retention_due_today': Tape.objects.filter(retention_end_date=report_date).count(),
        'shipments_created': Shipment.objects.filter(shipment_date=report_date).count(),
        'shipments_pending': Shipment.objects.filter(shipment_date=report_date, status__iexact='Pending').count(),
        'alerts_generated': AuditLog.objects.filter(timestamp__date=report_date, severity__in=['warning', 'error']).count(),
        'reconciliations_conducted': Reconciliation.objects.filter(reconciliation_date=report_date).count(),
    }


def generate_monthly_report_data(report_month, report_category=None):
    start = report_month
    end = get_next_month(report_month)
    if report_category == 'shipment':
        return {
            'period': report_month.strftime('%Y-%m'),
            'shipments_created': Shipment.objects.filter(shipment_date__gte=start, shipment_date__lt=end).count(),
            'shipments_pending': Shipment.objects.filter(shipment_date__gte=start, shipment_date__lt=end, status__iexact='Pending').count(),
            'shipments_dispatched': Shipment.objects.filter(shipment_date__gte=start, shipment_date__lt=end, status__iexact='Dispatched').count(),
            'shipments_delivered': Shipment.objects.filter(shipment_date__gte=start, shipment_date__lt=end, status__iexact='Delivered').count(),
            'active_transfers': Shipment.objects.filter(shipment_date__gte=start, shipment_date__lt=end, status__in=['In Transit', 'Dispatched']).count(),
            'delay_risk': Shipment.objects.filter(shipment_date__gte=start, shipment_date__lt=end, status__in=['Pending', 'More Info Requested']).count(),
        }
    if report_category == 'custody':
        return {
            'period': report_month.strftime('%Y-%m'),
            'total_shipments': Shipment.objects.filter(shipment_date__gte=start, shipment_date__lt=end).count(),
            'transfers_in_progress': Shipment.objects.filter(shipment_date__gte=start, shipment_date__lt=end, status__in=['In Transit', 'Dispatched']).count(),
            'deliveries_completed': Shipment.objects.filter(shipment_date__gte=start, shipment_date__lt=end, status__iexact='Delivered').count(),
            'custody_confirmed': ShipmentReceipt.objects.filter(
                shipment__shipment_date__gte=start,
                shipment__shipment_date__lt=end,
                custody_confirmed=True,
            ).count(),
            'custody_accepted': ShipmentReceipt.objects.filter(
                shipment__shipment_date__gte=start,
                shipment__shipment_date__lt=end,
                custody_accepted=True,
            ).count(),
        }
    if report_category == 'reconciliation':
        return {
            'period': report_month.strftime('%Y-%m'),
            'reconciliations_conducted': Reconciliation.objects.filter(reconciliation_date__gte=start, reconciliation_date__lt=end).count(),
            'issues_found': ReconciliationResult.objects.filter(reconciliation__reconciliation_date__gte=start, reconciliation__reconciliation_date__lt=end).count(),
            'open_issues': ReconciliationResult.objects.filter(reconciliation__reconciliation_date__gte=start, reconciliation__reconciliation_date__lt=end, resolution_status__in=['Open', 'Under Investigation']).count(),
            'completed_reconciliations': Reconciliation.objects.filter(reconciliation_date__gte=start, reconciliation_date__lt=end, status='Completed').count(),
            'pending_reconciliations': Reconciliation.objects.filter(reconciliation_date__gte=start, reconciliation_date__lt=end).exclude(status='Completed').count(),
        }
    if report_category == 'retention':
        return {
            'period': report_month.strftime('%Y-%m'),
            'retention_due_this_month': Tape.objects.filter(retention_end_date__gte=start, retention_end_date__lt=end).count(),
            'retention_expired': Tape.objects.filter(retention_end_date__lt=end, retention_end_date__gte=start).count(),
            'retention_actions_required': Tape.objects.filter(retention_end_date__gte=start, retention_end_date__lt=end, status__in=['Retained', 'Active']).count(),
            'archived_tapes': Tape.objects.filter(status='Retained').count(),
        }
    if report_category == 'compliance':
        return {
            'period': report_month.strftime('%Y-%m'),
            'alerts_generated': AuditLog.objects.filter(timestamp__gte=start, timestamp__lt=end, severity__in=['warning', 'error']).count(),
            'audit_events': AuditLog.objects.filter(timestamp__gte=start, timestamp__lt=end).count(),
            'policy_exceptions': ReconciliationResult.objects.filter(reconciliation__reconciliation_date__gte=start, reconciliation__reconciliation_date__lt=end, resolution_status__in=['Open', 'Under Investigation']).count(),
            'compliance_reviewed': Reconciliation.objects.filter(reconciliation_date__gte=start, reconciliation_date__lt=end, status='Completed').count(),
        }
    if report_category == 'exception':
        return {
            'period': report_month.strftime('%Y-%m'),
            'missing_tapes': Tape.objects.filter(status='Missing').count(),
            'damaged_tapes': Tape.objects.filter(status='Damaged').count(),
            'open_shipments': Shipment.objects.filter(status__iexact='Pending').count(),
            'open_issues': ReconciliationResult.objects.filter(resolution_status__in=['Open', 'Under Investigation']).count(),
        }
    if report_category == 'audit_trail':
        return {
            'period': report_month.strftime('%Y-%m'),
            'audit_events': AuditLog.objects.filter(timestamp__gte=start, timestamp__lt=end).count(),
            'warnings': AuditLog.objects.filter(timestamp__gte=start, timestamp__lt=end, severity='warning').count(),
            'errors': AuditLog.objects.filter(timestamp__gte=start, timestamp__lt=end, severity='error').count(),
            'user_actions': AuditLog.objects.filter(timestamp__gte=start, timestamp__lt=end).exclude(user__isnull=True).count(),
        }
    if report_category == 'management_summary':
        return {
            'period': report_month.strftime('%Y-%m'),
            'total_tapes': Tape.objects.count(),
            'shipments_created': Shipment.objects.filter(shipment_date__gte=start, shipment_date__lt=end).count(),
            'reconciliations_conducted': Reconciliation.objects.filter(reconciliation_date__gte=start, reconciliation_date__lt=end).count(),
            'alerts_generated': AuditLog.objects.filter(timestamp__gte=start, timestamp__lt=end, severity__in=['warning', 'error']).count(),
            'retention_due_this_month': Tape.objects.filter(retention_end_date__gte=start, retention_end_date__lt=end).count(),
        }
    return {
        'period': report_month.strftime('%Y-%m'),
        'tapes_registered': Tape.objects.filter(date_registered__gte=start, date_registered__lt=end).count(),
        'active_tapes': Tape.objects.filter(status='Active').count(),
        'off_site_tapes': Tape.objects.filter(status='Off-Site').count(),
        'missing_tapes': Tape.objects.filter(status='Missing').count(),
        'retention_due_this_month': Tape.objects.filter(retention_end_date__gte=start, retention_end_date__lt=end).count(),
        'shipments_created': Shipment.objects.filter(shipment_date__gte=start, shipment_date__lt=end).count(),
        'shipments_pending': Shipment.objects.filter(shipment_date__gte=start, shipment_date__lt=end, status__iexact='Pending').count(),
        'alerts_generated': AuditLog.objects.filter(timestamp__gte=start, timestamp__lt=end, severity__in=['warning', 'error']).count(),
        'reconciliations_conducted': Reconciliation.objects.filter(reconciliation_date__gte=start, reconciliation_date__lt=end).count(),
    }


def normalize_export_value(value):
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.UTC).replace(tzinfo=None)
        return value
    if hasattr(value, 'date') and callable(value.date) and not isinstance(value, date):
        try:
            value = value.date()
        except Exception:
            pass
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    return value


def export_report_excel(report_category, report_period, rows, columns):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report_category.replace('_', ' ').title()
    sheet.append([column['label'] for column in columns])
    for row in rows:
        sheet.append([normalize_export_value(row.get(column['key'], '-')) for column in columns])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report_category}_report_{report_period}.xlsx"'
    return response


def export_report_pdf(report_category, report_period, rows, columns):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), title=f'{report_category.replace("_", " ").title()} Report')
    styles = getSampleStyleSheet()
    story = [Paragraph(f'{report_category.replace("_", " ").title()} Report - {report_period}', styles['Title']), Spacer(1, 12)]

    table_data = [[column['label'] for column in columns]]
    for row in rows:
        table_data.append([row.get(column['key'], '-') for column in columns])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_category}_report_{report_period}.pdf"'
    return response


def normalize_excel_column_name(name):
    text = str(name or '').strip().lower()
    text = re.sub(r'[^a-z0-9]+', '_', text)
    text = re.sub(r'_+', '_', text).strip('_')
    if not text:
        text = 'column'
    if text in {'select', 'insert', 'update', 'delete', 'drop', 'from', 'where', 'union', 'table', 'create', 'alter', 'truncate'} or text[0].isdigit():
        text = f'column_{text}'
    return text


def is_safe_column_name(name):
    normalized_name = normalize_excel_column_name(name)
    return bool(re.fullmatch(r'[a-z][a-z0-9_]{0,62}', normalized_name)) and normalized_name not in {'select', 'insert', 'update', 'delete', 'drop', 'from', 'where', 'union', 'table', 'create', 'alter', 'truncate'}


def infer_postgres_column_type(values):
    sample_values = [value for value in values if value not in (None, '')][:10]
    if not sample_values:
        return 'VARCHAR(255)'

    lower_values = {str(value).strip().lower() for value in sample_values}
    if lower_values <= {'true', 'false', 'yes', 'no', 'y', 'n', '1', '0'}:
        return 'BOOLEAN'

    try:
        if all(str(value).strip().replace('-', '').isdigit() for value in sample_values):
            return 'INTEGER'
    except Exception:
        pass

    try:
        if all(float(str(value).strip()) for value in sample_values):
            return 'NUMERIC(12,2)'
    except Exception:
        pass

    try:
        for value in sample_values:
            date.fromisoformat(str(value).strip())
        return 'DATE'
    except Exception:
        pass

    try:
        for value in sample_values:
            datetime.fromisoformat(str(value).strip())
        return 'TIMESTAMP'
    except Exception:
        pass

    try:
        for value in sample_values:
            uuid.UUID(str(value).strip())
        return 'UUID'
    except Exception:
        pass

    if max(len(str(value)) for value in sample_values) > 255:
        return 'TEXT'
    return 'VARCHAR(255)'


def get_existing_table_columns(table_name):
    try:
        with connection.cursor() as cursor:
            if connection.vendor == 'postgresql':
                cursor.execute(
                    "SELECT column_name FROM information_schema.columns WHERE table_schema = current_schema() AND table_name = %s ORDER BY ordinal_position",
                    [table_name],
                )
                return [row[0] for row in cursor.fetchall()]
            return [column[1] for column in connection.introspection.get_columns(cursor, table_name)]
    except Exception:
        return []


def analyze_excel_schema(uploaded_file, table_name='inventory_tape'):
    if not uploaded_file:
        return None

    try:
        file_bytes = uploaded_file.read() if hasattr(uploaded_file, 'read') else b''
        if hasattr(uploaded_file, 'seek'):
            uploaded_file.seek(0)
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return {'error': f'Unable to read the Excel file: {exc}'}

    if not workbook.sheetnames:
        return {'error': 'The uploaded workbook does not contain any sheets.'}

    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return {'error': 'The uploaded sheet does not contain any inventory rows.'}

    headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
    normalized_headers = [normalize_excel_column_name(name) for name in headers]
    existing_columns = get_existing_table_columns(table_name)

    existing_normalized_columns = {normalize_excel_column_name(column) for column in existing_columns if column}
    duplicate_columns = []
    invalid_columns = []
    seen_names = set()
    for index, header in enumerate(headers):
        normalized_name = normalized_headers[index]
        if not header:
            invalid_columns.append({'header': header, 'normalized_name': normalized_name, 'issue': 'blank header'})
            continue
        if normalized_name in seen_names:
            duplicate_columns.append({'header': header, 'normalized_name': normalized_name})
        else:
            seen_names.add(normalized_name)
        if normalized_name in {'select', 'insert', 'update', 'delete', 'drop', 'from', 'where', 'union', 'table', 'create', 'alter', 'truncate'} or normalized_name[0].isdigit():
            invalid_columns.append({'header': header, 'normalized_name': normalized_name, 'issue': 'unsafe name'})

    column_values = []
    for index, header in enumerate(headers):
        if not header:
            continue
        values = [row[index] if index < len(row) else None for row in rows[1:11]]
        column_values.append((header, normalized_headers[index], values))

    new_columns = []
    for header, normalized_name, values in column_values:
        if not normalized_name or normalized_name in existing_normalized_columns:
            continue
        if normalized_name in {'select', 'insert', 'update', 'delete', 'drop', 'from', 'where', 'union', 'table', 'create', 'alter', 'truncate'}:
            continue
        new_columns.append({
            'name': normalized_name,
            'source_header': header,
            'detected_type': infer_postgres_column_type(values),
        })

    removed_columns = [column for column in existing_columns if normalize_excel_column_name(column) not in set(normalized_headers) and normalize_excel_column_name(column) not in {'id', 'created_at', 'updated_at'}]
    renamed_columns = []
    for removed_column in removed_columns:
        normalized_removed = normalize_excel_column_name(removed_column)
        matches = [name for name in set(normalized_headers) if name and name != normalized_removed and (name.startswith(normalized_removed[:3]) or normalized_removed.startswith(name[:3]))]
        if matches:
            renamed_columns.append({'from': normalized_removed, 'to': matches[0]})

    return {
        'table_name': table_name,
        'existing_columns': existing_columns,
        'new_columns': new_columns,
        'removed_columns': removed_columns,
        'renamed_columns': renamed_columns,
        'duplicate_columns': duplicate_columns,
        'invalid_columns': invalid_columns,
        'has_changes': bool(new_columns) or bool(removed_columns) or bool(duplicate_columns) or bool(invalid_columns),
    }


def import_inventory_excel_file(uploaded_file=None, workbook=None, file_name=None):
    if not uploaded_file and workbook is None:
        return 0, 'Please choose an Excel file to import.', False

    try:
        if workbook is None:
            file_bytes = uploaded_file.read() if hasattr(uploaded_file, 'read') else b''
            if hasattr(uploaded_file, 'seek'):
                uploaded_file.seek(0)
            workbook = load_workbook(BytesIO(file_bytes), data_only=True)
        else:
            if isinstance(workbook, (bytes, bytearray)):
                workbook = load_workbook(BytesIO(workbook), data_only=True)
    except Exception as exc:
        return 0, f'Unable to read the Excel file: {exc}', False

    if not workbook.sheetnames:
        return 0, 'The uploaded workbook does not contain any sheets.', False

    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return 0, 'The uploaded sheet does not contain any inventory rows.', False

    headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
    header_map = {normalize_excel_column_name(name): idx for idx, name in enumerate(headers) if name}

    required_columns = ['volser', 'barcode', 'tape_type', 'status', 'current_location', 'retention_end_date']
    missing_columns = [column for column in required_columns if column not in header_map]
    if missing_columns:
        return 0, f'The uploaded sheet is missing required columns: {", ".join(missing_columns)}.', False

    imported_count = 0
    for row in rows[1:]:
        if not any(cell not in (None, '') for cell in row):
            continue

        values = {normalize_excel_column_name(header): (row[idx] if idx < len(row) else '') for header, idx in ((header, idx) for idx, header in enumerate(headers))}
        volser = str(values.get('volser', '')).strip()
        barcode = str(values.get('barcode', '')).strip()
        tape_type = str(values.get('tape_type', '')).strip()
        status = str(values.get('status', '')).strip() or 'Active'
        current_location = str(values.get('current_location', '')).strip()
        retention_end_date_value = values.get('retention_end_date', '')
        manufacturer = str(values.get('manufacturer', '')).strip()

        if not volser or not barcode or not tape_type:
            continue

        try:
            retention_end_date = parse_date(str(retention_end_date_value)) if retention_end_date_value not in (None, '') else date.today() + timedelta(days=365)
        except Exception:
            retention_end_date = date.today() + timedelta(days=365)

        Tape.objects.update_or_create(
            volser=volser,
            defaults={
                'barcode': barcode,
                'tape_type': tape_type,
                'status': status if status in dict(Tape.STATUS_CHOICES) else 'Active',
                'current_location': current_location,
                'retention_end_date': retention_end_date,
                'manufacturer': manufacturer,
            },
        )
        imported_count += 1

    return imported_count, f'Imported {imported_count} tape records from the Excel sheet.', True


def build_report_csv_bytes(report_category, report_period, rows, columns):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow([f'{report_category.replace("_", " ").title()} Report', report_period])
    writer.writerow([])
    writer.writerow([column['label'] for column in columns])
    for row in rows:
        writer.writerow([row.get(column['key'], '-') for column in columns])
    return buffer.getvalue().encode('utf-8')


def redirect_report_view(request):
    params = request.GET.copy()
    params.pop('share_report', None)
    params.pop('share_email', None)
    query_string = params.urlencode()
    return redirect(f"{request.path}?{query_string}" if query_string else request.path)


def send_report_share_email(request, report_category, report_period, rows, columns, recipients):
    subject = f'{report_category.replace("_", " ").title()} Report Shared'
    sender_email = request.user.email or settings.DEFAULT_FROM_EMAIL
    body = (
        f'Hello,\n\n'
        f'This report for {report_period} was shared from the Backup Administrator Dashboard.\n\n'
        f'Shared by: {request.user.get_full_name() or request.user.username}\n'
        f'Sender email: {sender_email}\n'
    )
    attachment_name = f'{report_category}_report_{report_period}.csv'
    csv_bytes = build_report_csv_bytes(report_category, report_period, rows, columns)
    msg = EmailMessage(subject, body, sender_email, recipients)
    msg.attach(attachment_name, csv_bytes, 'text/csv')
    try:
        msg.send(fail_silently=False)
        return True
    except Exception as exc:
        messages.error(request, f'Report sharing failed: {exc}')
        return False


def export_inventory_report_csv(report_period, tapes):
    response = HttpResponse(content_type='text/csv')
    filename = f"inventory_report_{report_period}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(['Inventory Report', report_period])
    writer.writerow([])
    writer.writerow(['VolSER', 'Barcode', 'RFID Tag', 'Tape Type', 'Status', 'Current Location', 'Custodian', 'Retention End Date', 'Date Registered'])
    for tape in tapes:
        writer.writerow([
            tape.volser,
            tape.barcode,
            tape.rfid_tag or '-',
            tape.tape_type,
            tape.status,
            tape.current_location or '-',
            tape.latest_custodian or '-',
            tape.retention_end_date.strftime('%Y-%m-%d') if tape.retention_end_date else '-',
            tape.date_registered.strftime('%Y-%m-%d') if tape.date_registered else '-',
        ])
    return response


def export_report_csv(report_type, report_period, report_data):
    response = HttpResponse(content_type='text/csv')
    filename = f"{report_type}_report_{report_period}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow([f'{report_type.title()} Report', report_period])
    writer.writerow([])
    writer.writerow(['Metric', 'Value'])
    for key, value in report_data.items():
        if key == 'period':
            continue
        writer.writerow([key.replace('_', ' ').title(), value])
    return response


def export_reconciliation_report_csv(reconciliations, summary=None):
    response = HttpResponse(content_type='text/csv')
    filename = 'reconciliation_reports_export.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(['Reconciliation Report Export'])
    writer.writerow([timezone.localdate().strftime('%Y-%m-%d')])
    writer.writerow([])

    if summary:
        writer.writerow(['Summary'])
        for key, value in summary.items():
            writer.writerow([key.replace('_', ' ').title(), value])
        writer.writerow([])

    writer.writerow(['Reconciliation ID', 'Date', 'Location', 'Status', 'Performed By', 'Reviewed By', 'Approved By', 'Total Issues', 'Open Issues'])
    for reconciliation in reconciliations:
        performer = reconciliation.performed_by.username if reconciliation.performed_by else 'System'
        reviewer = reconciliation.reviewed_by.username if reconciliation.reviewed_by else '-'
        approver = reconciliation.approved_by.username if reconciliation.approved_by else '-'
        total_issues = reconciliation.results.count()
        open_issues = reconciliation.results.filter(resolution_status__in=['Open', 'Under Investigation']).count()
        writer.writerow([
            reconciliation.reconciliation_id,
            reconciliation.reconciliation_date.strftime('%Y-%m-%d'),
            reconciliation.location,
            reconciliation.status,
            performer,
            reviewer,
            approver,
            total_issues,
            open_issues,
        ])
    return response


ALLOWED_REPORT_ROLES = [
    'Backup Administrator',
    'Operations Manager',
    'Compliance Auditor',
    'Information Security Officer',
    'System Administrator',
]


def has_report_access(user):
    return user.is_authenticated and (
        user.is_superuser or user.groups.filter(name__in=ALLOWED_REPORT_ROLES).exists()
    )


def get_notification_recipients():
    recipients = set()
    backup_admins = User.objects.filter(is_active=True, groups__name='Backup Administrator').exclude(email='')
    superusers = User.objects.filter(is_active=True, is_superuser=True).exclude(email='')
    for user in backup_admins.iterator():
        recipients.add(user.email)
    for user in superusers.iterator():
        recipients.add(user.email)
    return sorted(recipients)


def send_email_alert(subject, message, recipients):
    if not recipients:
        return
    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, recipients, fail_silently=True)


def send_report_email(subject, message, recipients):
    if not recipients:
        return
    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, recipients, fail_silently=True)


def get_dr_team_users():
    return User.objects.filter(is_active=True).filter(
        Q(groups__name__iexact='DR Team') |
        Q(groups__name__iexact='dr team') |
        Q(groups__name__icontains='dr team') |
        Q(groups__name__icontains='disaster recovery') |
        Q(role__iexact='dr team') |
        Q(role__iexact='dr_team') |
        Q(role__iexact='drteam') |
        Q(role__iexact='disaster recovery team') |
        Q(role__iexact='disaster_recovery_team') |
        Q(role__icontains='dr team') |
        Q(role__icontains='disaster recovery')
    ).distinct()


def get_dr_team_email_recipients():
    return sorted({user.email for user in get_dr_team_users().exclude(email='') if user.email})


def send_dr_team_email_alert(subject, message):
    recipients = get_dr_team_email_recipients()
    if not recipients:
        return
    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, recipients, fail_silently=True)


def notify_email_alert(application_settings, subject, message):
    if application_settings.email_alerts_enabled:
        recipients = get_notification_recipients()
        send_email_alert(subject, message, recipients)


def send_courier_profile_email_alert(courier_profile, subject, message):
    if not courier_profile or not courier_profile.email:
        return
    application_settings = ApplicationSetting.objects.first() or ApplicationSetting.objects.create()
    if not application_settings.email_alerts_enabled:
        return
    send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [courier_profile.email], fail_silently=True)


# Create your views here.

User = get_user_model()

def index(request):
    return render(request, "index.html")


def get_user_role(user):
    if not user or not getattr(user, 'is_authenticated', False):
        return None
    role_key = (getattr(user, 'role', '') or '').strip().lower()
    if not role_key:
        return None
    return Role.objects.filter(Q(slug=role_key) | Q(dashboard_key=role_key), is_active=True).first()


def has_dashboard_feature_access(user, feature_key):
    if not user or not user.is_authenticated or not feature_key:
        return False
    if user.is_superuser:
        return True

    role = get_user_role(user)
    if role and RoleFeature.objects.filter(role=role, feature__feature_key=feature_key, is_active=True, feature__is_active=True).exists():
        return True

    if is_backup_administrator(user):
        return True

    if feature_key == 'warehouse_operations_dashboard' and is_warehouse_staff(user):
        return True

    if DashboardFeatureExemption.objects.filter(user=user, feature_key=feature_key, is_active=True).exists():
        return False

    group_ids = list(user.groups.values_list('id', flat=True))
    if not group_ids:
        return False

    return DashboardFeaturePermission.objects.filter(
        role_id__in=group_ids,
        feature_key=feature_key,
        can_view=True,
    ).exists()


def is_backup_administrator(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    role = get_user_role(user)
    if role and role.dashboard_key in {'backup', 'backup_administrator'}:
        return True
    role_name = (getattr(user, 'role', '') or '').strip().lower()
    if role_name in {'backup administrator', 'backup_administrator', 'backup', 'backup-admin'}:
        return True
    return user.groups.filter(name='Backup Administrator').exists()


def is_operations_manager(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    role = get_user_role(user)
    if role and role.dashboard_key in {'operations', 'operations_manager'}:
        return True
    if (getattr(user, 'role', '') or '').strip().lower() == 'operations_manager':
        return True
    return user.groups.filter(name='Operations Manager').exists()


def is_supreme_approver(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    role = get_user_role(user)
    if role and role.dashboard_key in {'supreme', 'supreme_approver'}:
        return True
    role_name = (getattr(user, 'role', '') or '').strip().lower()
    if role_name in {'supreme approver', 'supreme_approver', 'supremeapprover'}:
        return True
    return user.groups.filter(name__icontains='supreme').exists()


def is_warehouse_staff(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    role = get_user_role(user)
    if role and role.dashboard_key in {'warehouse', 'warehouse_operations'}:
        return True
    role_name = (getattr(user, 'role', '') or '').strip().lower()
    if role_name in {'warehouse', 'warehouse ops', 'warehouse_ops', 'warehouse_operator', 'warehouse operator'}:
        return True
    group_names = [group.name.lower() for group in user.groups.all()]
    return any('warehouse' in name for name in group_names)


def is_dr_team(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True

    role_name = (getattr(user, 'role', '') or '').strip().lower()
    if role_name in {'dr team', 'dr_team', 'drteam', 'disaster recovery team', 'disaster_recovery_team'}:
        return True

    group_names = [group.name.lower() for group in user.groups.all()]
    return any(
        name in {'dr team', 'dr_team', 'drteam', 'disaster recovery team', 'disaster_recovery_team'}
        or 'dr team' in name
        or 'drteam' in name
        or 'disaster recovery' in name
        for name in group_names
    )


def get_user_assigned_branch(user):
    if not user or not user.is_authenticated:
        return None
    return getattr(user, 'assigned_branch', None)


def get_branch_assignment_required_message():
    return 'You cannot create a shipment because no bank branch has been assigned to your account. Please contact the system administrator.'


def is_it_compliance_auditor(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True

    role = get_user_role(user)
    if role and role.dashboard_key in {'auditor', 'compliance_auditor'}:
        return True

    role_name = (getattr(user, 'role', '') or '').strip().lower()
    if role_name in {'auditor', 'it compliance auditor', 'compliance auditor', 'it auditor', 'compliance_auditor', 'it_compliance_auditor'}:
        return True

    group_names = [group.name.lower() for group in user.groups.all()]
    return any('auditor' in name for name in group_names)


def _build_auditor_context(request, page='dashboard'):
    now = timezone.localtime()
    tapes = Tape.objects.all()
    shipments = Shipment.objects.prefetch_related('tapes').all()
    exceptions = ShipmentException.objects.select_related('tape', 'shipment', 'reported_by').all()
    reconciliations = Reconciliation.objects.select_related('performed_by', 'reviewed_by').all()
    reconciliation_results = ReconciliationResult.objects.select_related('tape', 'reconciliation').all()
    audit_logs = AuditLog.objects.select_related('user').all().order_by('-timestamp')

    total_tapes = tapes.count()
    active_tapes = tapes.filter(status='Active').count()
    retained_tapes = tapes.filter(status='Retained').count()
    in_transit_tapes = tapes.filter(status='In Transit').count()
    missing_tapes = tapes.filter(status='Missing').count()
    damaged_tapes = tapes.filter(status='Damaged').count()
    legal_hold_tapes = tapes.filter(legal_hold=True).count()
    audit_hold_tapes = tapes.filter(audit_hold=True).count()
    open_exceptions = exceptions.filter(status__in=['Open', 'Investigating']).count()

    compliance_rate = 0
    if total_tapes:
        compliant_tapes = tapes.exclude(status__in=['Missing', 'Damaged']).count()
        compliance_rate = round((compliant_tapes / total_tapes) * 100)

    reconciliation_accuracy = 100
    if reconciliation_results.exists():
        clean_results = reconciliation_results.filter(issue_type='None').count()
        reconciliation_accuracy = round((clean_results / reconciliation_results.count()) * 100)

    completed_shipments = shipments.filter(status__in=['Delivered', 'Return Accepted']).count()
    shipment_sla_rate = 100
    if shipments.exists():
        shipment_sla_rate = round((completed_shipments / shipments.count()) * 100)

    retention_score = 100
    if total_tapes:
        retention_score = round((tapes.exclude(status__in=['Missing', 'Damaged']).filter(retention_end_date__gte=now.date()).count() / total_tapes) * 100)

    chain_of_custody_score = 100
    if shipments.exists():
        verified_shipments = shipments.filter(releasing_custodian__isnull=False, receiving_custodian__isnull=False).count()
        chain_of_custody_score = round((verified_shipments / shipments.count()) * 100)

    shipment_score = shipment_sla_rate
    reconciliation_score = reconciliation_accuracy
    audit_score = 100 if audit_logs.exists() else 0
    access_score = 100 if CustomUser.objects.filter(is_active=True).exists() else 50

    def score_status(score):
        if score >= 95:
            return 'Compliant'
        if score >= 80:
            return 'Warning'
        return 'Non-Compliant'

    health_cards = [
        {'name': 'Retention Compliance', 'score': retention_score, 'status': score_status(retention_score), 'violations': max(total_tapes - tapes.filter(retention_end_date__gte=now.date()).count(), 0), 'last_assessment_date': now.date().strftime('%Y-%m-%d')},
        {'name': 'Chain of Custody Compliance', 'score': chain_of_custody_score, 'status': score_status(chain_of_custody_score), 'violations': max(shipments.count() - shipments.filter(releasing_custodian__isnull=False, receiving_custodian__isnull=False).count(), 0), 'last_assessment_date': now.date().strftime('%Y-%m-%d')},
        {'name': 'Shipment Compliance', 'score': shipment_score, 'status': score_status(shipment_score), 'violations': max(shipments.count() - completed_shipments, 0), 'last_assessment_date': now.date().strftime('%Y-%m-%d')},
        {'name': 'Reconciliation Compliance', 'score': reconciliation_score, 'status': score_status(reconciliation_score), 'violations': max(reconciliation_results.count() - reconciliation_results.filter(issue_type='None').count(), 0), 'last_assessment_date': now.date().strftime('%Y-%m-%d')},
        {'name': 'Audit Logging Compliance', 'score': audit_score, 'status': score_status(audit_score), 'violations': 0 if audit_logs.exists() else 1, 'last_assessment_date': now.date().strftime('%Y-%m-%d')},
        {'name': 'User Access Compliance', 'score': access_score, 'status': score_status(access_score), 'violations': max(CustomUser.objects.filter(is_active=False).count(), 0), 'last_assessment_date': now.date().strftime('%Y-%m-%d')},
    ]

    kpi_cards = [
        {'label': 'Total Registered Tapes', 'value': total_tapes, 'trend': 'Stable', 'updated': now.strftime('%Y-%m-%d %H:%M')},
        {'label': 'Active Tapes', 'value': active_tapes, 'trend': 'Up', 'updated': now.strftime('%Y-%m-%d %H:%M')},
        {'label': 'Retained Tapes', 'value': retained_tapes, 'trend': 'Stable', 'updated': now.strftime('%Y-%m-%d %H:%M')},
        {'label': 'In Transit Tapes', 'value': in_transit_tapes, 'trend': 'Watch', 'updated': now.strftime('%Y-%m-%d %H:%M')},
        {'label': 'Missing Tapes', 'value': missing_tapes, 'trend': 'Critical', 'updated': now.strftime('%Y-%m-%d %H:%M')},
        {'label': 'Damaged Tapes', 'value': damaged_tapes, 'trend': 'Review', 'updated': now.strftime('%Y-%m-%d %H:%M')},
        {'label': 'Open Exceptions', 'value': open_exceptions, 'trend': 'Action', 'updated': now.strftime('%Y-%m-%d %H:%M')},
        {'label': 'Compliance Rate', 'value': f'{compliance_rate}%', 'trend': 'Green', 'updated': now.strftime('%Y-%m-%d %H:%M')},
        {'label': 'Reconciliation Accuracy', 'value': f'{reconciliation_accuracy}%', 'trend': 'Green', 'updated': now.strftime('%Y-%m-%d %H:%M')},
        {'label': 'Shipment SLA Compliance', 'value': f'{shipment_sla_rate}%', 'trend': 'Watch', 'updated': now.strftime('%Y-%m-%d %H:%M')},
        {'label': 'Tapes Under Legal Hold', 'value': legal_hold_tapes, 'trend': 'Protected', 'updated': now.strftime('%Y-%m-%d %H:%M')},
        {'label': 'Tapes Under Audit Hold', 'value': audit_hold_tapes, 'trend': 'Protected', 'updated': now.strftime('%Y-%m-%d %H:%M')},
    ]

    custody_rows = []
    for shipment in shipments:
        for tape in shipment.tapes.all():
            custody_rows.append({
                'volser': tape.volser,
                'barcode': tape.barcode,
                'current_custodian': shipment.receiving_custodian or shipment.releasing_custodian or tape.current_location,
                'previous_custodian': shipment.releasing_custodian or 'Unassigned',
                'transfer_date': shipment.release_datetime.date().strftime('%Y-%m-%d') if shipment.release_datetime else shipment.shipment_date.strftime('%Y-%m-%d'),
                'transfer_time': shipment.release_datetime.time().strftime('%H:%M') if shipment.release_datetime else '00:00',
                'transfer_type': shipment.shipment_type,
                'source_location': shipment.source_location or 'Unspecified',
                'destination_location': shipment.destination_location or 'Unspecified',
                'verification_status': 'Verified' if shipment.approved_by and shipment.delivery_date else 'Pending',
            })

    retention_rows = []
    for tape in tapes.order_by('retention_end_date'):
        retention_rows.append({
            'volser': tape.volser,
            'barcode': tape.barcode,
            'retention_start_date': tape.date_registered.date().strftime('%Y-%m-%d'),
            'retention_end_date': tape.retention_end_date.strftime('%Y-%m-%d'),
            'legal_hold': tape.legal_hold,
            'audit_hold': tape.audit_hold,
            'current_status': tape.status,
            'compliance_status': 'Compliant' if tape.retention_end_date >= now.date() and tape.status not in ['Missing', 'Damaged'] else 'Violation',
        })

    shipment_rows = []
    for shipment in shipments.order_by('-shipment_date'):
        shipment_rows.append({
            'shipment_id': shipment.shipment_id,
            'manifest_number': shipment.tracking_number or shipment.shipment_id,
            'shipment_type': shipment.shipment_type,
            'source_location': shipment.source_location or 'Unspecified',
            'destination_location': shipment.destination_location or 'Unspecified',
            'courier': shipment.courier_name or 'Pending',
            'dispatch_date': shipment.shipment_date.strftime('%Y-%m-%d'),
            'delivery_date': shipment.delivery_date.strftime('%Y-%m-%d') if shipment.delivery_date else 'Pending',
            'sla_status': 'On Track' if shipment.delivery_date and shipment.expected_delivery_date and shipment.delivery_date <= shipment.expected_delivery_date else 'At Risk',
            'manifest_complete': 'Yes' if shipment.is_manifest_complete() else 'No',
            'dual_custody_verified': 'Yes' if shipment.has_dual_custody() else 'No',
            'compliance_status': 'Compliant' if shipment.compliance_passed() else 'Review Required',
        })

    exception_rows = []
    for exc in exceptions.order_by('-reported_date'):
        exception_rows.append({
            'exception_id': exc.exception_id,
            'volser': exc.tape.volser if exc.tape else 'N/A',
            'exception_type': exc.exception_type,
            'severity': exc.severity,
            'description': exc.description,
            'date_reported': exc.reported_date.strftime('%Y-%m-%d %H:%M'),
            'reported_by': exc.reported_by.username if exc.reported_by else 'System',
            'status': exc.status,
            'resolution_date': exc.reported_date.strftime('%Y-%m-%d %H:%M') if exc.status in ['Resolved', 'Closed'] else 'Pending',
        })

    reconciliation_rows = []
    for result in reconciliation_results.order_by('-created_at'):
        reconciliation_rows.append({
            'reconciliation_id': result.reconciliation.reconciliation_id,
            'location': result.reconciliation.location,
            'date_performed': result.reconciliation.reconciliation_date.strftime('%Y-%m-%d'),
            'expected_tape_count': result.reconciliation.results.count(),
            'scanned_tape_count': result.reconciliation.results.count(),
            'missing_tapes': 1 if result.issue_type == 'Missing' else 0,
            'unexpected_tapes': 1 if result.issue_type == 'Unexpected' else 0,
            'duplicate_tapes': 1 if result.issue_type == 'Duplicate' else 0,
            'misplaced_tapes': 1 if result.issue_type == 'Misplaced' else 0,
            'compliance_status': 'Compliant' if result.issue_type == 'None' else 'Violation',
        })

    hold_rows = []
    for tape in tapes.filter(Q(legal_hold=True) | Q(audit_hold=True)).order_by('-date_registered'):
        hold_rows.append({
            'volser': tape.volser,
            'hold_type': 'Legal Hold' if tape.legal_hold and tape.audit_hold else ('Legal Hold' if tape.legal_hold else 'Audit Hold'),
            'reason': tape.remarks or 'Protected under policy',
            'date_applied': tape.date_registered.date().strftime('%Y-%m-%d'),
            'applied_by': 'System',
            'release_date': 'Pending',
            'status': 'Active',
        })

    alerts = []
    if missing_tapes:
        alerts.append({'timestamp': now.strftime('%Y-%m-%d %H:%M'), 'severity': 'Critical', 'category': 'Missing Tape', 'description': f'{missing_tapes} tape(s) are currently marked as missing.', 'status': 'Open'})
    if any(item['compliance_status'] == 'Violation' for item in retention_rows):
        alerts.append({'timestamp': now.strftime('%Y-%m-%d %H:%M'), 'severity': 'High', 'category': 'Retention Violation', 'description': 'One or more tapes exceed retention and compliance thresholds.', 'status': 'Investigating'})
    if shipments.filter(status__in=['Pending', 'In Transit']).exists():
        alerts.append({'timestamp': now.strftime('%Y-%m-%d %H:%M'), 'severity': 'Medium', 'category': 'Overdue Shipment', 'description': 'Pending and in-transit shipments require review.', 'status': 'Open'})
    if open_exceptions:
        alerts.append({'timestamp': now.strftime('%Y-%m-%d %H:%M'), 'severity': 'High', 'category': 'Exception Review', 'description': f'{open_exceptions} exception(s) remain unresolved.', 'status': 'Open'})

    chart_labels = []
    chart_counts = []
    for month in range(6):
        label_date = (now.date().replace(day=1) - timedelta(days=30 * month))
        chart_labels.append(label_date.strftime('%b'))
        chart_counts.append(max(0, total_tapes - month))
    chart_labels.reverse()
    chart_counts.reverse()

    context = {
        'request': request,
        'page': page,
        'auditor_name': request.user.get_full_name() or request.user.username,
        'role': 'IT Compliance Auditor',
        'last_login': request.user.last_login.strftime('%Y-%m-%d %H:%M') if request.user.last_login else 'First login',
        'current_datetime': now.strftime('%Y-%m-%d %H:%M:%S'),
        'is_read_only': True,
        'total_tapes': total_tapes,
        'active_tapes': active_tapes,
        'retained_tapes': retained_tapes,
        'in_transit_tapes': in_transit_tapes,
        'missing_tapes': missing_tapes,
        'damaged_tapes': damaged_tapes,
        'open_exceptions': open_exceptions,
        'compliance_rate': compliance_rate,
        'reconciliation_accuracy': reconciliation_accuracy,
        'shipment_sla_rate': shipment_sla_rate,
        'legal_hold_tapes': legal_hold_tapes,
        'audit_hold_tapes': audit_hold_tapes,
        'health_cards': health_cards,
        'kpi_cards': kpi_cards,
        'audit_logs': audit_logs[:10],
        'custody_rows': custody_rows[:20],
        'shipment_rows': shipment_rows[:20],
        'retention_rows': retention_rows[:20],
        'exception_rows': exception_rows[:20],
        'reconciliation_rows': reconciliation_rows[:20],
        'hold_rows': hold_rows[:20],
        'alerts': alerts[:8],
        'chart_labels': chart_labels,
        'chart_counts': chart_counts,
        'report_items': [
            {'name': 'Inventory Audit Report', 'type': 'inventory'},
            {'name': 'Shipment Compliance Report', 'type': 'shipments'},
            {'name': 'Chain of Custody Report', 'type': 'custody'},
            {'name': 'Retention Compliance Report', 'type': 'retention'},
            {'name': 'Exception Report', 'type': 'exceptions'},
            {'name': 'Reconciliation Report', 'type': 'reconciliation'},
            {'name': 'Audit Trail Report', 'type': 'audit'},
            {'name': 'Executive Compliance Summary', 'type': 'summary'},
        ],
    }
    return context


@login_required(login_url='signin')
@user_passes_test(is_it_compliance_auditor, login_url='signin')
def auditor_dashboard(request):
    view_name = request.GET.get('view')
    if view_name == 'audit-logs':
        return audit_logs_view(request)
    if view_name == 'reports':
        return compliance_reports_view(request)
    if view_name == 'exceptions':
        return exception_review_view(request)
    if view_name == 'shipments':
        return shipment_compliance_view(request)
    if view_name == 'retention':
        return retention_compliance_view(request)
    if view_name == 'reconciliation':
        return reconciliation_review_view(request)

    shipment_request_form = AuditorShipmentRequestForm(request.POST or None)
    if request.method == 'POST' and request.POST.get('form_type') == 'submit_shipment_request':
        if shipment_request_form.is_valid():
            branch_name = shipment_request_form.cleaned_data['branch_name'].strip()
            request_details = shipment_request_form.cleaned_data['request_details'].strip()
            shipment = Shipment.objects.create(
                shipment_date=timezone.localdate(),
                shipment_type='Off-Site Transfer',
                status='Pending',
                source_location=branch_name,
                releasing_custodian=request.user.get_full_name() or request.user.username,
                receiving_organization='Pending review',
                approval_remarks=request_details,
                created_by=request.user,
                last_updated_by=request.user,
            )
            AuditLog.objects.create(
                name='Shipment Request Submitted',
                action=f'Shipment request {shipment.shipment_id} created for {branch_name} by {request.user.username}',
                user=request.user,
                severity='info',
            )
            messages.success(request, 'Shipment request submitted to the backup administrator.')
            return redirect(reverse('auditor-dashboard'))

        messages.error(request, 'Please provide both the branch name and request details.')

    context = _build_auditor_context(request, page='dashboard')
    context['active_view'] = 'dashboard'
    context['shipment_request_form'] = shipment_request_form
    return render(request, 'auditor_dashboard.html', context)


@login_required(login_url='signin')
@user_passes_test(is_it_compliance_auditor, login_url='signin')
def audit_logs_view(request):
    context = _build_auditor_context(request, page='audit-logs')
    queryset, filters = build_audit_log_queryset(request)
    show_admin_history = 'show_admin_history' in request.GET
    show_audit_panel = False
    # If viewing admin history, restrict audit logs to the current user
    if show_admin_history:
        try:
            queryset = queryset.filter(user=request.user)
            filters['user'] = request.user.username
            show_audit_panel = True
        except Exception:
            pass
    # If admin requests their own history, filter logs to the current user
    if 'show_admin_history' in request.GET:
        show_audit_panel = True
        try:
            queryset = queryset.filter(user=request.user)
            filters['user'] = request.user.username
        except Exception:
            pass
    export_format = (request.GET.get('export') or '').strip().lower()
    if export_format in {'csv', 'excel', 'pdf'}:
        export_response = export_audit_logs(queryset, export_format)
        if export_response is not None:
            return export_response

    per_page = filters['per_page']
    paginator = Paginator(queryset, per_page)
    page = request.GET.get('page', 1)
    page_obj = paginator.get_page(page)

    context['audit_logs'] = page_obj.object_list
    context['page_obj'] = page_obj
    context['paginator'] = paginator
    context['filters'] = filters
    context['active_filters'] = [
        {'name': 'Search', 'value': filters['search']},
        {'name': 'Log Type', 'value': filters['log_type']},
        {'name': 'Module', 'value': filters['module']},
        {'name': 'Severity', 'value': filters['severity']},
        {'name': 'User', 'value': filters['user']},
        {'name': 'Status', 'value': filters['status']},
        {'name': 'Date Range', 'value': filters['date_range']},
        {'name': 'From', 'value': filters['date_from']},
        {'name': 'To', 'value': filters['date_to']},
    ]
    context['active_filters'] = [item for item in context['active_filters'] if item['value']]
    context['users'] = get_user_model().objects.order_by('username')
    context['log_types'] = ['Authentication', 'Request', 'Approval', 'Shipment', 'Inventory', 'Audit', 'Compliance', 'Exception', 'Reconciliation', 'Notification', 'API', 'Import', 'Export', 'Security', 'Configuration', 'System', 'Error']
    context['modules'] = ['Shipment', 'Inventory', 'Bank Branches', 'Couriers', 'Users', 'Vault', 'Reports', 'Dashboard', 'API', 'Notifications']
    context['severity_choices'] = ['info', 'success', 'warning', 'error']
    context['status_choices'] = ['Success', 'Failed', 'Pending', 'Approved', 'Rejected']
    context['date_ranges'] = [
        ('', 'All Dates'),
        ('today', 'Today'),
        ('yesterday', 'Yesterday'),
        ('last_7_days', 'Last 7 Days'),
        ('last_30_days', 'Last 30 Days'),
        ('this_month', 'This Month'),
        ('this_year', 'This Year'),
    ]
    return render(request, 'audit_logs.html', context)


@login_required(login_url='signin')
@user_passes_test(is_it_compliance_auditor, login_url='signin')
def compliance_reports_view(request):
    context = _build_auditor_context(request, page='reports')

    show_reports_panel = 'show_reports' in request.GET
    report_type = request.GET.get('report_type')
    report_category = request.GET.get('report_category')
    report_period = get_scoped_report_param(request, report_category, 'report_period')
    export_csv = get_scoped_report_flag(request, report_category, 'export_csv')
    export_pdf = get_scoped_report_flag(request, report_category, 'export_pdf')
    export_excel = get_scoped_report_flag(request, report_category, 'export_excel')
    report_search = get_scoped_report_param(request, report_category, 'report_search')
    report_filter_status = get_scoped_report_param(request, report_category, 'report_filter_status')
    report_date_from = get_scoped_report_param(request, report_category, 'report_date_from')
    report_date_to = get_scoped_report_param(request, report_category, 'report_date_to')
    report_sort = get_scoped_report_param(request, report_category, 'report_sort')
    report_order = (get_scoped_report_param(request, report_category, 'report_order', 'asc') or 'asc').lower()
    if report_order not in {'asc', 'desc'}:
        report_order = 'asc'

    report_categories = get_report_categories()
    current_month = timezone.localdate().strftime('%Y-%m')
    if show_reports_panel and not report_period:
        report_period = current_month

    inventory_report_tapes = []
    shipment_report_rows = []
    custody_report_rows = []
    reconciliation_report_rows = []
    retention_report_rows = []
    compliance_report_rows = []
    exception_report_rows = []
    audit_trail_report_rows = []
    management_summary_report_rows = []
    report_table_columns = []
    report_paginator = None
    report_page_obj = None

    if show_reports_panel and report_period:
        report_month = get_first_day_of_month(report_period) or get_first_day_of_month(current_month)
        if report_month:
            if report_category == 'inventory':
                qs = Tape.objects.filter(date_registered__gte=report_month, date_registered__lt=get_next_month(report_month))
                if report_search:
                    qs = qs.filter(Q(volser__icontains=report_search) | Q(barcode__icontains=report_search) | Q(rfid_tag__icontains=report_search) | Q(tape_type__icontains=report_search) | Q(status__icontains=report_search) | Q(current_location__icontains=report_search))
                if report_filter_status:
                    qs = qs.filter(status__iexact=report_filter_status)
                inventory_report_tapes = list(qs.order_by('volser'))
                for tape in inventory_report_tapes:
                    tape.latest_custodian = get_latest_custodian_for_tape(tape) or tape.current_location or '-'
                report_table_columns = [
                    {'key': 'volser', 'label': 'VolSER'},
                    {'key': 'barcode', 'label': 'Barcode'},
                    {'key': 'rfid_tag', 'label': 'RFID Tag'},
                    {'key': 'tape_type', 'label': 'Tape Type'},
                    {'key': 'status', 'label': 'Status'},
                    {'key': 'current_location', 'label': 'Current Location'},
                    {'key': 'latest_custodian', 'label': 'Custodian'},
                    {'key': 'retention_end_date', 'label': 'Retention End Date'},
                    {'key': 'date_registered', 'label': 'Date Registered'},
                ]
                if export_pdf:
                    return export_report_pdf(report_category, report_period, [
                        {
                            'volser': tape.volser,
                            'barcode': tape.barcode,
                            'rfid_tag': tape.rfid_tag or '-',
                            'tape_type': tape.tape_type,
                            'status': tape.status,
                            'current_location': tape.current_location or '-',
                            'latest_custodian': tape.latest_custodian,
                            'retention_end_date': tape.retention_end_date,
                            'date_registered': tape.date_registered,
                        }
                        for tape in inventory_report_tapes
                    ], report_table_columns)
                if export_excel:
                    return export_report_excel(report_category, report_period, [
                        {
                            'volser': tape.volser,
                            'barcode': tape.barcode,
                            'rfid_tag': tape.rfid_tag or '-',
                            'tape_type': tape.tape_type,
                            'status': tape.status,
                            'current_location': tape.current_location or '-',
                            'latest_custodian': tape.latest_custodian,
                            'retention_end_date': tape.retention_end_date,
                            'date_registered': tape.date_registered,
                        }
                        for tape in inventory_report_tapes
                    ], report_table_columns)
                if export_csv:
                    return export_inventory_report_csv(report_period, inventory_report_tapes)
            elif report_category == 'shipment':
                shipment_report_rows = list(Shipment.objects.filter(shipment_date__gte=report_month, shipment_date__lt=get_next_month(report_month)).order_by('shipment_id'))
                report_table_columns = [
                    {'key': 'shipment_id', 'label': 'Shipment ID'},
                    {'key': 'shipment_type', 'label': 'Shipment Type'},
                    {'key': 'source_location', 'label': 'Source Location'},
                    {'key': 'destination_location', 'label': 'Destination Location'},
                    {'key': 'courier_name', 'label': 'Courier'},
                    {'key': 'shipment_date', 'label': 'Dispatch Date'},
                    {'key': 'delivery_date', 'label': 'Delivery Date'},
                    {'key': 'status', 'label': 'Status'},
                    {'key': 'number_of_tapes', 'label': 'Number of Tapes'},
                ]
            elif report_category == 'custody':
                custody_report_rows = list(Shipment.objects.filter(shipment_date__gte=report_month, shipment_date__lt=get_next_month(report_month)).order_by('shipment_id'))
                for shipment in custody_report_rows:
                    shipment.transfer_date = shipment.shipment_date
                    shipment.transfer_time = shipment.release_datetime.time() if shipment.release_datetime else None
                    shipment.previous_custodian = shipment.releasing_custodian or '-'
                    shipment.new_custodian = shipment.receiving_custodian or '-'
                    shipment.location = shipment.destination_location or shipment.source_location or '-'
                    shipment.remarks = shipment.approval_remarks or shipment.delivery_notes or '-'
                report_table_columns = [
                    {'key': 'transfer_date', 'label': 'Transfer Date'},
                    {'key': 'transfer_time', 'label': 'Transfer Time'},
                    {'key': 'previous_custodian', 'label': 'Previous Custodian'},
                    {'key': 'new_custodian', 'label': 'New Custodian'},
                    {'key': 'location', 'label': 'Location'},
                    {'key': 'remarks', 'label': 'Remarks'},
                ]
            elif report_category == 'reconciliation':
                qs = Reconciliation.objects.filter(reconciliation_date__gte=report_month, reconciliation_date__lt=get_next_month(report_month)).order_by('-reconciliation_date')
                if report_search:
                    qs = qs.filter(Q(reconciliation_id__icontains=report_search) | Q(location__icontains=report_search) | Q(status__icontains=report_search))
                if report_filter_status:
                    qs = qs.filter(status__iexact=report_filter_status)
                reconciliation_report_rows = []
                for reconciliation in qs:
                    reconciliation_report_rows.append({
                        'pk': reconciliation.id,
                        'id': reconciliation.id,
                        'reconciliation_id': reconciliation.reconciliation_id,
                        'location': reconciliation.location,
                        'expected_tapes': reconciliation.total_tapes_expected,
                        'scanned_tapes': reconciliation.total_tapes_scanned,
                        'missing_tapes': reconciliation.results.filter(issue_type='Missing').count(),
                        'misplaced_tapes': reconciliation.results.filter(issue_type='Misplaced').count(),
                        'unexpected_tapes': reconciliation.results.filter(issue_type='Unexpected').count(),
                        'damaged_tapes': reconciliation.results.filter(issue_type='Damaged').count(),
                        'reconciliation_date': reconciliation.reconciliation_date,
                        'status': reconciliation.status,
                        'branch': reconciliation.location,
                        'operator': reconciliation.assigned_operator.username if reconciliation.assigned_operator else 'Unassigned',
                        'start_time': reconciliation.scan_started_at,
                        'end_time': reconciliation.scan_completed_at,
                        'comments': reconciliation.notes,
                    })
                report_table_columns = [
                    {'key': 'reconciliation_id', 'label': 'Reconciliation ID'},
                    {'key': 'location', 'label': 'Location'},
                    {'key': 'expected_tapes', 'label': 'Expected Tapes'},
                    {'key': 'scanned_tapes', 'label': 'Scanned Tapes'},
                    {'key': 'missing_tapes', 'label': 'Missing Tapes'},
                    {'key': 'misplaced_tapes', 'label': 'Misplaced Tapes'},
                    {'key': 'unexpected_tapes', 'label': 'Unexpected Tapes'},
                    {'key': 'reconciliation_date', 'label': 'Reconciliation Date'},
                    {'key': 'status', 'label': 'Status'},
                ]
                reconciliation_report_rows = sort_report_rows(reconciliation_report_rows, report_sort or None, report_order)
                report_paginator, report_page_obj = paginate_report_rows(request, reconciliation_report_rows, page_param=f'report_page_{report_category}')
                reconciliation_report_rows = list(report_page_obj.object_list)
            elif report_category == 'retention':
                qs = Tape.objects.filter(retention_end_date__gte=report_month, retention_end_date__lt=get_next_month(report_month)).order_by('retention_end_date')
                if report_search:
                    qs = qs.filter(Q(volser__icontains=report_search) | Q(barcode__icontains=report_search) | Q(status__icontains=report_search))
                if report_filter_status:
                    qs = qs.filter(status__iexact=report_filter_status)
                retention_report_rows = []
                for tape in qs:
                    retention_report_rows.append({
                        'volser': tape.volser,
                        'barcode': tape.barcode,
                        'retention_start_date': tape.date_registered.date() if getattr(tape, 'date_registered', None) else '-',
                        'retention_end_date': tape.retention_end_date,
                        'days_remaining': (tape.retention_end_date - timezone.localdate()).days if tape.retention_end_date else '-',
                        'legal_hold': 'Yes' if tape.legal_hold else 'No',
                        'audit_hold': 'Yes' if tape.audit_hold else 'No',
                        'status': tape.status,
                    })
                report_table_columns = [
                    {'key': 'volser', 'label': 'VolSER'},
                    {'key': 'barcode', 'label': 'Barcode'},
                    {'key': 'retention_start_date', 'label': 'Retention Start Date'},
                    {'key': 'retention_end_date', 'label': 'Retention End Date'},
                    {'key': 'days_remaining', 'label': 'Days Remaining'},
                    {'key': 'legal_hold', 'label': 'Legal Hold'},
                    {'key': 'audit_hold', 'label': 'Audit Hold'},
                    {'key': 'status', 'label': 'Status'},
                ]
                retention_report_rows = sort_report_rows(retention_report_rows, report_sort or None, report_order)
                report_paginator, report_page_obj = paginate_report_rows(request, retention_report_rows, page_param=f'report_page_{report_category}')
                retention_report_rows = list(report_page_obj.object_list)
            elif report_category == 'compliance':
                qs = ReconciliationResult.objects.filter(reconciliation__reconciliation_date__gte=report_month, reconciliation__reconciliation_date__lt=get_next_month(report_month)).order_by('-created_at')
                if report_search:
                    qs = qs.filter(Q(reconciliation__reconciliation_id__icontains=report_search) | Q(remarks__icontains=report_search) | Q(resolution_status__icontains=report_search))
                if report_filter_status:
                    qs = qs.filter(resolution_status__iexact=report_filter_status)
                compliance_report_rows = []
                for result in qs:
                    compliance_report_rows.append({
                        'compliance_id': result.reconciliation.reconciliation_id,
                        'policy_name': 'Tape Handling Policy',
                        'compliance_status': 'Compliant' if result.resolution_status == 'Resolved' else 'Needs Review',
                        'violations': result.issue_type,
                        'date_identified': result.created_at.date(),
                        'responsible_user': result.reconciliation.performed_by.username if result.reconciliation.performed_by else '-',
                        'resolution_status': result.resolution_status,
                    })
                report_table_columns = [
                    {'key': 'compliance_id', 'label': 'Compliance ID'},
                    {'key': 'policy_name', 'label': 'Policy Name'},
                    {'key': 'compliance_status', 'label': 'Compliance Status'},
                    {'key': 'violations', 'label': 'Violations'},
                    {'key': 'date_identified', 'label': 'Date Identified'},
                    {'key': 'responsible_user', 'label': 'Responsible User'},
                    {'key': 'resolution_status', 'label': 'Resolution Status'},
                ]
                compliance_report_rows = sort_report_rows(compliance_report_rows, report_sort or None, report_order)
                report_paginator, report_page_obj = paginate_report_rows(request, compliance_report_rows, page_param=f'report_page_{report_category}')
                compliance_report_rows = list(report_page_obj.object_list)
            elif report_category == 'exception':
                qs = ShipmentException.objects.filter(reported_date__date__gte=report_month, reported_date__date__lt=get_next_month(report_month)).order_by('-reported_date')
                if report_search:
                    qs = qs.filter(Q(exception_id__icontains=report_search) | Q(tape__volser__icontains=report_search) | Q(status__icontains=report_search))
                if report_filter_status:
                    qs = qs.filter(status__iexact=report_filter_status)
                exception_report_rows = []
                for exception in qs:
                    exception_report_rows.append({
                        'exception_id': exception.exception_id,
                        'tape_volser': exception.tape.volser if exception.tape else '-',
                        'exception_type': exception.exception_type,
                        'severity': exception.severity,
                        'reported_by': exception.reported_by.username if exception.reported_by else '-',
                        'date_reported': exception.reported_date.date(),
                        'status': exception.status,
                        'resolution_date': exception.reported_date.date(),
                    })
                report_table_columns = [
                    {'key': 'exception_id', 'label': 'Exception ID'},
                    {'key': 'tape_volser', 'label': 'Tape VolSER'},
                    {'key': 'exception_type', 'label': 'Exception Type'},
                    {'key': 'severity', 'label': 'Severity'},
                    {'key': 'reported_by', 'label': 'Reported By'},
                    {'key': 'date_reported', 'label': 'Date Reported'},
                    {'key': 'status', 'label': 'Status'},
                    {'key': 'resolution_date', 'label': 'Resolution Date'},
                ]
                exception_report_rows = sort_report_rows(exception_report_rows, report_sort or None, report_order)
                report_paginator, report_page_obj = paginate_report_rows(request, exception_report_rows, page_param=f'report_page_{report_category}')
                exception_report_rows = list(report_page_obj.object_list)
            elif report_category == 'audit_trail':
                qs = AuditLog.objects.filter(timestamp__date__gte=report_month, timestamp__date__lt=get_next_month(report_month)).order_by('-timestamp')
                if report_search:
                    qs = qs.filter(Q(name__icontains=report_search) | Q(action__icontains=report_search) | Q(message__icontains=report_search))
                if report_filter_status:
                    qs = qs.filter(severity__iexact=report_filter_status)
                audit_trail_report_rows = []
                for audit_entry in qs:
                    audit_trail_report_rows.append({
                        'audit_id': audit_entry.id,
                        'user': audit_entry.user.username if audit_entry.user else '-',
                        'action': audit_entry.action,
                        'module': audit_entry.name,
                        'record_affected': audit_entry.message,
                        'timestamp': audit_entry.timestamp,
                        'ip_address': '-',
                    })
                report_table_columns = [
                    {'key': 'audit_id', 'label': 'Audit ID'},
                    {'key': 'user', 'label': 'User'},
                    {'key': 'action', 'label': 'Action'},
                    {'key': 'module', 'label': 'Module'},
                    {'key': 'record_affected', 'label': 'Record Affected'},
                    {'key': 'timestamp', 'label': 'Timestamp'},
                    {'key': 'ip_address', 'label': 'IP Address'},
                ]
                audit_trail_report_rows = sort_report_rows(audit_trail_report_rows, report_sort or None, report_order)
                report_paginator, report_page_obj = paginate_report_rows(request, audit_trail_report_rows, page_param=f'report_page_{report_category}')
                audit_trail_report_rows = list(report_page_obj.object_list)
            elif report_category == 'management_summary':
                management_summary_report_rows = [{
                    'report_date': report_period,
                    'total_tapes': Tape.objects.count(),
                    'active_tapes': Tape.objects.filter(status='Active').count(),
                    'in_transit': Tape.objects.filter(status='In Transit').count(),
                    'missing_tapes': Tape.objects.filter(status='Missing').count(),
                    'damaged_tapes': Tape.objects.filter(status='Damaged').count(),
                    'open_exceptions': ShipmentException.objects.filter(status__in=['Open', 'Investigating']).count(),
                    'compliance_rate': '98.5%',
                    'reconciliation_accuracy': '99.2%',
                }]
                report_table_columns = [
                    {'key': 'report_date', 'label': 'Report Date'},
                    {'key': 'total_tapes', 'label': 'Total Tapes'},
                    {'key': 'active_tapes', 'label': 'Active Tapes'},
                    {'key': 'in_transit', 'label': 'In Transit'},
                    {'key': 'missing_tapes', 'label': 'Missing Tapes'},
                    {'key': 'damaged_tapes', 'label': 'Damaged Tapes'},
                    {'key': 'open_exceptions', 'label': 'Open Exceptions'},
                    {'key': 'compliance_rate', 'label': 'Compliance Rate'},
                    {'key': 'reconciliation_accuracy', 'label': 'Reconciliation Accuracy'},
                ]

    context.update({
        'active_view': 'reports',
        'show_reports_panel': show_reports_panel,
        'report_type': report_type,
        'report_category': report_category,
        'report_period': report_period,
        'report_categories': report_categories,
        'current_month': current_month,
        'report_search': report_search,
        'report_filter_status': report_filter_status,
        'report_date_from': report_date_from,
        'report_date_to': report_date_to,
        'report_sort': report_sort,
        'report_order': report_order,
        'inventory_report_tapes': inventory_report_tapes,
        'shipment_report_rows': shipment_report_rows,
        'custody_report_rows': custody_report_rows,
        'reconciliation_report_rows': reconciliation_report_rows,
        'retention_report_rows': retention_report_rows,
        'compliance_report_rows': compliance_report_rows,
        'exception_report_rows': exception_report_rows,
        'audit_trail_report_rows': audit_trail_report_rows,
        'management_summary_report_rows': management_summary_report_rows,
        'report_table_columns': report_table_columns,
        'report_paginator': report_paginator,
        'report_page_obj': report_page_obj,
    })
    return render(request, 'auditor_dashboard.html', context)


@login_required(login_url='signin')
@user_passes_test(is_it_compliance_auditor, login_url='signin')
def exception_review_view(request):
    context = _build_auditor_context(request, page='exceptions')
    context['active_view'] = 'exceptions'
    return render(request, 'auditor_dashboard.html', context)


@login_required(login_url='signin')
@user_passes_test(is_it_compliance_auditor, login_url='signin')
def shipment_compliance_view(request):
    context = _build_auditor_context(request, page='shipments')
    context['active_view'] = 'shipments'
    return render(request, 'auditor_dashboard.html', context)


@login_required(login_url='signin')
@user_passes_test(is_it_compliance_auditor, login_url='signin')
def retention_compliance_view(request):
    context = _build_auditor_context(request, page='retention')
    context['active_view'] = 'retention'
    return render(request, 'auditor_dashboard.html', context)


@login_required(login_url='signin')
@user_passes_test(is_it_compliance_auditor, login_url='signin')
def reconciliation_review_view(request):
    context = _build_auditor_context(request, page='reconciliation')
    context['active_view'] = 'reconciliation'
    return render(request, 'auditor_dashboard.html', context)


def _get_client_ip(request):
    forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR', '')
    if forwarded_for:
        return forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', 'Unknown')


def _get_client_browser_and_os(request):
    user_agent = request.META.get('HTTP_USER_AGENT', '') or ''
    browser = 'Unknown'
    operating_system = 'Unknown'
    lower_agent = user_agent.lower()

    if 'edg/' in lower_agent:
        browser = 'Edge'
    elif 'chrome' in lower_agent:
        browser = 'Chrome'
    elif 'firefox' in lower_agent:
        browser = 'Firefox'
    elif 'safari' in lower_agent:
        browser = 'Safari'
    elif 'opr/' in lower_agent or 'opera' in lower_agent:
        browser = 'Opera'

    if 'windows' in lower_agent:
        operating_system = 'Windows'
    elif 'macintosh' in lower_agent or 'mac os' in lower_agent:
        operating_system = 'macOS'
    elif 'linux' in lower_agent:
        operating_system = 'Linux'
    elif 'android' in lower_agent:
        operating_system = 'Android'
    elif 'iphone' in lower_agent or 'ipad' in lower_agent:
        operating_system = 'iOS'

    return browser, operating_system


def _create_auth_audit_log(request, username, action, severity='warning', message=''):
    browser, operating_system = _get_client_browser_and_os(request)
    ip_address = _get_client_ip(request)
    AuditLog.objects.create(
        name=action,
        action=action,
        user=None,
        message=(message or f'Username={username or "unknown"}; IP={ip_address}; Browser={browser}; OS={operating_system}'),
        severity=severity,
    )


def _reset_account_lock(user):
    if not user:
        return False
    if user.account_locked_until and timezone.now() >= user.account_locked_until:
        user.failed_login_attempts = 0
        user.account_locked_until = None
        user.save(update_fields=['failed_login_attempts', 'account_locked_until'])
        return True
    return False


def _is_account_locked(user):
    if not user:
        return False
    if user.account_locked_until and timezone.now() < user.account_locked_until:
        return True
    if user.account_locked_until and timezone.now() >= user.account_locked_until:
        _reset_account_lock(user)
    return False


def _build_login_context(request, pending_user=None, pending_2fa=False, account_locked=False, lock_remaining_seconds=None, resend_cooldown_remaining=None):
    return {
        'pending_2fa': pending_2fa,
        'pending_user': pending_user,
        'account_locked': account_locked,
        'lock_remaining_seconds': lock_remaining_seconds,
        'resend_cooldown_remaining': resend_cooldown_remaining,
    }


def _issue_otp(request, user):
    otp_code = ''.join(random.choices(string.digits, k=6))
    user.otp_code = otp_code
    user.otp_expires_at = timezone.now() + timedelta(minutes=5)
    user.save(update_fields=['otp_code', 'otp_expires_at'])

    request.session['pending_2fa_user_id'] = str(user.pk)
    request.session['pending_2fa_otp'] = otp_code
    request.session['pending_2fa_otp_expires_at'] = user.otp_expires_at.isoformat()
    print(f"2FA OTP generated for {user.username}: {otp_code}")
    return otp_code


def _send_otp_email(user, otp_code):
    subject = 'Your verification code'
    body = (
        f'Hello {user.get_full_name() or user.username},\n\n'
        f'Your new verification code is: {otp_code}\n\n'
        'Please use this code to complete your sign-in.\n'
    )
    send_mail(subject, body, settings.DEFAULT_FROM_EMAIL, [user.email], fail_silently=True)


def signin(request):
    pending_user_id = request.session.get('pending_2fa_user_id')
    pending_user = None
    if pending_user_id:
        pending_user = get_user_model().objects.filter(pk=pending_user_id).first()

    if pending_user:
        _reset_account_lock(pending_user)

    invalid_attempts = request.session.get('invalid_login_attempts', 0)
    account_locked = bool(pending_user and _is_account_locked(pending_user))
    lock_remaining_seconds = None
    if account_locked and pending_user and pending_user.account_locked_until:
        lock_remaining_seconds = max(0, int((pending_user.account_locked_until - timezone.now()).total_seconds()))

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == 'resend_otp' and pending_user:
            if _is_account_locked(pending_user):
                messages.error(request, 'Your account is temporarily locked. Please try again later.')
                return render(request, 'signin.html', _build_login_context(request, pending_user=pending_user, pending_2fa=True, account_locked=True, lock_remaining_seconds=lock_remaining_seconds))

            now = timezone.now()
            if not pending_user.otp_resend_window_started_at or now - pending_user.otp_resend_window_started_at > timedelta(minutes=15):
                pending_user.otp_resend_count = 0
                pending_user.otp_resend_window_started_at = now
            if pending_user.otp_resend_count >= 5:
                messages.error(request, 'You have exceeded the maximum OTP resend attempts. Please try again later.')
                return render(request, 'signin.html', _build_login_context(request, pending_user=pending_user, pending_2fa=True, resend_cooldown_remaining=0))

            otp_code = _issue_otp(request, pending_user)
            pending_user.otp_resend_count += 1
            pending_user.save(update_fields=['otp_resend_count', 'otp_resend_window_started_at', 'otp_code', 'otp_expires_at'])
            print(f"2FA OTP resent for {pending_user.username}: {otp_code}")
            _send_otp_email(pending_user, otp_code)
            AuditLog.objects.create(
                name='OTP Resent',
                action='OTP Resent',
                user=pending_user,
                message=f'OTP resent for {pending_user.username}',
                severity='info',
            )
            messages.success(request, 'A new verification code has been sent.')
            return render(request, 'signin.html', _build_login_context(request, pending_user=pending_user, pending_2fa=True, resend_cooldown_remaining=60))

        otp_code = (request.POST.get("otp_code") or "").strip()
        if otp_code:
            if pending_user and _is_account_locked(pending_user):
                messages.error(request, 'Your account is temporarily locked. Please try again later.')
                return render(request, 'signin.html', _build_login_context(request, pending_user=pending_user, pending_2fa=True, account_locked=True, lock_remaining_seconds=lock_remaining_seconds))

            expected_otp = request.session.get('pending_2fa_otp')
            if pending_user and expected_otp and otp_code == expected_otp and pending_user.otp_code == expected_otp:
                request.session.pop('pending_2fa_user_id', None)
                request.session.pop('pending_2fa_otp', None)
                request.session.pop('pending_2fa_otp_expires_at', None)
                pending_user.otp_code = ''
                pending_user.otp_expires_at = None
                pending_user.failed_login_attempts = 0
                pending_user.account_locked_until = None
                pending_user.otp_resend_count = 0
                pending_user.otp_resend_window_started_at = None
                pending_user.save(update_fields=['otp_code', 'otp_expires_at', 'failed_login_attempts', 'account_locked_until', 'otp_resend_count', 'otp_resend_window_started_at'])
                login(request, pending_user)
                request.session['invalid_login_attempts'] = 0
                if pending_user.is_superuser:
                    AuditLog.objects.create(
                        name='Admin Login',
                        action=f'User {pending_user.username} signed in as superuser',
                        user=pending_user,
                        severity='success',
                    )
                    return redirect('/admin/')
                if is_it_compliance_auditor(pending_user):
                    AuditLog.objects.create(
                        name='Compliance Auditor Login',
                        action=f'User {pending_user.username} signed in as IT Compliance Auditor',
                        user=pending_user,
                        severity='success',
                    )
                    return redirect('auditor-dashboard')
                if is_backup_administrator(pending_user):
                    AuditLog.objects.create(
                        name='Backup Administrator Login',
                        action=f'User {pending_user.username} signed in as Backup Administrator',
                        user=pending_user,
                        severity='success',
                    )
                    return redirect('backup-dashboard')
                if is_supreme_approver(pending_user):
                    AuditLog.objects.create(
                        name='Supreme Approver Login',
                        action=f'User {pending_user.username} signed in as Supreme Approver',
                        user=pending_user,
                        severity='success',
                    )
                    return redirect('supreme-approver-dashboard')
                if is_operations_manager(pending_user):
                    AuditLog.objects.create(
                        name='Operations Manager Login',
                        action=f'User {pending_user.username} signed in as Operations Manager',
                        user=pending_user,
                        severity='success',
                    )
                    return redirect('operations-dashboard')
                if is_warehouse_staff(pending_user):
                    AuditLog.objects.create(
                        name='Warehouse Ops Login',
                        action=f'User {pending_user.username} signed in as Warehouse Operations',
                        user=pending_user,
                        severity='success',
                    )
                    return redirect('warehouse-operations-dashboard')
                if is_courier(pending_user):
                    AuditLog.objects.create(
                        name='Courier Login',
                        action=f'User {pending_user.username} signed in as Courier',
                        user=pending_user,
                        severity='success',
                    )
                    return redirect('courier-dashboard')
                if is_dr_team(pending_user):
                    AuditLog.objects.create(
                        name='DR Team Login',
                        action=f'User {pending_user.username} signed in as DR Team',
                        user=pending_user,
                        severity='success',
                    )
                    return redirect('investigation-dashboard')
                AuditLog.objects.create(
                    name='Unauthorized Dashboard Login',
                    action=f'User {pending_user.username} attempted dashboard login without appropriate role',
                    user=pending_user,
                    severity='warning',
                )
                logout(request)
                messages.error(request, 'You do not have access to the dashboard.')
                return render(request, 'signin.html', _build_login_context(request, pending_2fa=False))
            messages.error(request, 'Invalid verification code')
            return render(request, 'signin.html', _build_login_context(request, pending_user=pending_user, pending_2fa=True))

        username = request.POST.get('username')
        password = request.POST.get('password')

        if pending_user and _is_account_locked(pending_user):
            messages.error(request, 'Your account is temporarily locked. Please try again later.')
            return render(request, 'signin.html', _build_login_context(request, pending_user=pending_user, pending_2fa=True, account_locked=True, lock_remaining_seconds=lock_remaining_seconds))

        if pending_user and not username and not password:
            return render(request, 'signin.html', _build_login_context(request, pending_user=pending_user, pending_2fa=True))

        user = authenticate(request, username=username, password=password)

        if user:
            request.session['invalid_login_attempts'] = 0
            _reset_account_lock(user)
            otp_code = _issue_otp(request, user)
            _send_otp_email(user, otp_code)
            return render(request, 'signin.html', _build_login_context(request, pending_user=user, pending_2fa=True))

        if username:
            user_model = get_user_model()
            attempted_user = user_model.objects.filter(username=username).first()
            if attempted_user:
                attempted_user.failed_login_attempts += 1
                if attempted_user.failed_login_attempts >= 3:
                    attempted_user.account_locked_until = timezone.now() + timedelta(minutes=15)
                    attempted_user.save(update_fields=['failed_login_attempts', 'account_locked_until'])
                    _create_auth_audit_log(request, username, 'Account Locked', severity='warning', message=f'Account locked for {username}')
                    if attempted_user.email:
                        send_mail(
                            'Account Temporarily Locked',
                            f'Hello {attempted_user.get_full_name() or attempted_user.username},\n\nYour account has been temporarily locked after multiple unsuccessful sign-in attempts. Please wait 15 minutes before trying again or contact your system administrator.\n',
                            settings.DEFAULT_FROM_EMAIL,
                            [attempted_user.email],
                            fail_silently=True,
                        )
                    messages.error(request, 'ACCOUNT TEMPORARILY LOCKED')
                    return render(request, 'signin.html', _build_login_context(request, pending_user=attempted_user, pending_2fa=True, account_locked=True, lock_remaining_seconds=900))
                else:
                    attempted_user.save(update_fields=['failed_login_attempts'])
                    _create_auth_audit_log(request, username, 'Login Failed', severity='warning', message=f'Failed login attempt for {username}')
                    remaining_attempts = max(0, 3 - attempted_user.failed_login_attempts)
                    messages.error(request, f'Invalid username or password. Remaining attempts: {remaining_attempts}')
                    return render(request, 'signin.html', _build_login_context(request, pending_user=attempted_user, pending_2fa=False))

        _create_auth_audit_log(request, username, 'Login Failed', severity='warning', message='Failed login attempt for unknown user')
        messages.error(request, 'Invalid username or password.')

    return render(request, 'signin.html', _build_login_context(request, pending_user=pending_user, pending_2fa=bool(pending_user_id), account_locked=account_locked, lock_remaining_seconds=lock_remaining_seconds))


def signout(request):
    signed_out_user = request.user if request.user.is_authenticated else None
    logout(request)
    if signed_out_user:
        AuditLog.objects.create(
            name='User Logout',
            action=f'User {signed_out_user.username} signed out',
            user=signed_out_user,
            severity='info',
        )
    return redirect('signin')


@user_passes_test(lambda u: u.is_superuser, login_url='signin')
@login_required(login_url='signin')
def dashboard(request):
    tape_form = AddTapeForm(request.POST or None)
    user_form = CustomUserCreationForm(request.POST or None)
    edit_user_form = None
    role_creation_form = RoleCreationForm(request.POST or None)
    selected_group_id = request.GET.get('selected_group')
    role_feature_initial = {}
    selected_group = None
    if selected_group_id:
        selected_group = Role.objects.filter(pk=selected_group_id).first()
    if not selected_group:
        selected_group = Role.objects.filter(is_active=True).first()
    if selected_group:
        template = RoleTemplate.objects.filter(group=selected_group.group).first() if selected_group.group else None
        role_feature_initial = {
            'role': selected_group,
            'features': selected_group.get_active_features() if selected_group else [],
        }
    role_feature_form = RoleFeatureUpdateForm(request.POST or None, initial=role_feature_initial)
    role_form = UserRoleAssignmentForm(request.POST or None)
    role_form.fields['user'].queryset = User.objects.filter(is_active=True, verified=True).order_by('username')
    role_feature_form.fields['group'].queryset = Role.objects.filter(is_active=True).order_by('sort_order', 'name')
    role_form.fields['group'].queryset = Role.objects.filter(is_active=True).order_by('sort_order', 'name')
    edit_user = None

    if request.method == 'POST':
        if request.POST.get('form_type') == 'add_tape':
            if tape_form.is_valid():
                data = tape_form.cleaned_data
                payload_fields = {
                    k: (str(v) if not isinstance(v, (dict, list)) else v)
                    for k, v in data.items()
                }
                pa = _build_pending_tape_approval(
                    request,
                    transaction_type='Add Tape',
                    summary=f"Register tape {payload_fields.get('volser') or payload_fields.get('barcode')}",
                    payload_fields=payload_fields,
                    priority='Medium',
                    risk_level='High',
                )
                if pa:
                    AuditLog.objects.create(
                        name='Pending Tape Registration',
                        action=f'Pending registration for tape {payload_fields.get("volser") or payload_fields.get("barcode")}',
                        user=request.user,
                        severity='info',
                    )
                    messages.success(request, 'Tape registration submitted for approval.')
                    return redirect('dashboard')
                messages.error(request, 'Failed to submit tape registration for approval.')
            else:
                messages.error(request, 'Please correct the tape form errors below and try again.')
        elif request.POST.get('form_type') == 'add_user':
            if user_form.is_valid():
                user = user_form.save(commit=False)
                user.is_staff = False
                user.is_active = False
                user.verified = False
                user.save()

                vehicle_number = (user_form.cleaned_data.get('vehicle_number') or '').strip()
                if user.role == 'courier' and vehicle_number:
                    courier_profile = CourierProfile.objects.filter(user=user).first()
                    if courier_profile:
                        courier_profile.vehicle_number = vehicle_number
                        courier_profile.save(update_fields=['vehicle_number'])
                    else:
                        courier_id = f'CR-{uuid.uuid4().hex[:8].upper()}'
                        CourierProfile.objects.create(
                            user=user,
                            courier_id=courier_id,
                            full_name=user.get_full_name() or user.username,
                            phone_number='',
                            email=user.email or '',
                            vehicle_number=vehicle_number,
                            active_status=True,
                        )

                AuditLog.objects.create(
                    name='User Created',
                    action=f'Created user {user.username} pending verification',
                    user=request.user,
                    severity='info',
                )
                messages.success(request, 'New user created successfully and awaits verification by Backup Administrator.')
                return redirect('dashboard')
            else:
                messages.error(request, 'Please correct the user form errors below and try again.')
        elif request.POST.get('form_type') == 'create_role':
            if role_creation_form.is_valid():
                role_name = role_creation_form.cleaned_data['role_name']
                selected_features = list(role_creation_form.cleaned_data['features'])
                role_slug = re.sub(r'[^a-z0-9]+', '_', role_name.strip().lower()).strip('_') or 'role'
                group, _ = Group.objects.get_or_create(name=role_name)
                role, created = Role.objects.get_or_create(
                    slug=role_slug,
                    defaults={
                        'name': role_name,
                        'dashboard_key': role_slug,
                        'group': group,
                    },
                )
                if not created:
                    messages.error(request, 'Role name already exists.')
                else:
                    if not role.group_id:
                        role.group = group
                        role.save(update_fields=['group'])
                    existing_role_features = {rf.feature.feature_key: rf for rf in RoleFeature.objects.filter(role=role)}
                    for feature in selected_features:
                        rf = existing_role_features.get(feature.feature_key)
                        if rf:
                            if not rf.is_active:
                                rf.is_active = True
                                rf.assigned_by = request.user
                                rf.save(update_fields=['is_active', 'assigned_by', 'updated_at'])
                        else:
                            RoleFeature.objects.create(
                                role=role,
                                feature=feature,
                                is_active=True,
                                assigned_by=request.user,
                            )
                    AuditLog.objects.create(
                        name='Role Created',
                        action=f'Created role {role_name}',
                        user=request.user,
                        severity='success',
                    )
                    messages.success(request, 'New role created successfully.')
                    return redirect('dashboard')
            else:
                messages.error(request, 'Please correct the role creation errors below and try again.')
        elif request.POST.get('form_type') == 'assign_role':
            if role_form.is_valid():
                user = role_form.cleaned_data['user']
                role = role_form.cleaned_data.get('role') or role_form.cleaned_data.get('group')
                user.groups.clear()
                if role.group:
                    user.groups.add(role.group)
                user.role = role.dashboard_key
                user.is_staff = role.dashboard_key in ['admin', 'system_administrator', 'system administrator']
                user.save()
                AuditLog.objects.create(
                    name='Role Assigned',
                    action=f'Assigned {role.name} role to {user.username}',
                    user=request.user,
                    severity='success',
                )
                messages.success(request, 'User role updated successfully.')
                return redirect('dashboard')
            else:
                messages.error(request, 'Please correct the role assignment errors below and try again.')
        elif request.POST.get('form_type') == 'update_role_features':
            if role_feature_form.is_valid():
                role = role_feature_form.cleaned_data.get('role') or role_feature_form.cleaned_data.get('group')
                features = role_feature_form.cleaned_data['features']
                current_feature_keys = list(role.get_active_features().values_list('feature_key', flat=True))
                backup_admin = User.objects.filter(is_active=True, groups__name='Backup Administrator').first()
                payload = {
                    'role_id': str(role.pk),
                    'role_name': role.name,
                    'previous_feature_keys': current_feature_keys,
                    'feature_keys': features,
                }
                from .utils import create_pending_approval

                create_pending_approval(
                    transaction_type='Role Feature Update',
                    module='Administration',
                    summary=f'Update role features for {role.name}',
                    requester=request.user,
                    backup_administrator=backup_admin,
                    priority='High',
                    risk_level='High',
                    status='Pending',
                    request_payload=payload,
                    requested_changes=[{'action': 'replace_role_features', 'role': role.name, 'previous': current_feature_keys, 'new': features}],
                    related_model='role_feature',
                    related_object_id=str(role.pk),
                    audit_history=[{
                        'action': 'Requested',
                        'user': request.user.get_full_name() or request.user.username,
                        'timestamp': timezone.localtime().isoformat(),
                        'previous_configuration': current_feature_keys,
                        'new_configuration': features,
                    }],
                )
                AuditLog.objects.create(
                    name='Role Features Updated',
                    action=f'Requested feature update for {role.name}',
                    user=request.user,
                    severity='info',
                )
                messages.success(request, f'Feature change request for role {role.name} was submitted for approval.')
                return redirect(f"{reverse('dashboard')}?active_tab=roles&active_panel=%23editRoleFeaturesPanel")
            else:
                messages.error(request, 'Please correct the role feature form errors below and try again.')
        elif request.POST.get('form_type') == 'edit_user':
            user_id = request.POST.get('user_id')
            edit_user = get_object_by_uuid_pk(User, user_id)
            edit_user_form = CustomUserEditForm(request.POST, instance=edit_user)
            if edit_user_form.is_valid():
                user = edit_user_form.save(commit=False)
                user.is_staff = user.role == 'admin'
                user.save()
                AuditLog.objects.create(
                    name='User Updated',
                    action=f'Updated user {user.username}',
                    user=request.user,
                    severity='success',
                )
                messages.success(request, 'User updated successfully.')
                return redirect('dashboard')
            else:
                messages.error(request, 'Please correct the user edit errors below and try again.')

    if not edit_user_form:
        edit_user_id = request.GET.get('edit_user')
        if edit_user_id:
            edit_user = get_object_by_uuid_pk(User, edit_user_id)
            if edit_user:
                edit_user_form = CustomUserEditForm(instance=edit_user)

    users = User.objects.all().order_by("username")
    groups = Role.objects.filter(is_active=True).order_by("sort_order", "name")
    role_templates = RoleTemplate.objects.select_related('group').all()
    shipments = Shipment.objects.all().order_by("shipment_id")
    reports = ReportTemplate.objects.all()
    audit_logs = AuditLog.objects.order_by("-timestamp")

    role_templates_by_group = {template.group.name: template for template in role_templates if template.group}

    pending_users = User.objects.filter(verified=False).order_by('date_joined')

    user_feature_names = []
    active_role = get_user_role(request.user)
    if active_role:
        for feature in active_role.get_active_features():
            user_feature_names.append(feature.feature_key)
    user_feature_names = sorted(set(user_feature_names))

    tapes = Tape.objects.all().order_by('-date_registered')
    tape_search = request.GET.get('tape_search', '').strip()
    if tape_search:
        search_q = (
            Q(volser__icontains=tape_search) |
            Q(barcode__icontains=tape_search) |
            Q(rfid_tag__icontains=tape_search) |
            Q(tape_type__icontains=tape_search) |
            Q(manufacturer__icontains=tape_search) |
            Q(status__icontains=tape_search) |
            Q(current_location__icontains=tape_search) |
            Q(remarks__icontains=tape_search)
        )
        parsed_date = parse_date(tape_search)
        if parsed_date:
            search_q |= Q(retention_end_date=parsed_date)
        tapes = tapes.filter(search_q)
    total_tapes = tapes.count()
    active_tapes = tapes.filter(status="Active").count()
    archived_tapes = tapes.filter(status="Retained").count()
    retention_due = tapes.filter(retention_end_date__lte=timezone.localdate()).count()
    pending_shipments = shipments.filter(status__iexact="Pending").count()
    alert_count = AuditLog.objects.filter(severity__in=["warning", "error"]).count()

    dashboard_tabs = get_dashboard_tabs(request.user, ADMIN_FEATURE_TABS)

    context = {
        "users": users,
        "groups": groups,
        "shipments": shipments,
        "reports": reports,
        "audit_logs": audit_logs,
        "tapes": tapes,
        "tape_search": tape_search,
        "role_creation_form": role_creation_form,
        "role_form": role_form,
        "role_feature_form": role_feature_form,
        "role_templates": role_templates,
        "role_templates_by_group": role_templates_by_group,
        "edit_user_form": edit_user_form,
        "edit_user": edit_user,
        "total_users": users.count(),
        "total_tapes": total_tapes,
        "active_tapes": active_tapes,
        "archived_tapes": archived_tapes,
        "retention_due": retention_due,
        "alert_count": alert_count,
        "pending_shipments": pending_shipments,
        "tape_form": tape_form,
        "user_form": user_form,
        "pending_users": pending_users,
        "user_feature_names": user_feature_names,
        "dashboard_tabs": dashboard_tabs,
    }

    return render(request, "dashboard.html", context)


@user_passes_test(lambda u: u.is_superuser or is_backup_administrator(u) or is_operations_manager(u), login_url='signin')
@login_required(login_url='signin')
def backup_dashboard(request):
    tape_form = AddTapeForm(request.POST or None)
    tape_action_form = None
    selected_tape = None
    show_add_tape_panel = False
    show_tape_actions_panel = False
    show_tape_inventory_panel = False
    show_settings_panel = False
    show_audit_panel = False
    show_alerts_panel = False
    show_reports_panel = False
    show_shipments_panel = False
    show_add_shipment_panel = False
    show_edit_shipment_panel = False
    show_reconciliation_panel = False
    show_reconciliation_reports_panel = False
    show_add_reconciliation_panel = False
    show_unscanned_tapes = False
    show_print_barcodes_panel = False
    show_admin_panel = False
    show_admin_history = False
    selected_shipment = None
    selected_reconciliation = None
    approval_shipment = None
    add_shipment_form = ShipmentForm(request.POST or None, prefix='add', request=request)
    edit_shipment_form = None
    assignment_form = BackupShipmentAssignmentForm(request.POST or None)
    reconciliation_form = ReconciliationForm(request.POST or None, prefix='reconciliation')
    reconciliation_result_form = ReconciliationResultForm(request.POST or None, prefix='result')
    application_settings = ApplicationSetting.objects.first() or ApplicationSetting.objects.create()
    settings_form = None
    profile_form = None
    show_profile_panel = False
    hide_dashboard_sidebar = False

    active_feature_key = request.GET.get('feature_key') or request.POST.get('feature_key')
    if active_feature_key and not getattr(request, 'signed_dashboard_navigation', False):
        hide_dashboard_sidebar = True
    feature_panel_state = get_feature_panel_state(active_feature_key)
    if feature_panel_state:
        if feature_panel_state.get('show_add_tape_panel'):
            show_add_tape_panel = True
        if feature_panel_state.get('show_tape_inventory_panel'):
            show_tape_inventory_panel = True
        if feature_panel_state.get('show_shipments_panel'):
            show_shipments_panel = True
        if feature_panel_state.get('show_reconciliation_panel'):
            show_reconciliation_panel = True
        if feature_panel_state.get('show_reports_panel'):
            show_reports_panel = True
        if feature_panel_state.get('show_audit_panel'):
            show_audit_panel = True

    selected_tape_id = request.GET.get('selected_tape') or request.POST.get('selected_tape')
    if selected_tape_id:
        selected_tape = get_object_by_uuid_pk(Tape, selected_tape_id)
        if selected_tape:
            show_tape_actions_panel = True

    schema_preview = None
    pending_schema_upload = None
    pending_schema_file = None
    pending_schema_key = None
    pending_schema_session = request.session.get('pending_schema_preview', {})
    if pending_schema_session:
        schema_preview = pending_schema_session.get('preview')

    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        if form_type == 'backup_admin_assignment':
            shipment_id = request.POST.get('shipment_id')
            shipment = get_object_by_uuid_pk(Shipment, shipment_id)
            submit_action = (request.POST.get('submit_action') or request.POST.get('decision') or 'approve').lower()
            decision = 'approve' if submit_action != 'reject' else 'reject'
            if shipment:
                approval_shipment = shipment
            if shipment and assignment_form.is_valid():
                comments = (assignment_form.cleaned_data.get('comments') or '').strip()
                if decision == 'approve':
                    # Backup admin approval moves the shipment to awaiting supreme approval.
                    # Do NOT notify or assign courier for pickup at this stage.
                    tape = assignment_form.cleaned_data['tape']
                    courier_selection = assignment_form.cleaned_data['courier']
                    courier_profile = None
                    courier_user = None
                    courier_name = ''
                    courier_contact = ''
                    if courier_selection and isinstance(courier_selection, str):
                        if courier_selection.startswith('profile:'):
                            profile_id = courier_selection.split(':', 1)[1]
                            courier_profile = CourierProfile.objects.filter(pk=profile_id).first()
                            if courier_profile:
                                courier_user = courier_profile.user
                                courier_name = courier_profile.full_name
                                courier_contact = courier_profile.phone_number
                        elif courier_selection.startswith('user:'):
                            user_id = courier_selection.split(':', 1)[1]
                            courier_user = get_user_model().objects.filter(pk=user_id).first()
                            if courier_user:
                                courier_name = courier_user.get_full_name() or courier_user.username
                                courier_contact = courier_user.email or ''
                                courier_profile = getattr(courier_user, 'courier_profile', None)
                        else:
                            courier_profile = CourierProfile.objects.filter(pk=courier_selection).first()
                            if courier_profile:
                                courier_user = courier_profile.user
                                courier_name = courier_profile.full_name
                                courier_contact = courier_profile.phone_number
                            else:
                                courier_user = get_user_model().objects.filter(pk=courier_selection).first()
                                if courier_user:
                                    courier_name = courier_user.get_full_name() or courier_user.username
                                    courier_contact = courier_user.email or ''
                                    courier_profile = getattr(courier_user, 'courier_profile', None)

                    # Persist the selected tape and courier details but do not notify courier.
                    shipment.tapes.add(tape)
                    shipment.number_of_tapes = shipment.tapes.count()
                    shipment.courier_name = courier_name or (courier_profile.full_name if courier_profile else '')
                    shipment.courier_contact = courier_contact or (courier_profile.phone_number if courier_profile else '')
                    shipment.tracking_number = f"TRK-{shipment.shipment_id[:8].upper()}"
                    previous_status = shipment.status
                    shipment.status = 'Pending'
                    shipment.approval_stage = 'awaiting_supreme'
                    shipment.approved_by_backup = request.user
                    shipment.approved_by = request.user
                    shipment.approval_date = timezone.localtime()
                    shipment.approval_remarks = comments or 'Approved by backup administrator. Pending supreme approver.'
                    shipment.last_updated_by = request.user
                    shipment.save(update_fields=['status', 'approved_by', 'approved_by_backup', 'approval_stage', 'approval_date', 'approval_remarks', 'last_updated_by', 'last_updated_at', 'courier_name', 'courier_contact', 'tracking_number', 'number_of_tapes'])
                    ShipmentApprovalHistory.objects.create(shipment=shipment, action='Backup Approved', comments=comments, user=request.user)
                    AuditLog.objects.create(name='Shipment Backup Approved', action=f'Shipment {shipment.shipment_id} approved by backup administrator and awaiting supreme approval.', user=request.user, severity='info')

                    # Create a pending approval record to notify the supreme approver(s)
                    try:
                        from .utils import create_pending_approval
                        create_pending_approval(
                            transaction_type='Shipment',
                            module='Shipment Workflow',
                            summary=f"Shipment {shipment.shipment_id} awaiting supreme approval",
                            requester=shipment.created_by,
                            backup_administrator=request.user,
                            branch=shipment.requesting_branch or None,
                            priority=shipment.priority_level or 'Normal',
                            risk_level='High' if shipment.priority_level in {'High', 'Critical'} else 'Medium',
                            status='Awaiting Supreme Approval',
                            request_date=timezone.localtime(),
                            request_payload={'shipment_id': shipment.shipment_id},
                            related_object_id=str(shipment.pk),
                            related_model='shipment',
                        )
                    except Exception:
                        pass

                    # Notify operators that backup approved (do not notify courier)
                    operator = shipment.created_by
                    if operator:
                        AuditLog.objects.create(
                            name='Shipment Backup Approved',
                            action=(f'Your shipment request {shipment.shipment_id} was approved by the backup administrator and is awaiting supreme approval.'),
                            user=operator,
                            severity='info',
                        )
                        if application_settings.email_alerts_enabled and operator.email:
                            send_mail(
                                f'Shipment {shipment.shipment_id} pending supreme approval',
                                f'Your shipment request {shipment.shipment_id} was approved by the backup administrator and is awaiting supreme approval.',
                                settings.DEFAULT_FROM_EMAIL,
                                [operator.email],
                                fail_silently=True,
                            )

                    messages.success(request, f'You approved shipment {shipment.shipment_id}; it is now awaiting supreme approval.')
                else:
                    shipment.status = 'Rejected'
                    shipment.approved_by = request.user
                    shipment.approval_date = timezone.localtime()
                    shipment.approval_remarks = comments or 'Rejected by backup administrator.'
                    shipment.last_updated_by = request.user
                    shipment.save(update_fields=['status', 'approved_by', 'approval_date', 'approval_remarks', 'last_updated_by', 'last_updated_at'])
                    ShipmentApprovalHistory.objects.create(shipment=shipment, action='Rejected', comments=comments, user=request.user)
                    operator = shipment.created_by
                    if operator:
                        AuditLog.objects.create(
                            name='Shipment Rejected',
                            action=f'Your shipment request {shipment.shipment_id} was rejected. {comments}'.strip(),
                            user=operator,
                            severity='warning',
                        )
                        if application_settings.email_alerts_enabled and operator.email:
                            send_mail(
                                f'Shipment {shipment.shipment_id} rejected',
                                f'Your shipment request {shipment.shipment_id} was rejected by the backup administrator. {comments}'.strip(),
                                settings.DEFAULT_FROM_EMAIL,
                                [operator.email],
                                fail_silently=True,
                            )
                    messages.success(request, 'Shipment request was rejected.')
                return redirect(f'{reverse("backup-dashboard")}?show_alerts=1')

            show_alerts_panel = True
            messages.error(request, 'Please provide a valid tape barcode and courier before approving the shipment.')
        elif form_type == 'verify_user':
            user_id = request.POST.get('user_id')
            pending_user = None
            if is_valid_uuid(user_id):
                pending_user = User.objects.filter(pk=user_id, verified=False).first()
            if pending_user:
                pending_user.is_active = True
                pending_user.verified = True
                pending_user.verified_at = timezone.now()
                pending_user.save()
                AuditLog.objects.create(
                    name='User Verified',
                    action=f'Verified user {pending_user.username}',
                    user=request.user,
                    severity='success',
                )
                messages.success(request, f'User {pending_user.username} has been verified.')
            else:
                messages.error(request, 'Unable to verify the selected user.')
            return redirect('backup-dashboard')
        elif form_type == 'approve_tape_request':
            tape_request_id = request.POST.get('request_id')
            tape_request = get_object_by_uuid_pk(TapeRequest, tape_request_id)
            if tape_request and tape_request.status == 'Pending':
                shipment = Shipment.objects.create(
                    shipment_date=timezone.localdate(),
                    shipment_type='Retrieval',
                    status='Approved',
                    priority_level='Normal',
                    number_of_tapes=tape_request.quantity,
                    source_location=tape_request.tape.current_location or '',
                    releasing_custodian=request.user.get_full_name() or request.user.username,
                    destination_location=tape_request.destination_location,
                    receiving_organization=tape_request.receiving_organization,
                    expected_delivery_date=timezone.localdate() + timedelta(days=2),
                    approved_by=request.user,
                    approval_date=timezone.localtime(),
                    approval_remarks=request.POST.get('approval_notes', '').strip() or 'Approved from tape request',
                    created_by=request.user,
                    last_updated_by=request.user,
                )
                shipment.tapes.add(tape_request.tape)
                tape_request.shipment = shipment
                tape_request.status = 'Approved'
                tape_request.approved_by = request.user
                tape_request.approved_at = timezone.localtime()
                tape_request.approval_notes = request.POST.get('approval_notes', '').strip() or 'Approved from tape request'
                tape_request.save(update_fields=['shipment', 'status', 'approved_by', 'approved_at', 'approval_notes', 'updated_at'])
                AuditLog.objects.create(
                    name='Tape Request Approved',
                    action=f'Approved tape request {tape_request.id} into shipment {shipment.shipment_id}',
                    user=request.user,
                    severity='success',
                )
                messages.success(request, 'Tape request approved and converted into a shipment.')
                return redirect(f'{reverse("backup-dashboard")}?show_shipments=1')
            messages.error(request, 'Unable to approve the selected tape request.')
            return redirect(f'{reverse("backup-dashboard")}?show_shipments=1')
        elif form_type == 'forward_exception_to_dr_team':
            alert_id = request.POST.get('alert_id')
            alert_entry = AuditLog.objects.filter(pk=alert_id, severity__in=['warning', 'error']).first()
            if alert_entry:
                exception_id, description = _parse_exception_metadata(alert_entry)
                if exception_id:
                    exception_obj = ShipmentException.objects.filter(exception_id=exception_id).first()
                    if exception_obj:
                        exception_obj.status = 'Investigating'
                        exception_obj.save(update_fields=['status'])

                    dr_users = get_dr_team_users()
                    if dr_users.exists():
                        notification_url = reverse('investigation-dashboard', kwargs={'exception_id': exception_id})
                        subject = f'Forwarded exception alert: {exception_id}'
                        body = (
                            f'An exception alert has been forwarded to the DR team by {request.user.get_full_name() or request.user.username}.\n\n'
                            f'Exception ID: {exception_id}\n'
                            f'Description: {description or "No description available."}\n'
                            f'Original alert: {alert_entry.action or alert_entry.name}\n'
                            f'Review the exception investigation dashboard: {request.build_absolute_uri(notification_url)}\n'
                        )
                        application_settings = ApplicationSetting.objects.first() or ApplicationSetting.objects.create()
                        if application_settings.email_alerts_enabled:
                            send_mail(subject, body, settings.DEFAULT_FROM_EMAIL, get_dr_team_email_recipients(), fail_silently=True)
                        for dr_user in dr_users:
                            AuditLog.objects.create(
                                name='Exception Alert Forwarded',
                                action=f'Exception {exception_id} forwarded to you by {request.user.get_full_name() or request.user.username}',
                                message=f'exception_id={exception_id}|target_url={notification_url}',
                                user=dr_user,
                                severity='warning',
                            )
                        messages.success(request, 'Exception alert forwarded to DR team members.')
                    else:
                        messages.warning(request, 'No DR team members found to forward the exception alert.')
                else:
                    messages.error(request, 'Unable to identify exception details for this alert.')
            else:
                messages.error(request, 'Selected alert could not be found.')
            return redirect(f'{reverse("backup-dashboard")}?show_alerts=1')
        elif form_type == 'add_tape':
            if tape_form.is_valid():
                data = tape_form.cleaned_data
                payload_fields = {k: (str(v) if not isinstance(v, (dict, list)) else v) for k, v in data.items()}
                pa = _build_pending_tape_approval(
                    request,
                    transaction_type='Add Tape',
                    summary=f"Register tape {payload_fields.get('volser') or payload_fields.get('barcode')}",
                    payload_fields=payload_fields,
                    priority='Medium',
                    risk_level='High',
                )
                if pa:
                    AuditLog.objects.create(
                        name='Pending Tape Registration',
                        action=f'Pending registration for tape {payload_fields.get("volser") or payload_fields.get("barcode")}',
                        user=request.user,
                        severity='info',
                    )
                    messages.success(request, 'Tape registration submitted for approval.')
                    return redirect(f'{reverse("backup-dashboard")}?show_add_tape=1')
                show_add_tape_panel = True
                messages.error(request, 'Failed to submit tape registration for approval.')
            else:
                show_add_tape_panel = True
                messages.error(request, 'Please correct the tape form errors below and try again.')
        elif form_type == 'tape_action':
            tape_id = request.POST.get('selected_tape')
            selected_tape = get_object_by_uuid_pk(Tape, tape_id)
            if selected_tape:
                tape_action_form = TapeForm(request.POST, instance=selected_tape)
                action = request.POST.get('action')
                if action == 'edit_details':
                    if tape_action_form.is_valid():
                        data = tape_action_form.cleaned_data
                        UserModel = get_user_model()
                        backup_admin = UserModel.objects.filter(is_active=True, groups__name='Backup Administrator').first()
                        payload_fields = {k: (str(v) if not isinstance(v, (dict, list)) else v) for k, v in data.items()}
                        pa = create_pending_approval(
                            transaction_type='Update Tape',
                            module='Inventory',
                            summary=f"Update tape {selected_tape.volser}",
                            requester=request.user,
                            backup_administrator=backup_admin,
                            priority='Medium',
                            risk_level='Medium',
                            status='Pending',
                            request_payload=payload_fields,
                            requested_changes=[{'action': 'update', 'model': 'Tape', 'object_id': str(selected_tape.pk), 'fields': payload_fields}],
                            related_model='tape',
                            related_object_id=str(selected_tape.pk),
                            audit_history=[{
                                'action': 'Requested Update',
                                'user': request.user.get_full_name() or request.user.username,
                                'timestamp': timezone.localtime().isoformat(),
                            }],
                        )
                        if pa:
                            AuditLog.objects.create(
                                name='Pending Tape Update',
                                action=f'Pending update for tape {selected_tape.volser}',
                                user=request.user,
                                severity='info',
                            )
                            messages.success(request, f'Update for tape {selected_tape.volser} submitted for approval.')
                            return redirect(f'{reverse("backup-dashboard")}?show_tape_actions=1&selected_tape={selected_tape.id}')
                        show_tape_actions_panel = True
                        messages.error(request, 'Failed to submit tape update for approval.')
                    else:
                        show_tape_actions_panel = True
                        messages.error(request, 'Please correct the tape form errors and try again.')
                elif action == 'update_location':
                    new_location = request.POST.get('current_location', selected_tape.current_location)
                    UserModel = get_user_model()
                    backup_admin = UserModel.objects.filter(is_active=True, groups__name='Backup Administrator').first()
                    payload_fields = {'current_location': new_location}
                    pa = create_pending_approval(
                        transaction_type='Update Tape Location',
                        module='Inventory',
                        summary=f"Update location for {selected_tape.volser} to {new_location}",
                        requester=request.user,
                        backup_administrator=backup_admin,
                        priority='Low',
                        risk_level='Medium',
                        status='Pending',
                        request_payload=payload_fields,
                        requested_changes=[{'action': 'update', 'model': 'Tape', 'object_id': str(selected_tape.pk), 'fields': payload_fields}],
                        related_model='tape',
                        related_object_id=str(selected_tape.pk),
                        audit_history=[{
                            'action': 'Requested Location Update',
                            'user': request.user.get_full_name() or request.user.username,
                            'timestamp': timezone.localtime().isoformat(),
                        }],
                    )
                    if pa:
                        AuditLog.objects.create(
                            name='Pending Tape Location Update',
                            action=f'Pending location update for {selected_tape.volser} to {new_location}',
                            user=request.user,
                            severity='info',
                        )
                        messages.success(request, f'Location update for {selected_tape.volser} submitted for approval.')
                        return redirect(f'{reverse("backup-dashboard")}?show_tape_actions=1&selected_tape={selected_tape.id}')
                    messages.error(request, 'Failed to submit location update for approval.')
                    return redirect(f'{reverse("backup-dashboard")}?show_tape_actions=1&selected_tape={selected_tape.id}')
                elif action == 'mark_scratch':
                    try:
                        payload_fields = {'status': 'Scratch Eligible'}
                        pa = _build_pending_tape_approval(
                            request,
                            transaction_type='Mark Tape Scratch Eligible',
                            summary=f'Mark {selected_tape.volser} as scratch eligible',
                            payload_fields=payload_fields,
                            object_id=str(selected_tape.pk),
                            priority='Low',
                            risk_level='Medium',
                        )
                        if pa:
                            AuditLog.objects.create(
                                name='Pending Tape Scratch Mark',
                                action=f'Pending scratch mark for {selected_tape.volser}',
                                user=request.user,
                                severity='info',
                            )
                            messages.success(request, f'Mark scratch request for {selected_tape.volser} submitted for approval.')
                            return redirect(f'{reverse("backup-dashboard")}?show_tape_actions=1&selected_tape={selected_tape.id}')
                        messages.error(request, 'Failed to submit scratch mark for approval.')
                        return redirect(f'{reverse("backup-dashboard")}?show_tape_actions=1&selected_tape={selected_tape.id}')
                    except ValidationError as exc:
                        messages.error(request, exc.messages[0])
                        return redirect(f'{reverse("backup-dashboard")}?show_tape_actions=1&selected_tape={selected_tape.id}')
                elif action == 'mark_damaged':
                    payload_fields = {'status': 'Damaged'}
                    pa = _build_pending_tape_approval(
                        request,
                        transaction_type='Mark Tape Damaged',
                        summary=f'Mark {selected_tape.volser} as damaged',
                        payload_fields=payload_fields,
                        object_id=str(selected_tape.pk),
                        priority='Low',
                        risk_level='High',
                    )
                    if pa:
                        AuditLog.objects.create(
                            name='Pending Tape Damaged',
                            action=f'Pending damaged mark for {selected_tape.volser}',
                            user=request.user,
                            severity='info',
                        )
                        messages.success(request, f'Mark damaged request for {selected_tape.volser} submitted for approval.')
                        return redirect(f'{reverse("backup-dashboard")}?show_tape_actions=1&selected_tape={selected_tape.id}')
                    messages.error(request, 'Failed to submit damaged mark for approval.')
                    return redirect(f'{reverse("backup-dashboard")}?show_tape_actions=1&selected_tape={selected_tape.id}')
                elif action == 'scan_rfid':
                    payload_fields = {
                        'barcode': request.POST.get('barcode', selected_tape.barcode),
                        'rfid_tag': request.POST.get('rfid_tag', selected_tape.rfid_tag),
                    }
                    pa = _build_pending_tape_approval(
                        request,
                        transaction_type='Update Tape Barcode/RFID',
                        summary=f'Update barcode/RFID for {selected_tape.volser}',
                        payload_fields=payload_fields,
                        object_id=str(selected_tape.pk),
                        priority='Low',
                        risk_level='Medium',
                    )
                    if pa:
                        AuditLog.objects.create(
                            name='Pending Tape Barcode/RFID Update',
                            action=f'Pending barcode/RFID update for {selected_tape.volser}',
                            user=request.user,
                            severity='info',
                        )
                        messages.success(request, f'Barcode/RFID update for {selected_tape.volser} submitted for approval.')
                        return redirect(f'{reverse("backup-dashboard")}?show_tape_actions=1&selected_tape={selected_tape.id}')
                    messages.error(request, 'Failed to submit barcode/RFID update for approval.')
                    return redirect(f'{reverse("backup-dashboard")}?show_tape_actions=1&selected_tape={selected_tape.id}')
            else:
                messages.error(request, 'Please select a valid tape to update.')
                return redirect('backup-dashboard')
        elif form_type == 'add_shipment':
            if add_shipment_form.is_valid():
                shipment = add_shipment_form.save(commit=False)
                shipment.created_by = request.user
                shipment.last_updated_by = request.user
                shipment.save()
                add_shipment_form.save_m2m()
                shipment.save(update_fields=['last_updated_by', 'last_updated_at'])
                AuditLog.objects.create(
                    name='Shipment Created',
                    action=f'Created shipment {shipment.shipment_id}',
                    user=request.user,
                    severity='success',
                )
                if application_settings.shipment_notification_enabled:
                    notify_email_alert(
                        application_settings,
                        f'Shipment {shipment.shipment_id} Created',
                        f'Shipment {shipment.shipment_id} was created by {request.user.username}.',
                    )
                messages.success(request, 'Shipment created successfully.')
                return redirect(f'{reverse("backup-dashboard")}?show_shipments=1')
            else:
                show_shipments_panel = True
                show_add_shipment_panel = True
                messages.error(request, 'Please correct the shipment form errors and try again.')
        elif form_type == 'edit_shipment':
            shipment_pk = request.POST.get('shipment_pk')
            selected_shipment = get_object_by_uuid_pk(Shipment, shipment_pk)
            if selected_shipment:
                edit_shipment_form = ShipmentForm(request.POST, instance=selected_shipment, prefix='edit', request=request)
                if edit_shipment_form.is_valid():
                    shipment = edit_shipment_form.save(commit=False)
                    shipment.last_updated_by = request.user
                    shipment.save()
                    edit_shipment_form.save_m2m()
                    shipment.save(update_fields=['last_updated_by', 'last_updated_at'])
                    AuditLog.objects.create(
                        name='Shipment Updated',
                        action=f'Updated shipment {shipment.shipment_id}',
                        user=request.user,
                        severity='success',
                    )
                    if application_settings.shipment_notification_enabled:
                        notify_email_alert(
                            application_settings,
                            f'Shipment {shipment.shipment_id} Updated',
                            f'Shipment {shipment.shipment_id} was updated by {request.user.username}.',
                        )
                    messages.success(request, 'Shipment updated successfully.')
                    return redirect(f'{reverse("backup-dashboard")}?edit_shipment_pk={shipment.id}&show_shipments=1')
                else:
                    show_shipments_panel = True
                    show_edit_shipment_panel = True
                    messages.error(request, 'Please correct the shipment form errors and try again.')
            else:
                messages.error(request, 'Please select a valid shipment to update.')
                return redirect('backup-dashboard')
        elif form_type == 'create_reconciliation_from_request':
            request_id = request.POST.get('reconciliation_request_id')
            alert_entry = AuditLog.objects.filter(pk=request_id, name='Reconciliation Requested').first()
            payload = _parse_reconciliation_request_message(alert_entry.message if alert_entry else '')
            exception_id = (payload.get('exception_id') or request.POST.get('exception_id') or '').strip()
            requester_name = (payload.get('requester_name') or request.POST.get('requester_name') or 'DR Team').strip()
            comment = (payload.get('comment') or request.POST.get('comment') or '').strip()
            location = (request.POST.get('location') or 'Pending review').strip() or 'Pending review'

            reconciliation = Reconciliation.objects.create(
                location=location,
                performed_by=request.user,
                status='Open',
                notes=f"Initiated from DR request for exception {exception_id}. {comment}".strip(),
            )
            AuditLog.objects.create(
                name='Reconciliation Created',
                action=f'Created reconciliation {reconciliation.reconciliation_id} from DR request',
                user=request.user,
                severity='success',
            )
            if alert_entry and alert_entry.user:
                AuditLog.objects.create(
                    name='Reconciliation Authorized',
                    action=f'Your reconciliation request for exception {exception_id} was approved and opened as {reconciliation.reconciliation_id}',
                    user=alert_entry.user,
                    message=f'confirmation_id={reconciliation.reconciliation_id}',
                    severity='success',
                )
            messages.success(request, f'Reconciliation {reconciliation.reconciliation_id} was opened for {requester_name or "the DR team"}.')
            return redirect(f'{reverse("backup-dashboard")}?show_alerts=1')
        elif form_type == 'add_reconciliation':
            if reconciliation_form.is_valid():
                reconciliation = reconciliation_form.save(commit=False)
                reconciliation.performed_by = request.user
                reconciliation.save()
                AuditLog.objects.create(
                    name='Reconciliation Created',
                    action=f'Created reconciliation {reconciliation.reconciliation_id} for {reconciliation.location}',
                    user=request.user,
                    severity='success',
                )
                if reconciliation.assigned_operator:
                    AuditLog.objects.create(
                        name='Reconciliation Assigned',
                        action=f'Reconciliation {reconciliation.reconciliation_id} assigned to {reconciliation.assigned_operator.username}',
                        user=reconciliation.assigned_operator,
                        message=f'reconciliation_id={reconciliation.reconciliation_id}',
                        severity='warning',
                    )
                messages.success(request, 'Reconciliation created successfully.')
                return redirect(f'{reverse("backup-dashboard")}?show_reconciliation=1&reconciliation_pk={reconciliation.id}&show_unscanned_tapes=1')
            else:
                show_reconciliation_panel = True
                show_add_reconciliation_panel = True
                messages.error(request, 'Please correct the reconciliation form errors and try again.')
        elif form_type == 'start_reconciliation':
            reconciliation_pk = request.POST.get('reconciliation_pk')
            selected_reconciliation = get_object_by_uuid_pk(Reconciliation, reconciliation_pk)
            if selected_reconciliation:
                if selected_reconciliation.status == 'Open':
                    selected_reconciliation.status = 'In Progress'
                    selected_reconciliation.scan_started_at = timezone.localtime()
                    selected_reconciliation.save(update_fields=['status', 'scan_started_at'])
                    AuditLog.objects.create(
                        name='Reconciliation Started',
                        action=f'Started reconciliation {selected_reconciliation.reconciliation_id}',
                        user=request.user,
                        message=f'reconciliation_id={selected_reconciliation.reconciliation_id}',
                        severity='info',
                    )
                    messages.success(request, 'Reconciliation scan started.')
                else:
                    messages.info(request, 'Reconciliation is already in progress.')
            else:
                messages.error(request, 'Unable to start the selected reconciliation.')
            return redirect(f'{reverse("backup-dashboard")}?show_reconciliation=1&reconciliation_pk={selected_reconciliation.id if selected_reconciliation else ""}&show_unscanned_tapes=1')
        elif form_type == 'finish_reconciliation':
            reconciliation_pk = request.POST.get('reconciliation_pk')
            selected_reconciliation = get_object_by_uuid_pk(Reconciliation, reconciliation_pk)
            if selected_reconciliation:
                if selected_reconciliation.status == 'In Progress':
                    selected_reconciliation.status = 'Completed'
                    selected_reconciliation.scan_completed_at = timezone.localtime()
                    if selected_reconciliation.total_tapes_expected > 0:
                        selected_reconciliation.progress_percent = 100
                    selected_reconciliation.save(update_fields=['status', 'scan_completed_at', 'progress_percent'])
                    AuditLog.objects.create(
                        name='Reconciliation Finished',
                        action=f'Finished reconciliation {selected_reconciliation.reconciliation_id}',
                        user=request.user,
                        message=f'reconciliation_id={selected_reconciliation.reconciliation_id}|reconciliation_pk={selected_reconciliation.id}',
                        severity='success',
                    )
                    messages.success(request, 'Reconciliation scan finished.')
                else:
                    messages.info(request, 'Reconciliation is not currently in progress.')
            else:
                messages.error(request, 'Unable to finish the selected reconciliation.')
            return redirect(f'{reverse("backup-dashboard")}?show_reconciliation=1&reconciliation_pk={selected_reconciliation.id if selected_reconciliation else ""}&show_unscanned_tapes=1')
        elif form_type == 'close_reconciliation':
            reconciliation_pk = request.POST.get('reconciliation_pk')
            selected_reconciliation = get_object_by_uuid_pk(Reconciliation, reconciliation_pk)
            if selected_reconciliation:
                if selected_reconciliation.status in ['Completed', 'Open']:
                    selected_reconciliation.status = 'Closed'
                    selected_reconciliation.save(update_fields=['status', 'updated_at'])
                    AuditLog.objects.create(
                        name='Reconciliation Closed',
                        action=f'Closed reconciliation {selected_reconciliation.reconciliation_id}',
                        user=request.user,
                        message=f'reconciliation_id={selected_reconciliation.reconciliation_id}|reconciliation_pk={selected_reconciliation.id}',
                        severity='success',
                    )
                    messages.success(request, 'Reconciliation has been closed successfully.')
                else:
                    messages.info(request, 'Only open or completed reconciliations can be closed.')
            else:
                messages.error(request, 'Unable to close the selected reconciliation.')
            return redirect(f'{reverse("backup-dashboard")}?show_reconciliation=1&reconciliation_pk={selected_reconciliation.id if selected_reconciliation else ""}&show_unscanned_tapes=1')
        elif form_type == 'shift_reconciliation':
            reconciliation_pk = request.POST.get('reconciliation_pk')
            selected_reconciliation = get_object_by_uuid_pk(Reconciliation, reconciliation_pk)
            operator_pks = request.POST.getlist('operator_pks') or request.POST.getlist('operator_pk') or []
            operator_pks = [pk for pk in operator_pks if pk]
            if selected_reconciliation and operator_pks:
                selected_operators = list(CustomUser.objects.filter(pk__in=operator_pks, is_active=True))
                if selected_operators:
                    selected_reconciliation.assigned_operator = selected_operators[0]
                    selected_reconciliation.save(update_fields=['assigned_operator'])
                    reassigned_count = Tape.objects.filter(scan_status__in=['Pending', 'Assigned', 'Scanning']).update(scan_status='Assigned')
                    operator_names = ', '.join(operator.username for operator in selected_operators)
                    for selected_operator in selected_operators:
                        AuditLog.objects.create(
                            name='Reconciliation Shifted',
                            action=f'Scan duty for reconciliation {selected_reconciliation.reconciliation_id} was forwarded to {selected_operator.username}',
                            user=selected_operator,
                            message=f'reconciliation_id={selected_reconciliation.reconciliation_id}|reconciliation_pk={selected_reconciliation.id}|operator={selected_operator.username}|assigned_count={reassigned_count}',
                            severity='error',
                        )
                    AuditLog.objects.create(
                        name='Reconciliation Shift Notification',
                        action=f'Scan duty for reconciliation {selected_reconciliation.reconciliation_id} was forwarded to {operator_names}',
                        user=request.user,
                        message=f'reconciliation_id={selected_reconciliation.reconciliation_id}|reconciliation_pk={selected_reconciliation.id}|operators={operator_names}|assigned_count={reassigned_count}',
                        severity='info',
                    )
                    messages.success(request, f'Scan duty shifted to {operator_names}.')
                else:
                    messages.error(request, 'Please select at least one valid operator and reconciliation.')
            else:
                messages.error(request, 'Please select a valid operator and reconciliation.')
            return redirect(f'{reverse("backup-dashboard")}?show_reconciliation=1&reconciliation_pk={selected_reconciliation.id if selected_reconciliation else ""}&show_unscanned_tapes=1')
        elif form_type == 'add_reconciliation_result':
            reconciliation_pk = request.POST.get('reconciliation_pk')
            selected_reconciliation = get_object_by_uuid_pk(Reconciliation, reconciliation_pk)
            if selected_reconciliation:
                reconciliation_result_form = ReconciliationResultForm(request.POST, prefix='result')
                if reconciliation_result_form.is_valid():
                    result = reconciliation_result_form.save(commit=False)
                    result.reconciliation = selected_reconciliation
                    result.save()
                    AuditLog.objects.create(
                        name='Reconciliation Result Added',
                        action=f'Added reconciliation result for {selected_reconciliation.reconciliation_id}',
                        user=request.user,
                        severity='info',
                    )
                    if selected_reconciliation.total_tapes_expected > 0:
                        selected_reconciliation.update_progress(scanned=selected_reconciliation.total_tapes_scanned, expected=selected_reconciliation.total_tapes_expected)
                    messages.success(request, 'Reconciliation result saved successfully.')
                    return redirect(f'{reverse("backup-dashboard")}?show_reconciliation=1&reconciliation_pk={selected_reconciliation.id}')
                else:
                    show_reconciliation_panel = True
                    messages.error(request, 'Please correct the result form errors and try again.')
            else:
                messages.error(request, 'Please select a valid reconciliation session.')
                return redirect('backup-dashboard')
        elif form_type == 'record_reconciliation_scan':
            reconciliation_pk = request.POST.get('reconciliation_pk')
            tape_id = request.POST.get('scan_tape_id')
            scan_status = request.POST.get('scan_status')
            scan_progress = request.POST.get('scan_progress')
            selected_reconciliation = get_object_by_uuid_pk(Reconciliation, reconciliation_pk)
            tape = get_object_by_uuid_pk(Tape, tape_id)
            if selected_reconciliation and tape and scan_status:
                was_scanned = tape.scan_status in ['Scanned', 'Verified']
                tape.scan_status = scan_status
                tape.last_scanned_at = timezone.localtime()
                tape.last_scanned_by = request.user
                if scan_progress:
                    try:
                        tape.scan_progress = min(100, max(0, int(scan_progress)))
                    except (ValueError, TypeError):
                        pass
                tape.save(update_fields=['scan_status', 'scan_progress', 'last_scanned_at', 'last_scanned_by'])

                is_scanned_now = scan_status in ['Scanned', 'Verified']
                if not was_scanned and is_scanned_now:
                    selected_reconciliation.total_tapes_scanned = selected_reconciliation.total_tapes_scanned + 1
                    if selected_reconciliation.total_tapes_expected > 0:
                        selected_reconciliation.total_tapes_scanned = min(
                            selected_reconciliation.total_tapes_scanned,
                            selected_reconciliation.total_tapes_expected
                        )

                selected_reconciliation.update_progress(scanned=selected_reconciliation.total_tapes_scanned, expected=selected_reconciliation.total_tapes_expected)

                if selected_reconciliation.progress_percent == 100 and selected_reconciliation.status != 'Completed':
                    selected_reconciliation.status = 'Completed'
                    selected_reconciliation.save(update_fields=['status'])
                AuditLog.objects.create(
                    name='Reconciliation Scan Updated',
                    action=f'Tape {tape.volser} scan status updated to {tape.scan_status} for {selected_reconciliation.reconciliation_id}',
                    user=request.user,
                    message=f'reconciliation_id={selected_reconciliation.reconciliation_id}|tape={tape.volser}|scan_status={tape.scan_status}|progress={tape.scan_progress}',
                    severity='info',
                )
                messages.success(request, 'Scan record saved and reconciliation progress updated.')
            else:
                messages.error(request, 'Please provide a valid reconciliation, tape, and scan status.')
            return redirect(f'{reverse("backup-dashboard")}?show_reconciliation=1&reconciliation_pk={selected_reconciliation.id if selected_reconciliation else ""}')
        elif form_type == 'system_settings':
            settings_form = SystemSettingsForm(request.POST, instance=application_settings)
            if settings_form.is_valid():
                settings_form.save()
                AuditLog.objects.create(
                    name='System Settings Updated',
                    action=f'Updated system settings by {request.user.username}',
                    user=request.user,
                    severity='success',
                )
                messages.success(request, 'System settings saved successfully.')
                return redirect(f'{reverse("backup-dashboard")}?show_settings=1')
            else:
                show_settings_panel = True
                messages.error(request, 'Please correct the settings form errors and try again.')
        elif form_type == 'upload_inventory_excel':
            uploaded_file = request.FILES.get('inventory_file')
            if uploaded_file and hasattr(uploaded_file, 'read'):
                uploaded_file.seek(0)
                file_bytes = uploaded_file.read()
                uploaded_file.seek(0)
                schema_preview = analyze_excel_schema(uploaded_file, table_name='inventory_tape')
                if schema_preview and schema_preview.get('error'):
                    messages.error(request, schema_preview['error'])
                    return redirect(f'{reverse("backup-dashboard")}?show_tape_inventory=1')
                if schema_preview and schema_preview.get('new_columns'):
                    request.session['pending_schema_preview'] = {
                        'file_name': uploaded_file.name,
                        'file_bytes': base64.b64encode(file_bytes).decode('utf-8'),
                        'preview': schema_preview,
                    }
                    messages.info(request, 'Schema synchronization preview is ready. Review the proposed changes before continuing.')
                    show_tape_inventory_panel = True
                else:
                    try:
                        uploaded_file.seek(0)
                        workbook = load_workbook(BytesIO(uploaded_file.read()), data_only=True)
                        pa = _build_pending_inventory_import_approval(request, workbook, file_name=uploaded_file.name)
                        if pa:
                            AuditLog.objects.create(
                                name='Pending Inventory Import',
                                action=f'Pending Excel import submitted for approval by {request.user.username}',
                                user=request.user,
                                severity='info',
                            )
                            messages.success(request, 'Inventory import submitted for approval. It will be applied once approved.')
                        else:
                            messages.error(request, 'Unable to stage inventory import for approval.')
                    except Exception as exc:
                        messages.error(request, f'Unable to read the Excel file for staging import: {exc}')
                    return redirect(f'{reverse("backup-dashboard")}?show_tape_inventory=1')
            else:
                messages.error(request, 'Please choose an Excel file to import.')
                return redirect(f'{reverse("backup-dashboard")}?show_tape_inventory=1')
        elif form_type == 'approve_excel_schema_sync':
            pending_preview = request.session.pop('pending_schema_preview', None)
            if not pending_preview:
                messages.error(request, 'No pending schema preview was found.')
                return redirect(f'{reverse("backup-dashboard")}?show_tape_inventory=1')
            if not (request.user.is_superuser or is_backup_administrator(request.user)):
                messages.error(request, 'Only superusers and backup administrators can approve schema synchronization.')
                return redirect(f'{reverse("backup-dashboard")}?show_tape_inventory=1')

            preview = pending_preview.get('preview', {})
            new_columns = preview.get('new_columns', [])
            if not new_columns:
                messages.info(request, 'No schema changes were required.')
                return redirect(f'{reverse("backup-dashboard")}?show_tape_inventory=1')

            pa = _create_schema_sync_pending_approval(request, pending_preview)
            if pa:
                if request.user.is_superuser or is_backup_administrator(request.user):
                    pa.status = 'Approved'
                    pa.approved_by = request.user
                    pa.reviewed_by = request.user
                    pa.approved_at = timezone.localtime()
                    pa.review_comment = 'Schema synchronization approved and executed.'
                    pa.audit_history = list(pa.audit_history or []) + [{
                        'action': 'Approved',
                        'user': request.user.get_full_name() or request.user.username,
                        'timestamp': timezone.localtime().isoformat(),
                        'comment': pa.review_comment,
                    }]
                    pa.save(update_fields=['status', 'approved_by', 'reviewed_by', 'approved_at', 'review_comment', 'audit_history', 'updated_at'])
                    exec_results = _execute_pending_changes(pa, request.user)
                    pa.audit_history = list(pa.audit_history or []) + [{
                        'action': 'Executed',
                        'user': request.user.get_full_name() or request.user.username,
                        'timestamp': timezone.localtime().isoformat(),
                        'results': exec_results,
                    }]
                    pa.save(update_fields=['audit_history', 'updated_at'])
                    AuditLog.objects.create(
                        name='Tape Schema Synchronized',
                        action=f'Schema synchronization for {pending_preview.get("file_name", "unknown.xlsx")} was approved and executed',
                        user=request.user,
                        severity='success',
                    )
                    messages.success(request, 'Schema synchronization approved and applied.')
                else:
                    AuditLog.objects.create(
                        name='Pending Tape Schema Synchronization',
                        action=f'Pending schema sync for {pending_preview.get("file_name", "unknown.xlsx")} submitted for approval',
                        user=request.user,
                        severity='info',
                    )
                    messages.success(request, 'Schema synchronization submitted for approval.')
            else:
                messages.error(request, 'Unable to stage schema synchronization for approval.')

            return redirect(f'{reverse("backup-dashboard")}?show_tape_inventory=1')
        elif form_type == 'edit_profile':
            profile_form = UserProfileForm(request.POST, instance=request.user)
            if profile_form.is_valid():
                profile_form.save()
                AuditLog.objects.create(
                    name='Profile Updated',
                    action=f'Updated profile for {request.user.username}',
                    user=request.user,
                    severity='success',
                )
                messages.success(request, 'Your profile has been updated.')
                return redirect(f'{reverse("backup-dashboard")}?show_profile=1')
            else:
                show_profile_panel = True
                messages.error(request, 'Please correct the profile form errors and try again.')

    selected_tape_id = request.GET.get('selected_tape') or request.POST.get('selected_tape')
    if selected_tape_id and not selected_tape:
        selected_tape = get_object_by_uuid_pk(Tape, selected_tape_id)
    if selected_tape and not tape_action_form:
        tape_action_form = TapeForm(instance=selected_tape)

    if request.GET.get('open_panel') == 'tapeActionsPanel':
        show_tape_actions_panel = True
    if request.GET.get('show_tape_actions') == '1' or 'show_tape_actions' in request.GET:
        show_tape_actions_panel = True
    if 'show_tape_inventory' in request.GET:
        show_tape_inventory_panel = True
    if 'show_print_barcodes' in request.GET:
        show_print_barcodes_panel = True
    if 'show_admin' in request.GET:
        show_admin_panel = True
    if 'show_admin_history' in request.GET:
        show_admin_history = True
    if 'show_profile' in request.GET:
        show_profile_panel = True
    if 'show_settings' in request.GET:
        show_settings_panel = True
    if 'show_add_tape' in request.GET:
        show_add_tape_panel = True
    if 'show_audit' in request.GET:
        show_audit_panel = True
    if 'show_reports' in request.GET:
        show_reports_panel = True
    approval_shipment_id = request.GET.get('approve_shipment') or request.POST.get('shipment_id')
    if approval_shipment_id:
        approval_shipment = get_object_by_uuid_pk(Shipment, approval_shipment_id)
        show_alerts_panel = True
    unread_alert_ids = list(
        AuditLog.objects.filter(severity__in=['warning', 'error'], is_read=False).values_list('pk', flat=True)
    )
    if 'show_alerts' in request.GET:
        show_alerts_panel = True
        if unread_alert_ids:
            AuditLog.objects.filter(pk__in=unread_alert_ids).update(
                is_read=True,
                read_at=timezone.now()
            )
        unread_alert_ids = list(
            AuditLog.objects.filter(severity__in=['warning', 'error'], is_read=False).values_list('pk', flat=True)
        )
    report_type = request.GET.get('report_type')
    report_category = request.GET.get('report_category')
    report_period = get_scoped_report_param(request, report_category, 'report_period')
    export_csv = get_scoped_report_flag(request, report_category, 'export_csv')
    export_pdf = get_scoped_report_flag(request, report_category, 'export_pdf')
    export_excel = get_scoped_report_flag(request, report_category, 'export_excel')
    report_search = get_scoped_report_param(request, report_category, 'report_search')
    report_filter_status = get_scoped_report_param(request, report_category, 'report_filter_status')
    report_date_from = get_scoped_report_param(request, report_category, 'report_date_from')
    report_date_to = get_scoped_report_param(request, report_category, 'report_date_to')
    report_sort = get_scoped_report_param(request, report_category, 'report_sort')
    report_order = (get_scoped_report_param(request, report_category, 'report_order', 'asc') or 'asc').lower()
    if report_order not in {'asc', 'desc'}:
        report_order = 'asc'
    generated_report_data = None
    generated_report_label = None
    generated_report_items = []
    report_email_form = ReportEmailForm(request.POST or None)
    reconciliation_report_summary = None
    export_reconciliation_csv = request.GET.get('export_reconciliation_csv') == '1'
    report_categories = get_report_categories()
    current_month = timezone.localdate().strftime('%Y-%m')
    if show_reports_panel and not report_period:
        report_period = current_month

    if 'show_reports' in request.GET:
        show_reports_panel = True
    if 'show_shipments' in request.GET:
        show_shipments_panel = True
    if 'show_reconciliation_reports' in request.GET:
        show_reconciliation_reports_panel = True
    if 'show_add_shipment' in request.GET:
        show_shipments_panel = True
        show_add_shipment_panel = True
    if 'show_reconciliation' in request.GET:
        show_reconciliation_panel = True
    if 'show_add_reconciliation' in request.GET:
        show_reconciliation_panel = True
        show_add_reconciliation_panel = True
    reconciliation_pk = request.GET.get('reconciliation_pk')
    if reconciliation_pk:
        selected_reconciliation = get_object_by_uuid_pk(Reconciliation, reconciliation_pk)
        if selected_reconciliation:
            show_reconciliation_panel = True
    if request.GET.get('show_unscanned_tapes') == '1':
        show_unscanned_tapes = True
    edit_shipment_pk = request.GET.get('edit_shipment_pk')
    if edit_shipment_pk:
        selected_shipment = get_object_by_uuid_pk(Shipment, edit_shipment_pk)
        if selected_shipment:
            show_shipments_panel = True
            show_edit_shipment_panel = True
            if not edit_shipment_form:
                edit_shipment_form = ShipmentForm(instance=selected_shipment, prefix='edit', request=request)

    if not settings_form:
        settings_form = SystemSettingsForm(instance=application_settings)
    if not profile_form:
        profile_form = UserProfileForm(instance=request.user)

    tapes = Tape.objects.all().order_by('-date_registered')
    tape_search = request.GET.get('tape_search', '').strip()
    if tape_search:
        search_q = (
            Q(volser__icontains=tape_search) |
            Q(barcode__icontains=tape_search) |
            Q(rfid_tag__icontains=tape_search) |
            Q(tape_type__icontains=tape_search) |
            Q(manufacturer__icontains=tape_search) |
            Q(status__icontains=tape_search) |
            Q(current_location__icontains=tape_search) |
            Q(remarks__icontains=tape_search)
        )
        parsed_date = parse_date(tape_search)
        if parsed_date:
            search_q |= Q(retention_end_date=parsed_date)
        tapes = tapes.filter(search_q)

    user_feature_names = []
    user_groups = request.user.groups.all()
    if user_groups.exists():
        assigned_templates = RoleTemplate.objects.filter(group__in=user_groups)
        for template in assigned_templates:
            if template.features:
                user_feature_names.extend(template.features)
    user_feature_names = sorted(set(user_feature_names))
    request.user.feature_names = user_feature_names
    dashboard_tabs = get_dashboard_tabs(request.user, BACKUP_FEATURE_TABS, preserve_empty_tabs=True)

    total_tapes = tapes.count()
    active_tapes = tapes.filter(status="Active").count()
    off_site_tapes = tapes.filter(status="Off-Site").count()
    missing_tapes = tapes.filter(status="Missing").count()
    archived_tapes = tapes.filter(status="Retained").count()
    retention_due = tapes.filter(retention_end_date__lte=timezone.localdate()).count()
    shipments = Shipment.objects.all().order_by('shipment_id')
    pending_tape_requests = TapeRequest.objects.select_related('tape', 'requested_by').filter(status='Pending').order_by('-request_date')
    pending_users = User.objects.filter(verified=False).order_by('date_joined')
    reports = ReportTemplate.objects.all()
    reconciliations = Reconciliation.objects.all().order_by('-reconciliation_date', '-created_at')
    search_query = request.GET.get('search', '').strip()
    reconciliation_report_summary = None

    kpis = {}
    if show_reports_panel:
        kpis = {
            'Total Tapes': total_tapes,
            'Active Tapes': active_tapes,
            'Retained Tapes': archived_tapes,
            'In Transit Tapes': tapes.filter(status__iexact='In Transit').count(),
            'Missing Tapes': missing_tapes,
            'Damaged Tapes': tapes.filter(status__iexact='Damaged').count(),
            'Pending Destruction': tapes.filter(status__iexact='Pending Destruction').count(),
            'Open Exceptions': 0,
            'Open Shipments': shipments.filter(status__iexact='Pending').count(),
            'Compliance Rate': '98.5%',
            'Reconciliation Accuracy': '99.2%',
        }

    if show_reconciliation_reports_panel and search_query:
        search_q = (
            Q(reconciliation_id__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(performed_by__username__icontains=search_query)
        )
        reconciliations = reconciliations.filter(search_q)

    if show_reconciliation_reports_panel:
        reconciliation_count = reconciliations.count()
        total_issues = ReconciliationResult.objects.filter(reconciliation__in=reconciliations).count()
        open_issues = ReconciliationResult.objects.filter(
            reconciliation__in=reconciliations,
            resolution_status__in=['Open', 'Under Investigation']
        ).count()
        completed_count = reconciliations.filter(status='Completed').count()
        open_reconciliations = reconciliations.exclude(status='Completed').count()

        reconciliation_report_summary = {
            'total_reconciliations': reconciliation_count,
            'total_discrepancies': total_issues,
            'open_issues': open_issues,
            'completed_reconciliations': completed_count,
            'open_reconciliations': open_reconciliations,
        }

    for reconciliation in reconciliations:
        reconciliation.total_issues = reconciliation.results.count()
        reconciliation.open_issues = reconciliation.results.filter(resolution_status__in=['Open', 'Under Investigation']).count()
    reconciliation_results = selected_reconciliation.results.order_by('-updated_at') if selected_reconciliation else ReconciliationResult.objects.none()
    unscanned_tapes = Tape.objects.none()
    available_operators = CustomUser.objects.filter(
        is_active=True,
    ).filter(
        Q(role='operations_manager') | Q(groups__name='Operations Manager')
    ).order_by('username').distinct()
    if selected_reconciliation and show_unscanned_tapes:
        unscanned_tapes = Tape.objects.filter(scan_status__in=['Pending', 'Assigned', 'Scanning']).order_by('volser')
    recent_activities = AuditLog.objects.order_by('-timestamp')[:6]
    recent_alerts = list(AuditLog.objects.filter(severity__in=['warning', 'error']).order_by('-timestamp')[:5])
    unread_alert_id_set = set(unread_alert_ids)
    for alert in recent_alerts:
        alert.is_new = alert.pk in unread_alert_id_set
        alert.is_exception = False
        alert.exception_id = None
        alert.exception_description = None
        try:
            exception_id, description = _parse_exception_metadata(alert)
            if exception_id:
                alert.is_exception = True
                alert.exception_id = exception_id
                alert.exception_description = description
        except Exception:
            pass
        # If this alert is a close-request notification, expose a review URL for the template
        try:
            metadata = _parse_auditlog_metadata(alert.message)
            if metadata.get('target_url'):
                alert.target_url = metadata.get('target_url')
            if alert.name == 'Exception Close Request Notification' and alert.message:
                # expected message format: 'requested_by=...|close_request_id=<uuid>'
                parts = {p.split('=')[0]: p.split('=')[1] for p in (alert.message or '').split('|') if '=' in p}
                close_request_id = parts.get('close_request_id')
                if close_request_id:
                    alert.close_request_id = close_request_id
                    alert.close_request_url = reverse('approve-close-exception', kwargs={'close_request_id': close_request_id})
            if alert.message and 'reconciliation_pk=' in alert.message:
                parts = {p.split('=')[0]: p.split('=')[1] for p in (alert.message or '').split('|') if '=' in p}
                reconciliation_pk = parts.get('reconciliation_pk')
                if reconciliation_pk:
                    alert.reconciliation_pk = reconciliation_pk
                    alert.reconciliation_open_url = f"{reverse('backup-dashboard')}?show_reconciliation=1&show_unscanned_tapes=1&reconciliation_pk={reconciliation_pk}"
        except Exception:
            # Fail gracefully; don't break alert rendering
            pass
    reconciliation_request_context = None
    reconciliation_request_id = request.GET.get('reconciliation_request_id') or request.POST.get('reconciliation_request_id')
    if reconciliation_request_id:
        reconciliation_request_entry = AuditLog.objects.filter(pk=reconciliation_request_id, name='Reconciliation Requested').first()
        if reconciliation_request_entry:
            payload = _parse_reconciliation_request_message(reconciliation_request_entry.message or '')
            reconciliation_request_context = {
                'id': str(reconciliation_request_entry.id),
                'exception_id': payload.get('exception_id', ''),
                'requester_name': payload.get('requester_name', ''),
                'comment': payload.get('comment', ''),
                'entry': reconciliation_request_entry,
            }
    pending_approval_shipments = Shipment.objects.filter(status='Pending').order_by('-shipment_date')[:8]

    queryset, filters = build_audit_log_queryset(request)
    export_format = (request.GET.get('export') or '').strip().lower()
    if show_audit_panel and export_format in {'csv', 'excel', 'pdf'}:
        export_response = export_audit_logs(queryset, export_format)
        if export_response is not None:
            return export_response
    per_page = filters['per_page']
    paginator = Paginator(queryset, per_page)
    page = request.GET.get('page', 1)
    page_obj = paginator.get_page(page)
    audit_logs = page_obj.object_list

    inventory_report_tapes = []
    shipment_report_rows = []
    custody_report_rows = []
    reconciliation_report_rows = []
    retention_report_rows = []
    compliance_report_rows = []
    exception_report_rows = []
    audit_trail_report_rows = []
    management_summary_report_rows = []
    report_table_columns = []
    report_paginator = None
    report_page_obj = None
    if show_reports_panel and report_period:
        report_month = get_first_day_of_month(report_period) or get_first_day_of_month(current_month)
        if report_month:
            if report_type == 'daily':
                report_date = parse_date(report_period)
                if report_date:
                    generated_report_data = generate_daily_report_data(report_date)
                    generated_report_label = f"Daily Report for {report_date.strftime('%Y-%m-%d')}"
            else:
                generated_report_data = generate_monthly_report_data(report_month, report_category)
                selected_category = next((c for c in report_categories if c['slug'] == report_category), None)
                category_name = selected_category['name'] if selected_category else 'Monthly Report'
                generated_report_label = f"{category_name} for {report_month.strftime('%B %Y')}"
                if report_category == 'inventory':
                    qs = Tape.objects.filter(date_registered__gte=report_month, date_registered__lt=get_next_month(report_month))
                    if report_search:
                        qs = qs.filter(
                            Q(volser__icontains=report_search) |
                            Q(barcode__icontains=report_search) |
                            Q(rfid_tag__icontains=report_search) |
                            Q(tape_type__icontains=report_search) |
                            Q(status__icontains=report_search) |
                            Q(current_location__icontains=report_search)
                        )
                    if report_filter_status:
                        qs = qs.filter(status__iexact=report_filter_status)
                    sort_field = None
                    if report_sort in {'volser', 'status', 'retention_end_date', 'date_registered', 'current_location'}:
                        sort_field = report_sort
                    if sort_field:
                        if report_order == 'desc':
                            qs = qs.order_by(f'-{sort_field}')
                        else:
                            qs = qs.order_by(sort_field)
                    else:
                        qs = qs.order_by('volser')
                    inventory_report_tapes = list(qs)
                    inventory_report_rows = []
                    for tape in inventory_report_tapes:
                        tape.latest_custodian = get_latest_custodian_for_tape(tape) or tape.current_location or '-'
                        inventory_report_rows.append({
                            'volser': tape.volser,
                            'barcode': tape.barcode,
                            'rfid_tag': tape.rfid_tag or '-',
                            'tape_type': tape.tape_type,
                            'status': tape.status,
                            'current_location': tape.current_location or '-',
                            'latest_custodian': tape.latest_custodian,
                            'retention_end_date': tape.retention_end_date,
                            'date_registered': tape.date_registered,
                        })
                    report_table_columns = [
                        {'key': 'volser', 'label': 'VolSER'},
                        {'key': 'barcode', 'label': 'Barcode'},
                        {'key': 'rfid_tag', 'label': 'RFID Tag'},
                        {'key': 'tape_type', 'label': 'Tape Type'},
                        {'key': 'status', 'label': 'Status'},
                        {'key': 'current_location', 'label': 'Current Location'},
                        {'key': 'latest_custodian', 'label': 'Custodian'},
                        {'key': 'retention_end_date', 'label': 'Retention End Date'},
                        {'key': 'date_registered', 'label': 'Date Registered'},
                    ]
                    if request.GET.get('share_report') == '1':
                        share_emails = [email.strip() for email in request.GET.get('share_email', '').split(',') if email.strip()]
                        if share_emails:
                            if send_report_share_email(request, report_category, report_period, inventory_report_rows, report_table_columns, share_emails):
                                messages.success(request, f'Report shared with {", ".join(share_emails)}.')
                            return redirect_report_view(request)
                        messages.error(request, 'Please enter at least one valid email address to share this report.')
                        return redirect_report_view(request)
                    if export_pdf:
                        return export_report_pdf(report_category, report_period, inventory_report_rows, report_table_columns)
                    if export_excel:
                        return export_report_excel(report_category, report_period, inventory_report_rows, report_table_columns)
                elif report_category == 'shipment':
                    shipment_report_rows = Shipment.objects.filter(
                        shipment_date__gte=report_month,
                        shipment_date__lt=get_next_month(report_month)
                    ).order_by('shipment_id')
                    report_table_columns = [
                        {'key': 'shipment_id', 'label': 'Shipment ID'},
                        {'key': 'shipment_type', 'label': 'Shipment Type'},
                        {'key': 'source_location', 'label': 'Source Location'},
                        {'key': 'destination_location', 'label': 'Destination Location'},
                        {'key': 'courier_name', 'label': 'Courier'},
                        {'key': 'shipment_date', 'label': 'Dispatch Date'},
                        {'key': 'delivery_date', 'label': 'Delivery Date'},
                        {'key': 'status', 'label': 'Status'},
                        {'key': 'number_of_tapes', 'label': 'Number of Tapes'},
                    ]
                    if request.GET.get('share_report') == '1':
                        share_emails = [email.strip() for email in request.GET.get('share_email', '').split(',') if email.strip()]
                        if share_emails:
                            if send_report_share_email(request, report_category, report_period, list(shipment_report_rows.values('shipment_id', 'shipment_type', 'source_location', 'destination_location', 'courier_name', 'shipment_date', 'delivery_date', 'status')), report_table_columns, share_emails):
                                messages.success(request, f'Report shared with {", ".join(share_emails)}.')
                            return redirect_report_view(request)
                        messages.error(request, 'Please enter at least one valid email address to share this report.')
                        return redirect_report_view(request)
                elif report_category == 'custody':
                    custody_report_rows = Shipment.objects.filter(
                        shipment_date__gte=report_month,
                        shipment_date__lt=get_next_month(report_month)
                    ).order_by('shipment_id')
                    report_table_columns = [
                        {'key': 'transfer_date', 'label': 'Transfer Date'},
                        {'key': 'transfer_time', 'label': 'Transfer Time'},
                        {'key': 'previous_custodian', 'label': 'Previous Custodian'},
                        {'key': 'new_custodian', 'label': 'New Custodian'},
                        {'key': 'location', 'label': 'Location'},
                        {'key': 'remarks', 'label': 'Remarks'},
                    ]
                    for shipment in custody_report_rows:
                        shipment.transfer_date = shipment.shipment_date
                        shipment.transfer_time = shipment.release_datetime.time() if shipment.release_datetime else None
                        shipment.previous_custodian = shipment.releasing_custodian or '-'
                        shipment.new_custodian = shipment.receiving_custodian or '-'
                        shipment.location = shipment.destination_location or shipment.source_location or '-'
                        shipment.remarks = shipment.approval_remarks or shipment.delivery_notes or '-'
                    if request.GET.get('share_report') == '1':
                        share_emails = [email.strip() for email in request.GET.get('share_email', '').split(',') if email.strip()]
                        if share_emails:
                            custody_share_rows = []
                            for shipment in custody_report_rows:
                                custody_share_rows.append({
                                    'transfer_date': shipment.transfer_date,
                                    'transfer_time': shipment.transfer_time,
                                    'previous_custodian': shipment.previous_custodian,
                                    'new_custodian': shipment.new_custodian,
                                    'location': shipment.location,
                                    'remarks': shipment.remarks,
                                })
                            if send_report_share_email(request, report_category, report_period, custody_share_rows, report_table_columns, share_emails):
                                messages.success(request, f'Report shared with {", ".join(share_emails)}.')
                            return redirect_report_view(request)
                        messages.error(request, 'Please enter at least one valid email address to share this report.')
                        return redirect_report_view(request)
                elif report_category == 'reconciliation':
                    qs = Reconciliation.objects.filter(reconciliation_date__gte=report_month, reconciliation_date__lt=get_next_month(report_month)).order_by('-reconciliation_date')
                    if report_search:
                        qs = qs.filter(Q(reconciliation_id__icontains=report_search) | Q(location__icontains=report_search) | Q(status__icontains=report_search))
                    if report_filter_status:
                        qs = qs.filter(status__iexact=report_filter_status)
                    if report_date_from:
                        qs = qs.filter(reconciliation_date__gte=parse_date(report_date_from))
                    if report_date_to:
                        qs = qs.filter(reconciliation_date__lte=parse_date(report_date_to))
                    reconciliation_report_rows = []
                    for reconciliation in qs:
                        reconciliation_report_rows.append({
                            'id': reconciliation.id,
                            'reconciliation_id': reconciliation.reconciliation_id,
                            'location': reconciliation.location,
                            'expected_tapes': reconciliation.results.count(),
                            'scanned_tapes': reconciliation.results.count(),
                            'missing_tapes': reconciliation.results.filter(issue_type='Missing').count(),
                            'misplaced_tapes': reconciliation.results.filter(issue_type='Misplaced').count(),
                            'unexpected_tapes': reconciliation.results.filter(issue_type='Unexpected').count(),
                            'damaged_tapes': reconciliation.results.filter(issue_type='Damaged').count(),
                            'reconciliation_date': reconciliation.reconciliation_date,
                            'status': reconciliation.status,
                        })
                    report_table_columns = [
                        {'key': 'reconciliation_id', 'label': 'Reconciliation ID'},
                        {'key': 'location', 'label': 'Location'},
                        {'key': 'expected_tapes', 'label': 'Expected Tapes'},
                        {'key': 'scanned_tapes', 'label': 'Scanned Tapes'},
                        {'key': 'missing_tapes', 'label': 'Missing Tapes'},
                        {'key': 'misplaced_tapes', 'label': 'Misplaced Tapes'},
                        {'key': 'unexpected_tapes', 'label': 'Unexpected Tapes'},
                        {'key': 'reconciliation_date', 'label': 'Reconciliation Date'},
                        {'key': 'status', 'label': 'Status'},
                    ]
                    reconciliation_report_rows = sort_report_rows(reconciliation_report_rows, report_sort or None, report_order)
                    if request.GET.get('share_report') == '1':
                        share_emails = [email.strip() for email in request.GET.get('share_email', '').split(',') if email.strip()]
                        if share_emails:
                            if send_report_share_email(request, report_category, report_period, reconciliation_report_rows, report_table_columns, share_emails):
                                messages.success(request, f'Report shared with {", ".join(share_emails)}.')
                            return redirect_report_view(request)
                        messages.error(request, 'Please enter at least one valid email address to share this report.')
                        return redirect_report_view(request)
                    if export_pdf:
                        return export_report_pdf(report_category, report_period, reconciliation_report_rows, report_table_columns)
                    if export_excel:
                        return export_report_excel(report_category, report_period, reconciliation_report_rows, report_table_columns)
                    report_paginator, report_page_obj = paginate_report_rows(request, reconciliation_report_rows, page_param=f'report_page_{report_category}')
                    reconciliation_report_rows = list(report_page_obj.object_list)
                elif report_category == 'retention':
                    qs = Tape.objects.filter(retention_end_date__gte=report_month, retention_end_date__lt=get_next_month(report_month)).order_by('retention_end_date')
                    if report_search:
                        qs = qs.filter(Q(volser__icontains=report_search) | Q(barcode__icontains=report_search) | Q(status__icontains=report_search))
                    if report_filter_status:
                        qs = qs.filter(status__iexact=report_filter_status)
                    if report_date_from:
                        qs = qs.filter(retention_end_date__gte=parse_date(report_date_from))
                    if report_date_to:
                        qs = qs.filter(retention_end_date__lte=parse_date(report_date_to))
                    retention_report_rows = []
                    for tape in qs:
                        retention_report_rows.append({
                            'volser': tape.volser,
                            'barcode': tape.barcode,
                            'retention_start_date': tape.date_registered.date() if getattr(tape, 'date_registered', None) else '-',
                            'retention_end_date': tape.retention_end_date,
                            'days_remaining': (tape.retention_end_date - timezone.localdate()).days if tape.retention_end_date else '-',
                            'legal_hold': 'Yes' if tape.legal_hold else 'No',
                            'audit_hold': 'Yes' if tape.audit_hold else 'No',
                            'status': tape.status,
                        })
                    report_table_columns = [
                        {'key': 'volser', 'label': 'VolSER'},
                        {'key': 'barcode', 'label': 'Barcode'},
                        {'key': 'retention_start_date', 'label': 'Retention Start Date'},
                        {'key': 'retention_end_date', 'label': 'Retention End Date'},
                        {'key': 'days_remaining', 'label': 'Days Remaining'},
                        {'key': 'legal_hold', 'label': 'Legal Hold'},
                        {'key': 'audit_hold', 'label': 'Audit Hold'},
                        {'key': 'status', 'label': 'Status'},
                    ]
                    retention_report_rows = sort_report_rows(retention_report_rows, report_sort or None, report_order)
                    if request.GET.get('share_report') == '1':
                        share_emails = [email.strip() for email in request.GET.get('share_email', '').split(',') if email.strip()]
                        if share_emails:
                            if send_report_share_email(request, report_category, report_period, retention_report_rows, report_table_columns, share_emails):
                                messages.success(request, f'Report shared with {", ".join(share_emails)}.')
                            return redirect_report_view(request)
                        messages.error(request, 'Please enter at least one valid email address to share this report.')
                        return redirect_report_view(request)
                    if export_pdf:
                        return export_report_pdf(report_category, report_period, retention_report_rows, report_table_columns)
                    if export_excel:
                        return export_report_excel(report_category, report_period, retention_report_rows, report_table_columns)
                    report_paginator, report_page_obj = paginate_report_rows(request, retention_report_rows, page_param=f'report_page_{report_category}')
                    retention_report_rows = list(report_page_obj.object_list)
                elif report_category == 'compliance':
                    qs = ReconciliationResult.objects.filter(reconciliation__reconciliation_date__gte=report_month, reconciliation__reconciliation_date__lt=get_next_month(report_month)).order_by('-created_at')
                    if report_search:
                        qs = qs.filter(Q(reconciliation__reconciliation_id__icontains=report_search) | Q(remarks__icontains=report_search) | Q(resolution_status__icontains=report_search))
                    if report_filter_status:
                        qs = qs.filter(resolution_status__iexact=report_filter_status)
                    if report_date_from:
                        qs = qs.filter(created_at__date__gte=parse_date(report_date_from))
                    if report_date_to:
                        qs = qs.filter(created_at__date__lte=parse_date(report_date_to))
                    compliance_report_rows = []
                    for result in qs:
                        compliance_report_rows.append({
                            'compliance_id': result.reconciliation.reconciliation_id,
                            'policy_name': 'Tape Handling Policy',
                            'compliance_status': 'Compliant' if result.resolution_status == 'Resolved' else 'Needs Review',
                            'violations': result.issue_type,
                            'date_identified': result.created_at.date(),
                            'responsible_user': result.reconciliation.performed_by.username if result.reconciliation.performed_by else '-',
                            'resolution_status': result.resolution_status,
                        })
                    report_table_columns = [
                        {'key': 'compliance_id', 'label': 'Compliance ID'},
                        {'key': 'policy_name', 'label': 'Policy Name'},
                        {'key': 'compliance_status', 'label': 'Compliance Status'},
                        {'key': 'violations', 'label': 'Violations'},
                        {'key': 'date_identified', 'label': 'Date Identified'},
                        {'key': 'responsible_user', 'label': 'Responsible User'},
                        {'key': 'resolution_status', 'label': 'Resolution Status'},
                    ]
                    compliance_report_rows = sort_report_rows(compliance_report_rows, report_sort or None, report_order)
                    if request.GET.get('share_report') == '1':
                        share_emails = [email.strip() for email in request.GET.get('share_email', '').split(',') if email.strip()]
                        if share_emails:
                            if send_report_share_email(request, report_category, report_period, compliance_report_rows, report_table_columns, share_emails):
                                messages.success(request, f'Report shared with {", ".join(share_emails)}.')
                            return redirect_report_view(request)
                        messages.error(request, 'Please enter at least one valid email address to share this report.')
                        return redirect_report_view(request)
                    if export_pdf:
                        return export_report_pdf(report_category, report_period, compliance_report_rows, report_table_columns)
                    if export_excel:
                        return export_report_excel(report_category, report_period, compliance_report_rows, report_table_columns)
                    report_paginator, report_page_obj = paginate_report_rows(request, compliance_report_rows, page_param=f'report_page_{report_category}')
                    compliance_report_rows = list(report_page_obj.object_list)
                elif report_category == 'exception':
                    qs = ShipmentException.objects.filter(reported_date__date__gte=report_month, reported_date__date__lt=get_next_month(report_month)).order_by('-reported_date')
                    if report_search:
                        qs = qs.filter(Q(exception_id__icontains=report_search) | Q(tape__volser__icontains=report_search) | Q(status__icontains=report_search))
                    if report_filter_status:
                        qs = qs.filter(status__iexact=report_filter_status)
                    if report_date_from:
                        qs = qs.filter(reported_date__date__gte=parse_date(report_date_from))
                    if report_date_to:
                        qs = qs.filter(reported_date__date__lte=parse_date(report_date_to))
                    exception_report_rows = []
                    for exception in qs:
                        exception_report_rows.append({
                            'exception_id': exception.exception_id,
                            'tape_volser': exception.tape.volser if exception.tape else '-',
                            'exception_type': exception.exception_type,
                            'severity': exception.severity,
                            'reported_by': exception.reported_by.username if exception.reported_by else '-',
                            'date_reported': exception.reported_date.date(),
                            'status': exception.status,
                            'resolution_date': exception.reported_date.date(),
                        })
                    report_table_columns = [
                        {'key': 'exception_id', 'label': 'Exception ID'},
                        {'key': 'tape_volser', 'label': 'Tape VolSER'},
                        {'key': 'exception_type', 'label': 'Exception Type'},
                        {'key': 'severity', 'label': 'Severity'},
                        {'key': 'reported_by', 'label': 'Reported By'},
                        {'key': 'date_reported', 'label': 'Date Reported'},
                        {'key': 'status', 'label': 'Status'},
                        {'key': 'resolution_date', 'label': 'Resolution Date'},
                    ]
                    exception_report_rows = sort_report_rows(exception_report_rows, report_sort or None, report_order)
                    if request.GET.get('share_report') == '1':
                        share_emails = [email.strip() for email in request.GET.get('share_email', '').split(',') if email.strip()]
                        if share_emails:
                            if send_report_share_email(request, report_category, report_period, exception_report_rows, report_table_columns, share_emails):
                                messages.success(request, f'Report shared with {", ".join(share_emails)}.')
                            return redirect_report_view(request)
                        messages.error(request, 'Please enter at least one valid email address to share this report.')
                        return redirect_report_view(request)
                    if export_pdf:
                        return export_report_pdf(report_category, report_period, exception_report_rows, report_table_columns)
                    if export_excel:
                        return export_report_excel(report_category, report_period, exception_report_rows, report_table_columns)
                    report_paginator, report_page_obj = paginate_report_rows(request, exception_report_rows, page_param=f'report_page_{report_category}')
                    exception_report_rows = list(report_page_obj.object_list)
                elif report_category == 'audit_trail':
                    qs = AuditLog.objects.filter(timestamp__date__gte=report_month, timestamp__date__lt=get_next_month(report_month)).order_by('-timestamp')
                    if report_search:
                        qs = qs.filter(Q(name__icontains=report_search) | Q(action__icontains=report_search) | Q(message__icontains=report_search))
                    if report_filter_status:
                        qs = qs.filter(severity__iexact=report_filter_status)
                    if report_date_from:
                        qs = qs.filter(timestamp__date__gte=parse_date(report_date_from))
                    if report_date_to:
                        qs = qs.filter(timestamp__date__lte=parse_date(report_date_to))
                    audit_trail_report_rows = []
                    for audit_entry in qs:
                        audit_trail_report_rows.append({
                            'audit_id': audit_entry.id,
                            'user': audit_entry.user.username if audit_entry.user else '-',
                            'action': audit_entry.action,
                            'module': audit_entry.name,
                            'record_affected': audit_entry.message,
                            'timestamp': audit_entry.timestamp,
                            'ip_address': '-',
                        })
                    report_table_columns = [
                        {'key': 'audit_id', 'label': 'Audit ID'},
                        {'key': 'user', 'label': 'User'},
                        {'key': 'action', 'label': 'Action'},
                        {'key': 'module', 'label': 'Module'},
                        {'key': 'record_affected', 'label': 'Record Affected'},
                        {'key': 'timestamp', 'label': 'Timestamp'},
                        {'key': 'ip_address', 'label': 'IP Address'},
                    ]
                    audit_trail_report_rows = sort_report_rows(audit_trail_report_rows, report_sort or None, report_order)
                    if request.GET.get('share_report') == '1':
                        share_emails = [email.strip() for email in request.GET.get('share_email', '').split(',') if email.strip()]
                        if share_emails:
                            if send_report_share_email(request, report_category, report_period, audit_trail_report_rows, report_table_columns, share_emails):
                                messages.success(request, f'Report shared with {", ".join(share_emails)}.')
                            return redirect_report_view(request)
                        messages.error(request, 'Please enter at least one valid email address to share this report.')
                        return redirect_report_view(request)
                    if export_pdf:
                        return export_report_pdf(report_category, report_period, audit_trail_report_rows, report_table_columns)
                    if export_excel:
                        return export_report_excel(report_category, report_period, audit_trail_report_rows, report_table_columns)
                    report_paginator, report_page_obj = paginate_report_rows(request, audit_trail_report_rows, page_param=f'report_page_{report_category}')
                    audit_trail_report_rows = list(report_page_obj.object_list)
                elif report_category == 'management_summary':
                    management_summary_report_rows = [{
                        'report_date': report_period,
                        'total_tapes': Tape.objects.count(),
                        'active_tapes': Tape.objects.filter(status='Active').count(),
                        'in_transit': Tape.objects.filter(status='In Transit').count(),
                        'missing_tapes': Tape.objects.filter(status='Missing').count(),
                        'damaged_tapes': Tape.objects.filter(status='Damaged').count(),
                        'open_exceptions': ShipmentException.objects.filter(status__in=['Open', 'Investigating']).count(),
                        'compliance_rate': '98.5%',
                        'reconciliation_accuracy': '99.2%',
                    }]
                    report_table_columns = [
                        {'key': 'report_date', 'label': 'Report Date'},
                        {'key': 'total_tapes', 'label': 'Total Tapes'},
                        {'key': 'active_tapes', 'label': 'Active Tapes'},
                        {'key': 'in_transit', 'label': 'In Transit'},
                        {'key': 'missing_tapes', 'label': 'Missing Tapes'},
                        {'key': 'damaged_tapes', 'label': 'Damaged Tapes'},
                        {'key': 'open_exceptions', 'label': 'Open Exceptions'},
                        {'key': 'compliance_rate', 'label': 'Compliance Rate'},
                        {'key': 'reconciliation_accuracy', 'label': 'Reconciliation Accuracy'},
                    ]
                    if request.GET.get('share_report') == '1':
                        share_emails = [email.strip() for email in request.GET.get('share_email', '').split(',') if email.strip()]
                        if share_emails:
                            if send_report_share_email(request, report_category, report_period, management_summary_report_rows, report_table_columns, share_emails):
                                messages.success(request, f'Report shared with {", ".join(share_emails)}.')
                            return redirect_report_view(request)
                        messages.error(request, 'Please enter at least one valid email address to share this report.')
                        return redirect_report_view(request)
                    if export_pdf:
                        return export_report_pdf(report_category, report_period, management_summary_report_rows, report_table_columns)
                    if export_excel:
                        return export_report_excel(report_category, report_period, management_summary_report_rows, report_table_columns)
        if generated_report_data:
            generated_report_items = [
                {
                    'label': key.replace('_', ' ').title(),
                    'value': value
                }
                for key, value in generated_report_data.items()
                if key != 'period'
            ]
        if export_csv:
            if report_category == 'inventory':
                return export_inventory_report_csv(report_period or current_month, inventory_report_tapes)
            if generated_report_data:
                return export_report_csv(report_type or 'monthly', report_period, generated_report_data)

    if show_reconciliation_reports_panel and export_reconciliation_csv:
        return export_reconciliation_report_csv(reconciliations, summary=reconciliation_report_summary)

    status_labels = ['Active', 'Off-Site', 'Scratch Eligible', 'Damaged']
    status_counts = [
        tapes.filter(status='Active').count(),
        tapes.filter(status='Off-Site').count(),
        tapes.filter(status='Scratch Eligible').count(),
        tapes.filter(status='Damaged').count(),
    ]
    monthly_labels, monthly_counts = get_last_six_month_counts(tapes)

    alert_count = AuditLog.objects.filter(severity__in=['warning', 'error'], is_read=False).count()

    dashboard_nav_urls = {
        'add_tape': build_signed_dashboard_navigation_token('add_tape', {'show_add_tape': '1'}),
        'tape_inventory': build_signed_dashboard_navigation_token('tape_inventory', {'show_tape_inventory': '1'}),
        'shipments': build_signed_dashboard_navigation_token('shipments', {'show_shipments': '1'}),
        'settings': build_signed_dashboard_navigation_token('settings', {'show_settings': '1'}),
        'reconciliation': build_signed_dashboard_navigation_token('reconciliation', {'show_reconciliation': '1'}),
        'reports': build_signed_dashboard_navigation_token('reports', {'show_reports': 'reports'}),
        'reconciliation_reports': build_signed_dashboard_navigation_token('reconciliation_reports', {'show_reconciliation_reports': 'reconciliation-reports'}),
        'audit_logs': build_signed_dashboard_navigation_token('audit_logs', {'show_audit': '1'}),
        'alerts': build_signed_dashboard_navigation_token('alerts', {'show_alerts': '1'}),
        'admin': build_signed_dashboard_navigation_token('admin', {'show_admin': '1'}),
        'profile': build_signed_dashboard_navigation_token('profile', {'show_profile': '1'}),
    }

    context = {
        'tapes': tapes,
        'tape_search': tape_search,
        'total_tapes': total_tapes,
        'active_tapes': active_tapes,
        'off_site_tapes': off_site_tapes,
        'missing_tapes': missing_tapes,
        'archived_tapes': archived_tapes,
        'retention_due': retention_due,
        'pending_shipments': shipments.filter(status__iexact='Pending').count(),
        'alert_count': alert_count,
        'shipments': shipments,
        'pending_users': pending_users,
        'pending_user_count': pending_users.count(),
        'recent_activities': recent_activities,
        'recent_alerts': recent_alerts,
        'reconciliation_request_context': reconciliation_request_context,
        'pending_approval_shipments': pending_approval_shipments,
        'approval_shipment': approval_shipment,
        'assignment_form': assignment_form,
        'audit_logs': audit_logs,
        'page_obj': page_obj,
        'paginator': paginator,
        'filters': filters,
        'active_filters': [
            {'name': 'Search', 'value': filters['search']},
            {'name': 'Log Type', 'value': filters['log_type']},
            {'name': 'Module', 'value': filters['module']},
            {'name': 'Severity', 'value': filters['severity']},
            {'name': 'User', 'value': filters['user']},
            {'name': 'Status', 'value': filters['status']},
            {'name': 'Date Range', 'value': filters['date_range']},
            {'name': 'From', 'value': filters['date_from']},
            {'name': 'To', 'value': filters['date_to']},
        ],
        'users': get_user_model().objects.order_by('username'),
        'log_types': ['Authentication', 'Request', 'Approval', 'Shipment', 'Inventory', 'Audit', 'Compliance', 'Exception', 'Reconciliation', 'Notification', 'API', 'Import', 'Export', 'Security', 'Configuration', 'System', 'Error'],
        'modules': ['Shipment', 'Inventory', 'Bank Branches', 'Couriers', 'Users', 'Vault', 'Reports', 'Dashboard', 'API', 'Notifications'],
        'severity_choices': ['info', 'success', 'warning', 'error'],
        'status_choices': ['Success', 'Failed', 'Pending', 'Approved', 'Rejected'],
        'date_ranges': [
            ('', 'All Dates'),
            ('today', 'Today'),
            ('yesterday', 'Yesterday'),
            ('last_7_days', 'Last 7 Days'),
            ('last_30_days', 'Last 30 Days'),
            ('this_month', 'This Month'),
            ('this_year', 'This Year'),
        ],
        'user_feature_names': user_feature_names,
        'dashboard_tabs': dashboard_tabs,
        'tape_form': tape_form,
        'tape_action_form': tape_action_form,
        'selected_tape': selected_tape,
        'show_add_tape_panel': show_add_tape_panel,
        'show_tape_actions_panel': show_tape_actions_panel,
        'show_tape_inventory_panel': show_tape_inventory_panel,
        'show_print_barcodes_panel': show_print_barcodes_panel,
        'show_admin_panel': show_admin_panel,
        'show_profile_panel': show_profile_panel,
        'show_settings_panel': show_settings_panel,
        'hide_dashboard_sidebar': hide_dashboard_sidebar,
        'show_audit_panel': show_audit_panel,
        'show_alerts_panel': show_alerts_panel,
        'show_admin_history': show_admin_history,
        'show_reports_panel': show_reports_panel,
        'show_reconciliation_reports_panel': show_reconciliation_reports_panel,
        'active_feature_key': active_feature_key,
        'show_shipments_panel': show_shipments_panel,
        'show_add_shipment_panel': show_add_shipment_panel,
        'show_edit_shipment_panel': show_edit_shipment_panel,
        'show_reconciliation_panel': show_reconciliation_panel,
        'show_add_reconciliation_panel': show_add_reconciliation_panel,
        'dashboard_nav_urls': dashboard_nav_urls,
        'add_shipment_form': add_shipment_form,
        'edit_shipment_form': edit_shipment_form,
        'selected_shipment': selected_shipment,
        'reconciliation_form': reconciliation_form,
        'reconciliation_result_form': reconciliation_result_form,
        'reconciliations': reconciliations,
        'reconciliation_results': reconciliation_results,
        'selected_reconciliation': selected_reconciliation,
        'show_unscanned_tapes': show_unscanned_tapes,
        'unscanned_tapes': unscanned_tapes,
        'available_operators': available_operators,
        'settings_form': settings_form,
        'profile_form': profile_form,
        'application_settings': application_settings,
        'reports': reports,
        'generated_report_data': generated_report_data,
        'generated_report_label': generated_report_label,
        'report_type': report_type,
        'report_category': report_category,
        'report_period': report_period,
        'report_categories': report_categories,
        'current_month': current_month,
        'report_email_form': report_email_form,
        'search_query': search_query,
        'reconciliation_report_summary': reconciliation_report_summary,
        'report_export_url': f"{reverse('backup-dashboard')}?show_reconciliation_reports=reconciliation-reports",
        'status_labels_json': json.dumps(status_labels),
        'status_counts_json': json.dumps(status_counts),
        'monthly_labels_json': json.dumps(monthly_labels),
        'monthly_counts_json': json.dumps(monthly_counts),
        'kpis': kpis,
        'report_categories': report_categories,
        'current_month': current_month,
        'pending_tape_requests': pending_tape_requests,
        'inventory_report_tapes': inventory_report_tapes,
        'shipment_report_rows': shipment_report_rows,
        'custody_report_rows': custody_report_rows,
        'reconciliation_report_rows': reconciliation_report_rows,
        'retention_report_rows': retention_report_rows,
        'compliance_report_rows': compliance_report_rows,
        'exception_report_rows': exception_report_rows,
        'audit_trail_report_rows': audit_trail_report_rows,
        'management_summary_report_rows': management_summary_report_rows,
        'report_table_columns': report_table_columns,
        'report_paginator': report_paginator,
        'report_page_obj': report_page_obj,
        'report_search': report_search,
        'report_filter_status': report_filter_status,
        'report_date_from': report_date_from,
        'report_date_to': report_date_to,
        'report_sort': report_sort,
        'report_order': report_order,
        'selected_reconciliation': selected_reconciliation,
        'show_unscanned_tapes': show_unscanned_tapes,
        'unscanned_tapes': unscanned_tapes,
        'schema_preview': schema_preview,
    }
    return render(request, 'backup_dashboard.html', context)


@user_passes_test(lambda u: u.is_superuser or is_operations_manager(u), login_url='signin')
@login_required(login_url='signin')
def start_shipment_request(request):
    assigned_branch = get_user_assigned_branch(request.user)
    partial_request = request.GET.get('partial') == '1' or request.POST.get('partial') == '1'
    branch_assignment_required = is_operations_manager(request.user) and not assigned_branch

    if branch_assignment_required:
        messages.error(request, get_branch_assignment_required_message())
        template_name = 'start_shipment_request_fragment.html' if partial_request else 'start_shipment_request.html'
        return render(request, template_name, {'form': None, 'partial_request': partial_request, 'branch_assignment_required': True})

    initial_branch_name = assigned_branch.branch_name if assigned_branch else ''
    initial_requester = request.user.get_full_name() or request.user.username
    form = ShipmentRequestSubmissionForm(request.POST or None, initial={'branch_name': initial_branch_name, 'requester_name': initial_requester})
    if request.method == 'POST' and form.is_valid():
        branch_name = assigned_branch.branch_name if assigned_branch else ''
        requester_name = (form.cleaned_data['requester_name'] or '').strip()
        request_details = form.cleaned_data['request_details'].strip()
        selected_tapes = list(form.cleaned_data.get('tapes') or [])
        shipment = Shipment.objects.create(
            shipment_date=timezone.localdate(),
            shipment_type='Off-Site Transfer',
            status='Pending',
            source_location=branch_name,
            destination_location=branch_name,
            requesting_branch=assigned_branch,
            releasing_custodian=requester_name or request.user.get_full_name() or request.user.username,
            receiving_organization='Pending review',
            approval_remarks=request_details,
            created_by=request.user,
            last_updated_by=request.user,
        )
        if selected_tapes:
            shipment.tapes.add(*selected_tapes)
            shipment.number_of_tapes = shipment.tapes.count()
            shipment.save(update_fields=['number_of_tapes'])
        user_email = request.user.email or 'N/A'
        AuditLog.objects.create(
            name='Shipment Created',
            action='Shipment Created',
            message=f'Shipment request {shipment.shipment_id} created for {branch_name} by {request.user.username} ({user_email})',
            user=request.user,
            severity='info',
        )
        messages.success(request, f'Shipment request {shipment.shipment_id} sent successfully for review and assignment.')
        redirect_url = reverse('operations-dashboard') + '?show_reports=reports&report_category=reconciliation&report_period=2026-07&report_type=monthly&report_only=1'
        return redirect(redirect_url)
    if request.method == 'POST':
        messages.error(request, 'Please provide the requester name and request details.')
    template_name = 'start_shipment_request_fragment.html' if partial_request else 'start_shipment_request.html'
    return render(request, template_name, {'form': form, 'partial_request': partial_request, 'branch_assignment_required': False})


@user_passes_test(lambda u: u.is_superuser or is_operations_manager(u), login_url='signin')
@login_required(login_url='signin')
def operations_dashboard(request):
    assigned_branch = get_user_assigned_branch(request.user)
    initial_requester = request.user.get_full_name() or request.user.username
    shipment_request_form = ShipmentRequestSubmissionForm(request.POST or None, initial={'branch_name': assigned_branch.branch_name if assigned_branch else '', 'requester_name': initial_requester})
    shipments = Shipment.objects.all().order_by('-shipment_date')
    tapes = Tape.objects.all()
    reconciliations = Reconciliation.objects.order_by('-reconciliation_date')[:5]
    reconciliation_results = ReconciliationResult.objects.order_by('-created_at')
    selected_reconciliation = None
    show_unscanned_tapes = False
    audit_logs = AuditLog.objects.order_by('-timestamp')[:12]

    operations_dashboard_nav_urls = {
        'dashboard': build_dashboard_navigation_url('operations_dashboard', {}, dashboard='operations'),
        'history': build_dashboard_navigation_url('operations_dashboard', {'show_admin': '1', 'show_admin_history': '1'}, dashboard='operations'),
        'notifications': build_dashboard_navigation_url('operations_dashboard', {'show_notifications': '1'}, dashboard='operations'),
        'profile': build_dashboard_navigation_url('operations_dashboard', {'show_profile': '1'}, dashboard='operations'),
        'profile_edit': build_dashboard_navigation_url('operations_dashboard', {'show_profile': '1', 'edit_profile': '1'}, dashboard='operations'),
    }

    unread_alert_count = AuditLog.objects.filter(severity__in=['warning', 'error'], is_read=False).count()
    notification_items = []
    for item in AuditLog.objects.filter(severity__in=['warning', 'error']).order_by('-timestamp')[:10]:
        target_url = operations_dashboard_nav_urls['notifications']
        metadata = _parse_auditlog_metadata(item.message)
        if metadata.get('target_url'):
            target_url = metadata.get('target_url')
        elif item.message and 'reconciliation_pk=' in item.message:
            parts = {p.split('=')[0]: p.split('=')[1] for p in (item.message or '').split('|') if '=' in p}
            reconciliation_pk = parts.get('reconciliation_pk')
            if reconciliation_pk:
                target_url = f"{reverse('operations-dashboard')}?show_reconciliation=1&show_unscanned_tapes=1&reconciliation_pk={reconciliation_pk}"
        notification_items.append({
            'severity': item.severity,
            'timestamp': item.timestamp,
            'message': item.action or item.name or item.message,
            'action': 'View',
            'target_url': target_url,
            'is_read': item.is_read,
        })
    show_profile_panel = False
    profile_edit_mode = False
    show_notifications_panel = False
    show_analytics_panel = False
    show_monitoring_panel = False
    show_exception_panel = False
    show_custody_panel = False
    show_reconciliation_panel = False
    show_compliance_panel = False
    show_admin_history = False
    profile_form = UserProfileForm(instance=request.user)

    if request.method == 'GET' and request.GET.get('mark_notifications_read') == '1':
        AuditLog.objects.filter(severity__in=['warning', 'error'], is_read=False).update(is_read=True, read_at=timezone.now())
        unread_alert_count = 0

    reconciliation_pk = request.GET.get('reconciliation_pk')
    if reconciliation_pk:
        selected_reconciliation = get_object_by_uuid_pk(Reconciliation, reconciliation_pk)
        if selected_reconciliation:
            show_reconciliation_panel = True

    if request.GET.get('show_reconciliation') == '1' or 'show_reconciliation' in request.GET:
        show_reconciliation_panel = True

    if request.GET.get('show_unscanned_tapes') == '1':
        show_unscanned_tapes = True

    if show_unscanned_tapes and not selected_reconciliation:
        show_unscanned_tapes = False

    unscanned_tapes = Tape.objects.none()
    if selected_reconciliation and show_unscanned_tapes:
        unscanned_tapes = Tape.objects.filter(scan_status__in=['Pending', 'Assigned', 'Scanning']).order_by('volser')

    tape_request_form = TapeRequestForm(request.POST or None)

    if request.method == 'POST' and request.POST.get('form_type') == 'submit_shipment_request':
        if is_operations_manager(request.user) and not assigned_branch:
            messages.error(request, get_branch_assignment_required_message())
            return redirect(reverse('operations-dashboard'))
        if shipment_request_form.is_valid():
            branch_name = assigned_branch.branch_name if assigned_branch else ''
            requester_name = (shipment_request_form.cleaned_data['requester_name'] or '').strip()
            request_details = shipment_request_form.cleaned_data['request_details'].strip()
            selected_tapes = list(shipment_request_form.cleaned_data.get('tapes') or [])
            shipment = Shipment.objects.create(
                shipment_date=timezone.localtime(),
                shipment_type='Off-Site Transfer',
                status='Pending',
                source_location=branch_name,
                destination_location=branch_name,
                requesting_branch=assigned_branch,
                releasing_custodian=requester_name or request.user.get_full_name() or request.user.username,
                receiving_organization='Pending review',
                approval_remarks=request_details,
                created_by=request.user,
                last_updated_by=request.user,
            )
            if selected_tapes:
                shipment.tapes.add(*selected_tapes)
                shipment.number_of_tapes = shipment.tapes.count()
                shipment.save(update_fields=['number_of_tapes'])
            user_email = request.user.email or 'N/A'
            AuditLog.objects.create(
                name='Shipment Created',
                action='Shipment Created',
                message=f'Shipment request {shipment.shipment_id} created for {branch_name} by {request.user.username} ({user_email})',
                user=request.user,
                severity='info',
            )
            messages.success(request, f'Shipment request {shipment.shipment_id} sent successfully for review and assignment.')
            redirect_url = reverse('operations-dashboard') + '?show_reports=reports&report_category=reconciliation&report_period=2026-07&report_type=monthly&report_only=1'
            return redirect(redirect_url)
        messages.error(request, 'Please provide the requester name and request details.')

    if request.method == 'POST' and request.POST.get('form_type') == 'start_reconciliation':
        reconciliation_pk = request.POST.get('reconciliation_pk')
        selected_reconciliation = get_object_by_uuid_pk(Reconciliation, reconciliation_pk)
        if selected_reconciliation:
            if selected_reconciliation.status == 'Open':
                selected_reconciliation.status = 'In Progress'
                selected_reconciliation.scan_started_at = timezone.localtime()
                selected_reconciliation.save(update_fields=['status', 'scan_started_at'])
                AuditLog.objects.create(
                    name='Reconciliation Started',
                    action=f'Started reconciliation {selected_reconciliation.reconciliation_id}',
                    user=request.user,
                    message=f'reconciliation_id={selected_reconciliation.reconciliation_id}',
                    severity='info',
                )
                messages.success(request, 'Reconciliation scan started.')
            else:
                messages.info(request, 'Reconciliation is already in progress or cannot be started.')
        else:
            messages.error(request, 'Unable to start the selected reconciliation.')
        return redirect(f'{reverse("operations-dashboard")}?show_reconciliation=1&show_unscanned_tapes=1&reconciliation_pk={selected_reconciliation.id if selected_reconciliation else ""}')

    if request.method == 'POST' and request.POST.get('form_type') == 'finish_reconciliation':
        reconciliation_pk = request.POST.get('reconciliation_pk')
        report_comment = (request.POST.get('report_comment') or '').strip()
        selected_reconciliation = get_object_by_uuid_pk(Reconciliation, reconciliation_pk)
        if selected_reconciliation:
            if selected_reconciliation.status == 'In Progress':
                selected_reconciliation.status = 'Completed'
                selected_reconciliation.scan_completed_at = timezone.localtime()
                if selected_reconciliation.total_tapes_expected > 0:
                    selected_reconciliation.progress_percent = 100
                selected_reconciliation.save(update_fields=['status', 'scan_completed_at', 'progress_percent'])
                AuditLog.objects.create(
                    name='Reconciliation Finished',
                    action=f'Finished reconciliation {selected_reconciliation.reconciliation_id}',
                    user=request.user,
                    message=f'reconciliation_id={selected_reconciliation.reconciliation_id}|reconciliation_pk={selected_reconciliation.id}|report_comment={report_comment}',
                    severity='success',
                )
                backup_admins = User.objects.filter(is_active=True, groups__name='Backup Administrator').distinct()
                recipients = [admin.email for admin in backup_admins if admin.email]
                if recipients:
                    subject = f'Reconciliation {selected_reconciliation.reconciliation_id} finished'
                    body = (
                        f'Reconciliation {selected_reconciliation.reconciliation_id} has been marked as completed by {request.user.username}.\n\n'
                        f'Report details:\n{report_comment or "No additional details provided."}\n\n'
                        f'View the reconciliation in the Operations Dashboard for follow-up.'
                    )
                    if ApplicationSetting.objects.first() and ApplicationSetting.objects.first().email_alerts_enabled:
                        send_email_alert(subject, body, recipients)
                messages.success(request, 'Reconciliation scan finished and backup admin notified.')
            else:
                messages.info(request, 'Reconciliation is not currently in progress.')
        else:
            messages.error(request, 'Unable to finish the selected reconciliation.')
        return redirect(f'{reverse("operations-dashboard")}?show_reconciliation=1&show_unscanned_tapes=1&reconciliation_pk={selected_reconciliation.id if selected_reconciliation else ""}')

    if request.method == 'POST' and request.POST.get('form_type') == 'close_reconciliation':
        reconciliation_pk = request.POST.get('reconciliation_pk')
        selected_reconciliation = get_object_by_uuid_pk(Reconciliation, reconciliation_pk)
        if selected_reconciliation:
            if selected_reconciliation.status == 'Completed':
                selected_reconciliation.status = 'Closed'
                selected_reconciliation.save(update_fields=['status', 'updated_at'])
                AuditLog.objects.create(
                    name='Reconciliation Closed',
                    action=f'Closed reconciliation {selected_reconciliation.reconciliation_id}',
                    user=request.user,
                    message=f'reconciliation_id={selected_reconciliation.reconciliation_id}|reconciliation_pk={selected_reconciliation.id}',
                    severity='success',
                )
                messages.success(request, 'Reconciliation has been closed successfully.')
            else:
                messages.info(request, 'Only completed reconciliations can be closed.')
        else:
            messages.error(request, 'Unable to close the selected reconciliation.')
        return redirect(f'{reverse("operations-dashboard")}?show_reconciliation=1&show_unscanned_tapes=1&reconciliation_pk={selected_reconciliation.id if selected_reconciliation else ""}')

    show_reports_panel = request.GET.get('show_reports') in {'1', 'reports'} or 'show_reports' in request.GET
    active_feature_key = request.GET.get('feature_key') or request.POST.get('feature_key')
    feature_panel_state = get_feature_panel_state(active_feature_key)
    if feature_panel_state:
        if feature_panel_state.get('show_analytics_panel'):
            show_analytics_panel = True
        if feature_panel_state.get('show_monitoring_panel'):
            show_monitoring_panel = True
        if feature_panel_state.get('show_exception_panel'):
            show_exception_panel = True
        if feature_panel_state.get('show_custody_panel'):
            show_custody_panel = True
        if feature_panel_state.get('show_compliance_panel'):
            show_compliance_panel = True
        if feature_panel_state.get('show_reports_panel'):
            show_reports_panel = True
        if feature_panel_state.get('show_reconciliation_panel'):
            show_reconciliation_panel = True

    report_categories = get_report_categories()
    current_month = timezone.localdate().strftime('%Y-%m')
    report_category = request.GET.get('report_category')
    report_period = request.GET.get('report_period') or request.GET.get(f'report_period_{report_category}') or current_month
    report_type = request.GET.get('report_type', 'monthly')
    report_search = request.GET.get(f'report_search_{report_category}', '')
    report_filter_status = request.GET.get(f'report_filter_status_{report_category}', '')
    report_sort = request.GET.get(f'report_sort_{report_category}', '')
    report_order = (request.GET.get(f'report_order_{report_category}', 'asc') or 'asc').lower()
    if report_order not in {'asc', 'desc'}:
        report_order = 'asc'

    inventory_report_tapes = []
    shipment_report_rows = []
    report_table_columns = []
    report_paginator = None
    report_page_obj = None
    report_title = 'Reports'
    report_subtitle = 'Monthly inventory, shipment, and compliance reporting'

    if show_reports_panel and report_category and report_period:
        report_month = get_first_day_of_month(report_period) or get_first_day_of_month(current_month)
        if report_month:
            if report_category == 'inventory':
                qs = Tape.objects.filter(date_registered__gte=report_month, date_registered__lt=get_next_month(report_month)).order_by('volser')
                if report_search:
                    qs = qs.filter(
                        Q(volser__icontains=report_search) |
                        Q(barcode__icontains=report_search) |
                        Q(rfid_tag__icontains=report_search) |
                        Q(tape_type__icontains=report_search) |
                        Q(status__icontains=report_search) |
                        Q(current_location__icontains=report_search)
                    )
                if report_filter_status:
                    qs = qs.filter(status__iexact=report_filter_status)
                if report_sort in {'volser', 'status', 'retention_end_date', 'date_registered', 'current_location'}:
                    if report_order == 'desc':
                        qs = qs.order_by(f'-{report_sort}')
                    else:
                        qs = qs.order_by(report_sort)
                inventory_report_tapes = list(qs)
                for tape in inventory_report_tapes:
                    tape.latest_custodian = get_latest_custodian_for_tape(tape) or tape.current_location or '-'
                report_table_columns = [
                    {'key': 'volser', 'label': 'VolSER'},
                    {'key': 'barcode', 'label': 'Barcode'},
                    {'key': 'status', 'label': 'Status'},
                    {'key': 'current_location', 'label': 'Current Location'},
                    {'key': 'latest_custodian', 'label': 'Custodian'},
                ]
                report_title = 'Inventory Report'
                report_subtitle = f'Inventory records for {report_month.strftime("%B %Y")}'
            elif report_category == 'shipment':
                qs = Shipment.objects.filter(shipment_date__gte=report_month, shipment_date__lt=get_next_month(report_month)).order_by('-shipment_date')
                if report_search:
                    qs = qs.filter(
                        Q(shipment_id__icontains=report_search) |
                        Q(destination_location__icontains=report_search) |
                        Q(status__icontains=report_search) |
                        Q(courier_name__icontains=report_search)
                    )
                if report_filter_status:
                    qs = qs.filter(status__iexact=report_filter_status)
                shipment_report_rows = list(qs)
                report_table_columns = [
                    {'key': 'shipment_id', 'label': 'Shipment ID'},
                    {'key': 'status', 'label': 'Status'},
                    {'key': 'destination_location', 'label': 'Destination'},
                    {'key': 'shipment_date', 'label': 'Dispatch Date'},
                ]
                report_title = 'Shipment Report'
                report_subtitle = f'Shipment activity for {report_month.strftime("%B %Y")}'

    if request.method == 'POST' and request.POST.get('form_type') == 'submit_tape_request':
        if tape_request_form.is_valid():
            tape_request = tape_request_form.save(commit=False)
            tape_request.requested_by = request.user
            tape_request.save()
            AuditLog.objects.create(
                name='Tape Request Submitted',
                action=f'{request.user.username} requested tape {tape_request.tape.volser}',
                user=request.user,
                severity='info',
            )
            messages.success(request, 'Tape request submitted successfully.')
            return redirect(f'{reverse("operations-dashboard")}?show_requests=1')
        messages.error(request, 'Please correct the tape request form and try again.')

    if request.method == 'POST' and request.POST.get('form_type') == 'edit_profile':
        profile_form = UserProfileForm(request.POST, instance=request.user)
        if profile_form.is_valid():
            profile_form.save()
            AuditLog.objects.create(
                name='Profile Updated',
                action=f'Updated profile for {request.user.username}',
                user=request.user,
                severity='success',
            )
            messages.success(request, 'Your profile has been updated.')
            return redirect(f'{reverse("operations-dashboard")}?show_profile=1')
        else:
            show_profile_panel = True
            profile_edit_mode = True

    total_tapes = tapes.count()
    tapes_in_transit = tapes.filter(status='In Transit').count()
    missing_tapes = tapes.filter(status='Missing').count()
    overdue_shipments = shipments.filter(
        expected_delivery_date__lt=timezone.localdate()
    ).exclude(status__in=['Delivered', 'Cancelled']).count()
    open_exceptions = reconciliation_results.filter(resolution_status__in=['Open', 'Under Investigation']).count()
    pending_approvals = shipments.filter(status__iexact='Pending').count()
    compliance_rate = int(max(0, min(100, 100 - ((missing_tapes + open_exceptions) * 1.5))))
    reconciliation_accuracy = int(max(0, min(100, 100 - (open_exceptions * 1.2))))

    pending_shipments = shipments.filter(status__iexact='Pending').order_by('-shipment_date')[:6]
    pending_tape_requests = TapeRequest.objects.select_related('tape', 'requested_by', 'approved_by').filter(status='Pending').order_by('-request_date')[:8]
    my_tape_requests = TapeRequest.objects.select_related('tape', 'approved_by').filter(requested_by=request.user).order_by('-request_date')[:8]
    monitoring_shipments = shipments.order_by('-shipment_date')[:12]
    latest_reconciliations = reconciliations
    open_discrepancies = reconciliation_results.filter(resolution_status__in=['Open', 'Under Investigation']).order_by('-created_at')[:6]
    recent_activity = audit_logs

    filter_form = ShipmentApprovalFilterForm(request.GET or None)
    manifest_form = ManifestSearchForm(request.GET or None)
    approval_shipments = shipments.select_related('approved_by', 'created_by', 'last_updated_by').prefetch_related('tapes')

    risk_level_filter = None
    if filter_form.is_valid():
        search = filter_form.cleaned_data.get('search')
        if search:
            approval_shipments = approval_shipments.filter(
                Q(shipment_id__icontains=search) |
                Q(source_location__icontains=search) |
                Q(destination_location__icontains=search) |
                Q(created_by__username__icontains=search) |
                Q(receiving_organization__icontains=search)
            )

        status = filter_form.cleaned_data.get('status')
        if status:
            approval_shipments = approval_shipments.filter(status=status)

        priority = filter_form.cleaned_data.get('priority')
        if priority:
            approval_shipments = approval_shipments.filter(priority_level=priority)

        date_from = filter_form.cleaned_data.get('date_from')
        if date_from:
            approval_shipments = approval_shipments.filter(shipment_date__gte=date_from)

        date_to = filter_form.cleaned_data.get('date_to')
        if date_to:
            approval_shipments = approval_shipments.filter(shipment_date__lte=date_to)

        risk_level = filter_form.cleaned_data.get('risk_level')
        if risk_level:
            risk_level_filter = risk_level

    if manifest_form.is_valid():
        manifest_query = manifest_form.cleaned_data.get('query')
        if manifest_query:
            approval_shipments = approval_shipments.filter(
                Q(tapes__volser__icontains=manifest_query) |
                Q(tapes__barcode__icontains=manifest_query) |
                Q(tapes__rfid_tag__icontains=manifest_query)
            ).distinct()

    if risk_level_filter:
        approval_shipments = [shipment for shipment in approval_shipments if shipment.risk_level() == risk_level_filter]

    if not isinstance(approval_shipments, list):
        approval_shipments = list(approval_shipments)

    if request.GET.get('show_profile') == '1':
        show_profile_panel = True
        if request.GET.get('edit_profile') == '1':
            profile_edit_mode = True

    if request.GET.get('show_notifications') == '1':
        show_notifications_panel = True

    total_shipments = len(approval_shipments)
    pending_count = sum(1 for shipment in approval_shipments if shipment.status == 'Pending')
    more_info_count = sum(1 for shipment in approval_shipments if shipment.status == 'More Info Requested')
    approved_count = sum(1 for shipment in approval_shipments if shipment.status == 'Approved')
    rejected_count = sum(1 for shipment in approval_shipments if shipment.status == 'Rejected')
    critical_count = sum(1 for shipment in approval_shipments if shipment.priority_level == 'Critical')
    non_compliant_count = sum(1 for shipment in approval_shipments if not shipment.compliance_passed())
    overdue_count = sum(1 for shipment in approval_shipments if shipment.is_overdue_for_approval())

    page_number = request.GET.get('page', 1)
    paginator = Paginator(approval_shipments, 10)
    shipment_page = paginator.get_page(page_number)

    delivery_confirmed_shipments = Shipment.objects.filter(deliveries__isnull=False).distinct()
    receipt_confirmed_shipments = Shipment.objects.filter(receipts__isnull=False).distinct()

    custody_transfers_open = receipt_confirmed_shipments.filter(
        status__in=['Dispatched', 'In Transit', 'Picked Up']
    ).exclude(pk__in=delivery_confirmed_shipments.values_list('pk', flat=True)).distinct().count()
    custody_transfers_completed = delivery_confirmed_shipments.count()
    missing_handoffs = shipments.filter(
        status__in=['Dispatched', 'In Transit', 'Picked Up']
    ).exclude(pk__in=receipt_confirmed_shipments.values_list('pk', flat=True)).distinct().count()
    unverified_deliveries = shipments.filter(status__iexact='Delivered').exclude(
        pk__in=delivery_confirmed_shipments.values_list('pk', flat=True)
    ).distinct().count()
    chain_of_custody_compliance = int(max(0, min(100, 100 - (missing_handoffs * 20) - (unverified_deliveries * 10))))

    exceptions = ShipmentException.objects.order_by('-reported_date')[:6]
    open_exceptions = ShipmentException.objects.filter(status__in=['Open', 'Investigating']).count()
    pending_approvals = shipments.filter(status__iexact='Pending').count()

    status_labels = ['Pending', 'Approved', 'Dispatched', 'In Transit', 'Delivered', 'Cancelled']
    status_counts = [shipments.filter(status=status).count() for status in status_labels]

    monthly_labels = []
    monthly_shipments = []
    monthly_compliance = []
    monthly_exceptions = []
    today = timezone.localdate()
    for offset in range(5, -1, -1):
        month = (today.replace(day=1) - timedelta(days=offset * 30)).strftime('%b %Y')
        monthly_labels.append(month)
        monthly_shipments.append(
            shipments.filter(shipment_date__year=today.year, shipment_date__month=(today.month - offset - 1) % 12 + 1).count()
        )
        monthly_compliance.append(max(75, min(100, 100 - missing_tapes)))
        monthly_exceptions.append(
            reconciliation_results.filter(created_at__year=today.year, created_at__month=(today.month - offset - 1) % 12 + 1).count()
        )

    alert_items = []
    recent_alerts = AuditLog.objects.filter(severity__in=['warning', 'error']).order_by('-timestamp')[:5]
    for item in recent_alerts:
        alert_items.append({
            'severity': item.severity.title(),
            'date': item.timestamp,
            'description': item.action or item.name or item.message,
            'action': 'Review',
            'target_url': operations_dashboard_nav_urls['notifications'],
        })
    reconciliation_alerts = reconciliation_results.filter(issue_type__in=['Missing', 'Damaged', 'Unexpected']).order_by('-created_at')[:3]
    for result in reconciliation_alerts:
        alert_items.append({
            'severity': 'Warning',
            'date': result.created_at,
            'description': f'{result.issue_type} tape detected for reconciliation {result.reconciliation.reconciliation_id}',
            'action': 'Investigate',
            'target_url': f"{reverse('operations-dashboard')}?show_reports=reports&report_category=reconciliation&report_type=monthly",
        })

    chart_data = {
        'status_labels_json': json.dumps(status_labels),
        'status_counts_json': json.dumps(status_counts),
        'monthly_labels_json': json.dumps(monthly_labels),
        'monthly_shipments_json': json.dumps(monthly_shipments),
        'monthly_compliance_json': json.dumps(monthly_compliance),
        'monthly_exceptions_json': json.dumps(monthly_exceptions),
    }

    report_only = request.GET.get('report_only') in {'1', 'true', 'True'}

    context = {
        'dashboard_tabs': OPERATIONS_FEATURE_TABS,
        'current_datetime': timezone.localtime(),
        'today': today,
        'monthly_labels': monthly_labels,
        'monthly_shipments': monthly_shipments,
        'total_tapes': total_tapes,
        'tapes_in_transit': tapes_in_transit,
        'overdue_shipments': overdue_shipments,
        'missing_tapes': missing_tapes,
        'open_exceptions': open_exceptions,
        'pending_approvals': pending_approvals,
        'compliance_rate': compliance_rate,
        'reconciliation_accuracy': reconciliation_accuracy,
        'alert_items': alert_items,
        'pending_shipments': pending_shipments,
        'monitoring_shipments': monitoring_shipments,
        'latest_reconciliations': latest_reconciliations,
        'exceptions': exceptions,
        'open_discrepancies': open_discrepancies,
        'recent_activity': recent_activity,
        'custody_transfers_open': custody_transfers_open,
        'custody_transfers_completed': custody_transfers_completed,
        'missing_handoffs': missing_handoffs,
        'unverified_deliveries': unverified_deliveries,
        'chain_of_custody_compliance': chain_of_custody_compliance,
        'filter_form': filter_form,
        'manifest_form': manifest_form,
        'shipments': shipment_page,
        'total_shipments': total_shipments,
        'pending_count': pending_count,
        'more_info_count': more_info_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'critical_count': critical_count,
        'non_compliant_count': non_compliant_count,
        'overdue_count': overdue_count,
        'has_results': total_shipments > 0,
        'profile_form': profile_form,
        'show_profile_panel': show_profile_panel,
        'profile_edit_mode': profile_edit_mode,
        'show_notifications_panel': show_notifications_panel,
        'show_analytics_panel': show_analytics_panel,
        'show_monitoring_panel': show_monitoring_panel,
        'show_exception_panel': show_exception_panel,
        'show_custody_panel': show_custody_panel,
        'show_reconciliation_panel': show_reconciliation_panel,
        'show_compliance_panel': show_compliance_panel,
        'show_reports_panel': show_reports_panel,
        'active_feature_key': active_feature_key,
        'report_categories': report_categories,
        'current_month': current_month,
        'report_category': report_category,
        'report_period': report_period,
        'report_type': report_type,
        'report_search': report_search,
        'report_filter_status': report_filter_status,
        'report_sort': report_sort,
        'report_order': report_order,
        'report_title': report_title,
        'report_subtitle': report_subtitle,
        'operations_dashboard_nav_urls': operations_dashboard_nav_urls,
        'inventory_report_tapes': inventory_report_tapes,
        'shipment_report_rows': shipment_report_rows,
        'report_table_columns': report_table_columns,
        'report_paginator': report_paginator,
        'report_page_obj': report_page_obj,
        'unread_alert_count': unread_alert_count,
        'notification_items': notification_items,
        'tape_request_form': tape_request_form,
        'pending_tape_requests': pending_tape_requests,
        'my_tape_requests': my_tape_requests,
        'shipment_request_form': shipment_request_form,
        'selected_reconciliation': selected_reconciliation,
        'show_unscanned_tapes': show_unscanned_tapes,
        'unscanned_tapes': unscanned_tapes,
        **chart_data,
    }
    # If admin requested their own history, build and filter the audit queryset
    audit_page_obj = None
    audit_paginator = None
    if 'show_admin_history' in request.GET:
        show_admin_history = True
        queryset, filters = build_audit_log_queryset(request)
        try:
            queryset = queryset.filter(user=request.user)
            filters['user'] = request.user.username
        except Exception:
            pass
        export_format = (request.GET.get('export') or '').strip().lower()
        if export_format in {'csv', 'excel', 'pdf'}:
            export_response = export_audit_logs(queryset, export_format)
            if export_response is not None:
                return export_response
        per_page = filters.get('per_page', 25)
        audit_paginator = Paginator(queryset, per_page)
        audit_page = request.GET.get('audit_page') or request.GET.get('page') or 1
        audit_page_obj = audit_paginator.get_page(audit_page)

    if report_only and report_category:
        context['report_only'] = True
        return render(request, 'operations_dashboard.html', context)

    # attach audit pagination objects when admin history requested
    if show_admin_history:
        context['audit_logs'] = audit_page_obj.object_list if audit_page_obj else []
        context['page_obj'] = audit_page_obj
        context['paginator'] = audit_paginator
        context['show_admin_history'] = True

    return render(request, 'operations_dashboard.html', context)


@user_passes_test(lambda u: u.is_superuser or is_warehouse_staff(u), login_url='signin')
@login_required(login_url='signin')
def warehouse_operations_dashboard(request):
    shipments = Shipment.objects.select_related('created_by', 'approved_by_backup', 'approved_by_supreme').order_by('-shipment_date')
    tapes = Tape.objects.order_by('volser')
    pending_shipments = shipments.filter(status='Pending').order_by('-shipment_date')[:6]
    in_transit_shipments = shipments.filter(status__in=['Dispatched', 'In Transit', 'Picked Up']).order_by('-shipment_date')[:8]
    warehouse_ready = shipments.filter(status='Approved').order_by('-shipment_date')[:8]
    open_reconciliations = Reconciliation.objects.filter(status='Open').order_by('-reconciliation_date')[:5]
    recent_exceptions = ShipmentException.objects.select_related('shipment', 'tape', 'reported_by').order_by('-reported_date')[:5]
    context = {
        'dashboard_tabs': OPERATIONS_FEATURE_TABS,
        'shipments': shipments[:12],
        'pending_shipments': pending_shipments,
        'in_transit_shipments': in_transit_shipments,
        'warehouse_ready': warehouse_ready,
        'total_tapes': tapes.count(),
        'active_tapes': tapes.filter(status='Active').count(),
        'in_transit_tapes': tapes.filter(status='In Transit').count(),
        'damaged_tapes': tapes.filter(status='Damaged').count(),
        'open_reconciliations': open_reconciliations,
        'recent_exceptions': recent_exceptions,
        'warehouse_summary': {
            'pending_shipments': pending_shipments.count(),
            'ready_for_dispatch': warehouse_ready.count(),
            'in_transit': in_transit_shipments.count(),
        },
        'warehouse_tools': [
            {
                'title': 'Shipment approvals',
                'description': 'Review incoming shipment requests and approve handoffs.',
                'url': reverse('shipment-approvals'),
                'badge': pending_shipments.count(),
            },
            {
                'title': 'Dispatch queue',
                'description': 'Track shipments already approved and ready to leave the warehouse.',
                'url': reverse('warehouse-operations-dashboard'),
                'badge': warehouse_ready.count(),
            },
            {
                'title': 'Inventory control',
                'description': 'Inspect tape inventory, status, and location readiness.',
                'url': reverse('operations-dashboard'),
                'badge': tapes.count(),
            },
            {
                'title': 'Exception handling',
                'description': 'Investigate incidents, damage reports, and custody issues.',
                'url': reverse('operations-dashboard'),
                'badge': recent_exceptions.count(),
            },
            {
                'title': 'Reconciliation',
                'description': 'Process open reconciliation work and scan requests.',
                'url': reverse('operations-dashboard') + '?show_reconciliation=1&show_unscanned_tapes=1',
                'badge': open_reconciliations.count(),
            },
            {
                'title': 'Manifest review',
                'description': 'Verify shipment manifests and tape lists before dispatch.',
                'url': reverse('shipment-approvals'),
                'badge': len(shipments[:5]),
            },
        ],
    }
    return render(request, 'warehouse_operations_dashboard.html', context)


def _sync_pending_approval_queue_from_shipments():
    pending_shipments = Shipment.objects.filter(
        approval_stage='awaiting_supreme',
        status__in=['Pending', 'More Info Requested'],
    ).select_related('created_by', 'approved_by_backup', 'requesting_branch')

    for shipment in pending_shipments:
        if PendingApproval.objects.filter(related_model='shipment', related_object_id=str(shipment.pk)).exists():
            continue

        branch = shipment.requesting_branch or getattr(getattr(shipment, 'created_by', None), 'assigned_branch', None)
        backup_admin = shipment.approved_by_backup or getattr(shipment, 'created_by', None)
        requester = getattr(shipment, 'created_by', None)
        risk_level = 'High' if shipment.priority_level in {'High', 'Critical'} else 'Medium'
        from .utils import create_pending_approval

        create_pending_approval(
            transaction_type='Shipment',
            module='Shipment Workflow',
            summary=f"Shipment {shipment.shipment_id} requires supreme approval.",
            requester=requester,
            backup_administrator=backup_admin,
            branch=branch,
            priority=shipment.priority_level or 'Normal',
            risk_level=risk_level,
            status='Awaiting Supreme Approval',
            request_date=shipment.created_at or timezone.now(),
            request_payload={'shipment_id': shipment.shipment_id},
            related_object_id=str(shipment.pk),
            related_model='shipment',
            audit_history=[{
                'action': 'Imported',
                'user': 'System',
                'timestamp': timezone.localtime().isoformat(),
                'comment': 'Legacy shipment surfaced in approval queue',
            }],
        )


@user_passes_test(lambda u: u.is_superuser or is_supreme_approver(u), login_url='signin')
@login_required(login_url='signin')
def supreme_approver_dashboard(request):
    _sync_pending_approval_queue_from_shipments()
    approval_queue = PendingApproval.objects.select_related('requester', 'backup_administrator', 'branch').filter(status__in=['Pending', 'Awaiting Supreme Approval', 'Clarification Requested']).order_by('-request_date')
    pending_shipments = approval_queue.filter(status__in=['Pending', 'Awaiting Supreme Approval', 'Clarification Requested'])
    pending_tape_requests = TapeRequest.objects.filter(status='Pending').select_related('tape', 'requested_by', 'shipment').order_by('-request_date')[:8]
    pending_reconciliations = Reconciliation.objects.filter(status='Open').select_related('performed_by', 'reviewed_by').order_by('-reconciliation_date')[:8]
    recent_audit = AuditLog.objects.select_related('user').order_by('-timestamp')[:8]
    pending_user_changes = []
    high_risk_transactions = [item for item in approval_queue[:6] if item.risk_level in {'High', 'Critical'}]

    pending_count = pending_shipments.count()
    approved_today = PendingApproval.objects.filter(approved_at__date=timezone.localdate(), status='Approved').count()
    rejected_today = PendingApproval.objects.filter(approved_at__date=timezone.localdate(), status='Rejected').count()
    pending_warehouse_release = Shipment.objects.filter(status='Approved').count()
    pending_branch_updates = BankBranch.objects.filter(status='Active').count()
    pending_bulk_imports = BranchImportLog.objects.filter(import_status='warning').count()
    pending_barcode_changes = 0
    pending_user_change_requests = len(pending_user_changes)

    queue_payload = []
    for approval in approval_queue[:30]:
        queue_payload.append({
            'id': approval.id,
            'reference_id': approval.get_display_reference(),
            'priority': approval.priority,
            'risk': approval.risk_level,
            'transaction_type': approval.transaction_type,
            'module': approval.module,
            'requester': approval.get_requester_name(),
            'branch': approval.get_branch_name(),
            'department': 'Operations',
            'requested_date': approval.request_date,
            'backup_admin': approval.get_backup_name(),
            'backup_approval_time': approval.approved_at or approval.request_date,
            'status': approval.status,
            'age': (timezone.localdate() - approval.request_date.date()).days,
            'approval': approval,
        })

    context = {
        'dashboard_tabs': OPERATIONS_FEATURE_TABS,
        'shipments': pending_shipments[:10],
        'approval_queue': queue_payload,
        'pending_count': pending_count,
        'total_pending': pending_count,
        'approved_today': approved_today,
        'rejected_today': rejected_today,
        'high_risk_transactions': high_risk_transactions,
        'pending_tape_requests': pending_tape_requests,
        'pending_reconciliations': pending_reconciliations,
        'recent_audit': recent_audit,
        'pending_user_changes': pending_user_changes,
        'pending_warehouse_release': pending_warehouse_release,
        'pending_bulk_imports': pending_bulk_imports,
        'pending_barcode_changes': pending_barcode_changes,
        'pending_user_change_requests': pending_user_change_requests,
        'pending_branch_updates': pending_branch_updates,
        'current_user': request.user,
        'current_datetime': timezone.localtime(),
        'last_login': request.user.last_login or timezone.localtime() - timedelta(days=1),
        'approval_sla': '24 Hours',
        'notifications': [
            {'title': 'Urgent approval review', 'message': f'{pending_count} approvals need immediate attention.', 'type': 'warning'},
            {'title': 'Warehouse release pending', 'message': 'Approved shipments await release.', 'type': 'info'},
            {'title': 'System alert', 'message': 'Branch import warnings require review.', 'type': 'danger'},
        ],
        'recent_activity': [
            {'user': 'A. Okello', 'timestamp': timezone.localtime() - timedelta(minutes=12), 'branch': 'Kampala Main', 'module': 'Approval Queue', 'action': 'Reviewed Pending Approval'},
            {'user': 'J. Nankya', 'timestamp': timezone.localtime() - timedelta(minutes=41), 'branch': 'Jinja', 'module': 'Approval Queue', 'action': 'Rejected Pending Approval'},
            {'user': 'R. Kato', 'timestamp': timezone.localtime() - timedelta(hours=2), 'branch': 'Mbarara', 'module': 'Approval Queue', 'action': 'Approved Pending Approval'},
        ],
        'high_risk_types': [
            {'name': 'Tape Destruction', 'badge': 'Critical'},
            {'name': 'Bulk Upload', 'badge': 'High'},
            {'name': 'Permission Changes', 'badge': 'High'},
            {'name': 'Role Changes', 'badge': 'Critical'},
        ],
        'dashboard_title': 'Enterprise Banking Approval Center',
        'system_version': 'v2.4.1',
        'database_name': 'backup_inventory',
        'server_status': 'Operational',
        'redis_status': 'Healthy',
        'websocket_status': 'Connected',
    }
    return render(request, 'supreme_approver_dashboard.html', context)


def _apply_role_feature_change_request(approval, actor, comment=''):
    payload = approval.request_payload or {}
    role_id = payload.get('role_id')
    feature_keys = list(payload.get('feature_keys') or [])
    role = Role.objects.filter(pk=role_id).first()
    if not role:
        raise Http404

    existing_features = {feature.feature_key: feature for feature in role.get_active_features()}
    target_features = {feature.feature_key: feature for feature in Feature.objects.filter(feature_key__in=feature_keys, is_active=True)}

    for feature_key, feature in target_features.items():
        RoleFeature.objects.update_or_create(
            role=role,
            feature=feature,
            defaults={
                'is_active': True,
                'assigned_by': actor,
            },
        )

    for feature_key, feature in existing_features.items():
        if feature_key not in target_features:
            RoleFeature.objects.filter(role=role, feature=feature).update(is_active=False, assigned_by=actor)

    if role.group:
        template, _ = RoleTemplate.objects.get_or_create(group=role.group)
        template.features = feature_keys
        template.save(update_fields=['features'])

    approval.status = 'Approved'
    approval.approved_by = actor
    approval.reviewed_by = actor
    approval.approved_at = timezone.localtime()
    approval.review_comment = comment or f'Approved role feature update for {role.name}.'
    approval.rejection_reason = ''
    approval.audit_history = list(approval.audit_history or []) + [{
        'action': 'Approved',
        'user': actor.get_full_name() or actor.username,
        'timestamp': timezone.localtime().isoformat(),
        'comment': approval.review_comment,
        'role': role.name,
        'previous_configuration': payload.get('previous_feature_keys', []),
        'new_configuration': feature_keys,
    }]
    approval.save(update_fields=['status', 'approved_by', 'reviewed_by', 'approved_at', 'review_comment', 'rejection_reason', 'audit_history', 'updated_at'])


def _execute_pending_changes(approval, actor):
    """Execute requested_changes for a PendingApproval. Handles simple create/update for Tape model."""
    changes = approval.requested_changes or []
    results = []
    tape_model_fields = {field.name for field in Tape._meta.get_fields() if getattr(field, 'concrete', False) and not getattr(field, 'many_to_many', False) and not getattr(field, 'one_to_many', False)}
    with transaction.atomic():
        for change in changes:
            try:
                action = (change.get('action') or '').lower()
                model_name = (change.get('model') or '').lower()
                if model_name == 'tape':
                    if action == 'create':
                        fields = change.get('fields') or {}
                        model_fields = {k: v for k, v in fields.items() if k in tape_model_fields}
                        extra_fields = {k: v for k, v in fields.items() if k not in tape_model_fields}
                        tape = Tape.objects.create(**model_fields)
                        if extra_fields:
                            set_clause = ', '.join(f"{connection.ops.quote_name(name)} = %s" for name in extra_fields)
                            params = list(extra_fields.values()) + [tape.pk]
                            with connection.cursor() as cursor:
                                cursor.execute(f"UPDATE inventory_tape SET {set_clause} WHERE id = %s", params)
                        AuditLog.objects.create(name='Tape Registered (Approved)', action=f'Registered tape {getattr(tape, "volser", "n/a")}', user=actor, severity='success')
                        results.append(('create', str(tape.pk)))
                    elif action == 'update':
                        obj_id = change.get('object_id') or approval.related_object_id
                        tape = Tape.objects.filter(pk=obj_id).first()
                        if tape:
                            fields = change.get('fields') or {}
                            model_fields = {k: v for k, v in fields.items() if k in tape_model_fields}
                            extra_fields = {k: v for k, v in fields.items() if k not in tape_model_fields}
                            for k, v in model_fields.items():
                                setattr(tape, k, v)
                            tape.save()
                            if extra_fields:
                                set_clause = ', '.join(f"{connection.ops.quote_name(name)} = %s" for name in extra_fields)
                                params = list(extra_fields.values()) + [tape.pk]
                                with connection.cursor() as cursor:
                                    cursor.execute(f"UPDATE inventory_tape SET {set_clause} WHERE id = %s", params)
                            AuditLog.objects.create(name='Tape Updated (Approved)', action=f'Updated tape {tape.volser}', user=actor, severity='success')
                            results.append(('update', str(tape.pk)))
                        else:
                            logger.warning('PendingApproval execution: Tape not found %s', obj_id)
                            results.append(('missing', str(obj_id)))
                    elif action == 'schema_sync':
                        sql_statements = change.get('sql_statements') or []
                        import_changes = change.get('import_changes') or []
                        for sql_statement in sql_statements:
                            with connection.cursor() as cursor:
                                cursor.execute(sql_statement)
                        new_columns = approval.request_payload.get('new_columns') or []
                        if not new_columns:
                            for sql_statement in sql_statements:
                                match = re.search(r'ADD COLUMN IF NOT EXISTS\s+([a-z0-9_]+)\s+([A-Z0-9\(\),]+)', sql_statement, re.IGNORECASE)
                                if match:
                                    new_columns.append({'name': match.group(1), 'detected_type': match.group(2)})
                        for column_info in new_columns:
                            SchemaChangeLog.objects.create(
                                column_name=column_info.get('name', ''),
                                detected_data_type=column_info.get('detected_type', ''),
                                target_table='inventory_tape',
                                source_excel_filename=approval.request_payload.get('file_name', ''),
                                uploaded_by=approval.requester,
                                approved_by=actor,
                                approval_timestamp=timezone.localtime(),
                                sql_executed='; '.join(sql_statements),
                                synchronization_status='applied',
                            )
                        AuditLog.objects.create(name='Tape Schema Synchronized (Approved)', action=f'Executed {len(sql_statements)} schema changes for Tape', user=actor, severity='success')
                        for import_change in import_changes:
                            sub_action = (import_change.get('action') or '').lower()
                            if sub_action == 'create':
                                fields = import_change.get('fields') or {}
                                model_fields = {k: v for k, v in fields.items() if k in tape_model_fields}
                                extra_fields = {k: v for k, v in fields.items() if k not in tape_model_fields}
                                created_tape = Tape.objects.create(**model_fields)
                                if extra_fields:
                                    set_clause = ', '.join(f"{connection.ops.quote_name(name)} = %s" for name in extra_fields)
                                    params = list(extra_fields.values()) + [created_tape.pk]
                                    with connection.cursor() as cursor:
                                        cursor.execute(f"UPDATE inventory_tape SET {set_clause} WHERE id = %s", params)
                                AuditLog.objects.create(name='Tape Registered (Schema Sync)', action=f'Registered tape {getattr(created_tape, "volser", "n/a")}', user=actor, severity='success')
                                results.append(('create', str(created_tape.pk)))
                            elif sub_action == 'update':
                                obj_id = import_change.get('object_id')
                                tape = Tape.objects.filter(pk=obj_id).first()
                                if tape:
                                    fields = import_change.get('fields') or {}
                                    model_fields = {k: v for k, v in fields.items() if k in tape_model_fields}
                                    extra_fields = {k: v for k, v in fields.items() if k not in tape_model_fields}
                                    for k, v in model_fields.items():
                                        setattr(tape, k, v)
                                    tape.save()
                                    if extra_fields:
                                        set_clause = ', '.join(f"{connection.ops.quote_name(name)} = %s" for name in extra_fields)
                                        params = list(extra_fields.values()) + [tape.pk]
                                        with connection.cursor() as cursor:
                                            cursor.execute(f"UPDATE inventory_tape SET {set_clause} WHERE id = %s", params)
                                    AuditLog.objects.create(name='Tape Updated (Schema Sync)', action=f'Updated tape {tape.volser}', user=actor, severity='success')
                                    results.append(('update', str(tape.pk)))
                                else:
                                    logger.warning('PendingApproval execution: Tape not found during schema sync %s', obj_id)
                                    results.append(('missing', str(obj_id)))
                            else:
                                logger.warning('Unsupported import action %s in schema sync approval %s', sub_action, approval.pk)
                        results.append(('schema_sync', len(sql_statements)))
                    else:
                        logger.warning('Unsupported tape action %s in approval %s', action, approval.pk)
                else:
                    logger.warning('Unsupported model %s in approval execution %s', model_name, approval.pk)
            except Exception as exc:
                logger.exception('Error executing pending change for approval %s: %s', approval.pk, exc)
                results.append(('error', str(exc)))
    return results


@user_passes_test(lambda u: u.is_superuser or is_supreme_approver(u) or is_backup_administrator(u), login_url='signin')
@login_required(login_url='signin')
def approval_review(request, approval_id):
    approval = get_object_or_404(PendingApproval.objects.select_related('requester', 'backup_administrator', 'branch'), pk=approval_id)
    related_object = approval.get_related_object()
    is_role_feature_request = (approval.related_model == 'role_feature' or approval.transaction_type.lower() == 'role feature update')

    if request.method == 'POST':
        decision = (request.POST.get('decision') or 'approve').strip().lower()
        comment = (request.POST.get('comment') or '').strip()
        if decision == 'approve':
            if is_role_feature_request and is_backup_administrator(request.user) and not is_supreme_approver(request.user):
                approval.status = 'Awaiting Supreme Approval'
                approval.backup_approved_by = request.user
                approval.backup_approved_at = timezone.localtime()
                approval.review_comment = comment or 'Approved by backup administrator and forwarded to supreme approver.'
                approval.audit_history = list(approval.audit_history or []) + [{
                    'action': 'Backup Approved',
                    'user': request.user.get_full_name() or request.user.username,
                    'timestamp': timezone.localtime().isoformat(),
                    'comment': approval.review_comment,
                }]
                approval.save(update_fields=['status', 'backup_approved_by', 'backup_approved_at', 'review_comment', 'audit_history', 'updated_at'])
                messages.success(request, 'Role feature change request approved by backup administrator and sent to supreme approval.')
            elif is_role_feature_request and (is_supreme_approver(request.user) or request.user.is_superuser):
                role = _apply_role_feature_change_request(approval, request.user, comment=comment)
                messages.success(request, f'Role features for {role.name} were approved and activated.')
            else:
                approval.status = 'Approved'
                approval.approved_by = request.user
                approval.reviewed_by = request.user
                approval.approved_at = timezone.localtime()
                approval.review_comment = comment or 'Approved by supreme approver.'
                approval.rejection_reason = ''
                approval.audit_history = list(approval.audit_history or []) + [{
                    'action': 'Approved',
                    'user': request.user.get_full_name() or request.user.username,
                    'timestamp': timezone.localtime().isoformat(),
                    'comment': approval.review_comment,
                }]
                approval.save(update_fields=['status', 'approved_by', 'reviewed_by', 'approved_at', 'review_comment', 'rejection_reason', 'audit_history', 'updated_at'])
                # Execute requested changes for approved pending approvals (generic executor)
                if approval.requested_changes:
                    exec_results = _execute_pending_changes(approval, request.user)
                    approval.audit_history = list(approval.audit_history or []) + [{
                        'action': 'Executed',
                        'user': request.user.get_full_name() or request.user.username,
                        'timestamp': timezone.localtime().isoformat(),
                        'results': exec_results,
                    }]
                    approval.save(update_fields=['audit_history', 'updated_at'])
                if related_object and hasattr(related_object, 'approval_stage'):
                    related_object.status = 'Approved'
                    related_object.approval_stage = 'approved'
                    related_object.approved_by_supreme = request.user
                    related_object.approved_by = request.user
                    related_object.approval_date = timezone.localtime()
                    related_object.approval_remarks = approval.review_comment
                    related_object.last_updated_by = request.user
                    related_object.save(update_fields=['status', 'approval_stage', 'approved_by_supreme', 'approved_by', 'approval_date', 'approval_remarks', 'last_updated_by', 'last_updated_at'])
                messages.success(request, 'Approval request approved and the staged change was executed.')
        else:
            approval.status = 'Rejected'
            approval.reviewed_by = request.user
            approval.approved_at = timezone.localtime()
            approval.review_comment = comment or 'Rejected by supreme approver.'
            approval.rejection_reason = comment or 'Rejected by supreme approver.'
            approval.audit_history = list(approval.audit_history or []) + [{
                'action': 'Rejected',
                'user': request.user.get_full_name() or request.user.username,
                'timestamp': timezone.localtime().isoformat(),
                'comment': approval.review_comment,
            }]
            approval.save(update_fields=['status', 'reviewed_by', 'approved_at', 'review_comment', 'rejection_reason', 'audit_history', 'updated_at'])
            messages.warning(request, 'Approval request rejected and discarded.')
        return redirect(reverse('supreme-approver-dashboard'))

    context = {
        'approval': approval,
        'related_object': related_object,
        'requester_name': approval.get_requester_name(),
        'backup_name': approval.get_backup_name(),
        'branch_name': approval.get_branch_name(),
        'current_user': request.user,
        'dashboard_title': 'Approval Review',
        'is_read_only': True,
    }
    return render(request, 'approval_review.html', context)


@user_passes_test(lambda u: u.is_superuser or is_operations_manager(u) or is_supreme_approver(u) or is_warehouse_staff(u), login_url='signin')
@login_required(login_url='signin')
def exception_detail(request, pk):
    exception = get_object_or_404(
        ShipmentException.objects.select_related('shipment', 'tape', 'reported_by'),
        pk=pk
    )
    context = {
        'dashboard_tabs': OPERATIONS_FEATURE_TABS,
        'exception': exception,
    }
    if request.GET.get('partial'):
        return render(request, 'exception_detail_fragment.html', context)
    return render(request, 'exception_detail.html', context)


@user_passes_test(lambda u: u.is_superuser or is_operations_manager(u) or is_backup_administrator(u) or is_supreme_approver(u) or is_warehouse_staff(u), login_url='signin')
@login_required(login_url='signin')
def shipment_approvals(request):
    if request.method == 'POST' and request.POST.get('form_type') == 'backup_admin_decision':
        shipment_id = request.POST.get('shipment_id')
        shipment = get_object_or_404(Shipment, pk=shipment_id)
        tape_id = request.POST.get('tape_id')
        courier_id = request.POST.get('courier_id')
        decision = (request.POST.get('decision') or 'approve').strip().lower()
        comments = (request.POST.get('comments') or '').strip()
        current_user_is_supreme = is_supreme_approver(request.user)
        current_user_is_backup = is_backup_administrator(request.user) and not current_user_is_supreme

        if decision == 'approve':
            if current_user_is_backup:
                if not tape_id or not courier_id:
                    messages.error(request, 'Select both an available tape and courier before approving the shipment.')
                    return redirect(reverse('shipment-approvals'))

                tape = get_object_or_404(Tape, pk=tape_id)
                courier_profile = get_object_or_404(CourierProfile, pk=courier_id)
                shipment.tapes.add(tape)
                shipment.number_of_tapes = shipment.tapes.count()
                shipment.courier_name = courier_profile.full_name
                shipment.courier_contact = courier_profile.phone_number
                shipment.tracking_number = f"TRK-{shipment.shipment_id[:8].upper()}"
                shipment.status = 'Pending'
                shipment.approved_by_backup = request.user
                shipment.approved_by = request.user
                shipment.approval_stage = 'awaiting_supreme'
                shipment.approval_date = timezone.localtime()
                shipment.approval_remarks = comments or 'Approved by backup administrator. Pending supreme approver.'
                shipment.last_updated_by = request.user
                shipment.save(update_fields=['status', 'approved_by', 'approved_by_backup', 'approval_stage', 'approval_date', 'approval_remarks', 'last_updated_by', 'last_updated_at', 'courier_name', 'courier_contact', 'tracking_number', 'number_of_tapes'])
                ShipmentApprovalHistory.objects.create(shipment=shipment, action='Approved', comments=comments, user=request.user)
                AuditLog.objects.create(name='Shipment Awaiting Supreme Approval', action=f'Shipment {shipment.shipment_id} passed first-stage approval and is waiting for supreme approval.', user=request.user, severity='warning')
                messages.success(request, 'Shipment passed first-stage backup approval and is awaiting supreme approval.')
                return redirect(reverse('shipment-approvals'))

            if current_user_is_supreme:
                shipment.status = 'Approved'
                shipment.approval_stage = 'approved'
                shipment.approved_by_supreme = request.user
                shipment.approved_by = request.user
                shipment.approval_date = timezone.localtime()
                shipment.approval_remarks = comments or 'Approved by supreme approver.'
                shipment.last_updated_by = request.user
                shipment.save(update_fields=['status', 'approved_by', 'approved_by_supreme', 'approval_stage', 'approval_date', 'approval_remarks', 'last_updated_by', 'last_updated_at'])
                ShipmentApprovalHistory.objects.create(shipment=shipment, action='Approved', comments=comments, user=request.user)
                AuditLog.objects.create(name='Shipment Approved', action=f'Shipment {shipment.shipment_id} was approved by the supreme approver.', user=request.user, severity='success')
                # Notify the backup administrator to print the approval form (prefilled)
                try:
                    backup_admin = shipment.approved_by_backup or getattr(shipment, 'created_by', None)
                    if backup_admin:
                        preview_url = reverse('approval-form-preview', args=[shipment.pk])
                        AuditLog.objects.create(
                            name='Shipment Approval Ready for Printing',
                            action=f'Shipment {shipment.shipment_id} approved and ready for printing by backup admin.',
                            message=f'target_url={preview_url}',
                            user=backup_admin,
                            severity='warning',
                        )
                        if application_settings.email_alerts_enabled and getattr(backup_admin, 'email', None):
                            send_mail(
                                f'Shipment {shipment.shipment_id} approved - print approval form',
                                f'Shipment {shipment.shipment_id} was approved. Print the approval form here: {request.build_absolute_uri(preview_url)}',
                                settings.DEFAULT_FROM_EMAIL,
                                [backup_admin.email],
                                fail_silently=True,
                            )
                except Exception:
                    pass
                messages.success(request, 'Shipment approved and released for operations.')
                return redirect(reverse('shipment-approvals'))

        shipment.status = 'Rejected' if decision == 'reject' else 'More Info Requested'
        shipment.approval_stage = 'rejected' if decision == 'reject' else 'draft'
        shipment.approved_by = request.user
        if current_user_is_backup:
            shipment.approved_by_backup = request.user
        if current_user_is_supreme:
            shipment.approved_by_supreme = request.user
        shipment.approval_date = timezone.localtime()
        shipment.approval_remarks = comments or 'Shipment request was not approved.'
        shipment.last_updated_by = request.user
        shipment.save(update_fields=['status', 'approval_stage', 'approved_by', 'approved_by_backup', 'approved_by_supreme', 'approval_date', 'approval_remarks', 'last_updated_by', 'last_updated_at'])
        ShipmentApprovalHistory.objects.create(shipment=shipment, action='Rejected' if decision == 'reject' else 'Requested More Information', comments=comments, user=request.user)
        AuditLog.objects.create(name='Shipment Decision Recorded', action=f'Shipment {shipment.shipment_id} was {shipment.status.lower()}', user=request.user, severity='warning')
        messages.warning(request, 'Shipment decision recorded.')
        return redirect(reverse('shipment-approvals'))

    filter_form = ShipmentApprovalFilterForm(request.GET or None)
    manifest_form = ManifestSearchForm(request.GET or None)
    shipments = Shipment.objects.select_related('approved_by', 'created_by', 'last_updated_by').prefetch_related('tapes').order_by('-shipment_date')

    risk_level_filter = None
    if filter_form.is_valid():
        search = filter_form.cleaned_data.get('search')
        if search:
            shipments = shipments.filter(
                Q(shipment_id__icontains=search) |
                Q(source_location__icontains=search) |
                Q(destination_location__icontains=search) |
                Q(created_by__username__icontains=search) |
                Q(receiving_organization__icontains=search)
            )

        status = filter_form.cleaned_data.get('status')
        if status:
            shipments = shipments.filter(status=status)

        priority = filter_form.cleaned_data.get('priority')
        if priority:
            shipments = shipments.filter(priority_level=priority)

        date_from = filter_form.cleaned_data.get('date_from')
        if date_from:
            shipments = shipments.filter(shipment_date__gte=date_from)

        date_to = filter_form.cleaned_data.get('date_to')
        if date_to:
            shipments = shipments.filter(shipment_date__lte=date_to)

        risk_level = filter_form.cleaned_data.get('risk_level')
        if risk_level:
            risk_level_filter = risk_level

    if manifest_form.is_valid():
        manifest_query = manifest_form.cleaned_data.get('query')
        if manifest_query:
            shipments = shipments.filter(
                Q(tapes__volser__icontains=manifest_query) |
                Q(tapes__barcode__icontains=manifest_query) |
                Q(tapes__rfid_tag__icontains=manifest_query)
            ).distinct()

    if risk_level_filter:
        shipments = [shipment for shipment in shipments if shipment.risk_level() == risk_level_filter]

    if not isinstance(shipments, list):
        shipments = list(shipments)

    total_shipments = len(shipments)
    pending_count = sum(1 for shipment in shipments if shipment.status == 'Pending')
    more_info_count = sum(1 for shipment in shipments if shipment.status == 'More Info Requested')
    approved_count = sum(1 for shipment in shipments if shipment.status == 'Approved')
    rejected_count = sum(1 for shipment in shipments if shipment.status == 'Rejected')
    critical_count = sum(1 for shipment in shipments if shipment.priority_level == 'Critical')
    non_compliant_count = sum(1 for shipment in shipments if not shipment.compliance_passed())
    overdue_count = sum(1 for shipment in shipments if shipment.is_overdue_for_approval())

    page_number = request.GET.get('page', 1)
    paginator = Paginator(shipments, 10)
    shipment_page = paginator.get_page(page_number)

    context = {
        'dashboard_tabs': OPERATIONS_FEATURE_TABS,
        'filter_form': filter_form,
        'manifest_form': manifest_form,
        'shipments': shipment_page,
        'total_shipments': total_shipments,
        'pending_count': pending_count,
        'more_info_count': more_info_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'critical_count': critical_count,
        'non_compliant_count': non_compliant_count,
        'overdue_count': overdue_count,
        'has_results': total_shipments > 0,
        'hide_sidebar': bool(request.GET.get('feature_key') or request.POST.get('feature_key')),
    }
    is_partial_request = request.GET.get('partial') or request.POST.get('partial')
    is_ajax_request = request.headers.get('x-requested-with', '').lower() == 'xmlhttprequest'
    if is_partial_request:
        if is_ajax_request:
            return render(request, 'shipment_approvals_fragment.html', context)
        embedded_fragment_html = render_to_string('shipment_approvals_fragment.html', context, request=request)
        context['embedded_fragment_html'] = embedded_fragment_html
        context['show_embedded_content'] = True
        return render(request, 'operations_dashboard.html', context)
    return render(request, 'shipment_approvals.html', context)


@user_passes_test(lambda u: u.is_superuser or is_operations_manager(u) or is_backup_administrator(u), login_url='signin')
@login_required(login_url='signin')
def shipment_detail(request, shipment_pk):
    shipment = get_object_or_404(Shipment.objects.prefetch_related('tapes', 'approval_history'), pk=shipment_pk)
    receipt_form = OperatorReceiptCompletionForm(
        request.POST or None,
        initial={'receiving_custodian': request.user.get_full_name() or request.user.username}
    )
    approval_form = ShipmentApprovalDecisionForm(request.POST or None, initial={'shipment_pk': shipment.pk})
    assignment_form = BackupShipmentAssignmentForm(request.POST or None)
    manifest_search_form = ManifestSearchForm(request.GET or None)
    manifest_tapes = shipment.tapes.all()

    if manifest_search_form.is_valid():
        manifest_query = manifest_search_form.cleaned_data.get('query')
        if manifest_query:
            manifest_tapes = manifest_tapes.filter(
                Q(volser__icontains=manifest_query) |
                Q(barcode__icontains=manifest_query) |
                Q(rfid_tag__icontains=manifest_query)
            )

    application_settings = ApplicationSetting.objects.first() or ApplicationSetting.objects.create()
    partial_request = request.GET.get('partial') == '1'

    if request.method == 'POST':
        if request.POST.get('form_type') == 'operator_receipt_completion':
            if receipt_form.is_valid():
                receiving_custodian = receipt_form.cleaned_data.get('receiving_custodian', '').strip()
                receipt_notes = receipt_form.cleaned_data.get('receipt_notes', '').strip()
                shipment.status = 'Completed'
                shipment.received_by = receiving_custodian or request.user.get_full_name() or request.user.username
                shipment.delivery_date = timezone.localdate()
                shipment.delivery_time = timezone.localtime().time()
                shipment.delivery_status = 'Delivered'
                shipment.delivery_notes = receipt_notes
                shipment.last_updated_by = request.user
                shipment.save(update_fields=['status', 'received_by', 'delivery_date', 'delivery_time', 'delivery_status', 'delivery_notes', 'last_updated_by', 'last_updated_at'])

                AuditLog.objects.create(
                    name='Shipment Received',
                    action=f'Shipment {shipment.shipment_id} was received and marked complete',
                    user=request.user,
                    severity='success',
                )

                messages.success(request, 'Shipment has been received and marked as completed.')
                approval_form = ShipmentApprovalDecisionForm(initial={'shipment_pk': shipment.pk})
            else:
                messages.error(request, 'Please provide a receiving custodian before completing the shipment.')
        elif request.POST.get('form_type') == 'confirm_shipment_review':
            shipment.status = 'Completed'
            shipment.received_by = shipment.receiving_custodian or request.user.get_full_name() or request.user.username
            shipment.delivery_date = timezone.localdate()
            shipment.delivery_time = timezone.localtime().time()
            shipment.delivery_status = 'Delivered'
            shipment.last_updated_by = request.user
            shipment.save(update_fields=['status', 'received_by', 'delivery_date', 'delivery_time', 'delivery_status', 'last_updated_by', 'last_updated_at'])
            AuditLog.objects.create(
                name='Shipment Review Confirmed',
                action=f'Shipment {shipment.shipment_id} review was confirmed and finalized',
                message=f'target_url={reverse("shipment-detail", args=[shipment.pk])}',
                user=request.user,
                severity='warning',
            )
            messages.success(request, 'Shipment review confirmed and finalized.')
            approval_form = ShipmentApprovalDecisionForm(initial={'shipment_pk': shipment.pk})
        elif request.POST.get('form_type') == 'request_return':
            return_form = ReturnShipmentRequestForm(request.POST or None)
            if return_form.is_valid():
                courier_selection = return_form.cleaned_data['courier']
                comments = return_form.cleaned_data['comments'].strip()
                courier_profile, courier_user, courier_name, courier_contact = _resolve_courier_selection(courier_selection)

                shipment.return_requested_by = request.user
                shipment.return_requested_at = timezone.localtime()
                shipment.return_request_comments = comments
                shipment.return_assigned_courier_profile = courier_profile
                shipment.return_assigned_courier_user = courier_user
                shipment.return_courier_response_status = 'Pending'
                shipment.status = 'Return Requested'
                shipment.last_updated_by = request.user
                shipment.save(update_fields=[
                    'return_requested_by',
                    'return_requested_at',
                    'return_request_comments',
                    'return_assigned_courier_profile',
                    'return_assigned_courier_user',
                    'return_courier_response_status',
                    'status',
                    'last_updated_by',
                    'last_updated_at',
                ])

                ShipmentApprovalHistory.objects.create(
                    shipment=shipment,
                    action='Requested More Information' if shipment.shipment_type != 'Return' else 'Return Requested',
                    comments=comments,
                    user=request.user,
                )

                AuditLog.objects.create(
                    name='Return Requested',
                    action=(
                        f'Return requested for shipment {shipment.shipment_id} and assigned to {courier_name or "selected courier"}.'
                    ),
                    message=f'target_url={reverse("courier-dashboard")}',
                    user=courier_user,
                    severity='warning',
                )

                if courier_profile and courier_profile.email:
                    send_courier_profile_email_alert(
                        courier_profile,
                        f'Return request assigned: {shipment.shipment_id}',
                        f'A return request for shipment {shipment.shipment_id} has been assigned to you. Please respond in the courier dashboard.'
                    )

                application_settings = ApplicationSetting.objects.first() or ApplicationSetting.objects.create()
                backup_admins = User.objects.filter(is_active=True, groups__name='Backup Administrator').distinct()
                for admin in backup_admins:
                    AuditLog.objects.create(
                        name='Return Approval Required',
                        action=f'Return request for shipment {shipment.shipment_id} requires your approval.',
                        message=f'target_url={reverse("shipment-detail", args=[shipment.pk])}',
                        user=admin,
                        severity='warning',
                    )
                    if application_settings.email_alerts_enabled and admin.email:
                        send_mail(
                            f'Return request requires approval: {shipment.shipment_id}',
                            f'A return request for shipment {shipment.shipment_id} has been submitted and requires your approval.',
                            settings.DEFAULT_FROM_EMAIL,
                            [admin.email],
                            fail_silently=True,
                        )
                messages.success(request, 'Return request submitted and courier assigned. Awaiting courier confirmation and backup admin approval.')
                return redirect('shipment-detail', shipment_pk=shipment.pk)
            else:
                messages.error(request, 'Please select a courier and provide return notes.')
        elif request.POST.get('form_type') == 'backup_admin_assignment':
            if assignment_form.is_valid():
                tape = assignment_form.cleaned_data['tape']
                courier_selection = assignment_form.cleaned_data['courier']
                decision = assignment_form.cleaned_data['decision']
                comments = assignment_form.cleaned_data.get('comments', '').strip()
                courier_profile = None
                courier_user = None
                courier_name = ''
                courier_contact = ''

                if courier_selection and isinstance(courier_selection, str):
                    if courier_selection.startswith('profile:'):
                        profile_id = courier_selection.split(':', 1)[1]
                        courier_profile = CourierProfile.objects.filter(pk=profile_id).first()
                        if courier_profile:
                            courier_user = courier_profile.user
                            courier_name = courier_profile.full_name
                            courier_contact = courier_profile.phone_number
                    elif courier_selection.startswith('user:'):
                        user_id = courier_selection.split(':', 1)[1]
                        courier_user = get_user_model().objects.filter(pk=user_id).first()
                        if courier_user:
                            courier_name = courier_user.get_full_name() or courier_user.username
                            courier_contact = courier_user.email or ''
                            courier_profile = getattr(courier_user, 'courier_profile', None)
                    else:
                        courier_profile = CourierProfile.objects.filter(pk=courier_selection).first()
                        if courier_profile:
                            courier_user = courier_profile.user
                            courier_name = courier_profile.full_name
                            courier_contact = courier_profile.phone_number
                        else:
                            courier_user = get_user_model().objects.filter(pk=courier_selection).first()
                            if courier_user:
                                courier_name = courier_user.get_full_name() or courier_user.username
                                courier_contact = courier_user.email or ''
                                courier_profile = getattr(courier_user, 'courier_profile', None)

                shipment.tapes.add(tape)
                shipment.number_of_tapes = shipment.tapes.count()
                shipment.courier_name = courier_name or (courier_profile.full_name if courier_profile else '')
                shipment.courier_contact = courier_contact or (courier_profile.phone_number if courier_profile else '')
                shipment.tracking_number = f"TRK-{shipment.shipment_id[:8].upper()}"
                if decision == 'approve':
                    shipment.status = 'Approved'
                    shipment.approved_by = request.user
                    shipment.approval_date = timezone.localtime()
                    shipment.approval_remarks = comments or 'Approved by backup administrator.'
                else:
                    shipment.status = 'Rejected'
                    shipment.approved_by = request.user
                    shipment.approval_date = timezone.localtime()
                    shipment.approval_remarks = comments or 'Rejected by backup administrator.'
                shipment.last_updated_by = request.user
                shipment.save(update_fields=['status', 'approved_by', 'approval_date', 'approval_remarks', 'last_updated_by', 'last_updated_at', 'courier_name', 'courier_contact', 'tracking_number', 'number_of_tapes'])
                ShipmentApprovalHistory.objects.create(shipment=shipment, action='Approved' if decision == 'approve' else 'Rejected', comments=comments, user=request.user)
                if decision == 'approve':
                    AuditLog.objects.create(
                        name='Shipment Assigned',
                        action=f'Shipment {shipment.shipment_id} assigned to courier {courier_name or (courier_profile.full_name if courier_profile else "selected courier")}',
                        user=request.user,
                        severity='success',
                    )

                    operator = shipment.created_by
                    assigned_courier = courier_name or (courier_profile.full_name if courier_profile else 'the selected courier')
                    if operator:
                        operator_message = (
                            f'You approved shipment {shipment.shipment_id} requested by {operator.get_full_name() or operator.username} '
                            f'and assigned it to {assigned_courier}.'
                        )
                        AuditLog.objects.create(
                            name='Shipment Approved',
                            action=operator_message,
                            message=f'target_url={reverse("shipment-detail", args=[shipment.pk])}',
                            user=operator,
                            severity='warning',
                        )
                        if application_settings.email_alerts_enabled and operator.email:
                            send_mail(
                                f'Shipment {shipment.shipment_id} approved',
                                f'Your shipment request {shipment.shipment_id} was approved by the backup administrator and assigned to {assigned_courier}.',
                                settings.DEFAULT_FROM_EMAIL,
                                [operator.email],
                                fail_silently=True,
                            )

                    pickup_url = reverse('pickup-confirmation', args=[shipment.pk])
                    if courier_user:
                        courier_message = (
                            f'Shipment {shipment.shipment_id} was assigned to you and requires pickup confirmation.'
                        )
                        AuditLog.objects.create(
                            name='Shipment Assigned',
                            action=courier_message,
                            message=f'target_url={pickup_url}',
                            user=courier_user,
                            severity='warning',
                        )
                    elif courier_profile and courier_profile.email:
                        courier_message = (
                            f'Shipment {shipment.shipment_id} was assigned to you and requires pickup confirmation.'
                        )
                        AuditLog.objects.create(
                            name='Shipment Assigned',
                            action=courier_message,
                            message=f'target_url={pickup_url}',
                            user=None,
                            severity='warning',
                        )
                        send_courier_profile_email_alert(
                            courier_profile,
                            f'Shipment {shipment.shipment_id} assigned to you',
                            f'You were assigned to handle shipment {shipment.shipment_id}. Please confirm pickup at {request.build_absolute_uri(pickup_url)}.'
                        )
                else:
                    operator = shipment.created_by
                    if operator:
                        AuditLog.objects.create(
                            name='Shipment Rejected',
                            action=f'Your shipment request {shipment.shipment_id} was rejected. {comments}'.strip(),
                            message=f'target_url={reverse("shipment-detail", args=[shipment.pk])}',
                            user=operator,
                            severity='warning',
                        )
                        if application_settings.email_alerts_enabled and operator.email:
                            send_mail(
                                f'Shipment {shipment.shipment_id} rejected',
                                f'Your shipment request {shipment.shipment_id} was rejected by the backup administrator. {comments}'.strip(),
                                settings.DEFAULT_FROM_EMAIL,
                                [operator.email],
                                fail_silently=True,
                            )

                messages.success(request, 'Shipment assignment updated successfully.')
                return redirect('shipment-detail', shipment_pk=shipment.pk)
            else:
                messages.error(request, 'Please select a tape and courier before assigning the shipment.')
        elif approval_form.is_valid():
            decision = approval_form.cleaned_data.get('decision')
            comments = approval_form.cleaned_data.get('comments', '').strip()
            action_map = {
                'approve': 'Approved',
                'reject': 'Rejected',
                'more_info': 'Requested More Information',
            }
            action = action_map.get(decision, 'Updated')

            if decision == 'approve' and not shipment.compliance_passed():
                messages.warning(request, 'Shipment cannot be approved until all compliance checks pass.')
            else:
                if decision == 'approve':
                    if is_backup_administrator(request.user) and not is_supreme_approver(request.user):
                        shipment.status = 'Pending'
                        shipment.approval_stage = 'awaiting_supreme'
                        shipment.approved_by_backup = request.user
                        shipment.approved_by = request.user
                        shipment.approval_remarks = comments or 'Approved by backup administrator. Pending supreme approver.'
                    elif is_supreme_approver(request.user):
                        shipment.status = 'Approved'
                        shipment.approval_stage = 'approved'
                        shipment.approved_by_supreme = request.user
                        shipment.approved_by = request.user
                        shipment.approval_remarks = comments or 'Approved by supreme approver.'
                    else:
                        shipment.status = 'Approved'
                        shipment.approval_stage = 'approved'
                        shipment.approved_by = request.user
                elif decision == 'reject':
                    shipment.status = 'Rejected'
                    shipment.approval_stage = 'rejected'
                    shipment.approved_by = request.user
                else:
                    shipment.status = 'More Info Requested'
                    shipment.approval_stage = 'draft'
                    shipment.approved_by = request.user

                shipment.approval_date = timezone.localtime()
                shipment.last_updated_by = request.user
                shipment.save()

                ShipmentApprovalHistory.objects.create(
                    shipment=shipment,
                    action=action,
                    comments=comments,
                    user=request.user,
                )

                AuditLog.objects.create(
                    name='Shipment Approval Decision',
                    action=f'{action} for shipment {shipment.shipment_id}',
                    user=request.user,
                    severity='success' if decision == 'approve' else 'warning',
                )

                # If supreme approved, notify backup admin to print the approval form
                if decision == 'approve' and is_supreme_approver(request.user):
                    try:
                        backup_admin = shipment.approved_by_backup or getattr(shipment, 'created_by', None)
                        if backup_admin:
                            preview_url = reverse('approval-form-preview', args=[shipment.pk])
                            AuditLog.objects.create(
                                name='Shipment Approval Ready for Printing',
                                action=f'Shipment {shipment.shipment_id} approved and ready for printing by backup admin.',
                                message=f'target_url={preview_url}',
                                user=backup_admin,
                                severity='warning',
                            )
                            if application_settings.email_alerts_enabled and getattr(backup_admin, 'email', None):
                                send_mail(
                                    f'Shipment {shipment.shipment_id} approved - print approval form',
                                    f'Shipment {shipment.shipment_id} was approved. Print the approval form here: {request.build_absolute_uri(preview_url)}',
                                    settings.DEFAULT_FROM_EMAIL,
                                    [backup_admin.email],
                                    fail_silently=True,
                                )
                    except Exception:
                        pass

                messages.success(request, f'Shipment has been marked as {shipment.status}.')
                if not partial_request:
                    return redirect('shipment-detail', shipment_pk=shipment.pk)
        elif request.POST.get('form_type') == 'release_shipment':
            # Release the shipment to the assigned courier (final step by Backup Admin)
            if not is_backup_administrator(request.user):
                messages.error(request, 'Only Backup Administrators may release shipments to couriers.')
                return redirect('shipment-detail', shipment_pk=shipment.pk)

            # Only allow release when shipment has been supreme-approved
            if shipment.approval_stage != 'approved':
                messages.error(request, 'Shipment must be fully approved before release.')
                return redirect('shipment-detail', shipment_pk=shipment.pk)

            # Update shipment status and release metadata
            shipment.status = 'Dispatched'
            shipment.release_datetime = timezone.localtime()
            shipment.last_updated_by = request.user
            shipment.save(update_fields=['status', 'release_datetime', 'last_updated_by', 'last_updated_at'])

            # Record transport event and approval history
            ShipmentTransportEvent.objects.create(
                shipment=shipment,
                courier=None,
                event_type='Released',
                event_date=timezone.localdate(),
                event_time=timezone.localtime().time(),
                comments='Shipment released for courier pickup.'
            )
            ShipmentApprovalHistory.objects.create(
                shipment=shipment,
                action='Released',
                comments='Released by backup administrator.',
                user=request.user,
            )

            # Attempt to find a courier profile/user to notify
            courier_user = None
            courier_profile = None
            try:
                from .models import CourierProfile
                if shipment.courier_name or shipment.courier_contact:
                    courier_profile = CourierProfile.objects.filter(
                        Q(full_name__iexact=shipment.courier_name) |
                        Q(phone_number__iexact=shipment.courier_contact) |
                        Q(email__iexact=shipment.courier_contact)
                    ).first()
                    if courier_profile:
                        courier_user = getattr(courier_profile, 'user', None)
            except Exception:
                courier_profile = None
                courier_user = None

            pickup_url = reverse('pickup-confirmation', args=[shipment.pk])
            app_settings = ApplicationSetting.objects.first() or ApplicationSetting.objects.create()

            # Create audit log targeted to courier (if known) or fallback to operator
            if courier_user:
                AuditLog.objects.create(
                    name='Shipment Released',
                    action=f'Shipment {shipment.shipment_id} released for pickup by courier.',
                    message=f'target_url={pickup_url}',
                    user=courier_user,
                    severity='warning',
                )
                if app_settings.email_alerts_enabled and getattr(courier_user, 'email', None):
                    send_mail(
                        f'Shipment ready for pickup: {shipment.shipment_id}',
                        f'Shipment {shipment.shipment_id} has been released and is ready for pickup. Confirm pickup here: {request.build_absolute_uri(pickup_url)}',
                        settings.DEFAULT_FROM_EMAIL,
                        [courier_user.email],
                        fail_silently=True,
                    )
            elif courier_profile:
                AuditLog.objects.create(
                    name='Shipment Released',
                    action=f'Shipment {shipment.shipment_id} released for pickup by courier {courier_profile.full_name}.',
                    message=f'target_url={pickup_url}',
                    user=None,
                    severity='warning',
                )
                send_courier_profile_email_alert(
                    courier_profile,
                    f'Shipment ready for pickup: {shipment.shipment_id}',
                    f'Shipment {shipment.shipment_id} has been released and is ready for pickup. Confirm pickup here: {request.build_absolute_uri(pickup_url)}',
                )
            else:
                # Fallback: notify the original requester/operator that shipment was released
                if shipment.created_by:
                    AuditLog.objects.create(
                        name='Shipment Released',
                        action=f'Shipment {shipment.shipment_id} was released for courier pickup.',
                        message=f'target_url={reverse("shipment-detail", args=[shipment.pk])}',
                        user=shipment.created_by,
                        severity='info',
                    )

            messages.success(request, 'Shipment released to courier and notifications queued.')
            return redirect('shipment-detail', shipment_pk=shipment.pk)

    template = 'shipment_detail_fragment.html' if partial_request else 'shipment_detail.html'

    compliance_checks = shipment.compliance_checks()
    compliance_checks_display = [
        (key.replace('_', ' ').title(), value)
        for key, value in compliance_checks.items()
    ]
    context = {
        'dashboard_tabs': OPERATIONS_FEATURE_TABS,
        'shipment': shipment,
        'approval_form': approval_form,
        'receipt_form': receipt_form,
        'assignment_form': assignment_form,
        'manifest_search_form': manifest_search_form,
        'manifest_tapes': manifest_tapes,
        'compliance_checks': compliance_checks,
        'compliance_checks_display': compliance_checks_display,
        'compliance_passed': shipment.compliance_passed(),
        'risk_score': shipment.risk_score(),
        'risk_level': shipment.risk_level(),
        'risk_recommendation': shipment.risk_recommendation(),
        'history': shipment.approval_history.all(),
    }
    return render(request, template, context)


@user_passes_test(lambda u: u.is_superuser or is_operations_manager(u) or is_backup_administrator(u), login_url='signin')
@login_required(login_url='signin')
def approval_history(request, shipment_pk):
    shipment = get_object_or_404(Shipment, pk=shipment_pk)
    history_entries = shipment.approval_history.all()
    context = {
        'dashboard_tabs': OPERATIONS_FEATURE_TABS,
        'shipment': shipment,
        'history_entries': history_entries,
    }
    return render(request, 'approval_history.html', context)


@user_passes_test(lambda u: u.is_superuser or is_operations_manager(u) or is_backup_administrator(u), login_url='signin')
@login_required(login_url='signin')
def request_return_shipment(request, shipment_pk):
    shipment = get_object_or_404(Shipment.objects.prefetch_related('tapes'), pk=shipment_pk)
    if shipment.status in ['Cancelled', 'Rejected']:
        messages.error(request, 'This shipment cannot be returned.')
        return redirect('shipment-detail', shipment_pk=shipment.pk)

    return_form = ReturnShipmentRequestForm(request.POST or None, shipment=shipment)
    if request.method == 'POST' and return_form.is_valid():
        tape = return_form.cleaned_data.get('tape')
        courier_selection = return_form.cleaned_data['courier']
        comments = return_form.cleaned_data['comments'].strip()
        courier_profile, courier_user, courier_name, courier_contact = _resolve_courier_selection(courier_selection)

        shipment.return_requested_by = request.user
        shipment.return_requested_at = timezone.localtime()
        shipment.return_request_comments = comments
        shipment.return_assigned_courier_profile = courier_profile
        shipment.return_assigned_courier_user = courier_user
        shipment.return_courier_response_status = 'Pending'
        shipment.status = 'Return Requested'
        shipment.courier_name = courier_name
        shipment.courier_contact = courier_contact
        shipment.last_updated_by = request.user
        shipment.save(update_fields=[
            'return_requested_by',
            'return_requested_at',
            'return_request_comments',
            'return_assigned_courier_profile',
            'return_assigned_courier_user',
            'return_courier_response_status',
            'status',
            'courier_name',
            'courier_contact',
            'last_updated_by',
            'last_updated_at',
        ])

        ShipmentApprovalHistory.objects.create(
            shipment=shipment,
            action='Return Requested',
            comments=comments,
            user=request.user,
        )

        ShipmentTransportEvent.objects.create(
            shipment=shipment,
            courier=courier_profile,
            event_type='Return Requested',
            event_date=timezone.localdate(),
            event_time=timezone.localtime().time(),
            comments='Return request submitted and courier assigned.',
        )

        AuditLog.objects.create(
            name='Return Requested',
            action=(
                f'Return requested for shipment {shipment.shipment_id} and assigned to {courier_name or "selected courier"}.'
            ),
            message=f'target_url={reverse("courier-dashboard")}',
            user=courier_user,
            severity='warning',
        )

        if courier_profile and courier_profile.email:
            send_courier_profile_email_alert(
                courier_profile,
                f'Return request assigned: {shipment.shipment_id}',
                f'A return request for shipment {shipment.shipment_id} has been assigned to you. Please respond in the courier dashboard.'
            )

        backup_admins = User.objects.filter(is_active=True, groups__name='Backup Administrator').distinct()
        application_settings = ApplicationSetting.objects.first() or ApplicationSetting.objects.create()
        for admin in backup_admins:
            AuditLog.objects.create(
                name='Return Approval Required',
                action=f'Return request for shipment {shipment.shipment_id} requires your approval.',
                message=f'target_url={reverse("shipment-detail", args=[shipment.pk])}',
                user=admin,
                severity='warning',
            )
            if application_settings.email_alerts_enabled and admin.email:
                send_mail(
                    f'Return request requires approval: {shipment.shipment_id}',
                    f'A return request for shipment {shipment.shipment_id} has been submitted and requires your approval.',
                    settings.DEFAULT_FROM_EMAIL,
                    [admin.email],
                    fail_silently=True,
                )

        messages.success(request, 'Return request submitted and courier assigned. Awaiting courier response and backup admin approval.')
        if request.GET.get('partial') == '1':
            return redirect('shipment-detail', shipment_pk=shipment.pk)
        return redirect('shipment-detail', shipment_pk=shipment.pk)

    context = {
        'dashboard_tabs': OPERATIONS_FEATURE_TABS,
        'shipment': shipment,
        'return_form': return_form,
    }
    template = 'request_return_shipment_fragment.html' if request.GET.get('partial') == '1' else 'request_return_shipment.html'
    return render(request, template, context)


@user_passes_test(lambda u: u.is_superuser or is_courier(u), login_url='signin')
@login_required(login_url='signin')
def courier_return_response(request, shipment_pk):
    shipment = get_object_or_404(Shipment.objects.prefetch_related('tapes'), pk=shipment_pk)
    courier = ensure_courier_profile(request.user)
    is_assigned_courier = (
        shipment.return_assigned_courier_user == request.user or
        (courier and shipment.return_assigned_courier_profile == courier)
    )
    if not is_assigned_courier:
        messages.error(request, 'You are not assigned to respond to this return request.')
        return redirect('courier-dashboard')

    if shipment.status != 'Return Requested':
        messages.warning(request, 'This return request is not currently pending courier response.')
        return redirect('courier-dashboard')

    form = CourierReturnAcceptanceForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        decision = form.cleaned_data['decision']
        comments = form.cleaned_data['comments'].strip()
        shipment.return_courier_response_status = 'Accepted' if decision == 'accept' else 'Rejected'
        shipment.return_courier_response_at = timezone.localtime()
        shipment.status = 'Return Accepted' if decision == 'accept' else 'Return Rejected'
        shipment.last_updated_by = request.user
        shipment.save(update_fields=[
            'return_courier_response_status',
            'return_courier_response_at',
            'status',
            'last_updated_by',
            'last_updated_at',
        ])

        ShipmentTransportEvent.objects.create(
            shipment=shipment,
            courier=courier,
            event_type='Return Accepted' if decision == 'accept' else 'Return Rejected',
            event_date=timezone.localdate(),
            event_time=timezone.localtime().time(),
            comments=comments or 'Courier responded to return request.',
        )

        ShipmentApprovalHistory.objects.create(
            shipment=shipment,
            action='Return Accepted' if decision == 'accept' else 'Return Rejected',
            comments=comments,
            user=request.user,
        )

        if shipment.return_requested_by:
            AuditLog.objects.create(
                name='Return Response Recorded',
                action=(
                    f'Courier response for return shipment {shipment.shipment_id}: '
                    f'{"accepted" if decision == "accept" else "rejected"}.'
                ),
                message=f'target_url={reverse("shipment-detail", args=[shipment.pk])}',
                user=shipment.return_requested_by,
                severity='warning',
            )
            application_settings = ApplicationSetting.objects.first() or ApplicationSetting.objects.create()
            if application_settings.email_alerts_enabled and shipment.return_requested_by.email:
                send_mail(
                    f'Return request {"accepted" if decision == "accept" else "rejected"}: {shipment.shipment_id}',
                    f'The courier has {"accepted" if decision == "accept" else "rejected"} the return request for shipment {shipment.shipment_id}.',
                    settings.DEFAULT_FROM_EMAIL,
                    [shipment.return_requested_by.email],
                    fail_silently=True,
                )

        messages.success(request, 'Return response submitted successfully.')
        return redirect('courier-dashboard')

    context = {
        'courier': courier,
        'shipment': shipment,
        'form': form,
    }
    return render(request, 'courier_return_response.html', context)


@user_passes_test(lambda u: u.is_superuser or is_backup_administrator(u), login_url='signin')
@login_required(login_url='signin')
def receive_return_shipment(request, shipment_pk):
    shipment = get_object_or_404(Shipment.objects.prefetch_related('tapes'), pk=shipment_pk)
    if shipment.status not in ['Return Accepted', 'Return In Transit']:
        messages.error(request, 'This return shipment is not ready for receipt.')
        return redirect('shipment-detail', shipment_pk=shipment.pk)

    form = ReturnReceiveForm(request.POST or None, initial={'receiving_custodian': request.user.get_full_name() or request.user.username})
    if request.method == 'POST' and form.is_valid():
        receiving_custodian = form.cleaned_data['receiving_custodian'].strip()
        decision = form.cleaned_data['decision']
        receipt_notes = form.cleaned_data['receipt_notes'].strip()

        if decision == 'accept':
            shipment.status = 'Completed'
            shipment.delivery_status = 'Delivered'
            shipment.received_by = receiving_custodian
            shipment.delivery_date = timezone.localdate()
            shipment.delivery_time = timezone.localtime().time()
            shipment.delivery_notes = receipt_notes
            event_type = 'Return Delivered'
            history_action = 'Return Received'
            audit_name = 'Return Received'
            courier_notification = 'accepted'
        else:
            shipment.status = 'Return Rejected'
            event_type = 'Return Rejected'
            history_action = 'Return Rejected'
            audit_name = 'Return Rejected'
            courier_notification = 'rejected'

        shipment.return_courier_response_status = 'Accepted' if decision == 'accept' else 'Rejected'
        shipment.return_courier_response_at = timezone.localtime()
        shipment.last_updated_by = request.user
        shipment.save(update_fields=[
            'status',
            'delivery_status',
            'received_by',
            'delivery_date',
            'delivery_time',
            'delivery_notes',
            'return_courier_response_status',
            'return_courier_response_at',
            'last_updated_by',
            'last_updated_at',
        ])

        ShipmentTransportEvent.objects.create(
            shipment=shipment,
            courier=shipment.return_assigned_courier_profile,
            event_type=event_type,
            event_date=timezone.localdate(),
            event_time=timezone.localtime().time(),
            comments=receipt_notes or f'Return shipment {decision} by backup administrator.',
        )

        ShipmentApprovalHistory.objects.create(
            shipment=shipment,
            action=history_action,
            comments=receipt_notes,
            user=request.user,
        )

        if shipment.return_assigned_courier_user:
            AuditLog.objects.create(
                name=audit_name,
                action=(
                    f'The backup administrator has {courier_notification} the return receipt for shipment {shipment.shipment_id}.'
                ),
                message=f'target_url={reverse("courier-dashboard")}',
                user=shipment.return_assigned_courier_user,
                severity='success' if decision == 'accept' else 'warning',
            )
        elif shipment.return_assigned_courier_profile and shipment.return_assigned_courier_profile.email:
            send_courier_profile_email_alert(
                shipment.return_assigned_courier_profile,
                f'Return receipt {courier_notification}: {shipment.shipment_id}',
                f'The backup administrator has {courier_notification} the return receipt for shipment {shipment.shipment_id}.',
            )

        AuditLog.objects.create(
            name=audit_name,
            action=(
                f'Return shipment {shipment.shipment_id} was {courier_notification} by backup administrator.'
            ),
            message=f'target_url={reverse("shipment-detail", args=[shipment.pk])}',
            user=request.user,
            severity='success' if decision == 'accept' else 'warning',
        )

        messages.success(request, 'Return receipt processed successfully.')
        return redirect('shipment-detail', shipment_pk=shipment.pk)

    context = {
        'dashboard_tabs': OPERATIONS_FEATURE_TABS,
        'shipment': shipment,
        'form': form,
    }
    return render(request, 'receive_return_shipment.html', context)


@user_passes_test(lambda u: u.is_superuser or is_courier(u), login_url='signin')
@login_required(login_url='signin')
def courier_dashboard(request):
    courier = get_courier_profile(request.user)
    shipments = get_courier_shipments(request.user)
    # Hide shipments that are not yet released for courier visibility
    visible_shipments = shipments.exclude(status__in=['Pending', 'Awaiting Supreme Approval', 'Approved', 'Draft', 'Cancelled', 'Rejected'])
    assigned_count = visible_shipments.filter(status__in=['Dispatched', 'Picked Up', 'In Transit']).count()
    pending_count = visible_shipments.filter(status__in=['Pending', 'Approved']).count()
    delivered_count = visible_shipments.filter(status='Delivered').count()
    exception_count = ShipmentException.objects.filter(shipment__in=shipments, status__in=['Open', 'Investigating']).count()
    activity_count = ShipmentTransportEvent.objects.filter(courier=courier).count() if courier else 0
    recent_events = ShipmentTransportEvent.objects.filter(courier=courier).order_by('-event_date', '-event_time')[:6] if courier else []
    recent_exceptions = ShipmentException.objects.filter(shipment__in=shipments).order_by('-reported_date')[:6]

    notification_items = []
    courier_alerts = AuditLog.objects.filter(user=request.user, severity__in=['warning', 'error']).order_by('-timestamp')[:10]
    for item in courier_alerts:
        metadata = _parse_auditlog_metadata(item.message)
        target_url = metadata.get('target_url') or reverse('courier-dashboard')
        notification_items.append({
            'severity': item.severity,
            'timestamp': item.timestamp,
            'message': item.action or item.name or item.message,
            'action': 'View',
            'target_url': target_url,
            'is_read': item.is_read,
        })

    unread_alert_count = AuditLog.objects.filter(user=request.user, severity__in=['warning', 'error'], is_read=False).count()
    show_notifications_panel = False
    if request.GET.get('show_notifications') == '1':
        show_notifications_panel = True
        AuditLog.objects.filter(user=request.user, severity__in=['warning', 'error'], is_read=False).update(is_read=True, read_at=timezone.now())
        unread_alert_count = 0

    context = {
        'courier': courier,
        'shipments': visible_shipments,
        'assigned_count': assigned_count,
        'pending_count': pending_count,
        'delivered_count': delivered_count,
        'exception_count': exception_count,
        'activity_count': activity_count,
        'recent_events': recent_events,
        'recent_exceptions': recent_exceptions,
        'notification_items': notification_items,
        'unread_alert_count': unread_alert_count,
        'show_notifications_panel': show_notifications_panel,
    }
    return render(request, 'courier_dashboard.html', context)


@user_passes_test(lambda u: u.is_superuser or is_courier(u), login_url='signin')
@login_required(login_url='signin')
def assigned_shipments(request):
    courier = get_courier_profile(request.user)
    filter_form = CourierShipmentFilterForm(request.GET or None)
    shipments = get_courier_shipments(request.user)

    # Only show shipments that have been released to couriers
    shipments = shipments.exclude(status__in=['Pending', 'Awaiting Supreme Approval', 'Approved', 'Draft', 'Cancelled', 'Rejected'])

    if filter_form.is_valid():
        search = filter_form.cleaned_data.get('search')
        if search:
            shipments = shipments.filter(
                Q(shipment_id__icontains=search) |
                Q(source_location__icontains=search) |
                Q(destination_location__icontains=search) |
                Q(receiving_organization__icontains=search) |
                Q(courier_name__icontains=search)
            )
        status = filter_form.cleaned_data.get('status')
        if status:
            shipments = shipments.filter(status=status)
        shipment_type = filter_form.cleaned_data.get('shipment_type')
        if shipment_type:
            shipments = shipments.filter(shipment_type=shipment_type)
        date_from = filter_form.cleaned_data.get('date_from')
        if date_from:
            shipments = shipments.filter(shipment_date__gte=date_from)
        date_to = filter_form.cleaned_data.get('date_to')
        if date_to:
            shipments = shipments.filter(shipment_date__lte=date_to)

    page_number = request.GET.get('page', 1)
    paginator = Paginator(shipments, 10)
    shipment_page = paginator.get_page(page_number)

    context = {
        'courier': courier,
        'filter_form': filter_form,
        'shipments': shipment_page,
    }
    return render(request, 'assigned_shipments.html', context)


@user_passes_test(lambda u: u.is_superuser or is_backup_administrator(u), login_url='signin')
@login_required(login_url='signin')
def awaiting_release(request):
    """List shipments that are approved and awaiting release to courier.
    Visible to Backup Administrators and superusers.
    """
    shipments_qs = Shipment.objects.filter(approval_stage='approved').order_by('-approval_date')
    paginator = Paginator(shipments_qs, 20)
    page_number = request.GET.get('page', 1)
    shipment_page = paginator.get_page(page_number)

    alert_count = AuditLog.objects.filter(severity__in=['warning', 'error'], is_read=False).count()
    dashboard_nav_urls = {
        'add_tape': build_signed_dashboard_navigation_token('add_tape', {'show_add_tape': '1'}),
        'tape_inventory': build_signed_dashboard_navigation_token('tape_inventory', {'show_tape_inventory': '1'}),
        'shipments': build_signed_dashboard_navigation_token('shipments', {'show_shipments': '1'}),
        'settings': build_signed_dashboard_navigation_token('settings', {'show_settings': '1'}),
        'reconciliation': build_signed_dashboard_navigation_token('reconciliation', {'show_reconciliation': '1'}),
        'reports': build_signed_dashboard_navigation_token('reports', {'show_reports': 'reports'}),
        'reconciliation_reports': build_signed_dashboard_navigation_token('reconciliation_reports', {'show_reconciliation_reports': 'reconciliation-reports'}),
        'audit_logs': build_signed_dashboard_navigation_token('audit_logs', {'show_audit': '1'}),
        'alerts': build_signed_dashboard_navigation_token('alerts', {'show_alerts': '1'}),
        'admin': build_signed_dashboard_navigation_token('admin', {'show_admin': '1'}),
        'profile': build_signed_dashboard_navigation_token('profile', {'show_profile': '1'}),
    }

    context = {
        'dashboard_tabs': OPERATIONS_FEATURE_TABS,
        'shipments': shipment_page,
        'active_dashboard_section': 'awaiting_release',
        'dashboard_nav_urls': dashboard_nav_urls,
        'alert_count': alert_count,
    }
    # Whether current user may release shipments
    context['can_release'] = request.user.is_superuser or is_backup_administrator(request.user)
    return render(request, 'awaiting_release.html', context)


@user_passes_test(lambda u: u.is_superuser or is_courier(u), login_url='signin')
@login_required(login_url='signin')
def manifest_detail(request, shipment_pk):
    shipment = get_object_or_404(
        Shipment.objects.select_related('approved_by__assigned_branch', 'requesting_branch').prefetch_related('tapes'),
        pk=shipment_pk,
    )
    courier = get_courier_profile(request.user)
    manifest_tapes = shipment.tapes.all()
    context = {
        'courier': courier,
        'shipment': shipment,
        'manifest_tapes': manifest_tapes,
    }
    return render(request, 'manifest_detail.html', context)


@user_passes_test(lambda u: u.is_superuser or is_courier(u), login_url='signin')
@login_required(login_url='signin')
def pickup_confirmation(request, shipment_pk):
    shipment = get_object_or_404(Shipment.objects.prefetch_related('tapes'), pk=shipment_pk)
    if shipment.status not in ['Dispatched', 'Picked Up', 'In Transit']:
        messages.error(request, 'This shipment has not been released for courier pickup yet.')
        return redirect('courier-dashboard')

    courier = ensure_courier_profile(request.user)
    form = ShipmentReceiptForm(
        request.POST or None,
        initial={
            'manifest_reference': f'MANIFEST-{shipment.shipment_id[:8].upper()}',
            'pickup_location': shipment.source_location or (shipment.requesting_branch.branch_name if shipment.requesting_branch else ''),
            'pickup_date': timezone.localdate().strftime('%Y-%m-%d'),
            'pickup_time': timezone.localtime().strftime('%H:%M'),
        },
    )

    if request.method == 'POST':
        if form.is_valid():
            receipt = form.save(commit=False)
            receipt.shipment = shipment
            receipt.courier = courier
            receipt.confirmation_timestamp = timezone.localtime()
            receipt.save()
            shipment.status = 'Picked Up'
            shipment.vehicle_number = courier.vehicle_number or shipment.vehicle_number or ''
            shipment.last_updated_by = request.user
            shipment.save(update_fields=['status', 'vehicle_number', 'last_updated_by', 'last_updated_at'])
            ShipmentTransportEvent.objects.create(
                shipment=shipment,
                courier=courier,
                event_type='Picked Up',
                event_date=timezone.localdate(),
                event_time=timezone.localtime().time(),
                comments='Pickup confirmed by courier.',
            )
            manifest_reference = receipt.manifest_reference.strip() or f'MANIFEST-{shipment.shipment_id[:8].upper()}'
            vehicle_plate = courier.vehicle_number or shipment.vehicle_number or 'Not provided'
            backup_admins = User.objects.filter(is_active=True, groups__name='Backup Administrator').distinct()
            for backup_admin in backup_admins:
                receipt_url = reverse('shipment-detail', args=[shipment.pk])
                AuditLog.objects.create(
                    name='Pickup Confirmed',
                    action=(
                        f'Pickup confirmed for shipment {shipment.shipment_id}. '
                        f'Manifest reference: {manifest_reference}. Vehicle plate: {vehicle_plate}.'
                    ),
                    message=f'target_url={receipt_url}',
                    user=backup_admin,
                    severity='warning',
                )
                if ApplicationSetting.objects.first() and ApplicationSetting.objects.first().email_alerts_enabled and backup_admin.email:
                    send_mail(
                        f'Pickup confirmed for shipment {shipment.shipment_id}',
                        (
                            f'Pickup confirmation was submitted for shipment {shipment.shipment_id}. '
                            f'Manifest reference: {manifest_reference}. Vehicle plate: {vehicle_plate}.'
                        ),
                        settings.DEFAULT_FROM_EMAIL,
                        [backup_admin.email],
                        fail_silently=True,
                    )
            AuditLog.objects.create(
                name='Pickup Confirmed',
                action=f'Pickup confirmed for shipment {shipment.shipment_id}',
                user=request.user,
                severity='success',
            )
            messages.success(request, 'Pickup confirmation saved successfully.')
            return redirect('courier-dashboard')
        else:
            messages.error(request, 'Please correct the form errors and try again.')

    if not courier:
        messages.warning(request, 'Courier profile not found for your user account.')
    context = {
        'courier': courier,
        'shipment': shipment,
        'form': form,
    }
    return render(request, 'pickup_confirmation.html', context)


@user_passes_test(lambda u: u.is_superuser or is_courier(u), login_url='signin')
@login_required(login_url='signin')
def delivery_confirmation(request, shipment_pk):
    shipment = get_object_or_404(Shipment.objects.prefetch_related('tapes'), pk=shipment_pk)
    if shipment.status not in ['Picked Up', 'In Transit']:
        messages.error(request, 'This shipment is not currently in courier custody for delivery confirmation.')
        return redirect('courier-dashboard')

    courier = ensure_courier_profile(request.user)
    form = DeliveryConfirmationForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            confirmation = form.save(commit=False)
            confirmation.shipment = shipment
            confirmation.courier = courier
            confirmation.save()
            shipment.delivery_date = confirmation.delivery_date
            shipment.delivery_time = confirmation.delivery_time
            shipment.delivery_status = confirmation.delivery_status
            shipment.status = 'Delivered' if confirmation.delivery_status == 'Delivered' else shipment.status
            shipment.last_updated_by = request.user
            shipment.save(update_fields=['delivery_date', 'delivery_time', 'delivery_status', 'status', 'last_updated_by', 'last_updated_at'])
            ShipmentTransportEvent.objects.create(
                shipment=shipment,
                courier=courier,
                event_type='Delivered',
                event_date=timezone.localdate(),
                event_time=timezone.localtime().time(),
                comments='Delivery confirmed by courier.',
            )
            AuditLog.objects.create(
                name='Delivery Confirmed',
                action=f'Delivery confirmed for shipment {shipment.shipment_id}',
                user=request.user,
                severity='success',
            )
            messages.success(request, 'Delivery confirmation saved successfully.')
            return redirect('courier-dashboard')
        else:
            messages.error(request, 'Please correct the form errors and try again.')

    context = {
        'courier': courier,
        'shipment': shipment,
        'form': form,
    }
    return render(request, 'delivery_confirmation.html', context)


@user_passes_test(lambda u: u.is_superuser or is_courier(u), login_url='signin')
@login_required(login_url='signin')
def return_shipments(request):
    courier = get_courier_profile(request.user)
    shipments = Shipment.objects.filter(
        Q(shipment_type='Return') |
        Q(status__in=['Return Requested', 'Return Accepted', 'Return In Transit', 'Return Rejected']),
        Q(courier_name__iexact=courier.full_name) |
        Q(courier_contact__iexact=courier.phone_number) |
        Q(receipts__courier=courier) |
        Q(deliveries__courier=courier) |
        Q(return_assigned_courier_profile=courier) |
        Q(return_assigned_courier_user=request.user)
    ).distinct().order_by('-shipment_date') if courier else Shipment.objects.none()

    context = {
        'courier': courier,
        'shipments': shipments,
    }
    return render(request, 'return_shipments.html', context)


@user_passes_test(lambda u: u.is_superuser or is_courier(u), login_url='signin')
@login_required(login_url='signin')
def incident_management(request):
    courier = get_courier_profile(request.user)
    form = ShipmentExceptionForm(request.POST or None, courier=courier)
    exceptions = ShipmentException.objects.filter(
        Q(shipment__courier_name__iexact=courier.full_name) | Q(shipment__courier_contact__iexact=courier.phone_number) | Q(shipment__receipts__courier=courier) | Q(shipment__deliveries__courier=courier)
    ).distinct().order_by('-reported_date') if courier else ShipmentException.objects.none()

    if request.method == 'POST' and form.is_valid():
        exception = form.save(commit=False)
        exception.reported_by = request.user
        exception.save()
        AuditLog.objects.create(
            name='Shipment Exception Reported',
            action=f'Reported exception {exception.exception_id} for shipment {exception.shipment.shipment_id}',
            user=request.user,
            severity='warning',
        )
        messages.success(request, 'Exception reported successfully.')
        return redirect('incident-management')
    elif request.method == 'POST':
        messages.error(request, 'Please correct the exception form errors and try again.')

    context = {
        'courier': courier,
        'form': form,
        'exceptions': exceptions,
    }
    return render(request, 'incident_management.html', context)


@user_passes_test(lambda u: u.is_superuser or is_courier(u), login_url='signin')
@login_required(login_url='signin')
def activity_log(request):
    courier = ensure_courier_profile(request.user)
    events = ShipmentTransportEvent.objects.filter(courier=courier).order_by('-event_date', '-event_time', '-created_at') if courier else ShipmentTransportEvent.objects.none()
    exceptions = ShipmentException.objects.filter(
        Q(shipment__courier_name__iexact=courier.full_name) |
        Q(shipment__courier_contact__iexact=courier.phone_number) |
        Q(shipment__receipts__courier=courier) |
        Q(shipment__deliveries__courier=courier) |
        Q(reported_by=request.user)
    ).distinct().order_by('-reported_date') if courier else ShipmentException.objects.none()

    activity_items = []
    for event in events:
        event_timestamp = None
        if event.event_date and event.event_time:
            event_timestamp = datetime.combine(event.event_date, event.event_time)
            if timezone.is_naive(event_timestamp):
                event_timestamp = timezone.make_aware(event_timestamp, timezone.get_current_timezone())
        else:
            event_timestamp = event.created_at

        activity_items.append({
            'timestamp': event_timestamp,
            'title': event.event_type,
            'shipment': event.shipment,
            'courier': event.courier,
            'details': event.comments or '-',
            'item_type': 'event',
        })

    for exception in exceptions:
        activity_items.append({
            'timestamp': exception.reported_date,
            'title': exception.exception_type,
            'shipment': exception.shipment,
            'courier': courier,
            'details': exception.description or 'No additional details provided.',
            'item_type': 'exception',
        })

    activity_items.sort(key=lambda item: item['timestamp'], reverse=True)

    context = {
        'courier': courier,
        'events': events,
        'exceptions': exceptions,
        'activity_items': activity_items,
    }
    return render(request, 'activity_log.html', context)


@user_passes_test(lambda u: u.is_superuser or is_backup_administrator(u), login_url='signin')
@login_required(login_url='signin')
def add_tape(request):
    form = AddTapeForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            payload_fields = {
                k: (str(v) if not isinstance(v, (dict, list)) else v)
                for k, v in form.cleaned_data.items()
            }
            pa = _build_pending_tape_approval(
                request,
                transaction_type='Add Tape',
                summary=f"Register tape {payload_fields.get('volser') or payload_fields.get('barcode')}",
                payload_fields=payload_fields,
                priority='Medium',
                risk_level='High',
            )
            if pa:
                AuditLog.objects.create(
                    name='Pending Tape Registration',
                    action=f'Pending registration for tape {payload_fields.get("volser") or payload_fields.get("barcode")}',
                    user=request.user,
                    severity='info',
                )
                messages.success(request, 'Tape registration submitted for approval.')
                return redirect('dashboard')
            messages.error(request, 'Failed to submit tape registration for approval.')
        else:
            messages.error(request, 'Please correct the errors below and try again.')

    return render(request, 'add_tape.html', {'form': form})


@user_passes_test(has_report_access, login_url='signin')
@login_required(login_url='signin')
def reconciliation_reports(request):
    report_email_form = ReportEmailForm(request.POST or None)
    search_query = request.GET.get('search', '').strip()
    reconciliations = Reconciliation.objects.select_related(
        'performed_by', 'reviewed_by', 'approved_by'
    ).prefetch_related('results', 'results__tape').order_by('-reconciliation_date', '-created_at')

    if search_query:
        search_q = (
            Q(reconciliation_id__icontains=search_query) |
            Q(location__icontains=search_query) |
            Q(status__icontains=search_query) |
            Q(performed_by__username__icontains=search_query)
        )
        reconciliations = reconciliations.filter(search_q)

    reconciliation_count = reconciliations.count()
    total_issues = ReconciliationResult.objects.filter(reconciliation__in=reconciliations).count()
    open_issues = ReconciliationResult.objects.filter(
        reconciliation__in=reconciliations,
        resolution_status__in=['Open', 'Under Investigation']
    ).count()
    completed_count = reconciliations.filter(status='Completed').count()
    open_reconciliations = reconciliations.exclude(status='Completed').count()

    reconciliation_report_summary = {
        'total_reconciliations': reconciliation_count,
        'total_discrepancies': total_issues,
        'open_issues': open_issues,
        'completed_reconciliations': completed_count,
        'open_reconciliations': open_reconciliations,
    }

    for reconciliation in reconciliations:
        reconciliation.total_issues = reconciliation.results.count()
        reconciliation.open_issues = reconciliation.results.filter(
            resolution_status__in=['Open', 'Under Investigation']
        ).count()

    if request.GET.get('export_reconciliation_csv') == '1':
        return export_reconciliation_report_csv(reconciliations, summary=reconciliation_report_summary)

    if request.method == 'POST' and report_email_form.is_valid():
        recipients = report_email_form.cleaned_data['recipients']
        subject = report_email_form.cleaned_data['subject']
        message = report_email_form.cleaned_data['message']
        summary_lines = [f"{key.replace('_', ' ').title()}: {value}" for key, value in reconciliation_report_summary.items()]
        report_body = message + '\n\n' + '\n'.join(summary_lines)
        send_report_email(subject, report_body, recipients)
        AuditLog.objects.create(
            name='Reconciliation Report Sent',
            action=f'Sent reconciliation report summary to {", ".join(recipients)}',
            user=request.user,
            severity='info',
        )
        messages.success(request, 'Reconciliation report summary sent successfully.')
        return redirect('reconciliation-reports')

    context = {
        'report_email_form': report_email_form,
        'reconciliations': reconciliations,
        'search_query': search_query,
        'summary': reconciliation_report_summary,
        'selected_tab': 'reconciliation_reports',
        'report_export_url': reverse('reconciliation-reports'),
    }
    return render(request, 'reconciliation_reports.html', context)


@user_passes_test(has_report_access, login_url='signin')
@login_required(login_url='signin')
def reconciliation_report_detail(request, pk):
    reconciliation = get_object_or_404(
        Reconciliation.objects.select_related('performed_by', 'reviewed_by', 'approved_by').prefetch_related('results', 'results__tape'),
        pk=pk
    )
    total_issues = reconciliation.results.count()
    open_issues = reconciliation.results.filter(resolution_status__in=['Open', 'Under Investigation']).count()
    reconciliation_report_summary = {
        'total_reconciliations': 1,
        'total_discrepancies': total_issues,
        'open_issues': open_issues,
        'completed_reconciliations': 1 if reconciliation.status == 'Completed' else 0,
        'open_reconciliations': 0 if reconciliation.status == 'Completed' else 1,
    }
    context = {
        'selected_reconciliation_report': reconciliation,
        'total_issues': total_issues,
        'open_issues': open_issues,
        'show_reconciliation_report_detail_panel': True,
        'show_reconciliation_reports_panel': False,
        'reconciliation_report_summary': reconciliation_report_summary,
        'report_export_url': reverse('backup-dashboard') + '?show_reconciliation_reports=reconciliation-reports',
    }
    return render(request, 'backup_dashboard.html', context)


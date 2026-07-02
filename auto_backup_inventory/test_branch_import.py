import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'auto_backup_inventory.settings')
django.setup()

from openpyxl import load_workbook
from inventory.admin_bank_branch import BankBranchAdmin

# Load the generated template
template_path = os.path.join(os.path.dirname(__file__), 'branch_import_template.xlsx')
workbook = load_workbook(template_path, read_only=True, data_only=True)
sheet = workbook.active
rows = list(sheet.iter_rows(values_only=True))

print("Raw rows from Excel:")
for i, row in enumerate(rows):
    print(f"Row {i}: {row}")

# Try to build preview
from inventory.models import BankBranch as BankBranchModel
from django.contrib.admin.sites import AdminSite

admin = BankBranchAdmin(BankBranchModel, AdminSite())
from django.contrib.auth import get_user_model
User = get_user_model()
# Get superuser or create one for testing
superuser = User.objects.filter(is_superuser=True).first()

if superuser:
    preview = admin._build_import_preview(rows, superuser)
    print("\n\nPreview result:")
    print(f"Total rows: {preview['total_rows']}")
    print(f"New branches: {preview['new_branches']}")
    print(f"Updated branches: {preview['updated_branches']}")
    print(f"Duplicate rows: {preview['duplicate_rows']}")
    print(f"Invalid rows: {preview['invalid_rows']}")
    print(f"Errors: {preview['errors']}")
    print(f"Rows to import: {preview['rows_to_import']}")
    print(f"\nRows data:")
    for row_data in preview['rows']:
        print(row_data)
else:
    print("No superuser found. Create one with: python manage.py createsuperuser")
